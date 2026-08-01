# CUI // SP-CTI
"""The BOM engine's REST surface.

Tests the handler directly — the decorator owns auth and error mapping and is
covered elsewhere; what matters here is the CONTRACT and the security posture.

Invented content. ICDEV is a public repo.
"""
from __future__ import annotations

import base64
import io

import openpyxl
import pytest

from tools.cortex.rest_v1 import api_v1_bom
from tools.cortex.service_keys import ALL_SCOPES, DEFAULT_SCOPES

# The undecorated logic.
_bom = api_v1_bom.__wrapped__


def _doc(name: str, rows: list[list], **extra) -> dict:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return {
        "filename": name,
        "content_base64": base64.b64encode(buf.getvalue()).decode(),
        **extra,
    }


_ROWS = [
    ["Item", "Part Number", "QTY", "Unit Price", "Extended"],
    ["Core switch", "CS-9500-16X", 2, 21000, 42000],
    ["Perimeter firewall", "FW-2110", 1, 10500, 10500],
]


class TestTheScope:
    def test_bom_is_a_real_scope(self):
        assert "cortex:bom" in ALL_SCOPES

    def test_it_is_never_in_the_default_grant(self):
        """The payload is the CONTENTS of somebody's bills of materials and quotes
        — the most commercially sensitive material they have. A key that can search
        the platform must not silently also be able to post a competitor's pricing
        into it.
        """
        assert "cortex:bom" not in DEFAULT_SCOPES


class TestTheContract:
    def test_a_single_bom_reconciles(self):
        out = _bom({"documents": [
            _doc("bom.xlsx", _ROWS, role="bom_claim", credibility_tier="authoritative"),
        ]})
        assert out["line_count"] == 2
        assert out["committed_total"] == 52500
        assert out["is_a_total"] is True

    def test_no_model_runs(self):
        """The REST surface is --no-llm. It cannot hallucinate, because there is
        nothing in it that could."""
        out = _bom({"documents": [_doc("bom.xlsx", _ROWS)]})
        assert out["llm_calls"] == 0

    def test_findings_cite_a_cell(self):
        rows = [
            ["Item", "QTY", "Unit Price", "Extended"],
            ["Widget A", 2, 100, None],
            ["Widget B", 1, None, None],       # priced at nothing, silently
        ]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"
        for r in rows:
            ws.append(r)
        ws["D2"] = "=B2*C2"
        ws["D3"] = "=B3*C3"
        buf = io.BytesIO()
        wb.save(buf)

        out = _bom({"documents": [{
            "filename": "b.xlsx",
            "content_base64": base64.b64encode(buf.getvalue()).decode(),
        }]})

        zeroed = [f for f in out["findings"] if f["type"] == "unpriced_line_zeroed"]
        assert zeroed
        ev = zeroed[0]["evidence"][0]
        assert ev["source_document"] == "b.xlsx"
        assert ev["locator"] == "D3"
        assert zeroed[0]["detector"] == "deterministic"


class TestSilenceIsNeverConfirmation:
    def test_what_the_caller_declares_is_binding(self):
        out = _bom({"documents": [
            _doc("bom.xlsx", _ROWS, role="bom_claim", credibility_tier="authoritative"),
        ]})
        src = out["sources"][0]
        assert src["credibility_tier"] == "authoritative"
        assert src["set_by"] == "human"

    def test_what_the_caller_omits_is_PROPOSED_and_says_so(self):
        out = _bom({"documents": [_doc("bom.xlsx", _ROWS)]})
        src = out["sources"][0]
        assert src["set_by"] == "ai_proposed"
        # And it explains itself, so somebody can disagree with a reason rather
        # than with a number.
        assert src["rationale"]


class TestASumOfEstimatesIsNotAnEstimate:
    def test_two_documents_claiming_the_same_project(self):
        out = _bom({"documents": [
            _doc("bom_one.xlsx", _ROWS, role="bom_claim"),
            _doc("bom_two.xlsx", [
                ["Item", "Part Number", "QTY", "Unit Price", "Extended"],
                ["Storage array", "ST-100", 1, 60000, 60000],
            ], role="bom_claim"),
        ]})
        # The caller is told, rather than shown a tidy number.
        assert out["is_a_total"] is False
        assert len(out["competing_claims"]) == 2


class TestTheSecurityPosture:
    def test_there_is_no_path_parameter_at_all(self):
        """A remote endpoint that accepted a filesystem path would be an
        arbitrary-file-read primitive wearing a convenience's clothes: the caller
        names /etc/passwd and we obligingly parse it into a bill of materials and
        hand it back. Bytes only. It is worth the base64.
        """
        with pytest.raises(Exception):
            _bom({"documents": [{"filename": "x.xlsx", "path": "/etc/passwd"}]})

    def test_a_filename_is_a_label_not_a_path(self):
        out = _bom({"documents": [{
            "filename": "../../../../etc/passwd",
            "content_base64": base64.b64encode(b"root:x:0:0").decode(),
        }]})
        # Basenamed into a temp dir, parsed as the nothing it is.
        assert out["line_count"] == 0

    def test_an_empty_document_is_refused(self):
        with pytest.raises(Exception, match="empty"):
            _bom({"documents": [{"filename": "x.xlsx", "content_base64": ""}]})

    def test_bad_base64_is_refused_by_name(self):
        with pytest.raises(Exception, match="base64"):
            _bom({"documents": [{"filename": "x.xlsx", "content_base64": "!!!not base64!!!"}]})

    def test_a_document_with_no_filename_is_refused(self):
        with pytest.raises(Exception, match="filename"):
            _bom({"documents": [{"content_base64": "QUJD"}]})

    def test_an_empty_corpus_is_refused(self):
        with pytest.raises(Exception, match="non-empty"):
            _bom({"documents": []})

    def test_nothing_the_caller_uploads_is_persisted(self, tmp_path):
        """Everything is parsed in a temporary directory and deleted.

        Persistence is the calling product's business, and that product knows whose
        data it is. This one does not.
        """
        import glob
        import tempfile

        before = set(glob.glob(f"{tempfile.gettempdir()}/icdev_bom_*"))
        _bom({"documents": [_doc("bom.xlsx", _ROWS)]})
        after = set(glob.glob(f"{tempfile.gettempdir()}/icdev_bom_*"))
        assert after == before
