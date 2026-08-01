# CUI // SP-CTI
"""The deterministic detectors. No model runs in any of this.

Content is invented — ICDEV is a public repo. What is real is the SHAPE of each
defect, which is what the engine actually keys on.
"""
from __future__ import annotations

import pytest

from tests.bom import fixtures
from tools.bom.extract_grid import extract_grid
from tools.bom.findings import Finding, analyze_document
from tools.bom.formula_graph import build_graph, missing_operands, parse_refs


def _by_type(findings, kind):
    return [f for f in findings if f.finding_type == kind]


class TestReferenceParsing:
    def test_a_plain_range(self):
        refs = parse_refs("=SUM(F4:F26)")
        assert len(refs) == 1
        assert (refs[0].sheet, refs[0].start, refs[0].end) == ("", "F4", "F26")
        assert len(refs[0].expand("S")) == 23

    def test_a_cross_sheet_reference(self):
        refs = parse_refs("=Networking!B4+Simulation!B4")
        assert {(r.sheet, r.start) for r in refs} == {
            ("Networking", "B4"), ("Simulation", "B4")
        }

    def test_a_quoted_sheet_name(self):
        refs = parse_refs("='Digital Twin'!C10")
        assert (refs[0].sheet, refs[0].start) == ("Digital Twin", "C10")

    def test_absolute_addressing(self):
        assert parse_refs("=$D$4*$E$4")[0].start == "D4"

    def test_a_sumif_criterion_is_not_a_cell_reference(self):
        """SUMIF(F3:F26,"<>—") — the quoted argument is a criterion, not an address.

        Reading it as a reference wires the graph to a cell that does not exist,
        and every detector downstream then reasons about a phantom.
        """
        refs = parse_refs('=SUMIF(F3:F26,"<>A1")')
        assert [(r.start, r.end) for r in refs] == [("F3", "F26")]


class TestMissingOperands:
    def test_a_formula_multiplying_by_a_cell_that_does_not_exist(self, tmp_path):
        p = fixtures.workbook_with_a_zeroed_line(tmp_path / "b.xlsx")
        g = build_graph(extract_grid(p).cells)
        assert missing_operands(g, ("BOM", "D3")) == [("BOM", "C3")]
        assert missing_operands(g, ("BOM", "D2")) == []


class TestTheZeroedLine:
    """A line that looks costed and costs nothing."""

    def test_it_is_found(self, tmp_path):
        p = fixtures.workbook_with_a_zeroed_line(tmp_path / "b.xlsx")
        found = _by_type(analyze_document(extract_grid(p)), "unpriced_line_zeroed")
        assert len(found) == 1
        assert found[0].severity == "high"
        assert found[0].evidence[0].locator == "D3"

    def test_it_refuses_to_invent_the_missing_price(self, tmp_path):
        """The impact is unknowable from this document, so it stays None.

        A plausible guess would be worse than nothing: it gets quoted back at
        somebody in a budget meeting, with our name on it.
        """
        p = fixtures.workbook_with_a_zeroed_line(tmp_path / "b.xlsx")
        found = _by_type(analyze_document(extract_grid(p)), "unpriced_line_zeroed")
        assert found[0].impact_usd is None
        assert "will not guess" in found[0].detail

    def test_it_survives_a_workbook_that_was_never_calculated(self, tmp_path):
        """A script-written workbook has formulas and NO cached values.

        data_only=True returns None for every formula cell, so there is no zero
        to notice. Keying on the cached result would find this in hand-saved files
        and miss it in generated ones — and a generated estimate is exactly the
        kind whose arithmetic most wants checking. The durable signal is the
        formula referencing a cell that is not there.
        """
        p = fixtures.workbook_with_a_zeroed_line(tmp_path / "gen.xlsx", cached=False)
        found = _by_type(analyze_document(extract_grid(p)), "unpriced_line_zeroed")
        assert len(found) == 1


class TestTheDoubleCount:
    """The same money reaching the total by two routes.

    This is NOT deduplication. An item may legitimately appear on two sheets as a
    cross-reference; the bug is when BOTH subtotals include it.
    """

    @pytest.fixture
    def found(self, tmp_path):
        p = fixtures.workbook_with_a_double_count(tmp_path / "dc.xlsx")
        return analyze_document(extract_grid(p))

    def test_it_is_found_and_it_is_critical(self, found):
        dc = _by_type(found, "intra_doc_double_count")
        assert len(dc) == 1
        # Critical, not high, precisely because somebody wrote the note. They
        # knew, and both subtotals still counted it.
        assert dc[0].severity == "critical"

    def test_the_impact_is_the_redundant_copy_only(self, found):
        """Counted twice at 10,000 means 10,000 too much — not 20,000."""
        assert _by_type(found, "intra_doc_double_count")[0].impact_usd == 10000.0

    def test_it_cites_both_occurrences(self, found):
        ev = _by_type(found, "intra_doc_double_count")[0].evidence
        assert {(e.sheet, e.locator) for e in ev} == {
            ("Networking", "B3"), ("Simulation", "B3")
        }

    def test_the_incriminating_note_is_quoted(self, found):
        """Somebody wrote "shared with Networking sheet" and it happened anyway.

        Quoting it verbatim is what turns a finding into an argument nobody can
        wave away.
        """
        assert "shared with Networking sheet" in _by_type(
            found, "intra_doc_double_count")[0].detail

    def test_a_note_does_not_hide_the_duplicate_it_describes(self, tmp_path):
        """Regression, and the subtlest bug in this module.

        Bucketing candidate duplicates on the whole row text fails on exactly the
        cases that matter: the second occurrence of a shared item is the one
        carrying the note, and the note makes the two rows look different. The
        annotation that PROVES it is a duplicate was the thing stopping it from
        being recognised as one. Key on the item label, not the row.
        """
        p = fixtures.workbook_with_a_double_count(tmp_path / "dc.xlsx")
        assert _by_type(analyze_document(extract_grid(p)), "intra_doc_double_count")


class TestQuantityTwoIsNotADoubleCount:
    """The distinction the formula graph exists for.

    Two occurrences feeding the SAME subtotal are a genuine quantity of two.
    Flagging that would be a false accusation, and a tool that cries wolf about
    the customer's arithmetic gets switched off.
    """

    def test_same_item_twice_under_one_subtotal_is_not_flagged(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Networking"
        ws.append(["Item", "Cost"])
        ws.append(["Switch", 5000])
        ws.append(["Switch", 5000])       # we are buying two. That is fine.
        ws["B4"] = "=SUM(B2:B3)"
        p = tmp_path / "two.xlsx"
        wb.save(p)

        assert not _by_type(
            analyze_document(extract_grid(p)), "intra_doc_double_count"
        )


class TestTheHardcodedRollup:
    def test_a_typed_number_among_formulas_is_found(self, tmp_path):
        p = fixtures.workbook_with_a_hardcoded_rollup(tmp_path / "h.xlsx")
        found = _by_type(analyze_document(extract_grid(p)), "hardcoded_rollup")
        assert len(found) == 1
        assert found[0].evidence[0].locator == "B4"
        assert found[0].impact_usd == 192000.0

    def test_a_plain_data_column_is_not_flagged(self, tmp_path):
        """Literal numbers are normal. It is only suspicious when the cells
        around it are computed and this one is not."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Item", "Cost"])
        for i in range(6):
            ws.append([f"Item {i}", 1000])
        p = tmp_path / "plain.xlsx"
        wb.save(p)

        assert not _by_type(analyze_document(extract_grid(p)), "hardcoded_rollup")


class TestTheSummaryBlockUsesPlainAddition:
    """Regression. Real summary blocks add cells directly.

    An earlier version only recognised a "computed neighbour" if it used SUM() or
    another aggregating function. Actual summary sheets are full of plain
    addition — "=Compute!E9+Compute!E16" — and requiring an aggregate function
    missed every hardcoded cell sitting among them.

    The tell was never which function was used. It is that the cells around this
    one are COMPUTED and this one was typed.
    """

    def test_a_literal_among_plain_addition_formulas_is_caught(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Category", "Amount"])
        ws.append(["A", None])
        ws.append(["B", None])
        ws.append(["C", None])
        ws.append(["D", None])
        ws["B2"] = "=Detail!B2+Detail!B3"    # plain addition, not SUM()
        ws["B3"] = "=Detail!B4+Detail!B5"
        ws["B4"] = 50000                     # <- typed. Will never update.
        ws["B5"] = "=Detail!B6+Detail!B7"

        d = wb.create_sheet("Detail")
        d.append(["Item", "Cost"])
        for i in range(6):
            d.append([f"i{i}", 1000])
        p = tmp_path / "s.xlsx"
        wb.save(p)

        found = _by_type(analyze_document(extract_grid(p)), "hardcoded_rollup")
        assert [e.locator for f in found for e in f.evidence] == ["B4"]


class TestSpeakingUpWhenTheGraphCannotProveIt:
    """The engine must not go quiet on the documents that need it most.

    On real evidence, the sheets carrying a shared item often have NO formula
    subtotals at all — the subtotal is a number somebody typed, which is exactly
    why it ALSO shows up as a hardcoded rollup. The money is untraceable because
    the workbook is broken in a second way.

    An engine that requires a clean formula graph before it will speak goes silent
    on precisely those files. So it reports, says it cannot verify the routing,
    says why, and files it as a DECISION for a human rather than a defect it
    cannot demonstrate.
    """

    def test_a_shared_line_is_reported_even_with_typed_in_subtotals(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()

        net = wb.active
        net.title = "Networking"
        net.append(["Item", "Cost"])
        net.append(["Simulation Licence, Enterprise", 10000])
        net.append(["SUBTOTAL", 10000])       # typed, not a formula

        dt = wb.create_sheet("Digital Twin")
        dt.append(["Item", "Cost"])
        dt.append(["Network Simulation Software — shared with Networking", 10000])
        dt.append(["SUBTOTAL", 10000])        # typed, not a formula

        p = tmp_path / "untraceable.xlsx"
        wb.save(p)

        found = _by_type(analyze_document(extract_grid(p)), "intra_doc_double_count")
        assert len(found) == 1
        f = found[0]
        # Not a 'defect' — we cannot prove it. But absolutely not silence either.
        assert f.kind == "decision"
        assert f.data["routing_traceable"] is False
        assert "cannot prove" in f.detail
        assert "typed-in numbers" in f.detail

    def test_it_reports_the_row_once_not_once_per_column(self, tmp_path):
        """A line item spans qty, unit and extended.

        Treating each numeric cell as a candidate reports the same duplicate three
        times and announces that "1.00 appears on both sheets" — which is the
        quantity column, and is nonsense. The row's money is its largest figure.
        """
        import openpyxl

        wb = openpyxl.Workbook()
        a = wb.active
        a.title = "Alpha"
        a.append(["Item", "Qty", "Unit", "Ext"])
        a.append(["Widget Licence", 1, 10000, 10000])

        b = wb.create_sheet("Beta")
        b.append(["Item", "Qty", "Unit", "Ext"])
        b.append(["Widget Licence — shared with Alpha", 1, 10000, 10000])

        p = tmp_path / "rows.xlsx"
        wb.save(p)

        found = _by_type(analyze_document(extract_grid(p)), "intra_doc_double_count")
        assert len(found) == 1
        assert found[0].impact_usd == 10000.0    # not 1.0


class TestPrecisionOverRecall:
    """A detector that cries wolf gets switched off, and then protects nothing."""

    def test_a_multi_year_licence_term_is_not_a_recurring_charge(self, tmp_path):
        """"DNA Essentials 3-yr" is a single purchase covering three years.

        An earlier version keyed on "3-yr" anywhere in the row and produced
        twenty-nine confident accusations about lines that were perfectly correct.
        """
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Item", "Cost"])
        ws.append(["Network Automation Licence, 3-yr term", 4800])
        ws.append(["Access Control, 250-endpoint, 3yr", 9500])
        p = tmp_path / "terms.xlsx"
        wb.save(p)

        assert not _by_type(
            analyze_document(extract_grid(p)), "capex_opex_conflation"
        )

    def test_an_explicit_monthly_charge_is_a_recurring_charge(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Item", "Cost"])
        ws.append(["Internet circuit, 3000/mo", 3000])
        p = tmp_path / "isp.xlsx"
        wb.save(p)

        found = _by_type(analyze_document(extract_grid(p)), "capex_opex_conflation")
        assert len(found) == 1
        # We do not know the term, so we do not invent one.
        assert found[0].impact_usd is None


class TestOrderingIsPartOfTheProduct:
    def test_most_severe_first_then_by_money(self, tmp_path):
        """An executive reads until they stop."""
        p = fixtures.workbook_with_a_double_count(tmp_path / "dc.xlsx")
        found = analyze_document(extract_grid(p))
        severities = [f.severity for f in found]
        assert severities == sorted(
            severities, key=lambda s: ["critical", "high", "medium", "low", "info"].index(s)
        )


class TestFindingIntegrity:
    def test_an_unknown_finding_type_is_refused_at_construction(self):
        with pytest.raises(ValueError, match="unknown finding_type"):
            Finding(finding_type="made_up", title="x")

    def test_the_fingerprint_is_stable_across_runs(self, tmp_path):
        """A disposition must survive the next upload.

        If the fingerprint moved, every 'waived' and 'accepted' a customer had
        recorded would silently detach and the register would refill with issues
        they already dealt with.
        """
        p = fixtures.workbook_with_a_double_count(tmp_path / "dc.xlsx")
        a = analyze_document(extract_grid(p))
        b = analyze_document(extract_grid(p))
        assert [f.fingerprint for f in a] == [f.fingerprint for f in b]
        assert all(len(f.fingerprint) == 16 for f in a)

    def test_every_finding_cites_a_cell(self, tmp_path):
        """A finding that cannot say where it came from is a rumour."""
        for build in (
            fixtures.workbook_with_a_double_count,
            fixtures.workbook_with_a_hardcoded_rollup,
            fixtures.workbook_with_a_zeroed_line,
        ):
            p = build(tmp_path / f"{build.__name__}.xlsx")
            for f in analyze_document(extract_grid(p)):
                assert f.evidence, f.title
                assert f.evidence[0].locator, f.title
                assert f.detector == "deterministic"
