# CUI // SP-CTI
"""Turning a grid of cells into line items that mean something.

Invented content. ICDEV is a public repo.
"""
from __future__ import annotations

import openpyxl
import pytest

from tests.bom import fixtures
from tools.bom.extract_grid import extract_grid
from tools.bom.lines import _match_role, extract_lines, find_header, load_config


def _wb(path, rows, title="BOM"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    for r in rows:
        ws.append(r)
    wb.save(path)
    return path


class TestColumnRoles:
    @pytest.mark.parametrize(
        ("header", "role"),
        [
            ("QTY", "quantity"),
            ("Quantity", "quantity"),
            ("UNIT MSRP", "unit_price"),
            ("EXT. MSRP", "extended_price"),
            ("EST. STREET", "street_price"),
            ("PART NUMBER", "part_number"),
            ("Manufacturer", "manufacturer"),
            ("PRODUCT / DESCRIPTION", "description"),
            ("NOTES / PURPOSE", "notes"),
        ],
    )
    def test_headers_resolve(self, header, role):
        assert _match_role(header, load_config()) == role

    def test_the_longest_synonym_wins(self):
        """A BOM routinely has three columns whose names all contain "price".

        Picking the shortest match swaps the unit cost for the line total, and
        every arithmetic check downstream then compares the wrong two numbers.
        """
        cfg = load_config()
        assert _match_role("Unit Price", cfg) == "unit_price"
        assert _match_role("Extended Price", cfg) == "extended_price"
        assert _match_role("Price", cfg) == "unit_price"

    def test_a_bare_hash_is_a_line_number_not_a_quantity(self):
        """Mapping "#" to quantity steals the role from the real QTY column beside
        it — which then goes unmapped, and every arithmetic check downstream
        compares a ROW NUMBER against a price."""
        cfg = load_config()
        assert _match_role("#", cfg) is None
        assert _match_role("PART #", cfg) == "part_number"


class TestWhatCountsAsABillOfMaterials:
    def test_a_sheet_must_both_name_a_thing_and_price_it(self, tmp_path):
        """Every table has numbers. Only a BOM says what you are buying.

        A power schedule ("QTY | AMP | Max Watt | Total") has a quantity and
        something called a total. An earlier version turned its rows into line
        items — inventing priced components out of an electrical calculation,
        which then reported themselves as costing nothing.
        """
        p = _wb(tmp_path / "power.xlsx", [
            ["QTY", "AMP", "Connection", "Max Watt", "Total Usable kW"],
            [41, 30, "L6-30", 7200, 236160],
            [6, 30, "L21-30", 10800, 51840],
        ], title="Power")
        assert extract_lines(extract_grid(p)) == []

    def test_an_inventory_prices_nothing_and_yields_no_lines(self, tmp_path):
        p = _wb(tmp_path / "inv.xlsx", [
            ["Model", "Serial", "Warranty Ended"],
            ["Node A", "6HYXHB2", "2019-06-02"],
        ], title="Assets")
        assert extract_lines(extract_grid(p)) == []

    def test_a_sheet_with_no_header_yields_nothing_rather_than_a_guess(self, tmp_path):
        """Not "best effort" lines. None.

        A schema this engine invented is a schema nobody can check, and every
        number downstream would inherit the invention.
        """
        p = _wb(tmp_path / "prose.xlsx", [
            ["Some notes about the project"],
            ["More notes"],
        ])
        assert find_header(extract_grid(p).cells, "BOM", load_config()) is None
        assert extract_lines(extract_grid(p)) == []

    def test_a_header_below_a_title_block_is_still_found(self, tmp_path):
        """Real workbooks put a title, a logo and a revision block on top."""
        p = _wb(tmp_path / "titled.xlsx", [
            ["ACME PROJECT — NETWORK BOM"],
            ["Revision 3, prepared 2026-01-01"],
            [],
            ["Item", "Part Number", "QTY", "Unit Price", "Extended"],
            ["Core Switch", "CS-9500", 2, 21000, 42000],
        ])
        lines = extract_lines(extract_grid(p))
        assert len(lines) == 1
        assert lines[0].description == "Core Switch"
        assert lines[0].qty == 2


class TestNoDescriptionColumnIsNormal:
    def test_a_bom_headed_manufacturer_and_model(self, tmp_path):
        """"Manufacturer | Model | QTY | Price | Total" never uses the word
        "description" — and it is a perfectly ordinary BOM.

        An earlier version only fell back to the leftmost text when a description
        column EXISTED but was empty, so a sheet like this yielded zero lines: an
        entire authoritative bill of materials read as blank, silently.
        """
        p = _wb(tmp_path / "mfr.xlsx", [
            ["Manufacturer", "Model", "QTY", "Price", "Total"],
            ["Acme", "SW-9336", 2, 37600, 75200],
            ["Beta", "FW-4145", 1, 200000, 200000],
        ])
        lines = extract_lines(extract_grid(p))
        assert len(lines) == 2
        assert "Acme" in lines[0].description
        assert lines[0].part_number == "SW-9336"
        assert lines[0].extended_price == 75200


class TestTotalRowsAreNotLineItems:
    def test_a_row_labelled_total_is_skipped(self, tmp_path):
        p = _wb(tmp_path / "t.xlsx", [
            ["Item", "QTY", "Unit Price", "Extended"],
            ["Switch", 1, 100, 100],
            ["TOTAL", None, None, 100],
        ])
        assert [ln.description for ln in extract_lines(extract_grid(p))] == ["Switch"]

    def test_a_subtotal_disguised_as_a_heading_is_skipped(self, tmp_path):
        """The one that gets past a text check.

        A section subtotal labelled "1.1 Compute Hosts" says nothing about being a
        total. But its price cell is a SUM, and that is a FACT rather than a guess:
        a rollup is not a line item, whatever it is called. Reading it as one and
        adding it back in double-counts everything above it — invisibly, because
        the number it produces looks exactly like the number it should have
        produced, only twice.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"
        ws.append(["Item", "QTY", "Unit Price", "Extended"])
        ws.append(["Server A", 2, 100, None])
        ws.append(["Server B", 1, 200, None])
        ws.append(["1.1 Compute Hosts", None, None, None])
        ws["D2"] = "=B2*C2"
        ws["D3"] = "=B3*C3"
        ws["D4"] = "=SUM(D2:D3)"      # a subtotal wearing a heading's clothes
        p = tmp_path / "sub.xlsx"
        wb.save(p)

        lines = extract_lines(extract_grid(p))
        assert [ln.description for ln in lines] == ["Server A", "Server B"]


class TestPriceBasisIsEvidenceNotAssumption:
    def test_the_workbook_says_so_in_its_own_hand(self, tmp_path):
        """A street column computed as "=E2*0.6" PROVES the price column is a list.

        Nothing a header says is stronger than the arithmetic somebody actually
        wrote.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"
        ws.append(["Item", "QTY", "Price", "Extended", "Est. Street"])
        ws.append(["Switch", 1, 21000, None, None])
        ws["D2"] = "=B2*C2"
        ws["E2"] = "=D2*0.6"
        p = tmp_path / "basis.xlsx"
        wb.save(p)

        ln = extract_lines(extract_grid(p))[0]
        assert ln.price_basis == "msrp"
        assert "0.6" in ln.price_basis_reason

    def test_the_header_wording_is_the_next_best_thing(self, tmp_path):
        p = _wb(tmp_path / "rom.xlsx", [
            ["Item", "QTY", "Unit Est.", "Total Est."],
            ["Server", 4, 25000, 100000],
        ])
        assert extract_lines(extract_grid(p))[0].price_basis == "rom"

    def test_an_unknown_basis_stays_unknown(self, tmp_path):
        """Not 'msrp', not 'street', not whatever is convenient.

        Averaging a list price against somebody's rough estimate produces a number
        that is not so much wrong as meaningless — so 'unknown' is a finding, and a
        default that quietly picked one would be a lie.
        """
        p = _wb(tmp_path / "u.xlsx", [
            ["Manufacturer", "Model", "QTY", "Price", "Total"],
            ["Acme", "SW-1", 1, 100, 100],
        ])
        ln = extract_lines(extract_grid(p))[0]
        assert ln.price_basis == "unknown"
        assert "refusing to assume" in ln.price_basis_reason


class TestTheZeroedLine:
    def test_a_quantity_with_no_price_is_flagged(self, tmp_path):
        p = fixtures.workbook_with_a_zeroed_line(tmp_path / "z.xlsx")
        lines = extract_lines(extract_grid(p))
        missing = [ln for ln in lines if ln.price_missing]
        assert len(missing) == 1
        assert missing[0].description == "Widget B"
        assert missing[0].qty == 1
        assert missing[0].unit_price is None


class TestLineHashSurvivesAParserImprovement:
    def test_the_hash_is_over_the_bytes_as_they_arrived(self, tmp_path):
        """A human's merge approval is keyed on this hash.

        If it were computed over PARSED fields, improving the parser would rewrite
        every hash and silently orphan every decision the customer ever made.
        """
        p = _wb(tmp_path / "h.xlsx", [
            ["Item", "QTY", "Unit Price", "Extended"],
            ["Switch", 1, 100, 100],
        ])
        a = extract_lines(extract_grid(p))[0]
        b = extract_lines(extract_grid(p))[0]
        assert a.line_hash == b.line_hash
        assert len(a.line_hash) == 64
