# CUI // SP-CTI
"""The circulated workbook.

The layout is borrowed from a real one. So is the bug it is fixing: that workbook's
Summary sheet had its category totals TYPED IN, so editing a category sheet moved
nothing and the headline figure had quietly stopped tracking its own inputs. Nobody
reading it could tell — a typed 192000 and a computed 192000 render identically.

The load-bearing test here is therefore not "the total is right". It is
**"the total is a formula"** — because a total that is right today and hardcoded is
a total that is wrong the first time anybody edits anything.

Invented content. ICDEV is a public repo.
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from tools.bom.export_categorized import categorized_workbook
from tools.bom.findings import Evidence, Finding
from tools.bom.pivot import Dataset, Row


def _row(desc, cat, qty, unit, *, committed=True, reason="", src="s1"):
    return Row(
        line_id=f"l-{desc}", cluster_id=f"c-{desc}", description=desc,
        dims={"category": cat, "source": src},
        qty=qty, unit_price=unit, extended_price=qty * unit,
        committed=committed, excluded_reason=reason,
    )


@pytest.fixture
def dataset():
    return Dataset(rows=[
        _row("Core switch", "Networking", 2, 18000),
        _row("Access switch", "Networking", 4, 3500),
        _row("Rack server", "Compute", 4, 25000),
        _row("Hypervisor licence", "Software", 12, 3500),
    ])


@pytest.fixture
def book(dataset):
    return openpyxl.load_workbook(
        io.BytesIO(categorized_workbook(dataset, title="LAB BOM")),
    )


class TestTheShapeTheyAlreadyRead:
    def test_one_sheet_per_category_plus_a_summary(self, book):
        assert book.sheetnames == ["Summary", "Compute", "Networking", "Software"]

    def test_each_category_sheet_opens_with_its_banner(self, book):
        assert book["Networking"]["A1"].value == "NETWORKING"

    def test_the_columns_are_the_ones_their_eye_expects(self, book):
        ws = book["Networking"]
        assert [ws.cell(row=3, column=c).value for c in range(1, 7)] == [
            "Item", "Description", "Qty", "Unit Est.", "Total Est.", "Notes",
        ]

    def test_lines_land_on_the_sheet_for_their_category(self, book):
        ws = book["Compute"]
        assert ws["A4"].value == "Rack server"


class TestEveryRollupIsAFormula:
    """The bug the source template shipped with, and it is invisible on the page."""

    def test_the_extended_price_is_computed_not_stored(self, book):
        """Change the quantity and the line total must move. Otherwise the
        spreadsheet is a picture of a spreadsheet."""
        assert book["Compute"]["E4"].value == "=C4*D4"

    def test_the_category_total_sums_its_own_rows(self, book):
        ws = book["Networking"]
        total = next(
            r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Total"
        )
        assert ws.cell(row=total, column=5).value.startswith("=SUM(E4:E")

    def test_the_summary_REFERENCES_the_category_sheets(self, book):
        """Not a number. A cross-sheet reference — this is the exact cell that was
        typed in, in the workbook this layout came from."""
        ws = book["Summary"]
        vals = [ws.cell(row=r, column=3).value for r in range(7, 11)]
        assert any(str(v).startswith("='Networking'!E") for v in vals), vals

    def test_the_grand_total_sums_the_categories(self, book):
        ws = book["Summary"]
        grand = next(
            r for r in range(1, ws.max_row + 1)
            if ws.cell(row=r, column=1).value == "TOTAL"
        )
        assert str(ws.cell(row=grand, column=3).value).startswith("=SUM(C")

    def test_not_one_money_cell_is_hardcoded(self, book):
        """The whole point, stated as an invariant. If any priced line or rollup is
        a literal, an edit somewhere else silently fails to reach it."""
        hardcoded = []
        for name in ("Compute", "Networking", "Software"):
            ws = book[name]
            for r in range(4, ws.max_row + 1):
                v = ws.cell(row=r, column=5).value
                if v is not None and not str(v).startswith("="):
                    hardcoded.append(f"{name}!E{r} = {v!r}")
        assert not hardcoded, hardcoded


class TestNothingIsSilentlyDropped:
    def test_an_excluded_line_still_appears_and_says_why(self):
        """A line that vanishes from a category sheet is a line nobody argues about
        again — which is how a disputed cost quietly becomes a settled one."""
        ds = Dataset(rows=[
            _row("Disputed firewall", "Security", 1, 40000,
                 committed=False, reason="two sources disagree 19x"),
        ])
        wb = openpyxl.load_workbook(io.BytesIO(categorized_workbook(ds)))
        ws = wb["Security"]
        assert ws["A4"].value == "Disputed firewall"
        assert "19x" in ws["F4"].value

    def test_an_uncategorized_line_gets_a_sheet_rather_than_disappearing(self):
        ds = Dataset(rows=[Row(line_id="x", cluster_id="x", description="Mystery",
                               qty=1, unit_price=5, extended_price=5)])
        wb = openpyxl.load_workbook(io.BytesIO(categorized_workbook(ds)))
        assert "Uncategorized" in wb.sheetnames

    def test_two_categories_that_truncate_alike_do_not_collide(self):
        """Excel caps a sheet name at 31 characters. Two long category names that
        share a prefix would silently become one sheet, and one category's money
        would simply be gone."""
        a = "Infrastructure Hardware and Racking"
        b = "Infrastructure Hardware and Cabling"
        ds = Dataset(rows=[_row("x", a, 1, 10), _row("y", b, 1, 20)])
        wb = openpyxl.load_workbook(io.BytesIO(categorized_workbook(ds)))
        cats = [s for s in wb.sheetnames if s != "Summary"]
        assert len(cats) == 2, cats
        assert len(set(cats)) == 2


class TestOpenItems:
    def test_findings_that_need_a_human_become_open_items(self):
        f = Finding(
            finding_type="asset_count_disputed", kind="decision", severity="high",
            title="Two sources disagree on how many servers exist",
            impact_usd=120000,
            evidence=[Evidence("bom.xlsx", "Compute", "A9", "")],
        )
        wb = openpyxl.load_workbook(
            io.BytesIO(categorized_workbook(Dataset(rows=[_row("s", "Compute", 1, 1)]), [f])),
        )
        ws = wb["Open Items"]
        assert ws["C4"].value.startswith("Two sources disagree")
        assert ws["D4"].value == 120000
        assert ws["E4"].value == "Decision"

    def test_no_open_findings_means_no_sheet_rather_than_an_empty_one(self, book):
        assert "Open Items" not in book.sheetnames
