# CUI // SP-CTI
"""Tests for tools/govcon/bom_generator.py.

Verify the BOM rollup:
  - classifies equipment categories from description / unit
  - aggregates IGCE line items + vendor quotes per procurement
  - rolls up totals by tier and by equipment category
  - filters by tier, fiscal year, procurement_id, and category

The tests use a temporary SQLite database so the production schema is
untouched.
"""

from __future__ import annotations

import sqlite3
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(_ROOT))


# ---------------------------------------------------------------------------
# Equipment-category classifier
# ---------------------------------------------------------------------------


class ClassifyEquipmentCategoryTests(unittest.TestCase):
    def setUp(self):
        from tools.govcon.bom_generator import classify_equipment_category
        self.classify = classify_equipment_category

    def test_labor_from_hour_unit(self):
        self.assertEqual(self.classify("", "hour"), "labor")
        # Unit is included in the haystack even when description is present
        self.assertEqual(self.classify("misc", "hrs"), "labor")

    def test_labor_from_keyword(self):
        self.assertEqual(self.classify("Senior Software Engineer", "day"), "labor")
        self.assertEqual(self.classify("Cloud Architect support", ""), "labor")
        self.assertEqual(self.classify("PM Support", ""), "labor")

    def test_computers(self):
        self.assertEqual(self.classify("Dell PowerEdge Server", "each"), "computers")
        self.assertEqual(self.classify("Lenovo ThinkPad Laptop", "each"), "computers")

    def test_networking(self):
        self.assertEqual(self.classify("Cisco Catalyst Switch", "each"), "networking")
        self.assertEqual(self.classify("Network firewall appliance", "each"), "networking")

    def test_storage(self):
        self.assertEqual(self.classify("NetApp SAN array", "each"), "storage")
        self.assertEqual(self.classify("10TB disk shelf", "each"), "storage")

    def test_software(self):
        self.assertEqual(self.classify("RedHat Enterprise Linux license", "each"), "software")
        self.assertEqual(self.classify("GitHub Enterprise SaaS subscription", "month"), "software")

    def test_services(self):
        # Labor keyword "engineer" wins over "installation" — engineer is more specific intent
        # For pure services descriptions the services branch still wins
        self.assertEqual(self.classify("Quarterly Maintenance Visit", "each"), "services")
        self.assertEqual(self.classify("24x7 Help Desk support", "month"), "services")
        self.assertEqual(self.classify("Onsite Installation", "day"), "services")

    def test_facilities(self):
        self.assertEqual(self.classify("Office space lease", "month"), "facilities")
        self.assertEqual(self.classify("Utilities reimbursement", "month"), "facilities")

    def test_other_fallback(self):
        self.assertEqual(self.classify("Miscellaneous widget", "each"), "other")

    def test_more_specific_phrase_wins(self):
        # "Network" should be networking, not other
        self.assertEqual(self.classify("Network Operations Center cabling", "lot"), "networking")
        # "license" should be software, not services
        self.assertEqual(self.classify("Calibration license", "year"), "software")


# ---------------------------------------------------------------------------
# BOM rollup against an in-memory schema
# ---------------------------------------------------------------------------


class BuildBomRollupTests(unittest.TestCase):
    """Exercise the end-to-end build_bom() against a temp SQLite DB."""

    def setUp(self):
        # Build a temp DB with the minimum required tables.
        tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        tmp.close()
        self.tmp_path = Path(tmp.name)
        self.conn = sqlite3.connect(self.tmp_path)
        # Use sqlite3.Row so the BOM module can read column names
        self.conn.row_factory = sqlite3.Row
        self._init_schema()

        # Patch get_connection so the BOM module reads from our temp DB.
        from tools.db import storage
        self._storage_patcher = patch.object(
            storage, "get_connection", lambda *a, **kw: self.conn
        )
        self._storage_patcher.start()

        from tools.govcon import bom_generator
        self.bom_module = bom_generator
        # also patch the module's _get_db
        bom_generator._get_db = lambda *a, **kw: self.conn

    def tearDown(self):
        self._storage_patcher.stop()
        self.conn.close()
        self.tmp_path.unlink(missing_ok=True)

    def _init_schema(self):
        c = self.conn.cursor()
        c.executescript(
            """
            CREATE TABLE cpmp_budget_allocations (
                id TEXT PRIMARY KEY,
                initiative_code TEXT NOT NULL,
                title TEXT NOT NULL,
                fiscal_year INTEGER NOT NULL,
                tier TEXT NOT NULL CHECK(tier IN ('tier_1','tier_2')),
                allocated_usd REAL NOT NULL DEFAULT 0,
                obligated_usd REAL NOT NULL DEFAULT 0,
                available_usd REAL NOT NULL DEFAULT 0,
                status TEXT NOT NULL DEFAULT 'active'
            );
            CREATE TABLE proc_procurements (
                id TEXT PRIMARY KEY,
                solicitation TEXT, title TEXT, agency TEXT,
                contract_type TEXT, allocation_id TEXT, status TEXT,
                created_at TEXT
            );
            CREATE TABLE proc_igce_line_items (
                id TEXT PRIMARY KEY,
                procurement_id TEXT NOT NULL, clin TEXT NOT NULL,
                description TEXT, unit TEXT, quantity REAL,
                unit_cost REAL, extended_cost REAL, basis TEXT,
                poc TEXT, notes TEXT, equipment_category TEXT,
                created_at TEXT, updated_at TEXT
            );
            CREATE TABLE proc_vendor_quotes (
                id TEXT PRIMARY KEY,
                procurement_id TEXT NOT NULL, vendor_name TEXT NOT NULL,
                quote_ref TEXT, clin TEXT, unit_price REAL,
                quantity REAL, total_price REAL, quote_date TEXT,
                valid_until TEXT, status TEXT, notes TEXT,
                created_at TEXT, updated_at TEXT
            );
            """
        )
        self.conn.commit()

    def _seed(self):
        c = self.conn.cursor()
        c.executemany(
            "INSERT INTO cpmp_budget_allocations VALUES (?,?,?,?,?,?,?,?,?)",
            [
                ("alloc-1", "INIT-100", "Cloud Migration", 2026, "tier_1",
                 1_000_000, 0, 1_000_000, "active"),
                ("alloc-2", "INIT-200", "Network Refresh", 2026, "tier_1",
                 500_000, 0, 500_000, "active"),
                ("alloc-3", "INIT-300", "AI Lab Pilot", 2027, "tier_2",
                 250_000, 0, 250_000, "active"),
            ],
        )
        c.executemany(
            "INSERT INTO proc_procurements VALUES (?,?,?,?,?,?,?,?)",
            [
                ("PROC-001", "W912DY-26-R-0001", "Cloud Servers", "USACE",
                 "ffp", "alloc-1", "open", "2026-05-01"),
                ("PROC-002", "W912DY-26-R-0002", "Network Switches", "USACE",
                 "ffp", "alloc-2", "open", "2026-05-02"),
                ("PROC-003", "W912DY-27-R-0003", "AI Pilot Hardware", "DARPA",
                 "ffp", "alloc-3", "open", "2026-05-03"),
                ("PROC-004", "W912DY-26-R-0004", "Engineering Labor", "USACE",
                 "ffp", "alloc-1", "open", "2026-05-04"),
            ],
        )
        c.executemany(
            "INSERT INTO proc_igce_line_items VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            [
                ("l1", "PROC-001", "0001", "Dell PowerEdge Server",
                 "each", 10.0, 8000.0, 80000.0, "GSA", "POC", "", "unspecified",
                 "2026-05-01", "2026-05-01"),
                ("l2", "PROC-001", "0002", "RedHat Enterprise Linux license",
                 "each", 10.0, 1000.0, 10000.0, "GSA", "POC", "", "unspecified",
                 "2026-05-01", "2026-05-01"),
                ("l3", "PROC-002", "0001", "Cisco Catalyst Switch",
                 "each", 20.0, 5000.0, 100000.0, "GSA", "POC", "", "networking",
                 "2026-05-01", "2026-05-01"),
                ("l4", "PROC-003", "0001", "AI GPU Server cluster",
                 "each", 2.0, 60000.0, 120000.0, "GSA", "POC", "", "unspecified",
                 "2026-05-01", "2026-05-01"),
                ("l5", "PROC-004", "0001", "Senior Cloud Engineer",
                 "hour", 1600.0, 175.0, 280000.0, "GSA CALC+", "POC", "", "unspecified",
                 "2026-05-01", "2026-05-01"),
            ],
        )
        c.executemany(
            "INSERT INTO proc_vendor_quotes VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            [
                ("q1", "PROC-001", "Acme Federal LLC", "AF-001", "0001",
                 7800.0, 10.0, 78000.0, "2026-05-15", "2026-06-15", "submitted",
                 "", "2026-05-15", "2026-05-15"),
                ("q2", "PROC-001", "Bravo Tech", "BT-001", "0001",
                 8200.0, 10.0, 82000.0, "2026-05-15", "2026-06-15", "submitted",
                 "", "2026-05-15", "2026-05-15"),
                ("q3", "PROC-002", "Cisco Direct", "CD-001", "0001",
                 5100.0, 20.0, 102000.0, "2026-05-15", "2026-06-15", "submitted",
                 "", "2026-05-15", "2026-05-15"),
                ("q4", "PROC-003", "GPU Vendor A", "GPU-A", "0001",
                 58000.0, 2.0, 116000.0, "2026-05-15", "2026-06-15", "submitted",
                 "", "2026-05-15", "2026-05-15"),
            ],
        )
        self.conn.commit()

    def test_classifies_uncategorized_lines(self):
        self._seed()
        result = self.bom_module.build_bom()
        self.assertEqual(result["status"], "ok")
        # line l1 was unspecified → should now be "computers"
        proc_001 = next(p for p in result["procurements"] if p["procurement_id"] == "PROC-001")
        line_0001 = next(ln for ln in proc_001["lines"] if ln["clin"] == "0001")
        self.assertEqual(line_0001["equipment_category"], "computers")
        line_0002 = next(ln for ln in proc_001["lines"] if ln["clin"] == "0002")
        self.assertEqual(line_0002["equipment_category"], "software")
        # l4 was unspecified with "AI GPU Server" — keyword "server" → computers
        proc_003 = next(p for p in result["procurements"] if p["procurement_id"] == "PROC-003")
        line_0001 = proc_003["lines"][0]
        self.assertEqual(line_0001["equipment_category"], "computers")
        # l5 was unspecified with "hour" → labor
        proc_004 = next(p for p in result["procurements"] if p["procurement_id"] == "PROC-004")
        self.assertEqual(proc_004["lines"][0]["equipment_category"], "labor")

    def test_vendor_quote_aggregation(self):
        self._seed()
        result = self.bom_module.build_bom()
        proc_001 = next(p for p in result["procurements"] if p["procurement_id"] == "PROC-001")
        line_0001 = next(ln for ln in proc_001["lines"] if ln["clin"] == "0001")
        self.assertEqual(line_0001["vendor_quote_count"], 2)
        self.assertEqual(line_0001["min_quote"], 78000.0)
        self.assertEqual(line_0001["max_quote"], 82000.0)
        self.assertEqual(line_0001["avg_quote"], 80000.0)
        # IGCE rolled up
        self.assertEqual(proc_001["igce_total"], 90000.0)

    def test_rollup_by_tier(self):
        self._seed()
        result = self.bom_module.build_bom()
        by_tier = result["by_tier"]
        self.assertIn("tier_1", by_tier)
        self.assertIn("tier_2", by_tier)
        # tier_1 contains PROC-001 (90k), PROC-002 (100k), PROC-004 (280k)
        self.assertEqual(by_tier["tier_1"]["procurement_count"], 3)
        self.assertEqual(by_tier["tier_1"]["igce_total"], 470000.0)
        self.assertEqual(by_tier["tier_2"]["procurement_count"], 1)
        self.assertEqual(by_tier["tier_2"]["igce_total"], 120000.0)

    def test_rollup_by_category(self):
        self._seed()
        result = self.bom_module.build_bom()
        by_cat = result["by_category"]
        # Expected categories: computers, software, networking, labor
        self.assertIn("computers", by_cat)
        self.assertIn("software", by_cat)
        self.assertIn("networking", by_cat)
        self.assertIn("labor", by_cat)
        # Each category should have tier_breakdown
        for cat_data in by_cat.values():
            self.assertIn("tier_breakdown", cat_data)

    def test_filter_by_tier(self):
        self._seed()
        result = self.bom_module.build_bom(tier="tier_1")
        self.assertEqual(result["totals"]["procurement_count"], 3)
        self.assertTrue(all(p["tier"] == "tier_1" for p in result["procurements"]))

    def test_filter_by_fiscal_year(self):
        self._seed()
        result = self.bom_module.build_bom(fiscal_year=2027)
        self.assertEqual(result["totals"]["procurement_count"], 1)
        self.assertEqual(result["procurements"][0]["procurement_id"], "PROC-003")

    def test_filter_by_procurement_id(self):
        self._seed()
        result = self.bom_module.build_bom(procurement_id="PROC-001")
        self.assertEqual(result["totals"]["procurement_count"], 1)
        self.assertEqual(result["procurements"][0]["procurement_id"], "PROC-001")

    def test_empty_db_returns_empty_rollup(self):
        # Don't seed; tables exist, but no rows.
        result = self.bom_module.build_bom()
        self.assertEqual(result["status"], "ok")
        self.assertEqual(result["totals"]["procurement_count"], 0)
        self.assertEqual(result["by_tier"], {})
        self.assertEqual(result["by_category"], {})
        self.assertEqual(result["procurements"], [])


# ---------------------------------------------------------------------------
# CSV / XLSX export
# ---------------------------------------------------------------------------


import csv
import io


def _build_seed_db():
    """Helper: construct a small in-memory DB with one procurement + 2 CLINs."""
    tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
    tmp.close()
    conn = sqlite3.connect(tmp.name)
    conn.row_factory = sqlite3.Row
    conn.executescript(
        """
        CREATE TABLE cpmp_budget_allocations (
            id TEXT PRIMARY KEY, initiative_code TEXT NOT NULL,
            title TEXT NOT NULL, fiscal_year INTEGER NOT NULL,
            tier TEXT NOT NULL CHECK(tier IN ('tier_1','tier_2')),
            allocated_usd REAL NOT NULL DEFAULT 0,
            obligated_usd REAL NOT NULL DEFAULT 0,
            available_usd REAL NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'active'
        );
        CREATE TABLE proc_procurements (
            id TEXT PRIMARY KEY, solicitation TEXT, title TEXT, agency TEXT,
            contract_type TEXT, allocation_id TEXT, status TEXT, created_at TEXT
        );
        CREATE TABLE proc_igce_line_items (
            id TEXT PRIMARY KEY, procurement_id TEXT NOT NULL, clin TEXT NOT NULL,
            description TEXT, unit TEXT, quantity REAL, unit_cost REAL,
            extended_cost REAL, basis TEXT, poc TEXT, notes TEXT,
            equipment_category TEXT, created_at TEXT, updated_at TEXT
        );
        CREATE TABLE proc_vendor_quotes (
            id TEXT PRIMARY KEY, procurement_id TEXT NOT NULL,
            vendor_name TEXT NOT NULL, quote_ref TEXT, clin TEXT,
            unit_price REAL, quantity REAL, total_price REAL, quote_date TEXT,
            valid_until TEXT, status TEXT, notes TEXT,
            created_at TEXT, updated_at TEXT
        );
        """
    )
    conn.execute(
        "INSERT INTO cpmp_budget_allocations VALUES (?,?,?,?,?,?,?,?,?)",
        ("alloc-1", "INIT-100", "Cloud Migration", 2026, "tier_1",
         1_000_000, 0, 1_000_000, "active"),
    )
    conn.execute(
        "INSERT INTO proc_procurements VALUES (?,?,?,?,?,?,?,?)",
        ("PROC-001", "W912DY-26-R-0001", "Cloud Servers", "USACE",
         "ffp", "alloc-1", "open", "2026-05-01"),
    )
    conn.executemany(
        "INSERT INTO proc_igce_line_items VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        [
            ("l1", "PROC-001", "0001", "Dell PowerEdge Server",
             "each", 10.0, 8000.0, 80000.0, "GSA", "POC", "", "unspecified",
             "2026-05-01", "2026-05-01"),
            ("l2", "PROC-001", "0002", "RedHat Enterprise Linux license",
             "each", 10.0, 1000.0, 10000.0, "GSA", "POC", "", "unspecified",
             "2026-05-01", "2026-05-01"),
        ],
    )
    conn.executemany(
        "INSERT INTO proc_vendor_quotes VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        [
            ("q1", "PROC-001", "Acme Federal LLC", "AF-001", "0001",
             7800.0, 10.0, 78000.0, "2026-05-15", "2026-06-15", "submitted",
             "", "2026-05-15", "2026-05-15"),
        ],
    )
    conn.commit()
    return conn, Path(tmp.name)


class BomExportTests(unittest.TestCase):
    """Verify CSV + XLSX export of the BOM rollup for procurement submission."""

    def setUp(self):
        self.conn, self.tmp_path = _build_seed_db()
        from tools.db import storage
        self._storage_patcher = patch.object(
            storage, "get_connection", lambda *a, **kw: self.conn
        )
        self._storage_patcher.start()
        from tools.govcon import bom_generator
        self.bom_module = bom_generator
        bom_generator._get_db = lambda *a, **kw: self.conn

    def tearDown(self):
        self._storage_patcher.stop()
        self.conn.close()
        self.tmp_path.unlink(missing_ok=True)

    def test_csv_export_returns_bytes_with_utf8_sig(self):
        result = self.bom_module.build_bom(procurement_id="PROC-001")
        payload = self.bom_module.export_bom_csv(result)
        # Must be bytes, not str — for direct file write to /download
        self.assertIsInstance(payload, bytes)
        # UTF-8 BOM at the start so Excel opens it cleanly
        self.assertEqual(payload[:3], b"\xef\xbb\xbf")

    def test_csv_export_has_proper_headers(self):
        result = self.bom_module.build_bom(procurement_id="PROC-001")
        payload = self.bom_module.export_bom_csv(result)
        text = payload.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        # Header + 2 CLINs
        self.assertEqual(len(rows), 3)
        header = rows[0]
        # Must include all 9 BOM line-item fields + procurement metadata
        for col in ("procurement_id", "solicitation", "agency", "clin",
                    "description", "quantity", "igce_unit_cost",
                    "igce_extended_cost", "equipment_category",
                    "min_quote", "max_quote", "tier", "fiscal_year"):
            self.assertIn(col, header, f"CSV missing column: {col}")

    def test_csv_export_row_contents(self):
        result = self.bom_module.build_bom(procurement_id="PROC-001")
        payload = self.bom_module.export_bom_csv(result)
        text = payload.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
        self.assertEqual(len(rows), 2)
        # Row 0: server (classified as computers)
        row0 = next(r for r in rows if r["clin"] == "0001")
        self.assertEqual(row0["procurement_id"], "PROC-001")
        self.assertEqual(row0["solicitation"], "W912DY-26-R-0001")
        self.assertEqual(row0["agency"], "USACE")
        self.assertEqual(row0["equipment_category"], "computers")
        self.assertEqual(row0["quantity"], "10.0")
        self.assertEqual(row0["igce_extended_cost"], "80000.0")
        self.assertEqual(row0["min_quote"], "78000.0")
        # Row 1: license → software
        row1 = next(r for r in rows if r["clin"] == "0002")
        self.assertEqual(row1["equipment_category"], "software")

    def test_csv_export_empty_rollup_has_header_only(self):
        # Build an empty rollup (no procurements)
        result = self.bom_module.build_bom(procurement_id="DOES-NOT-EXIST")
        payload = self.bom_module.export_bom_csv(result)
        text = payload.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        self.assertEqual(len(rows), 1)  # header only
        self.assertIn("procurement_id", rows[0])

    def test_xlsx_export_returns_bytes(self):
        result = self.bom_module.build_bom(procurement_id="PROC-001")
        payload = self.bom_module.export_bom_xlsx(result)
        self.assertIsInstance(payload, bytes)
        # XLSX files start with the PK zip signature
        self.assertEqual(payload[:2], b"PK")
        self.assertGreater(len(payload), 100)

    def test_xlsx_export_has_summary_and_lines_sheets(self):
        import openpyxl
        result = self.bom_module.build_bom(procurement_id="PROC-001")
        payload = self.bom_module.export_bom_xlsx(result)
        wb = openpyxl.load_workbook(io.BytesIO(payload))
        # Should have at least Lines + Summary sheets
        sheet_names = wb.sheetnames
        self.assertTrue(
            any("line" in s.lower() for s in sheet_names),
            f"XLSX missing Lines sheet: {sheet_names}",
        )
        self.assertTrue(
            any("summar" in s.lower() for s in sheet_names),
            f"XLSX missing Summary sheet: {sheet_names}",
        )

    def test_xlsx_export_lines_sheet_has_clin_rows(self):
        import openpyxl
        result = self.bom_module.build_bom(procurement_id="PROC-001")
        payload = self.bom_module.export_bom_xlsx(result)
        wb = openpyxl.load_workbook(io.BytesIO(payload))
        lines_sheet = next(s for s in wb.worksheets if "line" in s.title.lower())
        # Header row + 2 CLINs = 3 rows
        self.assertEqual(lines_sheet.max_row, 3)
        # Header should include procurement_id and clin
        header = [c.value for c in lines_sheet[1]]
        self.assertIn("procurement_id", header)
        self.assertIn("clin", header)
        self.assertIn("igce_extended_cost", header)

    def test_xlsx_export_summary_has_totals(self):
        import openpyxl
        result = self.bom_module.build_bom(procurement_id="PROC-001")
        payload = self.bom_module.export_bom_xlsx(result)
        wb = openpyxl.load_workbook(io.BytesIO(payload))
        summary_sheet = next(s for s in wb.worksheets if "summar" in s.title.lower())
        # Find the procurement_count or igce_total label
        all_text = []
        for row in summary_sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    all_text.append(str(cell).lower())
        joined = " ".join(all_text)
        self.assertIn("procurement_count", joined)
        self.assertIn("igce_total", joined)
        # The actual count and total should be present as values
        self.assertIn("1", all_text)  # 1 procurement
        # 80000 + 10000 = 90000 IGCE total
        self.assertTrue(
            any("90000" in t or "90,000" in t for t in all_text),
            f"XLSX summary missing 90000 IGCE total; got: {all_text}",
        )

    def test_export_bom_dispatches_to_csv(self):
        result = self.bom_module.build_bom(procurement_id="PROC-001")
        csv_bytes = self.bom_module.export_bom(result, format="csv")
        self.assertIsInstance(csv_bytes, bytes)
        self.assertEqual(csv_bytes[:3], b"\xef\xbb\xbf")

    def test_export_bom_dispatches_to_xlsx(self):
        result = self.bom_module.build_bom(procurement_id="PROC-001")
        xlsx_bytes = self.bom_module.export_bom(result, format="xlsx")
        self.assertEqual(xlsx_bytes[:2], b"PK")


if __name__ == "__main__":
    unittest.main()
