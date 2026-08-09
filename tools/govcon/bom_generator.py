#!/usr/bin/env python3
# CUI // SP-CTI
"""Bill of Materials (BOM) generator for procurement initiatives.

Aggregates IGCE line items + vendor quotes for every procurement linked to
a ``cpmp_budget_allocations`` (initiative) and rolls the result up two
ways: by *initiative tier* (tier_1 / tier_2) and by *equipment category*
(computers, networking, software, labor, services, other, ...).

Tiers are defined in [[tools/budget/initiative_allocator.py]] and stored
on the allocation row. Equipment categories live on each IGCE line item
(``proc_igce_line_items.equipment_category``) and are inferred from CLIN
descriptions when not explicitly assigned.

Schema read:
    - proc_procurements            (parent container, has allocation_id FK)
    - proc_igce_line_items         (one row per CLIN, equipment_category)
    - proc_vendor_quotes           (vendor pricing per CLIN)
    - cpmp_budget_allocations      (initiative, tier_1 / tier_2)

Usage:
    python tools/govcon/bom_generator.py --bom --json
    python tools/govcon/bom_generator.py --bom --tier tier_1 --json
    python tools/govcon/bom_generator.py --bom --fiscal-year 2026 --json
    python tools/govcon/bom_generator.py --bom --category computers --json
    python tools/govcon/bom_generator.py --procurement PROC-2026-001 --json
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(_ROOT))

from tools.db.storage import get_connection


# ----------------------------------------------------------------------------
# Equipment category taxonomy
# ----------------------------------------------------------------------------

# Canonical categories. When an IGCE line has equipment_category =
# 'unspecified' we re-classify by scanning the description.
CANONICAL_CATEGORIES = (
    "computers",
    "networking",
    "storage",
    "peripherals",
    "software",
    "labor",
    "services",
    "facilities",
    "other",
    "unspecified",
)

# Keyword → category. Order matters: more specific phrases first.
_CATEGORY_KEYWORDS = (
    # Labor first — labor roles are usually the dominant intent even
    # when a software/saas keyword co-occurs ("SaaS subscription engineer
    # hours" is still labor, not software).
    ("engineer", "labor"),
    ("developer", "labor"),
    ("analyst", "labor"),
    ("architect", "labor"),
    ("consultant", "labor"),
    ("technician", "labor"),
    ("subject matter expert", "labor"),
    ("program manager", "labor"),
    ("hour", "labor"),
    ("hrs", "labor"),
    ("hr", "labor"),
    ("staff", "labor"),
    # Services
    ("support", "services"),
    ("maintenance", "services"),
    ("training", "services"),
    ("installation", "services"),
    ("integration", "services"),
    ("deployment", "services"),
    # Hardware
    ("server", "computers"),
    ("workstation", "computers"),
    ("laptop", "computers"),
    ("desktop", "computers"),
    ("computer", "computers"),
    ("switch", "networking"),
    ("router", "networking"),
    ("firewall", "networking"),
    ("cisco", "networking"),
    ("network", "networking"),
    ("san", "storage"),
    ("nas", "storage"),
    ("storage", "storage"),
    ("disk", "storage"),
    ("monitor", "peripherals"),
    ("printer", "peripherals"),
    ("scanner", "peripherals"),
    # Software
    ("license", "software"),
    ("software", "software"),
    ("subscription", "software"),
    ("saas", "software"),
    ("paas", "software"),
    ("iaas", "software"),
    # Facilities
    ("facility", "facilities"),
    ("office", "facilities"),
    ("space", "facilities"),
    ("utilities", "facilities"),
)


def classify_equipment_category(description: str, unit: str = "") -> str:
    """Return the canonical category for a CLIN description.

    When ``description`` is empty the unit-of-measure is used (e.g. "hour"
    implies labor). When no keyword matches the result is "other" so
    the BOM never silently drops a line.
    """
    # Unit-of-measure takes precedence when description is empty.
    if not (description or "").strip():
        unit_norm = (unit or "").lower().strip()
        if unit_norm in ("hour", "hours", "hr", "hrs", "day", "days"):
            return "labor"
        if not unit_norm:
            return "other"
    desc = (description or "").lower()
    # Word-boundary matches for short ambiguous tokens like "pm"
    import re
    for word in ("pm", "sme"):
        if re.search(rf"\b{re.escape(word)}\b", desc):
            return "labor"
    haystack = f" {desc} {(unit or '').lower()} "
    for needle, category in _CATEGORY_KEYWORDS:
        if needle in haystack:
            return category
    return "other"


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------


def _get_db():
    conn = get_connection()
    # Service-layer read; clear RLS context to avoid d.classification
    # joins blowing up.
    try:
        conn.set_security_context(None)  # rls-bypass: read-only BOM rollup
    except Exception:
        pass
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def _row_to_dict(row) -> Dict[str, Any]:
    if row is None:
        return {}
    if hasattr(row, "keys"):
        try:
            return {k: row[k] for k in row.keys()}
        except (IndexError, TypeError):
            pass
    # Fall back to wrapping unknown row types
    return {"_row": row}


def _safe_float(value, default: float = 0.0) -> float:
    try:
        return float(value) if value is not None else default
    except (TypeError, ValueError):
        return default


# ----------------------------------------------------------------------------
# Core: BOM rollup
# ----------------------------------------------------------------------------


def _fetch_procurements(
    conn,
    tier: Optional[str] = None,
    fiscal_year: Optional[int] = None,
    procurement_id: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Return proc_procurements rows joined to their parent allocation.

    Procurements without an allocation are surfaced with tier='unassigned'
    so they can still be reported (and re-linked). Filtering by tier or
    FY is done at this layer.
    """
    sql = [
        "SELECT p.id, p.solicitation, p.title, p.agency, p.contract_type,",
        "       p.allocation_id, p.status, p.created_at,",
        "       a.initiative_code, a.title AS initiative_title,",
        "       a.tier, a.fiscal_year, a.allocated_usd, a.obligated_usd,",
        "       a.available_usd, a.status AS allocation_status",
        "FROM proc_procurements p",
        "LEFT JOIN cpmp_budget_allocations a ON a.id = p.allocation_id",
        "WHERE 1=1",
    ]
    args: List[Any] = []
    if procurement_id:
        sql.append("AND p.id = ?")
        args.append(procurement_id)
    if tier:
        sql.append("AND a.tier = ?")
        args.append(tier)
    if fiscal_year is not None:
        sql.append("AND a.fiscal_year = ?")
        args.append(fiscal_year)
    sql.append("ORDER BY a.tier, a.initiative_code, p.id")
    cur = conn.execute("\n".join(sql), tuple(args))
    return [_row_to_dict(r) for r in cur.fetchall()]


def _fetch_igce_for(conn, procurement_ids: List[str]) -> Dict[str, List[Dict[str, Any]]]:
    """Group IGCE line items by procurement_id."""
    if not procurement_ids:
        return {}
    placeholders = ",".join("?" * len(procurement_ids))
    sql = (
        "SELECT * FROM proc_igce_line_items "
        f"WHERE procurement_id IN ({placeholders}) "
        "ORDER BY procurement_id, clin"
    )
    cur = conn.execute(sql, tuple(procurement_ids))
    out: Dict[str, List[Dict[str, Any]]] = {pid: [] for pid in procurement_ids}
    for row in cur.fetchall():
        rec = _row_to_dict(row)
        out.setdefault(rec["procurement_id"], []).append(rec)
    return out


def _fetch_quotes_for(conn, procurement_ids: List[str]) -> Dict[str, List[Dict[str, Any]]]:
    if not procurement_ids:
        return {}
    placeholders = ",".join("?" * len(procurement_ids))
    sql = (
        "SELECT * FROM proc_vendor_quotes "
        f"WHERE procurement_id IN ({placeholders}) "
        "ORDER BY procurement_id, vendor_name, clin"
    )
    cur = conn.execute(sql, tuple(procurement_ids))
    out: Dict[str, List[Dict[str, Any]]] = {pid: [] for pid in procurement_ids}
    for row in cur.fetchall():
        rec = _row_to_dict(row)
        out.setdefault(rec["procurement_id"], []).append(rec)
    return out


def _aggregate_line(
    line: Dict[str, Any], quotes: List[Dict[str, Any]]
) -> Dict[str, Any]:
    """One CLIN → one BOM line record with stats rolled up from quotes."""
    cat = (line.get("equipment_category") or "unspecified").strip() or "unspecified"
    if cat == "unspecified":
        cat = classify_equipment_category(
            line.get("description", ""), line.get("unit", "")
        )

    line_quotes = [q for q in quotes if q.get("clin") == line.get("clin")]
    quote_prices = [_safe_float(q.get("total_price")) for q in line_quotes if q.get("total_price") is not None]
    min_quote = min(quote_prices) if quote_prices else None
    max_quote = max(quote_prices) if quote_prices else None
    avg_quote = round(sum(quote_prices) / len(quote_prices), 2) if quote_prices else None

    igce_total = _safe_float(line.get("extended_cost"))
    qty = _safe_float(line.get("quantity"))
    unit_cost = _safe_float(line.get("unit_cost"))

    return {
        "clin": line.get("clin", ""),
        "description": line.get("description", ""),
        "unit": line.get("unit", ""),
        "quantity": qty,
        "igce_unit_cost": unit_cost,
        "igce_extended_cost": igce_total,
        "equipment_category": cat,
        "vendor_quote_count": len(line_quotes),
        "min_quote": min_quote,
        "max_quote": max_quote,
        "avg_quote": avg_quote,
    }


def build_bom(
    tier: Optional[str] = None,
    fiscal_year: Optional[int] = None,
    procurement_id: Optional[str] = None,
) -> Dict[str, Any]:
    """Return the full BOM rollup. Structure::

        {
          "totals": { by_category: {...}, by_tier: {...} },
          "by_tier": { tier_1: { by_category: {...}, procurements: [...] }, ... },
          "procurements": [ ... ]
        }
    """
    conn = _get_db()
    procurements = _fetch_procurements(conn, tier=tier, fiscal_year=fiscal_year,
                                       procurement_id=procurement_id)
    if not procurements:
        return {
            "status": "ok",
            "filter": {"tier": tier, "fiscal_year": fiscal_year, "procurement_id": procurement_id},
            "totals": {"by_category": {}, "by_tier": {}, "procurement_count": 0, "igce_total": 0.0, "quote_min": None, "quote_max": None},
            "by_tier": {},
            "by_category": {},
            "procurements": [],
        }

    pids = [p["id"] for p in procurements]
    igce_by_proc = _fetch_igce_for(conn, pids)
    quotes_by_proc = _fetch_quotes_for(conn, pids)

    proc_records: List[Dict[str, Any]] = []
    for proc in procurements:
        pid = proc["id"]
        lines = igce_by_proc.get(pid, [])
        quotes = quotes_by_proc.get(pid, [])
        bom_lines = [_aggregate_line(line, quotes) for line in lines]
        igce_total = round(sum(b["igce_extended_cost"] for b in bom_lines), 2)
        quote_prices = [b["min_quote"] for b in bom_lines if b["min_quote"] is not None]
        proc_records.append({
            "procurement_id": pid,
            "solicitation": proc.get("solicitation", ""),
            "title": proc.get("title", ""),
            "agency": proc.get("agency", ""),
            "contract_type": proc.get("contract_type", ""),
            "allocation_id": proc.get("allocation_id"),
            "initiative_code": proc.get("initiative_code"),
            "initiative_title": proc.get("initiative_title"),
            "tier": proc.get("tier") or "unassigned",
            "fiscal_year": proc.get("fiscal_year"),
            "status": proc.get("status"),
            "line_count": len(bom_lines),
            "igce_total": igce_total,
            "min_quote": min(quote_prices) if quote_prices else None,
            "max_quote": max(quote_prices) if quote_prices else None,
            "lines": bom_lines,
        })

    # Roll up: by_tier → by_category
    by_tier: Dict[str, Dict[str, Any]] = {}
    by_category: Dict[str, Dict[str, Any]] = {}
    for proc in proc_records:
        t = proc["tier"] or "unassigned"
        bucket = by_tier.setdefault(t, {
            "procurement_count": 0,
            "igce_total": 0.0,
            "line_count": 0,
            "by_category": {},
            "procurement_ids": [],
        })
        bucket["procurement_count"] += 1
        bucket["igce_total"] = round(bucket["igce_total"] + proc["igce_total"], 2)
        bucket["line_count"] += proc["line_count"]
        bucket["procurement_ids"].append(proc["procurement_id"])
        for line in proc["lines"]:
            cat = line["equipment_category"]
            cb = bucket["by_category"].setdefault(cat, {
                "line_count": 0,
                "quantity_total": 0.0,
                "igce_total": 0.0,
            })
            cb["line_count"] += 1
            cb["quantity_total"] = round(cb["quantity_total"] + line["quantity"], 4)
            cb["igce_total"] = round(cb["igce_total"] + line["igce_extended_cost"], 2)
            # global
            gb = by_category.setdefault(cat, {
                "line_count": 0,
                "quantity_total": 0.0,
                "igce_total": 0.0,
                "tier_breakdown": {},
            })
            gb["line_count"] += 1
            gb["quantity_total"] = round(gb["quantity_total"] + line["quantity"], 4)
            gb["igce_total"] = round(gb["igce_total"] + line["igce_extended_cost"], 2)
            tb = gb["tier_breakdown"].setdefault(t, {"line_count": 0, "igce_total": 0.0})
            tb["line_count"] += 1
            tb["igce_total"] = round(tb["igce_total"] + line["igce_extended_cost"], 2)

    # Top-level totals
    total_igce = round(sum(p["igce_total"] for p in proc_records), 2)
    all_min = [p["min_quote"] for p in proc_records if p["min_quote"] is not None]
    all_max = [p["max_quote"] for p in proc_records if p["max_quote"] is not None]
    totals_by_tier = {t: {"procurement_count": v["procurement_count"],
                          "igce_total": v["igce_total"]}
                      for t, v in by_tier.items()}
    totals_by_category = {c: {"line_count": v["line_count"],
                              "igce_total": v["igce_total"]}
                          for c, v in by_category.items()}

    return {
        "status": "ok",
        "filter": {"tier": tier, "fiscal_year": fiscal_year, "procurement_id": procurement_id},
        "totals": {
            "procurement_count": len(proc_records),
            "line_count": sum(p["line_count"] for p in proc_records),
            "igce_total": total_igce,
            "quote_min": min(all_min) if all_min else None,
            "quote_max": max(all_max) if all_max else None,
            "by_tier": totals_by_tier,
            "by_category": totals_by_category,
        },
        "by_tier": by_tier,
        "by_category": by_category,
        "procurements": proc_records,
    }


# ----------------------------------------------------------------------------
# Export: CSV + XLSX for procurement submission
# ----------------------------------------------------------------------------

# Flat row schema for procurement submission. Procurement metadata is
# denormalised onto every CLIN row so the spreadsheet can be filtered /
# pivoted in Excel without needing a second table.
_CSV_COLUMNS: List[str] = [
    "procurement_id",
    "solicitation",
    "initiative_code",
    "initiative_title",
    "tier",
    "fiscal_year",
    "agency",
    "contract_type",
    "procurement_status",
    "clin",
    "description",
    "equipment_category",
    "unit",
    "quantity",
    "igce_unit_cost",
    "igce_extended_cost",
    "vendor_quote_count",
    "min_quote",
    "max_quote",
    "avg_quote",
    "basis",
    "poc",
    "notes",
]


def _flatten_bom_rows(bom: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Convert the nested build_bom() result into one dict per CLIN."""
    rows: List[Dict[str, Any]] = []
    for proc in bom.get("procurements", []):
        # Capture the full set of IGCE fields from the original line — the
        # build_bom() output already aggregates per-CLIN, but we re-derive
        # basis/poc/notes from the line records which contain them.
        for line in proc.get("lines", []):
            rows.append({
                "procurement_id": proc.get("procurement_id", ""),
                "solicitation": proc.get("solicitation", ""),
                "initiative_code": proc.get("initiative_code", "") or "",
                "initiative_title": proc.get("initiative_title", "") or "",
                "tier": proc.get("tier", "") or "",
                "fiscal_year": proc.get("fiscal_year", "") or "",
                "agency": proc.get("agency", "") or "",
                "contract_type": proc.get("contract_type", "") or "",
                "procurement_status": proc.get("status", "") or "",
                "clin": line.get("clin", ""),
                "description": line.get("description", "") or "",
                "equipment_category": line.get("equipment_category", "") or "",
                "unit": line.get("unit", "") or "",
                "quantity": line.get("quantity", 0) or 0,
                "igce_unit_cost": line.get("igce_unit_cost", 0) or 0,
                "igce_extended_cost": line.get("igce_extended_cost", 0) or 0,
                "vendor_quote_count": line.get("vendor_quote_count", 0) or 0,
                "min_quote": line.get("min_quote"),
                "max_quote": line.get("max_quote"),
                "avg_quote": line.get("avg_quote"),
                "basis": line.get("basis", "") or "",
                "poc": line.get("poc", "") or "",
                "notes": line.get("notes", "") or "",
            })
    return rows


def export_bom_csv(bom: Dict[str, Any]) -> bytes:
    """Render a BOM rollup as a CSV byte payload for procurement submission.

    Uses ``utf-8-sig`` so Microsoft Excel opens the file with the correct
    encoding (no mojibake on accented characters). One row per CLIN with
    procurement metadata denormalised onto every row.
    """
    rows = _flatten_bom_rows(bom)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_CSV_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        # Round numeric fields for readability
        for k, v in list(row.items()):
            if isinstance(v, float):
                row[k] = round(v, 2)
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")


def export_bom_xlsx(bom: Dict[str, Any]) -> bytes:
    """Render a BOM rollup as an XLSX workbook (Summary + Lines + By-Category).

    Suitable for direct submission to a contracting officer — opens in
    Excel without any further formatting. Returns a ``bytes`` blob; write
    with ``open(path, "wb")`` or return as ``application/vnd.openxmlformats-…``.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:  # pragma: no cover - openpyxl is a hard dep
        raise RuntimeError(
            "openpyxl is required for XLSX export; install with `pip install openpyxl`"
        ) from e

    wb = openpyxl.Workbook()

    # --- Sheet 1: Lines (one row per CLIN) ---
    lines_ws = wb.active
    lines_ws.title = "BOM Lines"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    rows = _flatten_bom_rows(bom)
    lines_ws.append(_CSV_COLUMNS)
    for cell in lines_ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        lines_ws.append([
            round(row[col], 2) if isinstance(row[col], float) else row[col]
            for col in _CSV_COLUMNS
        ])

    # Format money columns
    money_cols = {"igce_unit_cost", "igce_extended_cost", "min_quote", "max_quote", "avg_quote"}
    for col_idx, col_name in enumerate(_CSV_COLUMNS, start=1):
        if col_name in money_cols:
            for row_idx in range(2, lines_ws.max_row + 1):
                cell = lines_ws.cell(row=row_idx, column=col_idx)
                cell.number_format = '"$"#,##0.00'
    # Autofit column widths
    for col_idx, col_name in enumerate(_CSV_COLUMNS, start=1):
        max_len = len(col_name)
        for row in rows:
            val = row.get(col_name, "")
            if val is not None:
                max_len = max(max_len, len(str(val)))
        lines_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 50)
    lines_ws.freeze_panes = "A2"

    # --- Sheet 2: Summary ---
    summary_ws = wb.create_sheet("Summary")
    summary_ws["A1"] = "BOM Summary"
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws["A2"] = f"Generated: {datetime.now(timezone.utc).isoformat()}"
    summary_ws["A3"] = (
        f"Filter: tier={bom.get('filter', {}).get('tier') or '-'}, "
        f"FY={bom.get('filter', {}).get('fiscal_year') or '-'}, "
        f"procurement_id={bom.get('filter', {}).get('procurement_id') or '-'}"
    )

    totals = bom.get("totals", {}) or {}
    row_idx = 5
    summary_ws.cell(row=row_idx, column=1, value="Metric").font = Font(bold=True)
    summary_ws.cell(row=row_idx, column=2, value="Value").font = Font(bold=True)
    row_idx += 1
    for label in ("procurement_count", "line_count", "igce_total",
                  "quote_min", "quote_max"):
        summary_ws.cell(row=row_idx, column=1, value=label)
        value = totals.get(label)
        if label in ("igce_total", "quote_min", "quote_max") and value is not None:
            cell = summary_ws.cell(row=row_idx, column=2, value=float(value))
            cell.number_format = '"$"#,##0.00'
        else:
            summary_ws.cell(row=row_idx, column=2, value=value)
        row_idx += 1

    # By-category breakdown
    row_idx += 1
    summary_ws.cell(row=row_idx, column=1, value="By Equipment Category").font = Font(bold=True, size=12)
    row_idx += 1
    summary_ws.cell(row=row_idx, column=1, value="Category").font = Font(bold=True)
    summary_ws.cell(row=row_idx, column=2, value="Line Count").font = Font(bold=True)
    summary_ws.cell(row=row_idx, column=3, value="IGCE Total").font = Font(bold=True)
    row_idx += 1
    by_cat = bom.get("by_category", {}) or {}
    for cat_name in sorted(by_cat.keys(), key=lambda k: -(by_cat[k].get("igce_total") or 0)):
        cat = by_cat[cat_name]
        summary_ws.cell(row=row_idx, column=1, value=cat_name)
        summary_ws.cell(row=row_idx, column=2, value=cat.get("line_count", 0))
        total_cell = summary_ws.cell(row=row_idx, column=3, value=float(cat.get("igce_total") or 0))
        total_cell.number_format = '"$"#,##0.00'
        row_idx += 1

    # By-tier breakdown
    row_idx += 1
    summary_ws.cell(row=row_idx, column=1, value="By Initiative Tier").font = Font(bold=True, size=12)
    row_idx += 1
    summary_ws.cell(row=row_idx, column=1, value="Tier").font = Font(bold=True)
    summary_ws.cell(row=row_idx, column=2, value="Procurement Count").font = Font(bold=True)
    summary_ws.cell(row=row_idx, column=3, value="IGCE Total").font = Font(bold=True)
    row_idx += 1
    by_tier = totals.get("by_tier", {}) or {}
    for tier_name in sorted(by_tier.keys()):
        tier = by_tier[tier_name]
        summary_ws.cell(row=row_idx, column=1, value=tier_name)
        summary_ws.cell(row=row_idx, column=2, value=tier.get("procurement_count", 0))
        total_cell = summary_ws.cell(row=row_idx, column=3, value=float(tier.get("igce_total") or 0))
        total_cell.number_format = '"$"#,##0.00'
        row_idx += 1

    summary_ws.column_dimensions["A"].width = 30
    summary_ws.column_dimensions["B"].width = 22
    summary_ws.column_dimensions["C"].width = 22

    # --- Sheet 3: By Category detail ---
    cat_ws = wb.create_sheet("By Category")
    cat_ws.append(["Category", "Tier", "Line Count", "IGCE Total"])
    for cell in cat_ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for cat_name, cat in by_cat.items():
        for tier_name, tb in (cat.get("tier_breakdown") or {}).items():
            cell = cat_ws.cell(row=cat_ws.max_row + 1, column=1, value=cat_name)
            cat_ws.cell(row=cat_ws.max_row, column=2, value=tier_name)
            cat_ws.cell(row=cat_ws.max_row, column=3, value=tb.get("line_count", 0))
            money = cat_ws.cell(row=cat_ws.max_row, column=4, value=float(tb.get("igce_total") or 0))
            money.number_format = '"$"#,##0.00'
    for col_idx in range(1, 5):
        cat_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22

    # Serialize
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_bom(bom: Dict[str, Any], format: str = "json") -> bytes | str:
    """Dispatch helper: ``format`` is one of ``json|csv|xlsx``.

    Returns ``bytes`` for csv/xlsx (write to disk as binary) and a JSON
    string for ``format="json"``.
    """
    fmt = (format or "json").lower()
    if fmt == "csv":
        return export_bom_csv(bom)
    if fmt == "xlsx":
        return export_bom_xlsx(bom)
    if fmt == "json":
        return json.dumps(bom, indent=2, default=str)
    raise ValueError(f"Unknown export format: {format!r}; expected csv|xlsx|json")


# ----------------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------------


def _main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="BOM generator for procurement initiatives")
    ap.add_argument("--bom", action="store_true", help="Generate the full BOM rollup")
    ap.add_argument("--procurement", help="Restrict to a single procurement_id")
    ap.add_argument("--tier", choices=["tier_1", "tier_2"], help="Filter by initiative tier")
    ap.add_argument("--fiscal-year", type=int, help="Filter by allocation fiscal year")
    ap.add_argument("--category", choices=CANONICAL_CATEGORIES, help="Restrict to a single equipment category in the BOM lines")
    ap.add_argument("--json", action="store_true", help="Emit machine-readable JSON (default when stdout)")
    ap.add_argument("--format", choices=["json", "csv", "xlsx"],
                    help="Export format suitable for direct procurement submission. "
                         "csv = utf-8-sig (Excel-friendly); xlsx = multi-sheet workbook.")
    ap.add_argument("--output", "-o", help="Write payload to this file path (required for binary formats)")
    args = ap.parse_args(argv)

    if not (args.bom or args.procurement):
        ap.error("Provide --bom or --procurement PROC-ID")

    if args.procurement:
        result = build_bom(procurement_id=args.procurement)
    else:
        result = build_bom(tier=args.tier, fiscal_year=args.fiscal_year)

    if args.category and result.get("procurements"):
        for proc in result["procurements"]:
            proc["lines"] = [ln for ln in proc["lines"] if ln["equipment_category"] == args.category]

    fmt = (args.format or ("xlsx" if args.output and args.output.lower().endswith(".xlsx")
                           else "csv" if args.output and args.output.lower().endswith(".csv")
                           else "json")).lower()

    if fmt == "json":
        if args.output:
            Path(args.output).write_text(json.dumps(result, indent=2, default=str), encoding="utf-8", newline="")
        elif args.json or not args.format:
            json.dump(result, sys.stdout, indent=2, default=str)
            sys.stdout.write("\n")
        else:
            # --format json with no --output — write JSON to stdout
            json.dump(result, sys.stdout, indent=2, default=str)
            sys.stdout.write("\n")
        return 0

    # Binary formats
    if not args.output:
        ap.error(f"--format {fmt} requires --output PATH (-o PATH)")

    payload = export_bom(result, format=fmt)
    if not isinstance(payload, bytes):  # pragma: no cover — defensive
        raise RuntimeError(f"export_bom returned {type(payload).__name__}; expected bytes for {fmt}")
    Path(args.output).write_bytes(payload)
    print(f"BOM exported to {args.output} ({fmt.upper()}, {len(payload):,} bytes)", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(_main())
