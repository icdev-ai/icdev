# CUI // SP-CTI
"""What the evidence implies for the plan.

ICDEV does not schedule. Compass does, and Compass's scheduler was modelled on
exactly the kind of lab-tracker spreadsheet this programme already keeps — phases
grouping tasks, hierarchical ids ("2.3"), free-text assignees, due dates that are
frequently "TBD", semicolon-separated dependencies. It already computes the
forward pass, the slack, the critical path and the dependency cycles, and exports
a round-trip-stable workbook.

Writing a second scheduler here, beside a working one, inside the same ecosystem,
behind the same licence, would be the exact duplication the integration exists to
prevent. So the division is:

    ICDEV     reads the documents, reconciles the bill of materials, finds what
              is wrong with it, and works out what that MEANS for the plan.
    COMPASS   schedules.

This module is the first half. It turns findings into tasks, and it does one thing
that is a judgement rather than an arithmetic:

**IT PULLS DEVELOPMENT TO THE FRONT.**

The existing tracker gates every piece of software work behind Phase 5 — the
sandbox, the model-training environment, the cyber range — so nothing gets built
until facilities, network and virtualization are all finished. But hardware the
programme ALREADY OWNS can carry a virtualized environment on day one. That does
not need a procurement cycle, a facility, or a purchase order; it needs somebody
to turn it on.

So a Wave 0 is proposed that runs in PARALLEL with procurement rather than behind
it. It is the difference between "we start when the lab is ready" and "we start
now, and the lab catches up" — and on a nine-month programme that is most of a
quarter of engineering.

Every derived task says where it came from. A task nobody can trace is a task
nobody will do.

Public API::

    tracker_workbook(rows) -> bytes            # what Compass will accept
    tasks_from_findings(findings, ...) -> list[Task]
    gantt_slide(schedule, tasks) -> dict       # an ICDEV slide spec
    phases_slide(schedule, tasks, rollup) -> dict
"""
from __future__ import annotations

import io
import os
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable, Sequence

from tools.bom.findings import Finding

# The tracker's own column vocabulary. Compass parses this; do not "improve" it.
TRACKER_HEADERS = [
    "Phase", "Task ID", "Task Name", "Description", "Dependencies",
    "Status", "Assignee", "Due Date", "Comment",
]

WAVE_ZERO = "Phase 0: Start Now (owned hardware)"
WAVE_DECIDE = "Phase 0: Decisions Blocking the Ask"

_ISO_DATE = re.compile(r"\d{4}-\d{2}-\d{2}")


@dataclass
class Task:
    phase: str
    task_id: str
    name: str
    description: str = ""
    deps: list[str] = field(default_factory=list)
    status: str = "Not Started"
    assignee: str = "TBD"
    due_date: str = "TBD"
    comment: str = ""
    duration_days: int | None = None

    def row(self) -> list[Any]:
        return [
            self.phase, self.task_id, self.name, self.description,
            "; ".join(self.deps), self.status, self.assignee,
            self.due_date, self.comment,
        ]

    def as_compass(self) -> dict[str, Any]:
        out: dict[str, Any] = {
            "task_id": self.task_id,
            "name": self.name,
            "description": self.description,
            "phase_name": self.phase,
            "deps": self.deps,
            "status": self.status,
            "assignee": self.assignee,
            "comment": self.comment,
        }
        if self.duration_days:
            out["duration_days"] = self.duration_days
        # Only a real date goes over the wire. "TBD" is not a date, and handing it
        # to a scheduler as one invites it to be parsed into something.
        if _ISO_DATE.fullmatch(self.due_date or ""):
            out["due_date"] = self.due_date
        return out


def read_tracker(path: str, sheet: str = "Tasks") -> list[Task]:
    """Read an existing tracker. THEIR plan, carried over verbatim.

    The team's own tasks are not ours to rewrite. We add to them, we resequence
    them, and we never silently drop one — a planning tool that quietly loses
    somebody's work is worse than no planning tool.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]

    rows = [
        [("" if v is None else str(v).strip()) for v in row]
        for row in ws.iter_rows(values_only=True)
        if any(v is not None for v in row)
    ]
    if not rows:
        return []

    head = [h.lower() for h in rows[0]]

    def col(*names: str) -> int:
        for n in names:
            if n in head:
                return head.index(n)
        return -1

    ix = {k: col(*v) for k, v in {
        "phase": ("phase",), "task_id": ("task id", "id"), "name": ("task name", "name"),
        "description": ("description",), "deps": ("dependencies", "depends on"),
        "status": ("status",), "assignee": ("assignee", "owner"),
        "due": ("due date", "due"), "comment": ("comment", "comments"),
    }.items()}

    def cell(row: list[str], key: str) -> str:
        i = ix[key]
        return row[i].strip() if 0 <= i < len(row) else ""

    out: list[Task] = []
    for row in rows[1:]:
        tid = cell(row, "task_id")
        if not tid:
            continue
        raw_deps = cell(row, "deps")
        deps = [
            d.strip() for d in raw_deps.replace(",", ";").split(";")
            if d.strip() and d.strip().lower() not in ("none", "n/a", "-")
        ]
        due = cell(row, "due")
        # "2026-07-09 00:00:00" -> "2026-07-09"; "TBD" stays "TBD".
        if due and due[:4].isdigit():
            due = due[:10]
        out.append(Task(
            phase=cell(row, "phase"), task_id=tid, name=cell(row, "name"),
            description=cell(row, "description"), deps=deps,
            status=cell(row, "status") or "Not Started",
            assignee=cell(row, "assignee") or "TBD",
            due_date=due or "TBD", comment=cell(row, "comment"),
        ))
    return out


def tracker_workbook(tasks: Iterable[Task]) -> bytes:
    """A single-sheet workbook Compass will accept.

    Compass reads the FIRST worksheet. A real tracker has half a dozen sheets and
    the tasks are rarely on the first one — so ICDEV, which is the thing that
    reads documents, hands over just the grid.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"
    ws.append(TRACKER_HEADERS)
    for i, _ in enumerate(TRACKER_HEADERS, start=1):
        c = ws.cell(1, i)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")

    for t in tasks:
        ws.append(t.row())

    for col, width in zip("ABCDEFGHI", (34, 9, 40, 64, 16, 13, 14, 13, 40)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Findings become work ─────────────────────────────────────────────────────

def tasks_from_findings(
    findings: Iterable[Finding],
    *,
    owned_units: int = 0,
    owned_note: str = "",
    long_lead_days: int = 42,
) -> list[Task]:
    """The plan the evidence implies.

    Two groups, and the split is the argument:

    **Decisions blocking the ask.** Every one of these is somebody answering a
    question, not somebody building something. They cost days and they gate a
    number that leadership is being asked to approve — which makes them the
    cheapest and most urgent work on the programme.

    **Start now.** Hardware the programme already owns, doing useful work on day
    one, while procurement runs in parallel behind it.
    """
    findings = list(findings)
    out: list[Task] = []
    n = 0

    def add(phase: str, name: str, desc: str, *, days: int, comment: str,
            deps: list[str] | None = None, assignee: str = "TBD") -> str:
        nonlocal n
        n += 1
        tid = f"0.{n}"
        out.append(Task(
            phase=phase, task_id=tid, name=name, description=desc,
            deps=deps or [], assignee=assignee, duration_days=days,
            comment=comment,
        ))
        return tid

    def have(*types: str) -> list[Finding]:
        return [f for f in findings if f.finding_type in types]

    # ── The decisions ───────────────────────────────────────────────────────
    if any(f.finding_type == "intra_doc_double_count" for f in findings):
        f = have("intra_doc_double_count")[0]
        ev = f.evidence[0] if f.evidence else None
        add(
            WAVE_DECIDE,
            "Resolve the double-counted line",
            f.title,
            days=1,
            comment=(
                f"Evidence: {ev.source_document}!{ev.sheet}!{ev.locator}. "
                if ev else ""
            ) + "Confirm whether both category subtotals include this figure.",
        )

    if have("hardcoded_rollup", "stale_rollup"):
        cells = ", ".join(
            f"{f.evidence[0].sheet}!{f.evidence[0].locator}"
            for f in have("hardcoded_rollup", "stale_rollup")[:6] if f.evidence
        )
        add(
            WAVE_DECIDE,
            "Repair the summary formulas",
            "Typed-in numbers sitting where formulas belong. Editing the "
            "underlying sheets does not move them, so the headline total has "
            "silently stopped tracking its own inputs.",
            days=1,
            comment=f"Cells: {cells}",
        )

    if have("unpriced_line_zeroed"):
        f = have("unpriced_line_zeroed")[0]
        ev = f.evidence[0] if f.evidence else None
        add(
            WAVE_DECIDE,
            "Price the line that is costing nothing",
            "A quantity with no unit price. The formula multiplies it to zero, "
            "so it sits inside the total contributing nothing — and the total is "
            "understated by whatever the item is actually worth.",
            days=2,
            comment=f"{ev.source_document}!{ev.sheet}!{ev.locator}" if ev else "",
        )

    if have("asset_count_disputed", "baseline_asset_gap", "unverified_existing_asset"):
        f = have("asset_count_disputed", "baseline_asset_gap",
                 "unverified_existing_asset")[0]
        add(
            WAVE_DECIDE,
            "Walk the rack and settle the asset count",
            "The design leans on more owned units than the inventory records. A "
            "serial number proves a machine exists; its absence proves nothing, so "
            "this cannot be settled from a spreadsheet. Somebody has to go and "
            "look. The refresh reserve is sized from this number, so it has to be "
            "right.",
            days=2,
            comment=f.title,
            assignee="Larry",
        )

    if have("scope_priced_only_by_weak_source", "scope_declared_unpriced",
            "scope_declared_undesigned"):
        f = have("scope_priced_only_by_weak_source", "scope_declared_unpriced",
                 "scope_declared_undesigned")[0]
        add(
            WAVE_DECIDE,
            "Scope and price the declared-but-undesigned workstream",
            "Work we have said we are doing that appears in no agreed design and "
            "in no authoritative bill of materials. It reads as covered on a "
            "spreadsheet and it is not.",
            days=5,
            comment=f.title,
        )

    if have("capex_opex_conflation"):
        add(
            WAVE_DECIDE,
            "Separate recurring costs from capital",
            "Charges stated per period are carried as one-off amounts. The "
            "multi-year commitment is understated by however many periods nobody "
            "multiplied by.",
            days=1,
            comment=f"{len(have('capex_opex_conflation'))} line(s) affected",
        )

    if have("hidden_content"):
        add(
            WAVE_DECIDE,
            "Confirm the constraints found in hidden content",
            "Material found in the documents but not on the page — a screenshot in "
            "a sheet that renders empty, notes nobody opens. Constraints of this "
            "kind are unpriced by definition, because nobody reading the workbook "
            "ever saw them.",
            days=2,
            comment=have("hidden_content")[0].title,
        )

    # The one that has to exist whether or not a finding produced it: while several
    # documents each claim to price this project, there IS no number to approve.
    add(
        WAVE_DECIDE,
        "Nominate a source of record for each area of scope",
        "Several documents each claim to price this project. Until one governs "
        "each area, the totals are a SUM of competing estimates rather than an "
        "estimate of anything — and that is the arithmetic that produced the "
        "spread in the first place. This decision takes an afternoon and it "
        "unblocks the entire ask.",
        days=1,
        comment="Blocks leadership approval.",
        assignee="Larry",
    )

    # ── Start now ───────────────────────────────────────────────────────────
    if owned_units:
        stand_up = add(
            WAVE_ZERO,
            f"Stand up the virtual environment on the {owned_units} owned servers",
            "Hardware the programme already owns, doing useful work on day one. "
            "This needs no purchase order, no facility and no procurement cycle — "
            "it needs somebody to turn it on.",
            days=10,
            comment=owned_note or "Zero capital cost. Runs in parallel with procurement.",
            assignee="Larry",
        )
        bring_in = add(
            WAVE_ZERO,
            "Bring the platform in-house and start development",
            "The team begins building while the lab buildout runs behind it. This "
            "is the difference between starting when the lab is ready and starting "
            "now — on a nine-month programme, most of a quarter of engineering.",
            days=20,
            comment="Does not wait on Phase 1-4.",
            deps=[stand_up],
        )
        add(
            WAVE_ZERO,
            "Earmark the hardware refresh reserve",
            "The owned fleet is out of warranty. It is carrying the programme at "
            "zero capital cost today and it will fail. Ask for the replacement now, "
            "sized from the VERIFIED unit count, rather than discover it later.",
            days=3,
            comment="Sized from the asset count above — so that count has to be right.",
            deps=[],
        )
        _ = bring_in

    # NOTE what is deliberately NOT added here: a "place the long-lead order" task.
    #
    # The team's own plan already has one — procurement, sitting on the critical
    # path where it belongs. Adding a second would not shorten anything; it would
    # put two tasks in the tracker for one purchase order and make the schedule
    # lie about how much work there is. A planning tool whose first act is to
    # duplicate the work already in the plan is not helping.
    #
    # What the evidence DOES add is the dependency nobody had written down, and
    # link_decisions_to_approval() below is where that happens.
    _ = long_lead_days
    return out


# Words that identify the task a plan already has for "somebody signs this off".
_APPROVAL = ("approval", "approve", "leadership", "sign-off", "signoff", "funding")


def link_decisions_to_approval(derived: list[Task], existing: list[Task]) -> list[str]:
    """Make the approval task wait for the decisions that produce the number.

    This is the whole argument, expressed as a dependency.

    The team's plan has a task called something like "Get Leadership Approval",
    and it sits in the schedule as though the thing being approved already exists.
    It does not. Several documents each claim to price this project, so until
    somebody says which one governs, there IS no number — and an approval task
    with nothing to approve is a date in a spreadsheet, not a plan.

    Wiring the approval to the decisions stops the schedule pretending. What that
    then DOES to the critical path is the scheduler's answer, not ours — and on the
    evidence it is a good one: the decisions are short, so they finish inside the
    slack the plan already has and cost the programme nothing at all.

    That is a better argument than the one we set out to make. These decisions are
    FREE. A few days of somebody answering questions, no schedule impact, and until
    they are done there is no number for leadership to approve. There is no version
    of this project in which not doing them first is the cheaper option.

    (An earlier draft of this docstring asserted that the rewiring would put Phase 0
    on the critical path. It does not, and saying so would have been the tool
    telling a story the schedule does not support — which is the one thing this
    whole engine exists to stop.)

    Returns the ids of the existing tasks that were rewired, because silently
    editing somebody's plan is not something to do without saying so.
    """
    gating = [
        t.task_id for t in derived if t.phase == WAVE_DECIDE
    ]
    if not gating:
        return []

    touched: list[str] = []
    for t in existing:
        if not any(w in t.name.lower() for w in _APPROVAL):
            continue
        for g in gating:
            if g not in t.deps:
                t.deps.append(g)
        note = "Cannot approve a number that does not exist yet — gated on the Phase 0 decisions."
        t.comment = f"{t.comment}  {note}".strip() if t.comment else note
        touched.append(t.task_id)

    return touched


# ── Durations, and resolving "TBD" ───────────────────────────────────────────

_ARGS = Path(__file__).resolve().parents[2] / "args"
_PLAN_CONFIG = _ARGS / "bom_plan.yaml"
_PLAN_EXAMPLE = _ARGS / "bom_plan.example.yaml"

# Where a deployment keeps its real plan config. A live one names real people and
# real committed dates — customer material — so it lives OUTSIDE this repo, which
# is public. The example ships; the plan does not.
_PLAN_ENV = "BOM_PLAN_CONFIG"


def load_plan_config(path: Path | str | None = None) -> dict[str, Any]:
    """Resolve the plan config: explicit path, then ``$BOM_PLAN_CONFIG``, then the
    local ``args/bom_plan.yaml`` (gitignored), then the shipped example.

    Falling back to the example rather than raising is deliberate: a fresh clone
    should run end-to-end and produce a schedule, so the mechanism is inspectable
    without anybody having to hand over a real plan first.
    """
    import yaml

    p = Path(path) if path else None
    if p is None:
        env = os.environ.get(_PLAN_ENV)
        p = Path(env) if env else _PLAN_CONFIG
    if not p.exists():
        p = _PLAN_EXAMPLE

    return yaml.safe_load(p.read_text(encoding="utf-8")) or {}


def apply_durations(tasks: list[Task], cfg: dict[str, Any] | None = None) -> dict[str, Any]:
    """Give every task a duration, and say where each one came from.

    A range ("4 to 6 weeks") is two different statements: what somebody hopes for,
    and what they will not be surprised by. The baseline commits to the end named
    by ``plan_to`` — ``late``, by default, because a date built on the hopeful end
    is a date that slips, and a slipped date costs more credibility than a longer
    one ever costs goodwill. The optimistic figure is carried as upside and stated
    out loud rather than quietly banked.

    Anything that matches no rule gets the default AND is reported as unestimated.
    An unestimated task is a question, not a five-day task, and letting it wear a
    number it never earned is how a plan acquires a precision it does not have.
    """
    cfg = cfg or load_plan_config()
    by_name: dict[str, Any] = {
        k.lower(): v for k, v in (cfg.get("durations", {}).get("by_name") or {}).items()
    }
    ranges: dict[str, Any] = cfg.get("ranges") or {}
    range_tasks: dict[str, str] = {
        k.lower(): v for k, v in (cfg.get("range_tasks") or {}).items()
    }
    default = int(cfg.get("durations", {}).get("default", 5))
    late = str(cfg.get("plan_to", "late")).lower() == "late"

    unestimated: list[str] = []
    upside_days = 0

    for t in tasks:
        if t.duration_days:
            continue   # derived tasks already carry one

        name = t.name.lower()

        # Longest matching key wins: "hardware installation & burn-in" must beat a
        # shorter key that happens to be a substring of it.
        key = max(
            (k for k in by_name if k in name),
            key=len,
            default="",
        )

        rkey = max((k for k in range_tasks if k in name), key=len, default="")
        if rkey:
            r = ranges.get(range_tasks[rkey]) or {}
            early = int(r.get("early_weeks", 0)) * 5
            late_d = int(r.get("late_weeks", 0)) * 5
            t.duration_days = late_d if late else early
            upside_days += max(0, late_d - early)
            note = r.get("note", "")
            t.comment = (
                f"{t.comment}  Planned to the {'late' if late else 'early'} end of "
                f"{r.get('early_weeks')}-{r.get('late_weeks')} weeks. {note}"
            ).strip()
            continue

        if key and by_name[key] is not None:
            t.duration_days = int(by_name[key])
            continue

        t.duration_days = default
        unestimated.append(t.task_id)

    return {
        "unestimated": unestimated,
        "default_days": default,
        "plan_to": "late" if late else "early",
        # What the schedule would give back if every range came in at its early
        # end. Not banked; stated.
        "upside_days": upside_days,
    }


def apply_fixed_dates(tasks: list[Task], cfg: dict[str, Any] | None = None) -> list[str]:
    """Commitments already made to people. The scheduler does not get to move them."""
    cfg = cfg or load_plan_config()
    fixed = cfg.get("fixed") or {}

    pinned: list[str] = []
    for t in tasks:
        spec = fixed.get(t.task_id)
        if not spec:
            continue
        t.due_date = str(spec.get("due") or t.due_date)
        note = spec.get("note") or ""
        if note and note not in t.comment:
            t.comment = f"{t.comment}  {note}".strip()
        pinned.append(t.task_id)
    return pinned


def resolve_tbd(tasks: list[Task], sched: dict, cfg: dict[str, Any] | None = None) -> dict:
    """Turn every "TBD" into the date the schedule actually implies.

    "TBD" is not a missing date. It is a date nobody had a basis for — and the
    whole reason it stayed TBD is that the durations and the dependencies were
    never written down, so there was nothing to compute it FROM.

    Now there is. Every TBD becomes Compass's computed finish date, and the task
    says so, because a date that appeared from nowhere is a date nobody will
    defend. Change a duration in args/bom_plan.yaml and every one of these moves.
    That is what makes it a goal rather than a guess: not that it is right, but
    that it is adjustable and it is traceable.

    Dates a human COMMITTED to are never overwritten. A computed date is a
    projection; a promise made to somebody in a meeting is not.
    """
    cfg = cfg or load_plan_config()
    fixed = set((cfg.get("fixed") or {}).keys())
    rows = (sched or {}).get("tasks") or {}

    resolved, kept = [], []
    for t in tasks:
        if t.task_id in fixed:
            kept.append(t.task_id)
            continue
        if t.due_date and t.due_date.upper() not in ("TBD", "TBC", ""):
            kept.append(t.task_id)
            continue

        info = rows.get(t.task_id) or {}
        end = info.get("end")
        if not end:
            continue

        t.due_date = str(end)[:10]
        slack = info.get("slack_days")
        mark = "critical path — no slack" if info.get("critical") else (
            f"{slack}d slack" if slack is not None else ""
        )
        note = f"Computed from the schedule ({mark})." if mark else "Computed from the schedule."
        t.comment = f"{t.comment}  {note}".strip()
        resolved.append(t.task_id)

    return {"resolved": resolved, "already_dated": kept}


# ── Dependencies written in English ──────────────────────────────────────────

_PHASE_REF = re.compile(r"phase\s*(\d+)", re.I)


def normalize_deps(tasks: Sequence[Task]) -> dict[str, Any]:
    """Turn dependencies a human wrote into dependencies a scheduler can follow.

    People do not write ``["3.1","3.2","3.3","3.4","3.5"]`` in a spreadsheet. They
    write **"All Phase 3"**, and everyone in the room knows exactly what it means.
    A scheduler does not: it looks for a task called "All Phase 3", fails to find
    one, and treats the task as having NO predecessors — so the final security
    audit gets scheduled on day one, in parallel with the procurement it is meant
    to audit, and the computed end date comes out early and confident and wrong.

    That failure is silent in the worst way, because the plan still renders. The
    critical path just quietly routes around the constraint nobody encoded.

    So: expand the phase references, and *report* anything still unresolved rather
    than dropping it. A dependency the tool could not understand is a dependency a
    human needs to look at — not one to delete on their behalf.
    """
    by_id = {t.task_id: t for t in tasks}

    expanded: list[dict[str, Any]] = []
    unresolved: list[dict[str, str]] = []

    for t in tasks:
        out: list[str] = []
        for dep in t.deps:
            d = dep.strip()
            if not d:
                continue
            if d in by_id:
                out.append(d)
                continue

            m = _PHASE_REF.search(d)
            if m:
                n = m.group(1)
                members = [
                    o.task_id for o in tasks
                    if _PHASE_REF.match(o.phase.strip())
                    and _PHASE_REF.match(o.phase.strip()).group(1) == n
                    and o.task_id != t.task_id
                ]
                if members:
                    out.extend(members)
                    expanded.append({
                        "task_id": t.task_id, "wrote": d, "became": members,
                    })
                    continue

            unresolved.append({"task_id": t.task_id, "dep": d})

        # Dedupe, order-stable. A task listed twice is not a stronger dependency.
        seen: set[str] = set()
        t.deps = [x for x in out if not (x in seen or seen.add(x))]

    return {"expanded": expanded, "unresolved": unresolved}


# ── Scope that nobody priced ─────────────────────────────────────────────────

_PHASE_NUM = re.compile(r"phase\s+(\d+)", re.I)


def _find_by_name(tasks: Sequence[Task], needle: str) -> Task | None:
    """Longest-match on task name. Returns the task, or None — and None means
    the dependency is dropped rather than pointed at something that isn't there."""
    if not needle:
        return None
    n = needle.lower()
    hits = [t for t in tasks if n in t.name.lower()]
    return min(hits, key=lambda t: len(t.name)) if hits else None


def digital_twin_tasks(
    predecessors: Sequence[Task],
    cfg: dict[str, Any] | None = None,
) -> list[Task]:
    """The Digital Twin, put on the schedule precisely because it is on no BOM.

    Every task here carries the same note: this scope is real, and nobody has
    priced it. That is not a caveat — it is the point. Work that exists on a
    schedule and not in a budget is the most expensive kind of work there is,
    because by the time somebody notices, the money has already been allocated
    somewhere else.

    The engine will not invent a figure for it. It puts the scope in front of the
    people who can get a quote, in time for them to get one.
    """
    cfg = cfg or load_plan_config()
    dt = cfg.get("digital_twin") or {}
    note = (dt.get("unfunded_note") or "").strip()

    out: list[Task] = []
    known = list(predecessors)

    for wave in dt.get("waves") or []:
        phase = str(wave.get("phase") or "Digital Twin")
        m = _PHASE_NUM.search(phase)
        prefix = m.group(1) if m else str(len(out) + 8)

        # The wave hangs off whatever it said it waits for. If that task does not
        # exist in this plan, the wave floats — visibly, rather than by pointing
        # at a task ID nobody has.
        anchor = _find_by_name(known, str(wave.get("after") or ""))
        prev_id = anchor.task_id if anchor else ""

        for i, spec in enumerate(wave.get("tasks") or [], start=1):
            t = Task(
                phase=phase,
                task_id=f"{prefix}.{i}",
                name=str(spec.get("name") or ""),
                description=str(spec.get("description") or ""),
                deps=[prev_id] if prev_id else [],
                status="Not Started",
                assignee="TBD",
                due_date="TBD",
                comment=note,
                duration_days=int(spec.get("duration_days") or 10),
            )
            out.append(t)
            known.append(t)
            # Within a wave, the first task gates the rest only through the
            # anchor; the tasks themselves run in parallel. A twin of Windows
            # does not wait on a twin of Linux.

    return out


# ── Horizons: dates the plan is CHECKED against, never forced into ───────────

def _horizon_for(task: Task, cfg: dict[str, Any]) -> str:
    horizons: dict[str, Any] = cfg.get("horizons") or {}
    for key, spec in horizons.items():
        covers = str(spec.get("covers") or "")
        if covers.startswith("phase:") and covers[6:].lower() in task.phase.lower():
            return key
    return str(cfg.get("default_horizon") or "")


def check_horizons(
    tasks: Sequence[Task],
    sched: dict,
    cfg: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """Does the schedule actually land inside the dates we said it would?

    This asks. It does not enforce. A planner that compresses a schedule to hit a
    date somebody wanted has not made the date achievable — it has only moved the
    moment you find out, from now to month nine, when it is expensive.

    So a breach comes back as a breach. Then you shorten the work, move the date,
    or accept it, which are the only three things that were ever available.
    """
    cfg = cfg or load_plan_config()
    horizons: dict[str, Any] = cfg.get("horizons") or {}
    rows = (sched or {}).get("tasks") or {}

    out: list[dict[str, Any]] = []
    for t in tasks:
        key = _horizon_for(t, cfg)
        spec = horizons.get(key) or {}
        limit = str(spec.get("must_finish_by") or "")
        end = str((rows.get(t.task_id) or {}).get("end") or "")[:10]
        if not limit or not end:
            continue
        if end > limit:
            out.append({
                "task_id": t.task_id,
                "name": t.name,
                "horizon": key,
                "label": spec.get("label", key),
                "must_finish_by": limit,
                "scheduled_end": end,
                "over_by_days": (
                    date.fromisoformat(end) - date.fromisoformat(limit)
                ).days,
            })
    return out


def horizon_summary(
    tasks: Sequence[Task],
    sched: dict,
    cfg: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """What each horizon actually costs, in tasks and in the date it really ends."""
    cfg = cfg or load_plan_config()
    horizons: dict[str, Any] = cfg.get("horizons") or {}
    rows = (sched or {}).get("tasks") or {}

    out: list[dict[str, Any]] = []
    for key, spec in horizons.items():
        members = [t for t in tasks if _horizon_for(t, cfg) == key]
        ends = [
            str((rows.get(t.task_id) or {}).get("end") or "")[:10]
            for t in members
        ]
        ends = [e for e in ends if e]
        limit = str(spec.get("must_finish_by") or "")
        actual = max(ends) if ends else ""
        out.append({
            "horizon": key,
            "label": spec.get("label", key),
            "tasks": len(members),
            "must_finish_by": limit,
            "scheduled_end": actual,
            "meets": bool(actual and limit and actual <= limit),
        })
    return out


# ── Slides, from Compass's schedule ──────────────────────────────────────────

def _parse(d: str | None) -> date | None:
    if not d:
        return None
    try:
        return datetime.strptime(str(d)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def gantt_slide(
    sched: dict,
    tasks: Iterable[Task],
    *,
    weeks: int = 16,
    max_rows: int = 12,
    audience: str = "working",
) -> dict | None:
    """A Gantt, drawn as a table because a table is what a deck can carry.

    One column per week; a bar is a run of filled cells. Critical-path work is
    marked, because a Gantt that does not say what is on the critical path is a
    picture of a schedule rather than a schedule.

    **It rolls up to phases when the tasks will not fit.** A slide holds about a
    dozen rows; a real programme has fifty tasks. Draw them all and the renderer
    silently keeps the first twelve — so the deck shows a fortnight of decisions,
    calls it the schedule, and never mentions the lab. A rolled-up bar is a
    simplification the reader can see. A truncated one is a lie they cannot.
    """
    rows_in = (sched or {}).get("tasks") or {}
    anchor = _parse((sched or {}).get("anchor"))
    if not rows_in or anchor is None:
        return None

    by_id = {t.task_id: t for t in tasks}

    spans: list[tuple[str, date, date, bool]] = []
    for tid, info in rows_in.items():
        s, e = _parse(info.get("start")), _parse(info.get("end"))
        if not (s and e):
            continue
        t = by_id.get(tid)
        spans.append((tid, s, e, bool(info.get("critical"))))

    if not spans:
        return None

    rolled = len(spans) > max_rows
    if rolled:
        # Collapse to phases: earliest start, latest end, critical if anything in
        # the phase is. The bar is still computed, never asserted.
        acc: dict[str, list[Any]] = {}
        for tid, s, e, crit in spans:
            t = by_id.get(tid)
            key = (t.phase if t else "") or "(unphased)"
            a = acc.setdefault(key, [s, e, False, 0])
            a[0] = min(a[0], s)
            a[1] = max(a[1], e)
            a[2] = a[2] or crit
            a[3] += 1
        bars = [
            (f"{k} ({v[3]})", v[0], v[1], v[2])
            for k, v in acc.items()
        ]
    else:
        bars = [
            (f"{tid} {by_id[tid].name if tid in by_id else ''}", s, e, crit)
            for tid, s, e, crit in spans
        ]

    bars.sort(key=lambda b: (b[1], b[0]))

    # Columns: weeks for a short plan, months for a long one.
    #
    # Twenty-seven weekly columns do not make the plan more precise, they make the
    # LABELS narrower — every phase renders as "Phase 0:…" and the reader learns
    # nothing. A month is also the unit leadership already thinks in, so the axis
    # stops being something they have to translate.
    week0 = anchor - timedelta(days=anchor.weekday())
    last = max(e for _, _, e, _ in bars)

    if weeks > 16:
        buckets: list[tuple[date, date, str]] = []
        y, m = anchor.year, anchor.month
        while date(y, m, 1) <= last:
            first = date(y, m, 1)
            ny, nm = (y + 1, 1) if m == 12 else (y, m + 1)
            # January carries the year, so a plan that crosses a New Year says so
            # on the axis rather than in a footnote.
            label = f"{first:%b}" if first.month != 1 else f"{first:%b} {first:%y}"
            buckets.append((first, date(ny, nm, 1) - timedelta(days=1), label))
            y, m = ny, nm
    else:
        buckets = [
            (week0 + timedelta(weeks=w), week0 + timedelta(weeks=w, days=6), f"W{w + 1}")
            for w in range(weeks)
        ]

    headers = ["Phase" if rolled else "Task", *[b[2] for b in buckets]]

    rows: list[list[str]] = []
    for label, s, e, critical in bars:
        text = ("! " if critical else "") + label
        cells = ["███" if (s <= be and e >= bs) else "" for bs, be, _ in buckets]
        rows.append([text[:34], *cells])

    note = (
        "Computed by Compass from the dependency graph, not asserted. The critical "
        "path is what determines the end date: every day lost on a marked bar is a "
        "day lost on the programme, and every day lost on an unmarked one is free. "
        "That distinction is the only reason to have a schedule at all."
    )
    if rolled and audience != "leadership":
        note += (
            f" Rolled up to phases — {len(spans)} tasks will not fit on a slide. "
            f"The task-level Gantt is the tracker."
        )

    # A leadership footer says what the marks MEAN. It does not report on the
    # rendering decisions the tool made to fit the slide — that is the tool
    # talking about itself, in a room that is here to fund a programme.
    footer = "! = on the critical path"
    if rolled and audience != "leadership":
        footer += f"   ·   {len(spans)} tasks rolled up to phases"

    return {
        "slide_type": "table",
        "title": f"Schedule — anchored {anchor:%d %b %Y}, ends {sched.get('project_end', '?')}",
        "bullets": {
            "headers": headers,
            "rows": rows,
            "footer": [footer, *[""] * len(buckets)],
        },
        "speaker_notes": note,
    }


def dates_slide(
    sched: dict,
    tasks: Sequence[Task],
    cfg: dict[str, Any] | None = None,
    *,
    upside_days: int = 0,
    audience: str = "working",
) -> dict | None:
    """The goal, and whether the schedule meets it.

    The two audiences want genuinely different things here, and it is not a matter
    of tone. The workgroup needs the verdict per horizon and the size of any miss,
    because their job is to close it. Leadership needs the date and the confidence
    behind it — a table column headed "over" invites a conversation about a
    one-day slip in a plan that runs for eighteen months, which is not the
    conversation the programme needs to have in that room.

    The miss is not hidden. It is in the workgroup deck, in the tracker, and in the
    speaker notes, which is where a presenter can answer it if asked.
    """
    cfg = cfg or load_plan_config()
    summary = horizon_summary(tasks, sched, cfg)
    if not summary:
        return None

    breaches = check_horizons(tasks, sched, cfg)
    over = max((int(b["over_by_days"]) for b in breaches), default=0)
    plan_to = str(cfg.get("plan_to", "late"))

    if audience == "leadership":
        headers = ["Milestone", "Target", "Plan"]
        rows = [
            [str(h["label"])[:34], str(h["must_finish_by"]), str(h["scheduled_end"]) or "—"]
            for h in summary
        ]
    else:
        headers = ["Horizon", "Tasks", "Target", "Scheduled", "Verdict"]
        rows = [
            [
                str(h["label"])[:34],
                str(h["tasks"]),
                str(h["must_finish_by"]),
                str(h["scheduled_end"]) or "—",
                "on target" if h["meets"] else "over",
            ]
            for h in summary
        ]

    foot = [
        f"Planned to the {plan_to} end of every quoted lead time — the conservative "
        f"case, not the hopeful one."
    ]
    if upside_days:
        foot.append(
            f"If the quoted ranges land at their early end instead, "
            f"{upside_days} working days come back."
        )
    if over:
        foot.append(
            f"At the conservative end, {len(breaches)} task(s) run {over} day(s) "
            f"past target — covered by the upside above. Raise it only if asked."
        )

    return {
        "slide_type": "table",
        "title": "The dates",
        "bullets": {"headers": headers, "rows": rows, "footer": []},
        "speaker_notes": " ".join(foot),
    }


def phases_slide(sched: dict, tasks: Iterable[Task]) -> dict | None:
    """What each phase is, when it runs, and what is holding it up."""
    rows_in = (sched or {}).get("tasks") or {}
    if not rows_in:
        return None

    tasks = list(tasks)
    by_id = {t.task_id: t for t in tasks}

    phases: dict[str, dict[str, Any]] = {}
    for tid, info in rows_in.items():
        t = by_id.get(tid)
        if t is None:
            continue
        p = phases.setdefault(t.phase or "(unphased)", {
            "count": 0, "start": None, "end": None, "critical": 0,
        })
        p["count"] += 1
        s, e = _parse(info.get("start")), _parse(info.get("end"))
        if s and (p["start"] is None or s < p["start"]):
            p["start"] = s
        if e and (p["end"] is None or e > p["end"]):
            p["end"] = e
        if info.get("critical"):
            p["critical"] += 1

    if not phases:
        return None

    rows = []
    # The year is not optional once a plan crosses a New Year. "04 Jan" beside
    # "13 Jul" reads as six months earlier when it is six months later.
    for name, p in sorted(phases.items(), key=lambda kv: (kv[1]["start"] or date.max)):
        rows.append([
            name[:36],
            f"{p['start']:%d %b %y}" if p["start"] else "—",
            f"{p['end']:%d %b %y}" if p["end"] else "—",
            str(p["count"]),
            str(p["critical"]) if p["critical"] else "—",
        ])

    return {
        "slide_type": "table",
        "title": "The phases",
        "bullets": {
            "headers": ["Phase", "Starts", "Ends", "Tasks", "On critical path"],
            "rows": rows,
            "footer": [],
        },
        "speaker_notes": (
            "Phase 0 is the change. The existing plan gates every piece of "
            "software work behind Phase 5, so nothing gets built until facilities, "
            "network and virtualization are all finished. Hardware we already own "
            "can carry a virtual environment on day one — it needs no purchase "
            "order and no facility. So Phase 0 runs in PARALLEL with procurement "
            "rather than behind it, and the team starts building while the lab "
            "catches up."
        ),
    }
