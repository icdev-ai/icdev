# CUI // SP-CTI
"""The BOM in the shape the customer already reads.

`export_xlsx` produces the ENGINE's workbook: every line, every finding, every
source, full provenance. That is the audit trail, and a reviewer needs it.

It is not what gets circulated. What gets circulated looks like the workbook the
organisation already uses — a banner sheet per category, numbered subsections, and
the six columns everybody's eye already knows where to find. Handing a team a
differently-shaped spreadsheet and asking them to re-learn where the price column
is, is how a correct answer gets ignored.

So this module renders the reconciled data into THEIR layout. The structure is
config (`args/bom_xlsx_layout.yaml`); nothing here knows what a lab is.

**And it fixes the defect their template shipped with.** The workbook this layout
was taken from had a Summary sheet whose category totals were TYPED IN — four of
them — so editing a category sheet moved nothing and the headline figure had
silently stopped tracking its own inputs. Here every rollup is a real
``=SUM()`` across the sheet it claims to summarise. Change a quantity and the total
moves, which is the entire point of using a spreadsheet rather than a document.

Public API::

    categorized_workbook(dataset, findings, ...) -> bytes
"""
from __future__ import annotations

import io
from pathlib import Path
from typing import Any, Iterable

from tools.bom.findings import Finding
from tools.bom.pivot import Dataset, Row

_LAYOUT = Path(__file__).resolve().parents[2] / "args" / "bom_xlsx_layout.yaml"


def load_layout(path: Path | str | None = None) -> dict[str, Any]:
    import yaml

    p = Path(path) if path else _LAYOUT
    return yaml.safe_load(p.read_text(encoding="utf-8")) or {}


def _category_of(row: Row, key: str) -> str:
    return (row.dims.get(key) or "").strip() or "Uncategorized"


def _sheet_name(cat: str, used: set[str]) -> str:
    """Excel: 31 chars, no []:*?/\\ — and no duplicates.

    A silently truncated collision means two categories quietly become one sheet
    and one of them disappears.
    """
    clean = "".join(c for c in cat if c not in "[]:*?/\\")[:31].strip() or "Sheet"
    name, n = clean, 2
    while name.lower() in used:
        suffix = f" ({n})"
        name = clean[: 31 - len(suffix)] + suffix
        n += 1
    used.add(name.lower())
    return name


def categorized_workbook(
    dataset: Dataset,
    findings: Iterable[Finding] = (),
    *,
    title: str = "BILL OF MATERIALS",
    subtitle: str = "",
    prepared_by: str = "",
    category_dim: str = "category",
    layout: dict[str, Any] | None = None,
) -> bytes:
    """One sheet per category, plus a Summary that ADDS ITSELF UP."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    cfg = layout or load_layout()
    th = cfg.get("theme", {})
    banner_bg = str(th.get("banner_fill", "1F3864"))
    banner_fg = str(th.get("banner_text", "FFFFFF"))
    head_bg = str(th.get("header_fill", "D9E2F3"))
    money_fmt = str(th.get("money_format", '#,##0;[Red]-#,##0'))
    headers = [str(h) for h in cfg.get("columns", [])]
    widths = [int(w) for w in cfg.get("column_widths", [])]

    banner = PatternFill("solid", fgColor=banner_bg)
    head = PatternFill("solid", fgColor=head_bg)
    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Group. Excluded money is kept and LABELLED rather than dropped: a line that
    # silently vanishes from a category sheet is a line nobody argues about again.
    groups: dict[str, list[Row]] = {}
    for r in dataset.rows:
        groups.setdefault(_category_of(r, category_dim), []).append(r)

    summary = wb.create_sheet("Summary")
    used = {"summary"}
    per_sheet: list[tuple[str, str, int, int]] = []   # (cat, sheet, first, last)

    for cat in sorted(groups):
        rows = groups[cat]
        name = _sheet_name(cat, used)
        ws = wb.create_sheet(name)

        ws["A1"] = cat.upper()
        ws["A1"].font = Font(bold=True, size=14, color=banner_fg)
        ws["A1"].fill = banner
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

        hdr_row = 3
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=hdr_row, column=c, value=h)
            cell.font = Font(bold=True)
            cell.fill = head
            cell.border = box
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        first = hdr_row + 1
        r_i = first
        for row in sorted(rows, key=lambda x: -x.extended_price):
            note = row.excluded_reason if not row.committed else ""
            ws.cell(row=r_i, column=1, value=row.description)
            ws.cell(row=r_i, column=2, value=row.dims.get("source", ""))
            ws.cell(row=r_i, column=3, value=row.qty or None)
            ws.cell(row=r_i, column=4, value=row.unit_price or None)

            # A FORMULA, not a value. The extended price is qty x unit price, and
            # saying so in the cell is what lets somebody change a quantity and
            # watch the total move — which is the only reason this is a
            # spreadsheet and not a PDF.
            ws.cell(row=r_i, column=5,
                    value=f"=C{r_i}*D{r_i}" if (row.qty and row.unit_price)
                    else (row.extended_price or None))
            ws.cell(row=r_i, column=6, value=note)

            for c in range(1, len(headers) + 1):
                ws.cell(row=r_i, column=c).border = box
            for c in (4, 5):
                ws.cell(row=r_i, column=c).number_format = money_fmt
            r_i += 1

        last = r_i - 1

        total = ws.cell(row=r_i + 1, column=1, value="Total")
        total.font = Font(bold=True)
        t = ws.cell(row=r_i + 1, column=5,
                    value=f"=SUM(E{first}:E{last})" if last >= first else 0)
        t.font = Font(bold=True)
        t.number_format = money_fmt
        t.border = box

        for c, w in enumerate(widths[: len(headers)], start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = f"A{first}"

        per_sheet.append((cat, name, r_i + 1, len(rows)))

    # ── Summary: it adds itself up ───────────────────────────────────────────
    summary["A2"] = title
    summary["A2"].font = Font(bold=True, size=16, color=banner_fg)
    summary["A2"].fill = banner
    summary.merge_cells("A2:C2")
    if subtitle:
        summary["A3"] = subtitle
        summary["A3"].font = Font(italic=True)
    if prepared_by:
        summary["A4"] = f"Prepared by: {prepared_by}"

    r_i = 6
    summary.cell(row=r_i, column=1, value="CATEGORY").font = Font(bold=True)
    summary.cell(row=r_i, column=2, value="ITEMS").font = Font(bold=True)
    summary.cell(row=r_i, column=3, value="TOTAL").font = Font(bold=True)
    for c in range(1, 4):
        summary.cell(row=r_i, column=c).fill = head
        summary.cell(row=r_i, column=c).border = box

    first = r_i + 1
    for cat, sheet, total_row, n in per_sheet:
        r_i += 1
        summary.cell(row=r_i, column=1, value=cat)
        summary.cell(row=r_i, column=2, value=n)
        # Cross-sheet reference. THIS is the fix: the template this layout came
        # from had these cells typed in, so the categories and the headline had
        # quietly stopped being the same number.
        summary.cell(row=r_i, column=3, value=f"='{sheet}'!E{total_row}")
        summary.cell(row=r_i, column=3).number_format = money_fmt
        for c in range(1, 4):
            summary.cell(row=r_i, column=c).border = box

    r_i += 1
    summary.cell(row=r_i, column=1, value="TOTAL").font = Font(bold=True)
    g = summary.cell(row=r_i, column=3,
                     value=f"=SUM(C{first}:C{r_i - 1})" if per_sheet else 0)
    g.font = Font(bold=True)
    g.number_format = money_fmt
    for c in range(1, 4):
        summary.cell(row=r_i, column=c).border = box
        summary.cell(row=r_i, column=c).fill = head

    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 10
    summary.column_dimensions["C"].width = 16

    # ── Open Items: what is not settled ──────────────────────────────────────
    #
    # Their template had this sheet and it is the right instinct: a BOM that does
    # not say what is still open is asserting a confidence nobody has.
    open_items = [
        f for f in findings
        if f.kind in ("decision", "risk") or f.severity in ("critical", "high")
    ]
    if open_items:
        ws = wb.create_sheet("Open Items")
        ws["A1"] = "OPEN ITEMS & ASSUMPTIONS"
        ws["A1"].font = Font(bold=True, size=14, color=banner_fg)
        ws["A1"].fill = banner
        ws.merge_cells("A1:E1")

        cols = ["#", "Topic", "Detail", "Impact", "Status"]
        for c, h in enumerate(cols, start=1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = Font(bold=True)
            cell.fill = head
            cell.border = box

        for i, f in enumerate(
            sorted(open_items, key=lambda x: -(x.impact_usd or 0)), start=1,
        ):
            r = 3 + i
            ws.cell(row=r, column=1, value=i)
            ws.cell(row=r, column=2, value=f.finding_type.replace("_", " ").title())
            ws.cell(row=r, column=3, value=f.title)
            ws.cell(row=r, column=4, value=f.impact_usd or None).number_format = money_fmt
            ws.cell(row=r, column=5, value="Decision" if f.kind == "decision" else "Confirm")
            for c in range(1, 6):
                ws.cell(row=r, column=c).border = box
                ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True,
                                                               vertical="top")

        for c, w in zip("ABCDE", (5, 26, 62, 14, 12)):
            ws.column_dimensions[c].width = w
        ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["categorized_workbook", "load_layout"]
