# CUI // SP-CTI
"""The workbook you hand to somebody who does not believe you.

Which is the right audience to design for. A reconciled bill of materials is only
worth anything if a sceptic can take any number in it, follow it back to the cell
it came from, and satisfy themselves it is real. So every line carries its
provenance — the document, the sheet, the cell — and the sheets are ordered the
way an argument is: what you are asking for, what we found wrong, and who said so.

The Findings sheet is sorted by money, because that is the only ordering anybody
with a budget has ever used.

Public API::

    export(dataset, findings, sources, path=..., pivots=...) -> bytes
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any, Iterable

from tools.bom import constants as C
from tools.bom.findings import Finding
from tools.bom.lines import ExtractedLine
from tools.bom.pivot import Dataset, Pivot
from tools.bom.reconcile import Source

_MONEY = '#,##0.00;[Red]-#,##0.00'
_HEADER_FILL = "1F3864"
_WARN_FILL = "FFF2CC"
_BAD_FILL = "F8CBAD"


@dataclass
class Sheet:
    title: str
    headers: list[str]
    rows: list[list[Any]]
    money_cols: tuple[int, ...] = ()
    note: str = ""
    widths: tuple[int, ...] = ()


def _style(ws, sheet: Sheet, start_row: int) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    for i, _ in enumerate(sheet.headers, start=1):
        cell = ws.cell(start_row, i)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for col in sheet.money_cols:
        for row in range(start_row + 1, start_row + 1 + len(sheet.rows)):
            ws.cell(row, col).number_format = _MONEY

    widths = sheet.widths or tuple(
        min(60, max(12, len(str(h)) + 4)) for h in sheet.headers
    )
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze the header so a long BOM stays readable. Small thing; it is the
    # difference between a workbook somebody scrolls and one they close.
    ws.freeze_panes = ws.cell(start_row + 1, 1)


def _write(wb, sheet: Sheet) -> None:
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet(sheet.title[:31])
    row = 1

    if sheet.note:
        ws.cell(1, 1, sheet.note)
        ws.cell(1, 1).font = Font(italic=True, size=9)
        ws.cell(1, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=3, end_column=max(len(sheet.headers), 4),
        )
        ws.row_dimensions[1].height = 45
        row = 5

    ws.append([]) if row > 1 else None
    for i, h in enumerate(sheet.headers, start=1):
        ws.cell(row, i, h)
    for r in sheet.rows:
        row += 1
        for i, v in enumerate(r, start=1):
            ws.cell(row, i, v)

    _style(ws, sheet, row - len(sheet.rows))


def _the_ask(dataset: Dataset, pivots: Iterable[Pivot]) -> Sheet:
    """The number, and — when it is not a number — why not."""
    rows: list[list[Any]] = []

    if dataset.competing_claims:
        rows.append(["THIS IS NOT A TOTAL", None])
        rows.append([
            f"{len(dataset.claim_sources)} documents each claim to price this "
            f"project. Adding them together adds competing estimates of the same "
            f"project — the very arithmetic that produced the spread you are here "
            f"to resolve. Nominate a source of record and this becomes a number.",
            None,
        ])
        rows.append([None, None])

    rows.append(["Committed — agreed, and traceable to a source cell", dataset.committed_total])
    rows.append(["Open — still disputed; contributes NOTHING above", dataset.open_total])
    rows.append([None, None])
    rows.append(["Line items (one per reconciled item, not per document)", len(dataset.rows)])
    rows.append(["Still needing a decision", sum(1 for r in dataset.rows if not r.committed)])

    return Sheet(
        title="The Ask",
        headers=["", "Amount"],
        rows=rows,
        money_cols=(2,),
        widths=(78, 18),
        note=(
            "Every figure in this workbook traces to a cell in a source document. "
            "Anything still disputed contributes ZERO to the committed total — not "
            "its cheapest option, not an average — and is listed on the Findings "
            "sheet. A total that quietly absorbed the disputes would be tidier and "
            "would be a lie."
        ),
    )


def _lines(dataset: Dataset, by_id: dict[str, ExtractedLine]) -> Sheet:
    rows = []
    for r in sorted(dataset.rows, key=lambda r: -r.extended_price):
        ln = by_id.get(r.line_id)
        rows.append([
            r.description,
            ln.part_number if ln else "",
            ln.manufacturer if ln else "",
            r.qty or None,
            r.unit_price or None,
            r.extended_price or None,
            r.dims.get("price_basis", ""),
            r.dims.get("credibility", ""),
            "committed" if r.committed else "OPEN",
            r.excluded_reason,
            # Provenance. The whole point.
            ln.source_document if ln else "",
            f"{ln.source_sheet}!{ln.source_locator}" if ln else "",
        ])

    return Sheet(
        title="Bill of Materials",
        headers=[
            "Item", "Part number", "Manufacturer", "Qty", "Unit", "Extended",
            "Price basis", "Source credibility", "Status", "Why it is open",
            "Source document", "Source cell",
        ],
        rows=rows,
        money_cols=(5, 6),
        widths=(46, 18, 16, 7, 13, 14, 12, 16, 11, 34, 34, 18),
        note=(
            "One row per reconciled ITEM — not per source line. Four documents "
            "describing the same switch is one switch. 'Source cell' is where the "
            "winning figure actually came from; go and look."
        ),
    )


def _findings(findings: Iterable[Finding]) -> Sheet:
    ordered = sorted(
        findings,
        key=lambda f: (C.SEVERITY_RANK[f.severity], -(f.impact_usd or 0.0)),
    )
    rows = []
    for f in ordered:
        ev = f.evidence[0] if f.evidence else None
        rows.append([
            f.severity,
            f.kind,
            f.impact_usd,
            f.title,
            f.detail.replace("\n\n", "  ").replace("\n", " "),
            f.detector,
            ev.source_document if ev else "",
            f"{ev.sheet}!{ev.locator}" if ev else "",
        ])

    return Sheet(
        title="Findings",
        headers=[
            "Severity", "Kind", "Impact", "What", "Detail", "How we know",
            "Source document", "Source cell",
        ],
        rows=rows,
        money_cols=(3,),
        widths=(10, 11, 14, 52, 90, 14, 34, 18),
        note=(
            "Sorted by money, because that is the only ordering anybody with a "
            "budget has ever used. 'How we know' says whether a finding is "
            "arithmetic or a model's opinion — you are entitled to weigh those "
            "differently. A blank Impact means we could not price it from the "
            "evidence and refused to guess."
        ),
    )


def _sources(sources: dict[str, Source]) -> Sheet:
    rows = []
    for name, src in sorted(
        sources.items(), key=lambda kv: C.CREDIBILITY_RANK.get(kv[1].credibility_tier, 99)
    ):
        rows.append([
            name,
            src.credibility_tier,
            src.role,
            src.as_of or "",
        ])
    return Sheet(
        title="Sources",
        headers=["Document", "How much its word is worth", "What it is", "As of"],
        rows=rows,
        widths=(46, 26, 22, 14),
        note=(
            "The ladder the reconciliation used. When two documents disagreed, the "
            "one higher up this list won — and where two documents you both marked "
            "authoritative disagreed, NEITHER won and the item is still open."
        ),
    )


def _pivot_sheet(p: Pivot, title: str) -> Sheet:
    table = p.as_table()
    return Sheet(
        title=title,
        headers=[str(c) for c in table[0]],
        rows=[list(r) for r in table[1:]],
        money_cols=tuple(range(2, len(table[0]) + 1)),
        widths=(30, *([16] * (len(table[0]) - 1))),
        note=p.reconciliation_note,
    )


def export(
    dataset: Dataset,
    findings: Iterable[Finding] = (),
    sources: dict[str, Source] | None = None,
    lines: Iterable[ExtractedLine] = (),
    pivots: Iterable[tuple[str, Pivot]] = (),
) -> bytes:
    """The whole argument, as a workbook."""
    import openpyxl

    by_id = {ln.line_id: ln for ln in lines}
    sources = sources or {}
    pivots = list(pivots)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write(wb, _the_ask(dataset, [p for _, p in pivots]))
    _write(wb, _lines(dataset, by_id))
    _write(wb, _findings(findings))
    _write(wb, _sources(sources))
    for title, p in pivots:
        _write(wb, _pivot_sheet(p, title))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main() -> int:  # pragma: no cover
    import argparse

    ap = argparse.ArgumentParser(description="Export a reconciled BOM workbook.")
    ap.add_argument("--out", default="bom.xlsx")
    args = ap.parse_args()
    print(f"use tools.bom.export_xlsx.export(...) — writing an empty shell to {args.out}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
