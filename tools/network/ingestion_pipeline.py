# CUI // SP-CTI
"""ICDEV™ Network Canvas — Unified Ingestion Pipeline.

Central processing pipeline for all network data ingestion. Three input
channels (REST API, dashboard upload, folder watch) converge here.

Supported file types:
  - Diagrams: .drawio, .vdx, .svg
  - Configs: .conf, .cfg, .log, .txt (auto-detected as config)
  - Documents: .pdf, .docx, .txt, .md
  - Spreadsheets: .csv, .xlsx
  - Images: .jpg, .jpeg, .png (via vision LLM)

Air-gap safe: all processing uses local tools (Ollama for vision/embeddings).

Usage::

    from tools.network.ingestion_pipeline import ingest_file
    result = ingest_file("path/to/file.pdf", channel="upload")
"""

from __future__ import annotations
from tools.logging.icdev_logger import get_logger

import csv
import hashlib
import json
import logging
import sqlite3
import uuid
from tools.db.storage import get_connection
from pathlib import Path
from typing import Any, Dict, Optional

logger = get_logger("icdev.network.ingestion_pipeline")

_ICDEV_ROOT = Path(__file__).resolve().parents[2]
_DEFAULT_DB = _ICDEV_ROOT / "data" / "network_canvas.db"

# File extension → type mapping
_EXT_MAP = {
    # Diagrams
    ".drawio": "diagram",
    ".vdx": "diagram",
    ".svg": "diagram",
    # Configs (detected by content when ambiguous)
    ".conf": "config",
    ".cfg": "config",
    ".log": "config",
    # Documents
    ".pdf": "document",
    ".docx": "document",
    ".md": "document",
    # Spreadsheets
    ".csv": "spreadsheet",
    ".xlsx": "spreadsheet",
    ".xls": "spreadsheet",
    # Images
    ".jpg": "image",
    ".jpeg": "image",
    ".png": "image",
    ".gif": "image",
    ".webp": "image",
    ".tiff": "image",
    ".bmp": "image",
}

# .txt is ambiguous — could be doc or config. Classify by content.
_AMBIGUOUS_EXTS = {".txt"}


# ---------------------------------------------------------------------------
# File classification
# ---------------------------------------------------------------------------


def classify_file(file_path: str) -> str:
    """Classify a file by type based on extension and content heuristics.

    Returns:
        'diagram' | 'config' | 'document' | 'spreadsheet' | 'image'
    """
    p = Path(file_path)
    ext = p.suffix.lower()

    if ext in _EXT_MAP:
        return _EXT_MAP[ext]

    if ext in _AMBIGUOUS_EXTS:
        return _classify_text_content(p)

    # Unknown extension — try content detection
    return _classify_text_content(p)


def _classify_text_content(file_path: Path) -> str:
    """Heuristic: is this text file a device config or a document?"""
    try:
        with open(str(file_path), "r", encoding="utf-8", errors="replace") as f:
            head = f.read(4096)
    except Exception:
        return "document"

    from tools.network.config_parser import detect_vendor

    vendor = detect_vendor(head)
    if vendor != "generic":
        return "config"

    # Config indicators
    config_signals = [
        "hostname ", "interface ", "ip route ", "access-list ",
        "set interfaces", "set protocols", "set security",
        "feature ", "vlan ", "router ospf", "router bgp",
    ]
    hits = sum(1 for s in config_signals if s in head)
    if hits >= 2:
        return "config"

    return "document"


# ---------------------------------------------------------------------------
# Ingestion audit log
# ---------------------------------------------------------------------------


def _get_db(db_path: Optional[str] = None) -> sqlite3.Connection:
    """Get SQLite connection to network_canvas.db."""
    path = db_path or str(_DEFAULT_DB)
    conn = get_connection(str(path))
    return conn


def _log_ingestion(
    conn: sqlite3.Connection,
    channel: str,
    file_name: str,
    file_type: str,
    file_hash: str,
    status: str = "started",
    source_adapter: str = "",
    topology_id: str = "",
    project_id: str = "default",
    user_id: str = "",
    result_json: str = "{}",
    error: str = "",
) -> str:
    """Write to nc_ingestion_log (append-only). Returns log entry ID."""
    log_id = str(uuid.uuid4())[:12]
    conn.execute(
        "INSERT INTO nc_ingestion_log "
        "(id, channel, file_name, file_type, file_hash, source_adapter, "
        "status, result_json, error, topology_id, project_id, user_id) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (log_id, channel, file_name, file_type, file_hash, source_adapter,
         status, result_json, error, topology_id, project_id, user_id),
    )
    conn.commit()
    return log_id


def _complete_log(
    conn: sqlite3.Connection,
    log_id: str,
    status: str,
    result_json: str = "{}",
    error: str = "",
) -> None:
    """Mark ingestion log entry as completed/failed."""
    conn.execute(
        "UPDATE nc_ingestion_log SET status=?, result_json=?, error=?, "
        "completed_at=datetime('now') WHERE id=?",
        (status, result_json, error, log_id),
    )
    conn.commit()


# ---------------------------------------------------------------------------
# File hash helper
# ---------------------------------------------------------------------------


def _file_hash(file_path: Path) -> str:
    """SHA-256 hash of file contents."""
    h = hashlib.sha256()
    with open(str(file_path), "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Diagram ingestion
# ---------------------------------------------------------------------------


def ingest_diagram(
    file_path: str,
    topology_id: Optional[str] = None,
    project_id: str = "default",
    db_path: Optional[str] = None,
) -> Dict[str, Any]:
    """Ingest a diagram file (DrawIO, VDX, SVG).

    Delegates to existing NDC import functions.
    """
    p = Path(file_path)
    ext = p.suffix.lower()
    with open(str(p), "r", encoding="utf-8", errors="replace") as f:
        content = f.read()

    try:
        from tools.network.export_import import import_drawio, import_vdx, import_svg
    except ImportError:
        return {"error": "NDC export_import module not available"}

    if ext == ".drawio":
        graph = import_drawio(content)
    elif ext == ".vdx":
        graph = import_vdx(content)
    elif ext == ".svg":
        graph = import_svg(content)
    else:
        return {"error": "Unsupported diagram format: %s" % ext}

    nodes = graph.get("nodes", [])
    edges = graph.get("edges", [])

    return {
        "type": "diagram",
        "format": ext.lstrip("."),
        "nodes_count": len(nodes),
        "edges_count": len(edges),
        "graph": graph,
    }


# ---------------------------------------------------------------------------
# Config ingestion
# ---------------------------------------------------------------------------


def ingest_config(
    file_path: str,
    topology_id: Optional[str] = None,
    project_id: str = "default",
    db_path: Optional[str] = None,
) -> Dict[str, Any]:
    """Ingest a device configuration file.

    Dual-mode: raw text → RAG, structured parse → KG.
    """
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        raw = f.read()

    from tools.network.config_parser import ingest_config as _ingest

    result = _ingest(
        raw,
        topology_id=topology_id,
        project_id=project_id,
        db_path=db_path,
    )
    result["type"] = "config"
    return result


# ---------------------------------------------------------------------------
# Document ingestion
# ---------------------------------------------------------------------------


def ingest_document(
    file_path: str,
    project_id: str = "default",
    topology_id: Optional[str] = None,
    db_path: Optional[str] = None,
) -> Dict[str, Any]:
    """Ingest a network document (PDF, DOCX, TXT, MD).

    Extracts text, persists to nc_documents, returns for RAG chunking.
    """
    p = Path(file_path)
    ext = p.suffix.lower()
    fhash = _file_hash(p)
    file_size = p.stat().st_size

    # Check dedup
    conn = _get_db(db_path)
    existing = conn.execute(
        "SELECT id FROM nc_documents WHERE file_hash=?", (fhash,)
    ).fetchone()
    if existing:
        conn.close()
        return {
            "type": "document",
            "status": "duplicate",
            "existing_id": existing["id"],
            "file_hash": fhash,
        }

    # Extract text
    extracted_text = ""
    provider_used = ""
    page_count = 0

    if ext == ".pdf":
        try:
            from tools.rag.pdf_provider import extract_pdf_text
            result = extract_pdf_text(str(p))
            extracted_text = result.get("text", "")
            provider_used = result.get("provider", "")
            page_count = result.get("pages", 0)
        except ImportError:
            with open(str(p), "r", encoding="utf-8", errors="replace") as f:
                extracted_text = f.read()
            provider_used = "raw_read"
    elif ext == ".docx":
        try:
            from tools.requirements.document_extractor import extract_docx
            extracted_text = extract_docx(str(p))
            provider_used = "document_extractor"
        except ImportError:
            extracted_text = "(DOCX extraction requires python-docx)"
            provider_used = "unavailable"
    else:
        with open(str(p), "r", encoding="utf-8", errors="replace") as f:
            extracted_text = f.read()
        provider_used = "text_read"

    # Persist to nc_documents
    doc_id = str(uuid.uuid4())[:12]
    try:
        conn.execute(
            "INSERT INTO nc_documents "
            "(id, file_name, file_path, file_hash, file_size_bytes, "
            "doc_type, extracted_text, page_count, provider_used, "
            "topology_id, project_id, status) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (doc_id, p.name, str(p), fhash, file_size,
             "general", extracted_text, page_count, provider_used,
             topology_id or "", project_id, "ingested"),
        )
        conn.commit()
    except Exception as exc:
        logger.warning("Failed to persist document: %s", exc)
    finally:
        conn.close()

    return {
        "type": "document",
        "doc_id": doc_id,
        "file_name": p.name,
        "file_hash": fhash,
        "text_length": len(extracted_text),
        "page_count": page_count,
        "provider": provider_used,
        "status": "ingested",
    }


# ---------------------------------------------------------------------------
# Spreadsheet ingestion
# ---------------------------------------------------------------------------


def ingest_spreadsheet(
    file_path: str,
    topology_id: Optional[str] = None,
    project_id: str = "default",
    db_path: Optional[str] = None,
) -> Dict[str, Any]:
    """Ingest CSV/Excel for IP plans, VLAN tables, circuit inventories."""
    p = Path(file_path)
    ext = p.suffix.lower()

    rows = []  # type: List[Dict[str, str]]
    if ext == ".csv":
        with open(str(p), "r", encoding="utf-8", errors="replace") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
    elif ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                return {"type": "spreadsheet", "error": "No active worksheet"}
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
            wb.close()
        except ImportError:
            return {
                "type": "spreadsheet",
                "error": "Excel support requires openpyxl: pip install openpyxl",
            }

    # Also create text representation for RAG
    text_repr = "\n".join(
        ", ".join("%s=%s" % (k, v) for k, v in row.items() if v)
        for row in rows[:500]  # cap for very large files
    )

    return {
        "type": "spreadsheet",
        "file_name": p.name,
        "rows_count": len(rows),
        "columns": list(rows[0].keys()) if rows else [],
        "rows": rows,
        "text_repr": text_repr,
        "status": "parsed",
    }


# ---------------------------------------------------------------------------
# Image ingestion
# ---------------------------------------------------------------------------


def ingest_image(
    file_path: str,
    project_id: str = "default",
    db_path: Optional[str] = None,
) -> Dict[str, Any]:
    """Ingest an image via vision LLM (LLaVA via Ollama, air-gap safe).

    Attempts to extract network topology from diagram images.
    Falls back to general description for rack photos/whiteboards.
    """
    try:
        from tools.network.network_ingester import extract_diagram_via_vision
        result = extract_diagram_via_vision(file_path)
        return {
            "type": "image",
            "extraction_method": "vision_llm",
            "result": result,
            "status": "ingested",
        }
    except ImportError:
        pass
    except Exception as exc:
        logger.warning("Vision extraction failed: %s", exc)

    # Fallback: store as document with file reference
    return {
        "type": "image",
        "extraction_method": "none",
        "file_path": file_path,
        "status": "pending_vision",
        "note": "Vision LLM unavailable. Image stored for later processing.",
    }


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------


def ingest_file(
    file_path: str,
    channel: str = "api",
    project_id: str = "default",
    topology_id: Optional[str] = None,
    metadata: Optional[Dict[str, Any]] = None,
    db_path: Optional[str] = None,
    user_id: str = "",
) -> Dict[str, Any]:
    """Unified file ingestion entry point.

    Routes to the correct handler based on file type classification.
    Logs ingestion events to nc_ingestion_log (append-only audit trail).

    Args:
        file_path: Path to the file to ingest.
        channel: 'api' | 'upload' | 'folder_watch'
        project_id: Project context.
        topology_id: Link to existing topology (optional).
        metadata: Extra metadata dict (optional).
        db_path: Override DB path.
        user_id: Who initiated the ingestion.

    Returns:
        Ingestion result dict with type, status, and type-specific fields.
    """
    p = Path(file_path)
    if not p.exists():
        return {"error": "File not found: %s" % file_path}

    file_type = classify_file(file_path)
    fhash = _file_hash(p)

    # Log start
    conn = _get_db(db_path)
    log_id = _log_ingestion(
        conn, channel, p.name, file_type, fhash,
        topology_id=topology_id or "",
        project_id=project_id,
        user_id=user_id,
    )
    conn.close()

    # Route to handler
    try:
        if file_type == "diagram":
            result = ingest_diagram(file_path, topology_id, project_id, db_path)
        elif file_type == "config":
            result = ingest_config(file_path, topology_id, project_id, db_path)
        elif file_type == "document":
            result = ingest_document(file_path, project_id, topology_id, db_path)
        elif file_type == "spreadsheet":
            result = ingest_spreadsheet(file_path, topology_id, project_id, db_path)
        elif file_type == "image":
            result = ingest_image(file_path, project_id, db_path)
        else:
            result = {"error": "Unknown file type: %s" % file_type}

        # Complete log
        conn = _get_db(db_path)
        status = "failed" if "error" in result else "completed"
        _complete_log(conn, log_id, status, json.dumps(result, default=str))
        conn.close()

        result["log_id"] = log_id
        result["channel"] = channel
        result["file_hash"] = fhash
        return result

    except Exception as exc:
        conn = _get_db(db_path)
        _complete_log(conn, log_id, "failed", error=str(exc))
        conn.close()
        return {"error": str(exc), "log_id": log_id}


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> None:
    """CLI entry point."""
    import argparse

    parser = argparse.ArgumentParser(description="ICDEV NDC Ingestion Pipeline")
    parser.add_argument("files", nargs="+", help="Files to ingest")
    parser.add_argument("--channel", default="api", help="Channel: api, upload, folder_watch")
    parser.add_argument("--project-id", default="default", help="Project ID")
    parser.add_argument("--topology-id", default=None, help="Topology ID")
    parser.add_argument("--json", action="store_true", help="JSON output")
    args = parser.parse_args()

    results = []
    for fp in args.files:
        result = ingest_file(
            fp,
            channel=args.channel,
            project_id=args.project_id,
            topology_id=args.topology_id,
        )
        results.append(result)

    if args.json:
        print(json.dumps(results, indent=2, default=str))
    else:
        for r in results:
            status = r.get("status", r.get("error", "unknown"))
            print("%s: %s (%s)" % (r.get("type", "?"), status, r.get("file_name", r.get("file_hash", "")[:8])))


if __name__ == "__main__":
    main()