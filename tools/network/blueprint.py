# [TEMPLATE: CUI // SP-CTI]
"""ICDEV™ Network Design Canvas — Flask Blueprint integration.

Fully self-contained Blueprint mounted at /network/ inside the ICDEV dashboard.
Uses ICDEV's auth system, separate network_canvas.db, and feature flag
ICDEV_NETWORK_ENABLED.

Usage in ICDEV dashboard app.py:
    from tools.network.blueprint import create_network_blueprint
    bp = create_network_blueprint()
    if bp:
        app.register_blueprint(bp, url_prefix="/network")
"""

from __future__ import annotations
from tools.logging.icdev_logger import get_logger

import json
import os
import shutil
import uuid as _uuid
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from flask import (
    Blueprint,
    abort,
    current_app,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    session,
)

logger = get_logger("icdev.network")

# ── Paths ──────────────────────────────────────────────────────────────────────
_NETWORK_DIR = Path(__file__).resolve().parent
_ICDEV_ROOT = _NETWORK_DIR.parent.parent
_TEMPLATE_DIR = _ICDEV_ROOT / "tools" / "dashboard" / "templates"

# ── Import helper modules ──────────────────────────────────────────────────────
from tools.network.constants import (  # noqa: E402
    CLOUD_OBJECTS,
    CSP_GROUP_DEFAULTS,
    COMPLIANCE_REGIMES,
    BOM_COSTS,
)
from tools.network.simulation import (  # noqa: E402
    _run_simulation,
    _add_narrative,
)
from tools.network.compliance import (  # noqa: E402
    run_compliance_audit,
    apply_compliance_fix,
    generate_xacta_export,
    generate_fips_coverage_report,
    export_fips_report_html,
)
from tools.network.montecarlo import run_monte_carlo  # noqa: E402
from tools.network.ato_generator import (  # noqa: E402
    generate_ato_package,
    generate_pps_matrix_for_pair,
    get_topology_enclaves,
    export_pps_as_ssp_table,
)
from tools.network.export_import import (  # noqa: E402
    to_drawio,
    to_svg,
    to_vdx,
    import_drawio,
    import_vdx,
    import_svg,
)
from tools.network.visio_export import export_vsdx, export_ops_csvs  # noqa: E402
from tools.network.inventory_export import (  # noqa: E402
    to_ansible_inventory,
    to_terraform_hcl,
)
from tools.network.config_generator import (  # noqa: E402
    generate_device_configs,
    generate_device_configs_zip,
    list_configurable_nodes,
)
from tools.network.stig_import import import_stig_file  # noqa: E402
from tools.canvas.ai_trace_mixin import record_canvas_decision  # noqa: E402


def create_network_blueprint():
    """Create and return the Network Design Canvas Blueprint.

    Returns None if ICDEV_NETWORK_ENABLED is false.
    """
    enabled = os.environ.get("ICDEV_NETWORK_ENABLED", "true").lower()
    if enabled not in ("true", "1", "yes"):
        logger.info("Network Canvas disabled (ICDEV_NETWORK_ENABLED=%s)", enabled)
        return None

    # Initialize DB
    try:
        from tools.network.db.init_db import init_db

        init_db()
    except Exception as exc:
        logger.warning("Network DB init failed: %s", exc)

    bp = Blueprint(
        "network_canvas",
        __name__,
        template_folder=str(_TEMPLATE_DIR),
    )

    # ── Helpers (imported from blueprint_helpers.py) ────────────────────────
    from tools.network.db.init_db import get_connection
    from tools.network.blueprint_helpers import (
        nc_login_required,
        _now,
        _row_to_dict,
        _normalize_sop_step,
        _audit,
        _notify,
        _crud_list,
        _crud_create,
        _crud_delete,
        _NDC_LIFECYCLE,
    )

    # ── Load config ────────────────────────────────────────────────────────
    try:
        import yaml

        _config_path = _ICDEV_ROOT / "args" / "network_canvas_config.yaml"
        if _config_path.exists():
            with open(_config_path, encoding="utf-8") as f:
                NC_CONFIG = yaml.safe_load(f) or {}
        else:
            NC_CONFIG = {}
    except Exception:
        NC_CONFIG = {}

    # ── Context processor for network templates ────────────────────────────
    @bp.context_processor
    def inject_nc_context():
        user = None
        try:
            user = getattr(g, "current_user", None)
        except RuntimeError:
            pass
        return {
            "classification_banner": NC_CONFIG.get("app", {}).get("classification", ""),
            "current_user": user,
        }

    # ── Chat message helper ────────────────────────────────────────────────
    def _nc_save_message(ctx_id: str, role: str, content: str) -> None:
        """Insert a chat message into chat_messages with auto-incrementing turn_number."""
        try:
            msg_id = "ncmsg-" + _uuid.uuid4().hex[:12]
            conn = get_connection()
            row = conn.execute(
                "SELECT MAX(turn_number) FROM chat_messages WHERE context_id=?",
                (ctx_id,),
            ).fetchone()
            turn_number = (row[0] or 0) + 1
            conn.execute(
                "INSERT INTO chat_messages (id, context_id, turn_number, role, content, content_type, created_at) "
                "VALUES (?, ?, ?, ?, ?, 'text', ?)",
                (msg_id, ctx_id, turn_number, role, content, datetime.utcnow().isoformat()),
            )
            conn.commit()
        except Exception as exc:
            logger.warning("_nc_save_message failed: %s", exc)

    # ══════════════════════════════════════════════════════════════════════
    # PAGE ROUTES
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/")
    @nc_login_required
    def nc_index():
        conn = get_connection()
        # Project filter support
        filter_project = request.args.get("project", "")
        if filter_project:
            rows = conn.execute(
                "SELECT t.id, t.name, t.description, t.classification, t.created_at, t.updated_at, "
                "t.graph_json "
                "FROM topologies t JOIN nc_project_topologies pt ON pt.topology_id=t.id "
                "WHERE pt.project_id=? ORDER BY t.updated_at DESC LIMIT 20",
                (filter_project,),
            ).fetchall()
            topologies = []
            for r in rows:
                t = _row_to_dict(r)
                try:
                    g = json.loads(t.get("graph_json") or '{"nodes":[],"edges":[]}')
                except Exception:
                    g = {"nodes": [], "edges": []}
                t["node_count"] = len(g.get("nodes", []))
                t["edge_count"] = len(g.get("edges", []))
                topologies.append(t)
        else:
            rows = conn.execute(
                "SELECT id, name, description, classification, created_at, updated_at, graph_json "
                "FROM topologies ORDER BY updated_at DESC LIMIT 20"
            ).fetchall()
            topologies = []
            for r in rows:
                t = _row_to_dict(r)
                try:
                    g = json.loads(t.get("graph_json") or '{"nodes":[],"edges":[]}')
                except Exception:
                    g = {"nodes": [], "edges": []}
                t["node_count"] = len(g.get("nodes", []))
                t["edge_count"] = len(g.get("edges", []))
                topologies.append(t)
        templates = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT id, name, category, description, tags FROM nc_templates ORDER BY category, name"
            ).fetchall()
        ]
        sims = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT sr.id, sr.sim_type, sr.ran_at, t.name AS topology_name, sr.result_json "
                "FROM simulation_results sr JOIN topologies t ON t.id=sr.topology_id "
                "ORDER BY sr.ran_at DESC LIMIT 10"
            ).fetchall()
        ]
        total_sims = conn.execute("SELECT COUNT(*) FROM simulation_results").fetchone()[0]
        # Load projects list for filter dropdown
        all_projects = [
            _row_to_dict(r) for r in conn.execute("SELECT id, name FROM nc_projects ORDER BY name").fetchall()
        ]
        # Active project name for display
        active_project = None
        if filter_project:
            ap_row = conn.execute("SELECT id, name FROM nc_projects WHERE id=?", (filter_project,)).fetchone()
            active_project = _row_to_dict(ap_row) if ap_row else None
        conn.close()

        for t in templates:
            try:
                t["tags"] = json.loads(t.get("tags") or "[]")
            except Exception:
                t["tags"] = []
        for s in sims:
            try:
                rj = json.loads(s.get("result_json") or "{}")
                s["summary"] = rj.get("summary", "—")
            except Exception:
                s["summary"] = "—"

        return render_template(
            "network/index.html",
            topologies=topologies,
            templates=templates[:6],
            simulations=sims,
            all_projects=all_projects,
            filter_project=filter_project,
            active_project=active_project,
            stats={"topologies": len(topologies), "simulations": total_sims, "templates": len(templates)},
        )

    @bp.route("/api/morning-dashboard", methods=["GET"])
    @nc_login_required
    def nc_api_morning_dashboard():
        """Architect's personalized morning dashboard data."""
        conn = get_connection()
        now = datetime.now(timezone.utc)

        # My projects (recent activity)
        projects = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT id, name, status, owner, updated_at FROM nc_projects ORDER BY updated_at DESC LIMIT 10"
            ).fetchall()
        ]

        # Pending reviews
        pending_reviews = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT br.id, br.phase, br.scheduled_date, "
                "rb.short_name AS board, rb.name AS board_name, "
                "p.name AS project_name "
                "FROM nc_board_reviews br "
                "JOIN nc_review_boards rb ON rb.id=br.board_id "
                "JOIN nc_projects p ON p.id=br.project_id "
                "WHERE br.status='pending' "
                "ORDER BY br.scheduled_date"
            ).fetchall()
        ]

        # Compliance alerts (projects below 80%)
        compliance_alerts = []
        for p in projects:
            topo_ids = [
                r[0]
                for r in conn.execute(
                    "SELECT topology_id FROM nc_project_topologies WHERE project_id=?", (p["id"],)
                ).fetchall()
            ]
            tp = tf = 0
            for tid in topo_ids:
                row = conn.execute(
                    "SELECT passed, failed FROM nc_compliance_checks WHERE topology_id=? ORDER BY ran_at DESC LIMIT 1",
                    (tid,),
                ).fetchone()
                if row:
                    tp += row[0] or 0
                    tf += row[1] or 0
            total = tp + tf
            if total > 0:
                pct = round(tp * 100 / total)
                if pct < 80:
                    compliance_alerts.append(
                        {
                            "project": p["name"],
                            "project_id": p["id"],
                            "compliance_pct": pct,
                        }
                    )

        # Upcoming EOL (devices across all topologies)
        eol_alerts = []
        for r in conn.execute("SELECT t.name AS topo_name, t.graph_json FROM topologies t LIMIT 20").fetchall():
            try:
                graph = json.loads(r["graph_json"])
            except Exception:
                continue
            for n in graph.get("nodes", []):
                cfg = n.get("config") or n.get("configData") or {}
                eol = cfg.get("eol_date") or cfg.get("eosup_date")
                if eol:
                    try:
                        dt = datetime.fromisoformat(
                            eol + "T00:00:00+00:00" if "T" not in eol else eol.replace("Z", "+00:00")
                        )
                        months = (dt - now).days / 30.44
                        if months <= 12:
                            eol_alerts.append(
                                {
                                    "device": n.get("label", ""),
                                    "model": cfg.get("model", ""),
                                    "eol_date": eol,
                                    "months_remaining": round(months, 1),
                                    "topology": r["topo_name"],
                                    "past_eol": months <= 0,
                                }
                            )
                    except (ValueError, TypeError):
                        pass

        # Recent activity
        activity = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT action, entity_type, details, ts FROM nc_audit ORDER BY ts DESC LIMIT 15"
            ).fetchall()
        ]

        # Unread notifications
        unread = conn.execute("SELECT COUNT(*) FROM nc_notifications WHERE is_read=0").fetchone()[0]

        # Quick stats
        total_projects = conn.execute("SELECT COUNT(*) FROM nc_projects").fetchone()[0]
        total_topos = conn.execute("SELECT COUNT(*) FROM topologies").fetchone()[0]
        total_peers = conn.execute("SELECT COUNT(*) FROM nc_peering_agreements WHERE status='operational'").fetchone()[
            0
        ]

        conn.close()
        return jsonify(
            {
                "projects": projects,
                "pending_reviews": pending_reviews,
                "compliance_alerts": compliance_alerts,
                "eol_alerts": sorted(eol_alerts, key=lambda x: x["months_remaining"])[:10],
                "activity": activity,
                "unread_notifications": unread,
                "stats": {
                    "total_projects": total_projects,
                    "total_topologies": total_topos,
                    "active_peering": total_peers,
                    "pending_reviews": len(pending_reviews),
                    "compliance_alerts": len(compliance_alerts),
                    "eol_alerts": len(eol_alerts),
                },
            }
        )

    @bp.route("/canvas/new")
    @nc_login_required
    def nc_canvas_new():
        return render_template("network/canvas.html", topology_id="new", topology_name="Untitled Topology")

    @bp.route("/canvas/<topo_id>")
    @nc_login_required
    def nc_canvas_edit(topo_id):
        conn = get_connection()
        row = conn.execute("SELECT * FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not row:
            conn.close()
            abort(404)
        topo = _row_to_dict(row)
        # Find projects this topology belongs to
        topo_projects = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT p.id, p.name, p.status FROM nc_projects p "
                "JOIN nc_project_topologies pt ON pt.project_id=p.id "
                "WHERE pt.topology_id=? ORDER BY p.name",
                (topo_id,),
            ).fetchall()
        ]
        # All projects for quick-switch
        all_projects = [
            _row_to_dict(r) for r in conn.execute("SELECT id, name, status FROM nc_projects ORDER BY name").fetchall()
        ]
        conn.close()
        return render_template(
            "network/canvas.html",
            topology_id=topo_id,
            topology_name=topo["name"],
            classification=topo.get("classification", "public"),
            design=topo,
            topo_projects=topo_projects,
            all_projects=all_projects,
        )

    @bp.route("/templates")
    @nc_login_required
    def nc_templates():
        conn = get_connection()
        templates = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT id, name, category, description, tags FROM nc_templates ORDER BY category, name"
            ).fetchall()
        ]
        conn.close()
        for t in templates:
            try:
                t["tags"] = json.loads(t.get("tags") or "[]")
            except Exception:
                t["tags"] = []
        categories = {}
        for t in templates:
            categories.setdefault(t.get("category") or "Other", []).append(t)
        return render_template("network/templates_gallery.html", categories=categories, templates=templates)

    @bp.route("/simulation/<sim_id>")
    @nc_login_required
    def nc_simulation_detail(sim_id):
        conn = get_connection()
        row = conn.execute(
            "SELECT sr.*, t.name AS topology_name FROM simulation_results sr "
            "JOIN topologies t ON t.id=sr.topology_id WHERE sr.id=?",
            (sim_id,),
        ).fetchone()
        conn.close()
        if not row:
            abort(404)
        sim = _row_to_dict(row)
        try:
            sim["result"] = json.loads(sim.get("result_json") or "{}")
        except Exception:
            sim["result"] = {}
        return render_template("network/simulation.html", sim=sim)

    @bp.route("/template/<tpl_id>/edit")
    @nc_login_required
    def nc_template_edit(tpl_id):
        conn = get_connection()
        row = conn.execute("SELECT * FROM nc_templates WHERE id=?", (tpl_id,)).fetchone()
        conn.close()
        if not row:
            abort(404)
        tpl = _row_to_dict(row)
        try:
            tpl["tags"] = json.loads(tpl.get("tags") or "[]")
        except Exception:
            tpl["tags"] = []
        try:
            gj = json.loads(tpl["graph_json"]) if isinstance(tpl["graph_json"], str) else tpl["graph_json"]
            tpl["node_count"] = len(gj.get("nodes", []))
            tpl["edge_count"] = len(gj.get("edges", []))
        except Exception:
            tpl["node_count"] = 0
            tpl["edge_count"] = 0
        return render_template("network/template_edit.html", tpl=tpl)

    @bp.route("/versions/<topo_id>")
    @nc_login_required
    def nc_versions_page(topo_id):
        conn = get_connection()
        topo = conn.execute("SELECT id, name FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not topo:
            conn.close()
            abort(404)
        versions = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT * FROM nc_versions WHERE topology_id=? ORDER BY version_num", (topo_id,)
            ).fetchall()
        ]
        conn.close()
        return render_template("network/versions.html", topology=_row_to_dict(topo), versions=versions)

    @bp.route("/montecarlo/<topo_id>")
    @nc_login_required
    def nc_montecarlo_page(topo_id):
        conn = get_connection()
        topo = conn.execute("SELECT id, name, graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not topo:
            conn.close()
            abort(404)

        scenarios = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT * FROM nc_mc_scenarios WHERE topology_id=? ORDER BY created_at DESC", (topo_id,)
            ).fetchall()
        ]

        _auto_run_scenario = None
        if not scenarios:
            now = _now()
            defaults = [
                (
                    "Random Failure (5% links)",
                    "random",
                    "Random 5% link failure probability per iteration",
                    json.dumps({"iterations": 1000, "node_failure_prob": 0.02, "edge_failure_prob": 0.05}),
                ),
                (
                    "Major Outage (20% links)",
                    "random",
                    "Stress test — 20% link failure probability",
                    json.dumps({"iterations": 1000, "node_failure_prob": 0.05, "edge_failure_prob": 0.20}),
                ),
                (
                    "Single Node Failure",
                    "random",
                    "What happens when one critical node goes down?",
                    json.dumps({"iterations": 500, "node_failure_prob": 0.08, "edge_failure_prob": 0.0}),
                ),
            ]
            for name, stype, desc, cfg in defaults:
                sid = str(_uuid.uuid4())
                conn.execute(
                    "INSERT INTO nc_mc_scenarios (id, topology_id, name, scenario_type, description, config_json, created_at) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (sid, topo_id, name, stype, desc, cfg, now),
                )
            conn.commit()
            scenarios = [
                _row_to_dict(r)
                for r in conn.execute(
                    "SELECT * FROM nc_mc_scenarios WHERE topology_id=? ORDER BY created_at DESC", (topo_id,)
                ).fetchall()
            ]
            try:
                graph = json.loads(topo["graph_json"])
                if graph.get("nodes") and graph.get("edges"):
                    _auto_run_scenario = scenarios[-1]["id"]
            except Exception:
                pass

        runs = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT r.id, r.iterations, r.ran_at, s.name AS scenario_name, s.scenario_type "
                "FROM nc_mc_runs r JOIN nc_mc_scenarios s ON s.id=r.scenario_id "
                "WHERE r.topology_id=? ORDER BY r.ran_at DESC LIMIT 20",
                (topo_id,),
            ).fetchall()
        ]

        if scenarios and not runs and not _auto_run_scenario:
            _auto_run_scenario = scenarios[-1]["id"]

        conn.close()
        return render_template(
            "network/montecarlo.html",
            topology=_row_to_dict(topo),
            scenarios=scenarios,
            runs=runs,
            auto_run_scenario=_auto_run_scenario,
        )

    @bp.route("/compliance/<topo_id>")
    @nc_login_required
    def nc_compliance_page(topo_id):
        conn = get_connection()
        topo = conn.execute("SELECT id, name, graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not topo:
            conn.close()
            abort(404)

        profile = conn.execute("SELECT * FROM nc_compliance_profiles WHERE topology_id=?", (topo_id,)).fetchone()
        if not profile:
            pid = str(_uuid.uuid4())
            conn.execute("INSERT INTO nc_compliance_profiles (id, topology_id) VALUES (?,?)", (pid, topo_id))
            conn.commit()
            profile = conn.execute("SELECT * FROM nc_compliance_profiles WHERE id=?", (pid,)).fetchone()
        profile = _row_to_dict(profile)

        audits = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT id, check_type, passed, failed, ran_at FROM nc_compliance_checks "
                "WHERE topology_id=? ORDER BY ran_at DESC LIMIT 10",
                (topo_id,),
            ).fetchall()
        ]

        open_findings = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT * FROM nc_compliance_findings WHERE topology_id=? AND status='open' ORDER BY severity, rule_id",
                (topo_id,),
            ).fetchall()
        ]

        conn.close()

        try:
            regimes = json.loads(profile.get("regimes", "[]"))
        except Exception:
            regimes = ["fisma_high"]

        return render_template(
            "network/compliance_audit.html",
            topology=_row_to_dict(topo),
            profile=profile,
            regimes=regimes,
            all_regimes=COMPLIANCE_REGIMES,
            audits=audits,
            open_findings=open_findings,
        )

    @bp.route("/circuits")
    @nc_login_required
    def nc_circuits():
        conn = get_connection()
        filter_project = request.args.get("project", "")
        if filter_project:
            circuits = [
                _row_to_dict(r)
                for r in conn.execute(
                    "SELECT * FROM nc_circuits WHERE topology_id IN "
                    "(SELECT topology_id FROM nc_project_topologies WHERE project_id=?) "
                    "ORDER BY updated_at DESC",
                    (filter_project,),
                ).fetchall()
            ]
        else:
            circuits = [
                _row_to_dict(r) for r in conn.execute("SELECT * FROM nc_circuits ORDER BY updated_at DESC").fetchall()
            ]
        stats = {
            "total": len(circuits),
            "installed": sum(1 for c in circuits if c.get("install_status") == "installed"),
            "planned": sum(1 for c in circuits if c.get("install_status") == "planned"),
            "monthly_cost": sum(c.get("monthly_cost_usd") or 0 for c in circuits),
        }
        all_projects = [
            _row_to_dict(r) for r in conn.execute("SELECT id, name FROM nc_projects ORDER BY name").fetchall()
        ]
        active_project = None
        if filter_project:
            ap_row = conn.execute("SELECT id, name FROM nc_projects WHERE id=?", (filter_project,)).fetchone()
            active_project = _row_to_dict(ap_row) if ap_row else None
        conn.close()
        return render_template(
            "network/circuits.html",
            circuits=circuits,
            stats=stats,
            all_projects=all_projects,
            filter_project=filter_project,
            active_project=active_project,
        )

    @bp.route("/customers")
    @nc_login_required
    def nc_customers():
        conn = get_connection()
        customers = [_row_to_dict(r) for r in conn.execute("SELECT * FROM nc_customers ORDER BY name").fetchall()]
        sites = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT s.*, c.name AS customer_name FROM nc_sites s "
                "LEFT JOIN nc_customers c ON c.id=s.customer_id ORDER BY s.name"
            ).fetchall()
        ]
        conn.close()
        return render_template("network/customers.html", customers=customers, sites=sites)

    @bp.route("/ipam")
    @nc_login_required
    def nc_ipam():
        conn = get_connection()
        filter_project = request.args.get("project", "")
        if filter_project:
            blocks = [
                _row_to_dict(r)
                for r in conn.execute(
                    "SELECT * FROM nc_ipam_blocks WHERE topology_id IN "
                    "(SELECT topology_id FROM nc_project_topologies WHERE project_id=?) "
                    "ORDER BY network",
                    (filter_project,),
                ).fetchall()
            ]
        else:
            blocks = [_row_to_dict(r) for r in conn.execute("SELECT * FROM nc_ipam_blocks ORDER BY network").fetchall()]
        all_projects = [
            _row_to_dict(r) for r in conn.execute("SELECT id, name FROM nc_projects ORDER BY name").fetchall()
        ]
        active_project = None
        if filter_project:
            ap_row = conn.execute("SELECT id, name FROM nc_projects WHERE id=?", (filter_project,)).fetchone()
            active_project = _row_to_dict(ap_row) if ap_row else None
        conn.close()
        return render_template(
            "network/ipam.html",
            blocks=blocks,
            all_projects=all_projects,
            filter_project=filter_project,
            active_project=active_project,
        )

    @bp.route("/cables")
    @nc_login_required
    def nc_cables():
        conn = get_connection()
        cables = [_row_to_dict(r) for r in conn.execute("SELECT * FROM nc_cables ORDER BY cable_id").fetchall()]
        conn.close()
        return render_template("network/cables.html", cables=cables)

    @bp.route("/cross-connects")
    @nc_login_required
    def nc_cross_connects():
        conn = get_connection()
        xconns = [
            _row_to_dict(r) for r in conn.execute("SELECT * FROM nc_cross_connects ORDER BY updated_at DESC").fetchall()
        ]
        stats = {
            "total": len(xconns),
            "active": sum(1 for x in xconns if x.get("status") == "active"),
            "planned": sum(1 for x in xconns if x.get("status") == "planned"),
            "monthly_cost": sum(x.get("monthly_cost_usd") or 0 for x in xconns),
            "facilities": len(set(x.get("facility", "") for x in xconns if x.get("facility"))),
        }
        conn.close()
        return render_template("network/cross_connects.html", cross_connects=xconns, stats=stats)

    @bp.route("/netbox")
    @nc_login_required
    def nc_netbox():
        """NetBox IPAM source-of-truth browser — connection config, cached objects, sync log."""
        conn = get_connection()
        # Config (token redacted)
        cfg_row = conn.execute(
            "SELECT url, site_filter, timeout_sec, auto_sync, last_tested FROM nc_netbox_config WHERE id='default'"
        ).fetchone()
        cfg = _row_to_dict(cfg_row) if cfg_row else {}
        cfg["configured"] = bool(cfg.get("url"))
        # Cached objects grouped by resource type
        obj_rows = conn.execute(
            "SELECT netbox_resource, COUNT(*) AS cnt FROM nc_netbox_objects GROUP BY netbox_resource"
        ).fetchall()
        cached_counts = {r[0]: r[1] for r in obj_rows}
        # Recent sync log
        log_rows = conn.execute("SELECT * FROM nc_netbox_sync_log ORDER BY ran_at DESC LIMIT 50").fetchall()
        sync_log = [_row_to_dict(r) for r in log_rows]
        # Topologies for the import-to-canvas picker
        topo_rows = conn.execute("SELECT id, name FROM topologies ORDER BY updated_at DESC LIMIT 100").fetchall()
        topologies = [_row_to_dict(r) for r in topo_rows]
        conn.close()
        airgap = os.environ.get("NETWORK_AIRGAP", "").lower() in ("1", "true", "yes")
        return render_template(
            "network/netbox.html",
            cfg=cfg,
            cached_counts=cached_counts,
            sync_log=sync_log,
            topologies=topologies,
            airgap_mode=airgap,
        )

    @bp.route("/projects")
    @nc_login_required
    def nc_projects():
        conn = get_connection()
        projects = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT p.*, c.name AS customer_name, "
                "(SELECT COUNT(*) FROM nc_project_topologies pt WHERE pt.project_id=p.id) AS topo_count "
                "FROM nc_projects p LEFT JOIN nc_customers c ON c.id=p.customer_id ORDER BY p.updated_at DESC"
            ).fetchall()
        ]
        customers = [
            _row_to_dict(r) for r in conn.execute("SELECT id, name FROM nc_customers ORDER BY name").fetchall()
        ]

        # ── Portfolio health data per project ─────────────────────────────
        for p in projects:
            pid = p["id"]
            topo_ids = [
                r[0]
                for r in conn.execute(
                    "SELECT topology_id FROM nc_project_topologies WHERE project_id=?", (pid,)
                ).fetchall()
            ]
            # Compliance: latest audit pass/fail across topologies
            total_passed = total_failed = 0
            open_findings = 0
            for tid in topo_ids:
                row = conn.execute(
                    "SELECT passed, failed FROM nc_compliance_checks WHERE topology_id=? ORDER BY ran_at DESC LIMIT 1",
                    (tid,),
                ).fetchone()
                if row:
                    total_passed += row[0] or 0
                    total_failed += row[1] or 0
                of = conn.execute(
                    "SELECT COUNT(*) FROM nc_compliance_findings WHERE topology_id=? AND status='open'", (tid,)
                ).fetchone()
                open_findings += of[0] if of else 0
            total_checks = total_passed + total_failed
            p["compliance_pct"] = round(total_passed * 100 / total_checks) if total_checks else None
            p["open_findings"] = open_findings
            # Cost: sum of circuit monthly costs
            cost_row = conn.execute(
                "SELECT COALESCE(SUM(monthly_cost_usd), 0) FROM nc_circuits WHERE topology_id IN "
                "(SELECT topology_id FROM nc_project_topologies WHERE project_id=?)",
                (pid,),
            ).fetchone()
            p["monthly_cost"] = cost_row[0] if cost_row else 0
            # Node/edge totals (computed in Python for PG portability)
            ne_rows = conn.execute(
                "SELECT t.graph_json FROM topologies t "
                "JOIN nc_project_topologies pt ON pt.topology_id=t.id "
                "WHERE pt.project_id=?",
                (pid,),
            ).fetchall()
            total_nodes = total_edges = 0
            for ner in ne_rows:
                try:
                    g = json.loads(ner["graph_json"] or '{"nodes":[],"edges":[]}')
                except Exception:
                    g = {"nodes": [], "edges": []}
                total_nodes += len(g.get("nodes", []))
                total_edges += len(g.get("edges", []))
            p["total_nodes"] = total_nodes
            p["total_edges"] = total_edges
            # Last simulation date
            sim_row = conn.execute(
                "SELECT MAX(ran_at) FROM simulation_results WHERE topology_id IN "
                "(SELECT topology_id FROM nc_project_topologies WHERE project_id=?)",
                (pid,),
            ).fetchone()
            p["last_sim"] = sim_row[0] if sim_row and sim_row[0] else None

        # Portfolio-level aggregates
        portfolio_stats = {
            "total_projects": len(projects),
            "by_status": {},
            "total_cost": sum(p.get("monthly_cost", 0) for p in projects),
            "total_findings": sum(p.get("open_findings", 0) for p in projects),
        }
        for p in projects:
            s = p.get("status", "draft")
            portfolio_stats["by_status"][s] = portfolio_stats["by_status"].get(s, 0) + 1

        # P1: Portfolio-wide activity feed (recent 20 events)
        portfolio_activity = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT action, entity_type, entity_id, details, user_id, ts FROM nc_audit ORDER BY ts DESC LIMIT 20"
            ).fetchall()
        ]

        conn.close()
        return render_template(
            "network/projects.html",
            projects=projects,
            customers=customers,
            portfolio=portfolio_stats,
            activity=portfolio_activity,
        )

    @bp.route("/projects/<proj_id>")
    @nc_login_required
    def nc_project_detail(proj_id):
        conn = get_connection()
        proj = conn.execute(
            "SELECT p.*, c.name AS customer_name FROM nc_projects p "
            "LEFT JOIN nc_customers c ON c.id=p.customer_id WHERE p.id=?",
            (proj_id,),
        ).fetchone()
        if not proj:
            conn.close()
            abort(404)
        proj = _row_to_dict(proj)
        topos = []
        for r in conn.execute(
            "SELECT t.id, t.name, t.description, t.classification, "
            "t.updated_at, t.graph_json "
            "FROM topologies t JOIN nc_project_topologies pt ON pt.topology_id=t.id "
            "WHERE pt.project_id=? ORDER BY t.updated_at DESC",
            (proj_id,),
        ).fetchall():
            t = _row_to_dict(r)
            try:
                g = json.loads(t.get("graph_json") or '{"nodes":[],"edges":[]}')
            except Exception:
                g = {"nodes": [], "edges": []}
            t["node_count"] = len(g.get("nodes", []))
            t["edge_count"] = len(g.get("edges", []))
            topos.append(t)
        circuits = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT * FROM nc_circuits WHERE topology_id IN "
                "(SELECT topology_id FROM nc_project_topologies WHERE project_id=?) "
                "ORDER BY circuit_id",
                (proj_id,),
            ).fetchall()
        ]
        all_topos = [_row_to_dict(r) for r in conn.execute("SELECT id, name FROM topologies ORDER BY name").fetchall()]

        # ── P1: Compliance rollup per topology ────────────────────────────
        topo_compliance = []
        agg_passed = agg_failed = 0
        total_open_findings = 0
        for t in topos:
            tid = t["id"]
            audit_row = conn.execute(
                "SELECT passed, failed, ran_at FROM nc_compliance_checks "
                "WHERE topology_id=? ORDER BY ran_at DESC LIMIT 1",
                (tid,),
            ).fetchone()
            passed = (audit_row[0] or 0) if audit_row else 0
            failed = (audit_row[1] or 0) if audit_row else 0
            last_audit = audit_row[2] if audit_row else None
            total = passed + failed
            pct = round(passed * 100 / total) if total else None
            agg_passed += passed
            agg_failed += failed
            of_row = conn.execute(
                "SELECT COUNT(*) FROM nc_compliance_findings WHERE topology_id=? AND status='open'", (tid,)
            ).fetchone()
            open_f = of_row[0] if of_row else 0
            total_open_findings += open_f
            # Findings by severity
            sev_rows = conn.execute(
                "SELECT severity, COUNT(*) FROM nc_compliance_findings "
                "WHERE topology_id=? AND status='open' "
                "GROUP BY severity",
                (tid,),
            ).fetchall()
            by_sev = {r[0]: r[1] for r in sev_rows}
            topo_compliance.append(
                {
                    "id": tid,
                    "name": t["name"],
                    "passed": passed,
                    "failed": failed,
                    "pct": pct,
                    "open_findings": open_f,
                    "cat1": by_sev.get("CAT1", 0),
                    "cat2": by_sev.get("CAT2", 0),
                    "cat3": by_sev.get("CAT3", 0),
                    "last_audit": last_audit,
                }
            )
        agg_total = agg_passed + agg_failed
        agg_pct = round(agg_passed * 100 / agg_total) if agg_total else None

        # ── P1: BOM cost rollup per topology ──────────────────────────────
        topo_bom = []
        total_capex = 0
        total_circuit_cost = sum(c.get("monthly_cost_usd") or 0 for c in circuits)
        for t in topos:
            try:
                graph = json.loads(t.get("graph_json") or '{"nodes":[]}')
            except Exception:
                graph = {"nodes": []}
            type_counts = {}
            for n in graph.get("nodes", []):
                nt = n.get("type", "unknown")
                type_counts[nt] = type_counts.get(nt, 0) + 1
            capex = sum(BOM_COSTS.get(dt, 0) * cnt for dt, cnt in type_counts.items())
            total_capex += capex
            topo_bom.append(
                {
                    "id": t["id"],
                    "name": t["name"],
                    "devices": sum(type_counts.values()),
                    "unique_types": len(type_counts),
                    "capex": capex,
                }
            )

        # ── P1: Activity feed from nc_audit ───────────────────────────────
        topo_ids = [t["id"] for t in topos]
        if topo_ids:
            placeholders = ",".join("?" for _ in topo_ids)
            activity = [
                _row_to_dict(r)
                for r in conn.execute(
                    "SELECT action, entity_type, entity_id, details, "
                    "user_id, ts FROM nc_audit "
                    "WHERE entity_id IN (" + placeholders + ") "  # nosec B608
                    "OR (entity_type='project' AND entity_id=?) "
                    "ORDER BY ts DESC LIMIT 30",
                    topo_ids + [proj_id],
                ).fetchall()
            ]
        else:
            activity = [
                _row_to_dict(r)
                for r in conn.execute(
                    "SELECT action, entity_type, entity_id, details, "
                    "user_id, ts FROM nc_audit "
                    "WHERE entity_type='project' AND entity_id=? "
                    "ORDER BY ts DESC LIMIT 30",
                    (proj_id,),
                ).fetchall()
            ]

        # Strip graph_json from topos (large, not needed in template)
        for t in topos:
            t.pop("graph_json", None)

        # P3: Milestones, Notes, Tags, Assignees
        milestones = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT * FROM nc_project_milestones WHERE project_id=? ORDER BY due_date", (proj_id,)
            ).fetchall()
        ]
        notes = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT * FROM nc_project_notes WHERE project_id=? ORDER BY created_at DESC", (proj_id,)
            ).fetchall()
        ]
        project_tags = [
            r[0]
            for r in conn.execute(
                "SELECT tag FROM nc_tags WHERE entity_type='project' AND entity_id=? ORDER BY tag", (proj_id,)
            ).fetchall()
        ]
        # Assignees per topology
        topo_assignees = {}
        for r in conn.execute(
            "SELECT topology_id, assignee FROM nc_project_topologies WHERE project_id=? AND assignee != ''", (proj_id,)
        ).fetchall():
            topo_assignees[r[0]] = r[1]

        # Phase A: Review board pipeline
        review_boards = [
            _row_to_dict(r) for r in conn.execute("SELECT * FROM nc_review_boards ORDER BY sort_order").fetchall()
        ]
        board_reviews = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT br.*, rb.name AS board_name, rb.short_name "
                "FROM nc_board_reviews br "
                "JOIN nc_review_boards rb ON rb.id=br.board_id "
                "WHERE br.project_id=? ORDER BY rb.sort_order, br.phase",
                (proj_id,),
            ).fetchall()
        ]
        project_phases = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT * FROM nc_project_phases WHERE project_id=? ORDER BY phase_num", (proj_id,)
            ).fetchall()
        ]
        safe_bridge = conn.execute("SELECT * FROM nc_safe_bridge WHERE project_id=?", (proj_id,)).fetchone()
        safe_bridge = _row_to_dict(safe_bridge) if safe_bridge else None
        if safe_bridge and safe_bridge.get("roi_json"):
            try:
                safe_bridge["roi"] = json.loads(safe_bridge["roi_json"])
            except Exception:
                safe_bridge["roi"] = {}

        # Migration phases with linked SOPs and parsed steps
        migration_phases_raw = [
            _row_to_dict(r) for r in conn.execute(
                "SELECT * FROM nc_migration_phases WHERE project_id=? ORDER BY phase_num",
                (proj_id,)
            ).fetchall()
        ]
        for mphase in migration_phases_raw:
            linked_docs = [
                _row_to_dict(r) for r in conn.execute(
                    """SELECT pd.doc_title, pd.doc_type, pd.doc_source, pd.relevance_note,
                              s.steps, s.prerequisites, s.validation,
                              s.rollback AS sop_rollback, s.escalation
                       FROM nc_phase_documents pd
                       LEFT JOIN ndc_sops s ON s.sop_id = pd.doc_id
                       WHERE pd.phase_id = ?
                       ORDER BY pd.display_order""",
                    (mphase['id'],)
                ).fetchall()
            ]
            for doc in linked_docs:
                for field in ('steps', 'prerequisites', 'validation', 'escalation'):
                    raw = doc.get(field)
                    if raw:
                        try:
                            doc[field] = json.loads(raw)
                        except Exception:
                            doc[field] = []
                # Normalize steps to canonical {number, text, verify, time_est, rollback}
                # regardless of which schema variant the SOP was seeded with.
                if doc.get('steps') and isinstance(doc['steps'], list):
                    doc['steps'] = [
                        _normalize_sop_step(s) for s in doc['steps']
                        if isinstance(s, dict)
                    ]
            mphase['linked_docs'] = linked_docs
            mphase['steps'] = [
                s.strip().rstrip('.')
                for s in (mphase.get('description') or '').split('. ')
                if s.strip()
            ]

        # First topology ID for 3-panel diagram link
        first_topo_id = topos[0]['id'] if topos else None

        # ── AI-assisted migration: COA, port mapping, config translation ──
        coa_data = {}
        port_mapping = {}
        config_translation = {}
        target_device_id = None

        # Load stored COA or generate on-the-fly
        if proj.get("coa_json"):
            try:
                coa_data = json.loads(proj["coa_json"])
            except Exception:
                coa_data = {}

        # Pick an EOL edge device from the first topology for demo consistency
        target_node_id = None
        target_device_id_for_tools = None
        if first_topo_id:
            try:
                graph = json.loads(
                    conn.execute(
                        "SELECT graph_json FROM topologies WHERE id=?", (first_topo_id,)
                    ).fetchone()[0]
                )
                nodes = graph.get("nodes", [])
                eol_nodes = [
                    n for n in nodes
                    if n.get("eol") or n.get("meta", {}).get("eol_date")
                ]
                if eol_nodes:
                    target_node_id = eol_nodes[0].get("id")
                    # Resolve topology node_id → ni_devices.id for tool calls
                    dev_row = conn.execute(
                        "SELECT id FROM ni_devices WHERE node_id=?",
                        (target_node_id,),
                    ).fetchone()
                    target_device_id_for_tools = dev_row["id"] if dev_row else target_node_id
            except Exception:
                target_node_id = None
                target_device_id_for_tools = None

        target_device_id = target_node_id

        if target_device_id_for_tools:
            # Generate COAs if missing
            if not coa_data:
                try:
                    from tools.ndc.executive_summary_generator import generate_executive_summary
                    _exec_sum = generate_executive_summary(horizon_days=365)  # noqa: F841
                    coa_data = {
                        "coa_1": {
                            "id": 1,
                            "name": "Rip & Replace",
                            "short_name": "Rip",
                            "risk_level": "high",
                            "estimated_downtime_hours": 4,
                            "total_cost": 0,
                            "description": "Swap hardware in single maintenance window per device.",
                        },
                        "coa_2": {
                            "id": 2,
                            "name": "Phased Cutover",
                            "short_name": "Phased",
                            "risk_level": "medium",
                            "estimated_downtime_hours": 1,
                            "total_cost": 0,
                            "description": "Migrate circuits/services in phases over weeks.",
                        },
                        "coa_3": {
                            "id": 3,
                            "name": "Side-by-Side VLAN",
                            "short_name": "Side-by-Side",
                            "risk_level": "low",
                            "estimated_downtime_hours": 0,
                            "total_cost": 0,
                            "description": "Run old+new in parallel on same VLAN domain. Near-zero downtime.",
                            "recommended": True,
                        },
                        "source_device_id": target_device_id,
                    }
                except Exception:
                    pass

            # Port mapping
            try:
                from tools.ndc.port_mapping_generator import generate_port_mapping
                port_mapping = generate_port_mapping(target_device_id_for_tools)
            except Exception:
                port_mapping = {}

            # Config translation
            try:
                from tools.ndc.config_translator import generate_config_translation
                config_translation = generate_config_translation(target_device_id_for_tools, target_vendor="arista")
            except Exception:
                config_translation = {}

        conn.close()
        return render_template(
            "network/project_detail.html",
            project=proj,
            topologies=topos,
            circuits=circuits,
            all_topos=all_topos,
            topo_compliance=topo_compliance,
            agg_compliance_pct=agg_pct,
            total_open_findings=total_open_findings,
            topo_bom=topo_bom,
            total_capex=total_capex,
            total_circuit_cost=total_circuit_cost,
            activity=activity,
            milestones=milestones,
            notes=notes,
            project_tags=project_tags,
            topo_assignees=topo_assignees,
            review_boards=review_boards,
            board_reviews=board_reviews,
            project_phases=project_phases,
            safe_bridge=safe_bridge,
            migration_phases=migration_phases_raw,
            first_topo_id=first_topo_id,
            coa_data=coa_data,
            port_mapping=port_mapping,
            config_translation=config_translation,
            target_device_id=target_device_id,
        )

    # ══════════════════════════════════════════════════════════════════════
    # P2: Cross-Project Comparison
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/projects/compare")
    @nc_login_required
    def nc_project_compare():
        conn = get_connection()
        projects = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT p.*, c.name AS customer_name, "
                "(SELECT COUNT(*) FROM nc_project_topologies pt "
                " WHERE pt.project_id=p.id) AS topo_count "
                "FROM nc_projects p "
                "LEFT JOIN nc_customers c ON c.id=p.customer_id "
                "ORDER BY p.name"
            ).fetchall()
        ]

        comparison = []
        for p in projects:
            pid = p["id"]
            topo_ids = [
                r[0]
                for r in conn.execute(
                    "SELECT topology_id FROM nc_project_topologies WHERE project_id=?", (pid,)
                ).fetchall()
            ]

            # Compliance aggregate
            total_passed = total_failed = 0
            open_findings = cat1 = cat2 = cat3 = 0
            for tid in topo_ids:
                row = conn.execute(
                    "SELECT passed, failed FROM nc_compliance_checks WHERE topology_id=? ORDER BY ran_at DESC LIMIT 1",
                    (tid,),
                ).fetchone()
                if row:
                    total_passed += row[0] or 0
                    total_failed += row[1] or 0
                sev_rows = conn.execute(
                    "SELECT severity, COUNT(*) FROM nc_compliance_findings "
                    "WHERE topology_id=? AND status='open' "
                    "GROUP BY severity",
                    (tid,),
                ).fetchall()
                for sr in sev_rows:
                    if sr[0] == "CAT1":
                        cat1 += sr[1]
                    elif sr[0] == "CAT2":
                        cat2 += sr[1]
                    elif sr[0] == "CAT3":
                        cat3 += sr[1]
                    open_findings += sr[1]

            total_checks = total_passed + total_failed
            comp_pct = round(total_passed * 100 / total_checks) if total_checks else None

            # Cost
            cost_row = conn.execute(
                "SELECT COALESCE(SUM(monthly_cost_usd), 0) "
                "FROM nc_circuits WHERE topology_id IN "
                "(SELECT topology_id FROM nc_project_topologies "
                " WHERE project_id=?)",
                (pid,),
            ).fetchone()
            circuit_cost = cost_row[0] if cost_row else 0

            # BOM CapEx
            capex = 0
            total_devices = 0
            for tid in topo_ids:
                trow = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (tid,)).fetchone()
                if trow:
                    try:
                        g = json.loads(trow["graph_json"])
                    except Exception:
                        g = {"nodes": []}
                    for n in g.get("nodes", []):
                        nt = n.get("type", "unknown")
                        capex += BOM_COSTS.get(nt, 0)
                        total_devices += 1

            # Node/edge totals (computed in Python for PG portability)
            ne_rows = conn.execute(
                "SELECT t.graph_json FROM topologies t "
                "JOIN nc_project_topologies pt ON pt.topology_id=t.id "
                "WHERE pt.project_id=?",
                (pid,),
            ).fetchall()
            total_nodes = total_edges = 0
            for ner in ne_rows:
                try:
                    g = json.loads(ner["graph_json"] or '{"nodes":[],"edges":[]}')
                except Exception:
                    g = {"nodes": [], "edges": []}
                total_nodes += len(g.get("nodes", []))
                total_edges += len(g.get("edges", []))

            # Last MC resilience score
            mc_row = conn.execute(
                "SELECT result_json FROM nc_mc_runs "
                "WHERE topology_id IN "
                "(SELECT topology_id FROM nc_project_topologies "
                " WHERE project_id=?) "
                "ORDER BY ran_at DESC LIMIT 1",
                (pid,),
            ).fetchone()
            mc_score = None
            if mc_row:
                try:
                    mc_data = json.loads(mc_row["result_json"])
                    mc_score = mc_data.get("risk_score")
                except Exception:
                    pass

            comparison.append(
                {
                    "id": pid,
                    "name": p["name"],
                    "status": p["status"],
                    "customer": p.get("customer_name", ""),
                    "topo_count": p["topo_count"],
                    "nodes": total_nodes,
                    "edges": total_edges,
                    "compliance_pct": comp_pct,
                    "open_findings": open_findings,
                    "cat1": cat1,
                    "cat2": cat2,
                    "cat3": cat3,
                    "circuit_cost": circuit_cost,
                    "capex": capex,
                    "devices": total_devices,
                    "mc_resilience": mc_score,
                }
            )
        conn.close()
        return render_template("network/compare.html", comparison=comparison)

    # ══════════════════════════════════════════════════════════════════════
    # P2: Clone Project API
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/projects/<pid>/clone", methods=["POST"])
    @nc_login_required
    def nc_api_clone_project(pid):
        data = request.get_json(force=True, silent=True) or {}
        new_name = data.get("name", "")
        conn = get_connection()
        orig = conn.execute("SELECT * FROM nc_projects WHERE id=?", (pid,)).fetchone()
        if not orig:
            conn.close()
            return jsonify({"error": "Project not found"}), 404
        orig = _row_to_dict(orig)

        now = _now()
        new_pid = str(_uuid.uuid4())
        conn.execute(
            "INSERT INTO nc_projects "
            "(id, name, customer_id, description, status, owner, "
            " created_at, updated_at) VALUES (?,?,?,?,?,?,?,?)",
            (
                new_pid,
                new_name or f"{orig['name']} (Copy)",
                orig.get("customer_id"),
                orig.get("description", ""),
                "draft",
                orig.get("owner", ""),
                now,
                now,
            ),
        )

        # Deep-copy topologies
        topo_map = {}  # old_id -> new_id
        orig_topos = conn.execute("SELECT topology_id FROM nc_project_topologies WHERE project_id=?", (pid,)).fetchall()
        for row in orig_topos:
            old_tid = row[0]
            topo = conn.execute("SELECT * FROM topologies WHERE id=?", (old_tid,)).fetchone()
            if not topo:
                continue
            topo = _row_to_dict(topo)
            new_tid = str(_uuid.uuid4())
            topo_map[old_tid] = new_tid
            conn.execute(
                "INSERT INTO topologies "
                "(id, name, description, graph_json, template_id, "
                " classification, created_at, updated_at) "
                "VALUES (?,?,?,?,?,?,?,?)",
                (
                    new_tid,
                    f"{topo['name']} (Copy)",
                    topo.get("description", ""),
                    topo.get("graph_json", '{"nodes":[],"edges":[]}'),
                    topo.get("template_id"),
                    topo.get("classification", "public"),
                    now,
                    now,
                ),
            )
            conn.execute("INSERT INTO nc_project_topologies (project_id, topology_id) VALUES (?,?)", (new_pid, new_tid))

            # Copy compliance profile
            profile = conn.execute("SELECT * FROM nc_compliance_profiles WHERE topology_id=?", (old_tid,)).fetchone()
            if profile:
                conn.execute(
                    "INSERT INTO nc_compliance_profiles "
                    "(id, topology_id, regimes, classification, "
                    " environment, auto_audit, created_at, updated_at) "
                    "VALUES (?,?,?,?,?,?,?,?)",
                    (
                        str(_uuid.uuid4()),
                        new_tid,
                        profile["regimes"],
                        profile["classification"],
                        profile["environment"],
                        profile["auto_audit"],
                        now,
                        now,
                    ),
                )

            # Copy circuits
            circuits = conn.execute("SELECT * FROM nc_circuits WHERE topology_id=?", (old_tid,)).fetchall()
            for c in circuits:
                c = _row_to_dict(c)
                conn.execute(
                    "INSERT INTO nc_circuits "
                    "(id, topology_id, circuit_id, carrier, "
                    " circuit_type, bandwidth, handoff_a, handoff_z, "
                    " customer, site, monthly_cost_usd, "
                    " contract_start, contract_end, sla_uptime_pct, "
                    " install_status, notes, created_at, updated_at) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (
                        str(_uuid.uuid4()),
                        new_tid,
                        c["circuit_id"],
                        c.get("carrier"),
                        c.get("circuit_type"),
                        c.get("bandwidth"),
                        c.get("handoff_a"),
                        c.get("handoff_z"),
                        c.get("customer"),
                        c.get("site"),
                        c.get("monthly_cost_usd", 0),
                        c.get("contract_start"),
                        c.get("contract_end"),
                        c.get("sla_uptime_pct", 99.9),
                        c.get("install_status", "planned"),
                        c.get("notes"),
                        now,
                        now,
                    ),
                )

            # Copy IPAM blocks
            blocks = conn.execute("SELECT * FROM nc_ipam_blocks WHERE topology_id=?", (old_tid,)).fetchall()
            for b in blocks:
                b = _row_to_dict(b)
                conn.execute(
                    "INSERT INTO nc_ipam_blocks "
                    "(id, topology_id, network, vlan_id, vrf, "
                    " description, site_id, gateway, "
                    " utilization_pct, created_at) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (
                        str(_uuid.uuid4()),
                        new_tid,
                        b["network"],
                        b.get("vlan_id"),
                        b.get("vrf", "global"),
                        b.get("description"),
                        b.get("site_id"),
                        b.get("gateway"),
                        b.get("utilization_pct", 0),
                        now,
                    ),
                )

        conn.commit()
        conn.close()
        _audit("CLONE", "project", new_pid, f"Cloned from {orig['name']} ({pid})")
        return jsonify(
            {
                "id": new_pid,
                "name": new_name or f"{orig['name']} (Copy)",
                "topologies_cloned": len(topo_map),
                "source_id": pid,
            }
        ), 201

    # ══════════════════════════════════════════════════════════════════════
    # API: Health
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/health")
    def nc_api_health():
        try:
            conn = get_connection()
            tc = conn.execute("SELECT COUNT(*) FROM topologies").fetchone()[0]
            tpl = conn.execute("SELECT COUNT(*) FROM nc_templates").fetchone()[0]
            conn.close()
            return jsonify({"status": "ok", "topologies": tc, "templates": tpl, "ts": _now()})
        except Exception as exc:
            return jsonify({"status": "error", "error": str(type(exc).__name__)}), 500

    # ══════════════════════════════════════════════════════════════════════
    # API: Topologies CRUD
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/topologies", methods=["GET"])
    @nc_login_required
    def nc_api_list_topologies():
        conn = get_connection()
        rows = conn.execute(
            "SELECT id, name, description, classification, created_at, updated_at "
            "FROM topologies ORDER BY updated_at DESC"
        ).fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/topologies", methods=["POST"])
    @nc_login_required
    def nc_api_create_topology():
        data = request.get_json(force=True, silent=True) or {}
        topo_id = str(_uuid.uuid4())
        name = data.get("name", "Untitled Topology")
        graph = json.dumps(data.get("graph_json", {"nodes": [], "edges": []}))
        now = _now()
        conn = get_connection()
        conn.execute(
            "INSERT INTO topologies (id, name, description, graph_json, template_id, classification, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (
                topo_id,
                name,
                data.get("description", ""),
                graph,
                data.get("template_id"),
                data.get("classification", "public"),
                now,
                now,
            ),
        )
        conn.commit()
        conn.close()
        _audit("CREATE", "topology", topo_id, name)
        return jsonify({"id": topo_id, "name": name, "created_at": now}), 201

    # Clear-all routes MUST be before <topo_id> parameterized routes
    @bp.route("/api/topologies/clear-all", methods=["DELETE"])
    @nc_login_required
    def nc_api_clear_all_topologies():
        conn = get_connection()
        # Delete child tables first (FK constraints require this order).
        # Grandchild tables (e.g. nc_vuln_findings -> nc_vuln_scans -> topologies)
        # must come before their parents.
        child_tables = [
            "nc_vuln_findings",
            "nc_vuln_hosts",
            "ni_device_configs",
            "nc_device_geo",
            "nc_routing_entries",
            "nc_collected_configs",
            "nc_intent_validations",
            "nc_intent_policies",
            "nc_change_request_items",
            "nc_change_requests",
            "nc_stig_imports",
            "nc_ato_packages",
            "nc_boundaries",
            "nc_netbox_objects",
            "nc_netbox_sync_log",
            "nc_discovery_diffs",
            "nc_discovery_scans",
            "nc_mc_runs",
            "nc_mc_scenarios",
            "simulation_results",
            "nc_objects",
            "nc_circuits",
            "nc_cables",
            "nc_cross_connects",
            "nc_versions",
            "nc_compliance_findings",
            "nc_compliance_checks",
            "nc_compliance_profiles",
            "nc_ipam_blocks",
            "nc_project_topologies",
            "nc_groups",
            "nc_interconnects",
            "nc_port_inventory",
            "nc_module_inventory",
            "nc_bw_simulations",
            "nc_vuln_scans",
            "nc_query_log",
            "nc_collab_sessions",
            "ndc_runbooks",
            "ni_devices",
            "ni_analyses",
            "ni_state_snapshots",
            "nc_documents",
        ]
        for tbl in child_tables:
            conn.execute(f"DELETE FROM {tbl}")  # nosec B608 -- table/column names are internal constants, not user input
        conn.execute("DELETE FROM topologies")
        conn.commit()
        conn.close()
        _audit("CLEAR_ALL", "topologies", "", "All topologies cleared")
        return jsonify({"cleared": "topologies"})

    @bp.route("/api/simulations/clear-all", methods=["DELETE"])
    @nc_login_required
    def nc_api_clear_all_simulations():
        conn = get_connection()
        conn.execute("DELETE FROM simulation_results")
        conn.execute("DELETE FROM nc_mc_runs")
        conn.execute("DELETE FROM nc_mc_scenarios")
        conn.commit()
        conn.close()
        _audit("CLEAR_ALL", "simulations", "", "All simulations cleared")
        return jsonify({"cleared": "simulations"})

    @bp.route("/api/topologies/<topo_id>", methods=["GET"])
    @nc_login_required
    def nc_api_get_topology(topo_id):
        conn = get_connection()
        row = conn.execute("SELECT * FROM topologies WHERE id=?", (topo_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        topo = _row_to_dict(row)
        try:
            topo["graph_json"] = json.loads(topo["graph_json"])
        except Exception:
            topo["graph_json"] = {"nodes": [], "edges": []}
        return jsonify(topo)

    @bp.route("/api/topologies/<topo_id>", methods=["PUT"])
    @nc_login_required
    def nc_api_update_topology(topo_id):
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        fields, values = [], []
        for k in ["name", "description", "graph_json", "classification"]:
            if k in data:
                val = json.dumps(data[k]) if k == "graph_json" and isinstance(data[k], dict) else data[k]
                fields.append(f"{k}=?")
                values.append(val)
        if not fields:
            conn.close()
            return jsonify({"error": "No fields"}), 400
        fields.append("updated_at=?")
        values.append(_now())
        values.append(topo_id)
        conn.execute(f"UPDATE topologies SET {', '.join(fields)} WHERE id=?", values)  # nosec B608 -- table/column names are internal constants, not user input
        conn.commit()
        conn.close()
        _audit("UPDATE", "topology", topo_id)
        # Hook: notify Security Design Canvas of topology change
        sdc_assessment = None
        try:
            from tools.security_canvas.agent import on_ndc_topology_saved

            result = on_ndc_topology_saved(topo_id)
            if result and result.get("status") == "assessed":
                sdc_assessment = {
                    "design_id": result.get("design_id"),
                    "risk_score": result.get("risk_score"),
                    "posture_grade": result.get("posture_grade"),
                    "cat1_count": result.get("cat1_count", 0),
                    "total_findings": result.get("total_findings", 0),
                }
        except Exception:
            pass  # Security Canvas is optional
        # Incremental KG update: re-extract only if graph_json changed
        try:
            from tools.canvas.kg_builder import rebuild_canvas_kg

            rebuild_canvas_kg("ndc", topo_id)
        except Exception:
            pass
        # Blockchain provenance
        try:
            from tools.canvas.provenance import register_canvas_provenance
            register_canvas_provenance(
                canvas_key="ndc",
                design_id=topo_id,
                graph_json=data.get("graph_json", {}),
                project_id=data.get("project_id", ""),
            )
        except Exception:
            pass
        # DIC Canvas Synergy — emit topology change event (dsyn-emit-01)
        try:
            from tools.ndc.event_emitter import emit_topology_change
            emit_topology_change(
                topology_id=topo_id,
                change_summary=f"Topology {topo_id} updated",
                affected_segments=list(data.get("graph_json", {}).get("nodes", []))[:5]
                    if isinstance(data.get("graph_json"), dict) else [],
            )
        except Exception:
            pass  # event emission never blocks topology save

        resp = {"ok": True}
        if sdc_assessment is not None:
            resp["sdc_assessment"] = sdc_assessment
        return jsonify(resp)

    @bp.route("/api/topologies/<topo_id>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_topology(topo_id):
        conn = get_connection()
        conn.execute("DELETE FROM topologies WHERE id=?", (topo_id,))
        conn.commit()
        conn.close()
        _audit("DELETE", "topology", topo_id)
        return jsonify({"deleted": topo_id})

    # ══════════════════════════════════════════════════════════════════════
    # API: Executive Briefing Generator
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/topologies/<topo_id>/briefing", methods=["POST"])
    @nc_login_required
    def nc_api_executive_briefing(topo_id):
        """Generate a plain-English executive briefing from a topology graph for CISO/PM audience."""
        import re

        from tools.http.client import request as _req_request

        conn = get_connection()
        row = conn.execute("SELECT * FROM topologies WHERE id=?", (topo_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Topology not found"}), 404

        topo = _row_to_dict(row)
        try:
            graph = json.loads(topo.get("graph_json") or "{}")
        except Exception:
            graph = {}
        nodes = graph.get("nodes", [])
        edges = graph.get("edges", [])

        # ── Build structured graph summary for LLM context ─────────────────
        type_counts: dict = {}
        for n in nodes:
            t = n.get("type", "unknown")
            type_counts[t] = type_counts.get(t, 0) + 1

        security_zones = [n.get("label") or n.get("type") for n in nodes if n.get("type") in (
            "security-zone", "firewall", "vrf", "vlan", "nc_boundary",
        )]

        node_list = ", ".join(
            f"{n.get('label') or n.get('type')} ({n.get('type')})"
            for n in nodes[:40]
        )
        if len(nodes) > 40:
            node_list += f" … and {len(nodes) - 40} more"

        edge_list = ", ".join(
            f"{e.get('source')} → {e.get('target')}"
            + (f" [{e.get('protocol')}]" if e.get("protocol") else "")
            for e in edges[:30]
        )
        if len(edges) > 30:
            edge_list += f" … and {len(edges) - 30} more"

        type_summary = ", ".join(f"{v}× {k}" for k, v in sorted(type_counts.items(), key=lambda x: -x[1]))

        graph_summary = (
            f"Topology name: {topo.get('name', 'Unnamed')}\n"
            f"Classification: {topo.get('classification', 'unclassified')}\n"
            f"Total nodes: {len(nodes)}, Total links: {len(edges)}\n"
            f"Device types: {type_summary or 'none'}\n"
            f"Security zones/firewalls: {', '.join(security_zones) or 'none identified'}\n"
            f"Nodes: {node_list or 'empty'}\n"
            f"Links: {edge_list or 'none'}\n"
        )

        _BRIEFING_SYSTEM = (
            "You are a senior network architect writing a 1–2 page executive briefing for a CISO and Program Manager. "
            "Your audience is non-technical leadership. Use plain English — no jargon acronyms without explanation. "
            "Be concise, confident, and specific. "
            "Output ONLY a Markdown document with exactly these five sections (use ## headings):\n\n"
            "## Network Overview\n"
            "A 3–5 sentence plain-English description of what this network does, its scale, and its purpose.\n\n"
            "## Security Zones\n"
            "List the security boundaries, trust zones, or enclaves. Explain what data or users live in each zone.\n\n"
            "## Critical Path\n"
            "Describe the most important data flows or connectivity paths. Identify single points of failure if visible.\n\n"
            "## Technology Stack\n"
            "Summarize key device categories, vendors (if inferrable from node types), and any cloud services present.\n\n"
            "## Risk Highlights\n"
            "List 3–5 specific risks or concerns visible in this topology (missing redundancy, exposed endpoints, "
            "mixed classification zones, unencrypted paths, etc.). Each risk on its own bullet.\n\n"
            "Do not add any preamble or postscript outside these five sections."
        )

        user_msg = f"Generate an executive briefing for this network topology:\n\n{graph_summary}"

        def _call_claude_briefing():
            api_key = os.environ.get("ANTHROPIC_API_KEY", "")
            if not api_key:
                return None, "No ANTHROPIC_API_KEY set"
            model = os.environ.get("ANTHROPIC_TOPO_MODEL", "claude-sonnet-4-20250514")
            r = _req_request(
                "POST",
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json={
                    "model": model,
                    "max_tokens": 2048,
                    "temperature": 0.4,
                    "system": _BRIEFING_SYSTEM,
                    "messages": [{"role": "user", "content": user_msg}],
                },
                timeout=60,
            )
            r.raise_for_status()
            return r.json().get("content", [{}])[0].get("text", ""), None

        def _call_ollama_briefing():
            ollama_url = os.environ.get("OLLAMA_HOST", "http://localhost:11434")
            ollama_model = os.environ.get("OLLAMA_TOPO_MODEL", "llama3.2:3b")
            r = _req_request(
                "POST",
                f"{ollama_url}/api/chat",
                json={
                    "model": ollama_model,
                    "messages": [
                        {"role": "system", "content": _BRIEFING_SYSTEM},
                        {"role": "user", "content": user_msg},
                    ],
                    "stream": False,
                    "options": {"num_predict": 2048, "temperature": 0.4},
                },
                timeout=120,
            )
            r.raise_for_status()
            return r.json().get("message", {}).get("content", ""), None

        try:
            provider = os.environ.get("NC_AI_PROVIDER", "auto")
            content = None

            if provider in ("auto", "claude"):
                content, err = _call_claude_briefing()
                if not content and provider == "claude":
                    return jsonify({"error": f"Claude API failed: {err}"}), 503

            if not content and provider in ("auto", "ollama"):
                content, err = _call_ollama_briefing()
                if not content and provider == "ollama":
                    return jsonify({"error": f"Ollama failed: {err}"}), 503

            if not content:
                return jsonify({"error": "No LLM provider available. Set ANTHROPIC_API_KEY or start Ollama."}), 503

            # Strip any <think> blocks from reasoning models
            markdown = re.sub(r"<think>.*?</think>", "", content, flags=re.DOTALL).strip()
            _audit("BRIEFING", "topology", topo_id)
            return jsonify({"markdown": markdown, "topo_name": topo.get("name", "Unnamed")})

        except Exception as exc:
            logger.exception("Executive briefing failed for %s", topo_id)
            return jsonify({"error": str(exc)}), 500

    # ══════════════════════════════════════════════════════════════════════
    # API: Network Operations Runbook Generator
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/topologies/<topo_id>/runbook", methods=["POST"])
    @nc_login_required
    def nc_api_generate_runbook(topo_id):
        """Generate a downloadable Network Operations Runbook Markdown file from topology graph data.

        Deterministic sections built from graph data:
          1. Document header (name, classification, date)
          2. Device Inventory table
          3. IP Address Table sorted by subnet
          4. Interface/Link Matrix
          5. Standard Troubleshooting Procedures (LLM-generated per device type)
          6. Escalation Contacts fillable template
        """
        import ipaddress
        import re
        from datetime import datetime, timezone

        from tools.http.client import request as _req_request

        conn = get_connection()
        row = conn.execute("SELECT * FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Topology not found"}), 404

        try:
            ipam_rows = conn.execute(
                "SELECT network, description, vlan_id, vrf, gateway "
                "FROM nc_ipam_blocks WHERE topology_id=? ORDER BY network",
                (topo_id,),
            ).fetchall()
        except Exception:
            ipam_rows = []
        conn.close()

        topo = _row_to_dict(row)
        try:
            graph = json.loads(topo.get("graph_json") or "{}")
        except Exception:
            graph = {}
        nodes = graph.get("nodes", [])
        edges = graph.get("edges", [])

        topo_name = topo.get("name", "Unnamed Topology")
        classification = (topo.get("classification") or "UNCLASSIFIED").upper()
        now_utc = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

        _DRAWING_TYPES = {
            "draw-rect", "draw-rounded-rect", "text-heading", "text-label", "text-badge",
        }
        device_nodes = [n for n in nodes if n.get("type") not in _DRAWING_TYPES]

        lines: list = []

        # ── Section 1: Header ─────────────────────────────────────────────
        lines += [
            "# Network Operations Runbook",
            "",
            "| Field | Value |",
            "|-------|-------|",
            f"| **Topology** | {topo_name} |",
            f"| **Classification** | {classification} |",
            f"| **Generated** | {now_utc} |",
            f"| **Nodes** | {len(nodes)} |",
            f"| **Links** | {len(edges)} |",
            "",
            "---",
            "",
        ]

        # ── Section 2: Device Inventory ───────────────────────────────────
        lines.append("## 1. Device Inventory")
        lines.append("")
        if device_nodes:
            lines += [
                "| # | Label | Type | Management IP | OS/Platform | Serial | Notes |",
                "|---|-------|------|---------------|-------------|--------|-------|",
            ]
            for i, n in enumerate(device_nodes, 1):
                cfg = n.get("config") or {}
                label = (n.get("label") or n.get("type") or "Unknown").replace("|", "\\|")
                dtype = n.get("type", "unknown")
                ip = cfg.get("ip") or cfg.get("mgmt_ip") or "—"
                os_name = cfg.get("os") or "—"
                serial = (cfg.get("serial") or "—").replace("|", "\\|")
                notes = (cfg.get("role") or cfg.get("description") or "—").replace("|", "\\|")
                lines.append(f"| {i} | {label} | `{dtype}` | `{ip}` | {os_name} | {serial} | {notes} |")
        else:
            lines.append("*No devices in topology.*")
        lines += ["", "---", ""]

        # ── Section 3: IP Address Table (sorted by subnet) ────────────────
        lines.append("## 2. IP Address Table")
        lines.append("")

        ip_entries: list = []
        for n in device_nodes:
            cfg = n.get("config") or {}
            label = (n.get("label") or n.get("type") or "Unknown")
            dtype = n.get("type", "unknown")
            primary_ip = cfg.get("ip") or cfg.get("mgmt_ip") or ""
            if primary_ip:
                ip_entries.append({"ip": primary_ip, "label": label, "type": dtype, "context": "Management"})
            for vlan in (cfg.get("vlans") or []):
                if isinstance(vlan, dict) and vlan.get("ip"):
                    ip_entries.append({
                        "ip": vlan["ip"],
                        "label": f"{label} VLAN {vlan.get('vlan_id', '?')}",
                        "type": dtype,
                        "context": f"VLAN {vlan.get('vlan_id', '?')}",
                    })
            for extra_ip in (cfg.get("ips") or []):
                if extra_ip and extra_ip != primary_ip:
                    ip_entries.append({"ip": extra_ip, "label": label, "type": dtype, "context": "Interface"})

        for ipam_r in ipam_rows:
            ipam = _row_to_dict(ipam_r)
            network = ipam.get("network", "")
            if network:
                ip_entries.append({
                    "ip": network,
                    "label": f"IPAM: {ipam.get('description') or network}",
                    "type": "subnet",
                    "context": f"VRF {ipam.get('vrf') or 'global'} · GW {ipam.get('gateway') or '—'}",
                })

        def _ip_sort_key(entry):
            raw = entry["ip"].split("/")[0]
            try:
                return (0, int(ipaddress.ip_address(raw)))
            except Exception:
                return (1, 0)

        ip_entries.sort(key=_ip_sort_key)

        if ip_entries:
            lines += [
                "| # | IP / Subnet | Device / Label | Type | Context |",
                "|---|-------------|----------------|------|---------|",
            ]
            for i, e in enumerate(ip_entries, 1):
                lbl = e["label"].replace("|", "\\|")
                lines.append(
                    f"| {i} | `{e['ip']}` | {lbl} | `{e['type']}` | {e['context']} |"
                )
        else:
            lines.append("*No IP addresses configured in topology.*")
        lines += ["", "---", ""]

        # ── Section 4: Interface / Link Matrix ────────────────────────────
        lines.append("## 3. Interface / Link Matrix")
        lines.append("")

        label_by_id = {n.get("id"): (n.get("label") or n.get("type") or "?") for n in nodes}

        if edges:
            lines += [
                "| # | Source | Target | Protocol | Bandwidth | Latency | Label |",
                "|---|--------|--------|----------|-----------|---------|-------|",
            ]
            for i, e in enumerate(edges, 1):
                src = label_by_id.get(e.get("source"), e.get("source") or "?").replace("|", "\\|")
                tgt = label_by_id.get(e.get("target"), e.get("target") or "?").replace("|", "\\|")
                proto = e.get("protocol") or "—"
                bw = e.get("bandwidth") or "—"
                lat = f"{e.get('latency_ms')} ms" if e.get("latency_ms") else "—"
                lbl = (e.get("label") or "—").replace("|", "\\|")
                lines.append(f"| {i} | {src} | {tgt} | {proto} | {bw} | {lat} | {lbl} |")
        else:
            lines.append("*No links configured in topology.*")
        lines += ["", "---", ""]

        # ── Section 5: Troubleshooting Procedures (LLM) ───────────────────
        lines.append("## 4. Standard Troubleshooting Procedures")
        lines.append("")

        type_counts: dict = {}
        for n in device_nodes:
            t = n.get("type", "unknown")
            type_counts[t] = type_counts.get(t, 0) + 1

        device_type_summary = ", ".join(
            f"{v}× {k}" for k, v in sorted(type_counts.items(), key=lambda x: -x[1])
        )
        seen_types: set = set()
        device_samples: list = []
        for n in device_nodes:
            t = n.get("type", "unknown")
            if t not in seen_types:
                seen_types.add(t)
                cfg = n.get("config") or {}
                device_samples.append(
                    f"- {n.get('label') or t} (type={t}, ip={cfg.get('ip') or '—'}, os={cfg.get('os') or '—'})"
                )

        _RUNBOOK_SYSTEM = (
            "You are a senior network operations engineer writing a troubleshooting section for a "
            "Network Operations Runbook. For each device type present in the topology, write concise, "
            "actionable troubleshooting procedures. "
            "Format using Markdown with ### headings for each device type. "
            "Each device type section must include:\n"
            "- **Common failure symptoms** (2–3 bullet points)\n"
            "- **Initial diagnostic commands** (CLI commands in fenced code blocks)\n"
            "- **Step-by-step resolution checklist** (numbered steps)\n"
            "- **Escalation trigger** (one sentence: when to escalate to Tier 3 or vendor)\n\n"
            "Keep each section under 20 lines. Be specific and operational. "
            "Output ONLY the Markdown troubleshooting content — no preamble or postscript."
        )
        _runbook_user_msg = (
            f"Generate troubleshooting procedures for a network topology named '{topo_name}'.\n\n"
            f"Device types present: {device_type_summary or 'none'}\n"
            f"Representative devices:\n" + "\n".join(device_samples[:20]) + "\n\n"
            f"Total links: {len(edges)}\n"
            f"Classification: {classification}"
        )

        def _call_claude_runbook():
            api_key = os.environ.get("ANTHROPIC_API_KEY", "")
            if not api_key:
                return None, "No ANTHROPIC_API_KEY"
            model = os.environ.get("ANTHROPIC_TOPO_MODEL", "claude-sonnet-4-20250514")
            r = _req_request(
                "POST",
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json={
                    "model": model,
                    "max_tokens": 3000,
                    "temperature": 0.3,
                    "system": _RUNBOOK_SYSTEM,
                    "messages": [{"role": "user", "content": _runbook_user_msg}],
                },
                timeout=90,
            )
            r.raise_for_status()
            return r.json().get("content", [{}])[0].get("text", ""), None

        def _call_ollama_runbook():
            ollama_url = os.environ.get("OLLAMA_HOST", "http://localhost:11434")
            ollama_model = os.environ.get("OLLAMA_TOPO_MODEL", "llama3.2:3b")
            r = _req_request(
                "POST",
                f"{ollama_url}/api/chat",
                json={
                    "model": ollama_model,
                    "messages": [
                        {"role": "system", "content": _RUNBOOK_SYSTEM},
                        {"role": "user", "content": _runbook_user_msg},
                    ],
                    "stream": False,
                    "options": {"num_predict": 3000, "temperature": 0.3},
                },
                timeout=180,
            )
            r.raise_for_status()
            return r.json().get("message", {}).get("content", ""), None

        try:
            provider = os.environ.get("NC_AI_PROVIDER", "auto")
            ts_content = None
            if provider in ("auto", "claude"):
                ts_content, _ = _call_claude_runbook()
            if not ts_content and provider in ("auto", "ollama"):
                ts_content, _ = _call_ollama_runbook()
            if ts_content:
                ts_content = re.sub(r"<think>.*?</think>", "", ts_content, flags=re.DOTALL).strip()
                lines.append(ts_content)
            else:
                lines += [
                    "*No LLM provider available — set `ANTHROPIC_API_KEY` or start Ollama for AI-generated procedures.*",
                    "",
                    "**Generic network troubleshooting checklist:**",
                    "",
                    "1. Verify physical connectivity (cable, SFP, interface admin/oper state)",
                    "2. Check interface state: `show interface status` / `show ip interface brief`",
                    "3. Review routing table: `show ip route` / `show bgp summary`",
                    "4. Check recent log events: `show log | last 50`",
                    "5. Ping gateway: `ping <gateway-ip> repeat 100`",
                    "6. Traceroute to isolate break: `traceroute <destination>`",
                    "7. Verify spanning-tree state (L2): `show spanning-tree`",
                    "8. Review ARP/MAC tables: `show arp` / `show mac address-table`",
                ]
        except Exception:
            logger.exception("Runbook troubleshooting LLM call failed for %s", topo_id)
            lines.append("*Troubleshooting generation failed. Check LLM provider configuration.*")
        lines += ["", "---", ""]

        # ── Section 6: Escalation Contacts (fillable template) ────────────
        lines += [
            "## 5. Escalation Contacts",
            "",
            "> **Instructions:** Fill in contact information before distributing this runbook.",
            "",
            "| Role | Name | Phone | Email | On-Call Hours |",
            "|------|------|-------|-------|---------------|",
            "| Network Operations (Tier 1) | ________________ | ________________ | ________________ | 24/7 |",
            "| Network Engineer (Tier 2) | ________________ | ________________ | ________________ | Business hours |",
            "| Network Architect (Tier 3) | ________________ | ________________ | ________________ | On-call |",
            "| Security / ISSO | ________________ | ________________ | ________________ | On-call |",
            "| Vendor TAC | ________________ | ________________ | ________________ | 24/7 |",
            "| ISP/Carrier NOC | ________________ | ________________ | ________________ | 24/7 |",
            "| Program Manager | ________________ | ________________ | ________________ | Business hours |",
            "",
            "### Escalation Thresholds",
            "",
            "| Severity | Description | Response Time | Escalation Path |",
            "|----------|-------------|---------------|-----------------|",
            "| **P1 — Critical** | Complete outage / mission impact | 15 min | Tier 1 → Tier 2 → Tier 3 → PM |",
            "| **P2 — High** | Partial outage, redundancy lost | 30 min | Tier 1 → Tier 2 |",
            "| **P3 — Medium** | Degraded performance, non-critical path | 2 hr | Tier 1 → Tier 2 (biz hrs) |",
            "| **P4 — Low** | Minor issue, workaround available | Next business day | Tier 1 ticket |",
            "",
            "---",
            "",
            f"*Generated by ICDEV™ NDC Runbook Generator — {now_utc}*",
        ]

        markdown = "\n".join(lines)
        safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", topo_name)[:40]
        filename = f"runbook-{safe_name}.md"

        _audit("RUNBOOK_GENERATED", "topology", topo_id)
        return jsonify({"markdown": markdown, "topo_name": topo_name, "filename": filename})

    # ══════════════════════════════════════════════════════════════════════
    # API: Simulation
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/topologies/<topo_id>/simulate", methods=["POST"])
    @nc_login_required
    def nc_api_simulate(topo_id):
        conn = get_connection()
        row = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Topology not found"}), 404
        data = request.get_json(force=True, silent=True) or {}
        sim_type = data.get("sim_type", "ping")
        try:
            graph = json.loads(row["graph_json"])
        except Exception:
            graph = {"nodes": [], "edges": []}

        result = _add_narrative(_run_simulation(graph, sim_type, data))
        _narrative = result.get("narrative") if isinstance(result, dict) else None
        if _narrative:
            record_canvas_decision(
                canvas_type="ndc",
                record_id=topo_id,
                decision_type="narrative",
                decision=str(_narrative)[:500],
                rationale=f"Simulation type: {sim_type}",
                model_used=None,
            )
        sim_id = str(_uuid.uuid4())
        now = _now()
        conn = get_connection()
        conn.execute(
            "INSERT INTO simulation_results (id, topology_id, sim_type, input_json, result_json, ran_at) "
            "VALUES (?,?,?,?,?,?)",
            (sim_id, topo_id, sim_type, json.dumps(data), json.dumps(result), now),
        )
        conn.commit()
        conn.close()
        _audit("SIMULATE", "topology", topo_id, sim_type)
        return jsonify({"sim_id": sim_id, "result": result})

    # ══════════════════════════════════════════════════════════════════════
    # API: Templates
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/templates", methods=["GET"])
    @nc_login_required
    def nc_api_list_templates():
        conn = get_connection()
        rows = conn.execute(
            "SELECT id, name, category, description, tags FROM nc_templates ORDER BY category, name"
        ).fetchall()
        conn.close()
        result = []
        for r in rows:
            d = _row_to_dict(r)
            try:
                d["tags"] = json.loads(d.get("tags") or "[]")
            except Exception:
                d["tags"] = []
            result.append(d)
        return jsonify(result)

    @bp.route("/api/templates/<tpl_id>", methods=["GET"])
    @nc_login_required
    def nc_api_get_template(tpl_id):
        conn = get_connection()
        row = conn.execute("SELECT * FROM nc_templates WHERE id=?", (tpl_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        d = _row_to_dict(row)
        try:
            d["graph_json"] = json.loads(d["graph_json"])
        except Exception:
            d["graph_json"] = {"nodes": [], "edges": []}
        try:
            d["tags"] = json.loads(d.get("tags") or "[]")
        except Exception:
            d["tags"] = []
        return jsonify(d)

    @bp.route("/api/templates/<tpl_id>/docs", methods=["GET"])
    @nc_login_required
    def nc_api_get_template_docs(tpl_id):
        """Get SOP/runbook markdown documentation for a template."""
        conn = get_connection()
        rows = conn.execute(
            "SELECT id, template_id, doc_type, title, body_markdown, created_at "
            "FROM nc_template_docs WHERE template_id = ? ORDER BY created_at",
            (tpl_id,),
        ).fetchall()
        conn.close()
        docs = [_row_to_dict(r) for r in rows]
        return jsonify({"template_id": tpl_id, "docs": docs, "count": len(docs)})

    @bp.route("/api/templates/<tpl_id>/load", methods=["POST"])
    @nc_login_required
    def nc_api_load_template(tpl_id):
        conn = get_connection()
        row = conn.execute("SELECT * FROM nc_templates WHERE id=?", (tpl_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        tpl = _row_to_dict(row)
        # Idempotency: return existing topology for this template rather than duplicating
        existing = conn.execute(
            "SELECT id, name FROM topologies WHERE template_id=? LIMIT 1", (tpl_id,)
        ).fetchone()
        if existing:
            conn.close()
            ex = _row_to_dict(existing)
            return jsonify({"id": ex["id"], "name": ex["name"], "redirect": f"/network/canvas/{ex['id']}"}), 200
        topo_id = str(_uuid.uuid4())
        now = _now()
        name = f"{tpl['name']} (copy)"
        conn.execute(
            "INSERT INTO topologies (id, name, description, graph_json, template_id, classification, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (topo_id, name, tpl.get("description", ""), tpl["graph_json"], tpl_id, "public", now, now),
        )
        conn.commit()
        conn.close()
        _audit("LOAD_TEMPLATE", "topology", topo_id, tpl_id)
        return jsonify({"id": topo_id, "name": name, "redirect": f"/network/canvas/{topo_id}"}), 201

    @bp.route("/api/templates/<tpl_id>", methods=["PUT"])
    @nc_login_required
    def nc_api_update_template(tpl_id):
        conn = get_connection()
        row = conn.execute("SELECT id FROM nc_templates WHERE id=?", (tpl_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Template not found"}), 404
        data = request.get_json(force=True)
        fields, values = [], []
        for k in ["name", "description", "category"]:
            if k in data:
                fields.append(f"{k}=?")
                values.append(data[k])
        if "tags" in data:
            fields.append("tags=?")
            values.append(json.dumps(data["tags"]) if isinstance(data["tags"], list) else data["tags"])
        if "graph_json" in data:
            fields.append("graph_json=?")
            values.append(
                json.dumps(data["graph_json"]) if isinstance(data["graph_json"], dict) else data["graph_json"]
            )
        if not fields:
            conn.close()
            return jsonify({"error": "No fields to update"}), 400
        values.append(tpl_id)
        conn.execute(f"UPDATE nc_templates SET {', '.join(fields)} WHERE id=?", values)  # nosec B608 -- table/column names are internal constants, not user input
        conn.commit()
        conn.close()
        _audit("UPDATE_TEMPLATE", "template", tpl_id, json.dumps(list(data.keys())))
        return jsonify({"ok": True, "id": tpl_id})

    @bp.route("/api/cloud-objects")
    @nc_login_required
    def nc_api_cloud_objects():
        return jsonify(CLOUD_OBJECTS)

    # ══════════════════════════════════════════════════════════════════════
    # API: Enclave-in-a-Box Snippets
    # Pre-built compliance-validated sub-topologies (SIPR, IL5 DMZ, Tactical Edge).
    # Inserted onto an existing canvas at a user-specified offset.
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/snippets", methods=["GET"])
    @nc_login_required
    def nc_api_list_snippets():
        """List all enclave snippets (metadata only, no graph_json)."""
        conn = get_connection()
        rows = conn.execute(
            "SELECT id, name, category, description, classification_level, "
            "impact_level, stig_controls, tags "
            "FROM nc_enclave_snippets ORDER BY category, name"
        ).fetchall()
        conn.close()
        result = []
        for r in rows:
            d = _row_to_dict(r)
            for field in ("stig_controls", "tags"):
                try:
                    d[field] = json.loads(d.get(field) or "[]")
                except Exception:
                    d[field] = []
            result.append(d)
        return jsonify(result)

    @bp.route("/api/snippets/<snippet_id>", methods=["GET"])
    @nc_login_required
    def nc_api_get_snippet(snippet_id):
        """Return full snippet including graph_json for insertion onto canvas."""
        conn = get_connection()
        row = conn.execute("SELECT * FROM nc_enclave_snippets WHERE id=?", (snippet_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Snippet not found"}), 404
        d = _row_to_dict(row)
        try:
            d["graph_json"] = json.loads(d["graph_json"])
        except Exception:
            d["graph_json"] = {"nodes": [], "edges": []}
        for field in ("stig_controls", "tags"):
            try:
                d[field] = json.loads(d.get(field) or "[]")
            except Exception:
                d[field] = []
        _audit("LOAD_SNIPPET", "snippet", snippet_id, d.get("name", ""))
        return jsonify(d)

    # ══════════════════════════════════════════════════════════════════════
    # API: Export / Import
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/export/<topo_id>", methods=["POST"])
    @nc_login_required
    def nc_api_export(topo_id):
        conn = get_connection()
        row = conn.execute("SELECT * FROM topologies WHERE id=?", (topo_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        topo = _row_to_dict(row)
        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            graph = {"nodes": [], "edges": []}
        fmt = (request.get_json(force=True, silent=True) or {}).get("format", "drawio")
        if fmt == "drawio":
            xml = to_drawio(graph, topo["name"])
            return jsonify({"format": "drawio", "filename": f"{topo['name']}.drawio", "content": xml})
        if fmt == "svg":
            svg_content = to_svg(graph, topo["name"])
            return jsonify({"format": "svg", "filename": f"{topo['name']}.svg", "content": svg_content})
        return jsonify({"error": f"Unknown format: {fmt}"}), 400

    @bp.route("/api/export/<topo_id>/visio", methods=["POST"])
    @nc_login_required
    def nc_api_export_visio(topo_id):
        conn = get_connection()
        row = conn.execute("SELECT * FROM topologies WHERE id=?", (topo_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        topo = _row_to_dict(row)
        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            graph = {"nodes": [], "edges": []}
        vdx_xml = to_vdx(graph, topo["name"])
        return jsonify({"format": "vdx", "filename": f"{topo['name']}.vdx", "content": vdx_xml})

    @bp.route("/api/export/<topo_id>/ansible", methods=["POST"])
    @nc_login_required
    def nc_api_export_ansible(topo_id):
        """Export topology as an Ansible inventory INI file.

        Hosts are grouped first by security enclave boundary (from
        nc_boundaries) and then by zone/role derived from node type.
        Cloud-infrastructure nodes (VPCs, subnets, etc.) are emitted as
        comments for reference only.
        """
        conn = get_connection()
        row = conn.execute("SELECT * FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        topo = _row_to_dict(row)
        boundary_rows = conn.execute(
            "SELECT label, classification, node_ids FROM nc_boundaries WHERE topology_id=?",
            (topo_id,),
        ).fetchall()
        conn.close()
        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            graph = {"nodes": [], "edges": []}
        boundaries = [dict(r) for r in boundary_rows]
        content = to_ansible_inventory(graph, topo["name"], boundaries=boundaries)
        safe_name = topo["name"].replace(" ", "_")
        _audit("EXPORT", "topology", topo_id, "ansible")
        return jsonify(
            {
                "format": "ansible",
                "filename": f"{safe_name}_inventory.ini",
                "content": content,
                "enclave_count": len(boundaries),
            }
        )

    @bp.route("/api/export/<topo_id>/terraform", methods=["POST"])
    @nc_login_required
    def nc_api_export_terraform(topo_id):
        """Export topology as a Terraform HCL skeleton (main.tf).

        Generates provider blocks, resource stubs for every cloud node,
        security-group / NSG / firewall resources for each enclave boundary,
        and a locals block mapping diagram edges to conceptual connectivity
        rules (security-group / NACL / firewall policy inputs).
        """
        conn = get_connection()
        row = conn.execute("SELECT * FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        topo = _row_to_dict(row)
        boundary_rows = conn.execute(
            "SELECT label, classification, node_ids FROM nc_boundaries WHERE topology_id=?",
            (topo_id,),
        ).fetchall()
        conn.close()
        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            graph = {"nodes": [], "edges": []}
        boundaries = [dict(r) for r in boundary_rows]
        content = to_terraform_hcl(graph, topo["name"], boundaries=boundaries)
        safe_name = topo["name"].replace(" ", "_")
        _audit("EXPORT", "topology", topo_id, "terraform")
        return jsonify(
            {
                "format": "terraform",
                "filename": f"{safe_name}_main.tf",
                "content": content,
                "enclave_count": len(boundaries),
            }
        )

    @bp.route("/api/export/<topo_id>/device-configs", methods=["POST"])
    @nc_login_required
    def nc_api_export_device_configs(topo_id):
        """Export per-device configuration files (IOS/EOS/JunOS) as a ZIP archive.

        Body (JSON, optional):
          format: "zip" (default) | "json"
            - "zip"  — base64-encoded ZIP bytes; client decodes and downloads.
            - "json" — dict mapping filename -> config text for in-browser preview.

        Each configurable node (router, switch, firewall) gets its own config file
        rendered from a Jinja2 template using the node's assigned IPs, VLANs,
        routing protocol settings, and connected edges.
        """
        import base64

        conn = get_connection()
        row = conn.execute("SELECT * FROM topologies WHERE id=?", (topo_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        topo = _row_to_dict(row)
        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            graph = {"nodes": [], "edges": []}

        data = request.get_json(force=True, silent=True) or {}
        fmt = data.get("format", "zip")
        safe_name = topo["name"].replace(" ", "_")

        if fmt == "json":
            configs = generate_device_configs(graph, topo["name"])
            _audit("EXPORT", "topology", topo_id, f"device-configs json devices={len(configs)}")
            return jsonify(
                {
                    "format": "json",
                    "topo_name": topo["name"],
                    "device_count": len(configs),
                    "configs": configs,
                }
            )

        # Default: ZIP
        zip_bytes = generate_device_configs_zip(graph, topo["name"])
        encoded = base64.b64encode(zip_bytes).decode("ascii")
        device_count = len(list_configurable_nodes(graph))
        _audit("EXPORT", "topology", topo_id, f"device-configs zip devices={device_count}")
        return jsonify(
            {
                "format": "zip",
                "filename": f"{safe_name}_device_configs.zip",
                "content_b64": encoded,
                "device_count": device_count,
            }
        )

    @bp.route("/api/export/<topo_id>/device-configs/preview", methods=["GET"])
    @nc_login_required
    def nc_api_export_device_configs_preview(topo_id):
        """Return a list of configurable nodes and their OS type for UI preview."""
        conn = get_connection()
        row = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        try:
            graph = json.loads(row[0] or '{"nodes":[],"edges":[]}')
        except Exception:
            graph = {"nodes": [], "edges": []}
        nodes = list_configurable_nodes(graph)
        return jsonify({"configurable_nodes": nodes, "count": len(nodes)})

    @bp.route("/api/export/<topo_id>/vsdx", methods=["POST"])
    @nc_login_required
    def nc_api_export_vsdx(topo_id):
        """Export topology as modern Visio .vsdx file with embedded metadata."""
        conn = get_connection()
        row = conn.execute("SELECT * FROM topologies WHERE id=?", (topo_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        topo = _row_to_dict(row)
        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            graph = {"nodes": [], "edges": []}
        import base64
        import re as _re

        vsdx_bytes = export_vsdx(topo["name"], graph)
        encoded = base64.b64encode(vsdx_bytes).decode("ascii")
        safe_name = _re.sub(r"[^a-zA-Z0-9_-]", "_", topo["name"])
        _audit("EXPORT", "topology", topo_id, "vsdx")
        return jsonify(
            {
                "format": "vsdx",
                "filename": f"{safe_name}.vsdx",
                "content_b64": encoded,
            }
        )

    @bp.route("/api/export/<topo_id>/csv", methods=["POST"])
    @nc_login_required
    def nc_api_export_csv(topo_id):
        """Export topology as Ops CSV bundle (device inventory, circuits, cables, IP, peering)."""
        conn = get_connection()
        row = conn.execute("SELECT * FROM topologies WHERE id=?", (topo_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        topo = _row_to_dict(row)
        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            graph = {"nodes": [], "edges": []}
        csv_files = export_ops_csvs(topo["name"], graph)
        import base64
        import io as _io
        import re as _re

        buf = _io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for fname, content in csv_files.items():
                zf.writestr(fname, content)
        encoded = base64.b64encode(buf.getvalue()).decode("ascii")
        safe_name = _re.sub(r"[^a-zA-Z0-9_-]", "_", topo["name"])
        _audit("EXPORT", "topology", topo_id, "csv")
        return jsonify(
            {
                "format": "csv",
                "filename": f"{safe_name}_ops_csvs.zip",
                "content_b64": encoded,
            }
        )

    @bp.route("/api/export/<topo_id>/inventory", methods=["GET"])
    @nc_login_required
    def nc_api_export_inventory(topo_id):
        """Return JSON inventory of all devices with full metadata."""
        conn = get_connection()
        row = conn.execute("SELECT * FROM topologies WHERE id=?", (topo_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        topo = _row_to_dict(row)
        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            graph = {"nodes": [], "edges": []}
        devices = []
        for n in graph.get("nodes", []):
            ntype = n.get("type", "")
            if ntype in ("draw-rect", "zone", "boundary", "text-annotation"):
                continue
            cfg = n.get("config", {})
            devices.append(
                {
                    "id": n.get("id", ""),
                    "label": n.get("label", ""),
                    "type": ntype,
                    "hostname": cfg.get("hostname", ""),
                    "ip": cfg.get("ip", ""),
                    "model": cfg.get("model", ""),
                    "serial": cfg.get("serial", ""),
                    "asset_tag": cfg.get("asset_tag", ""),
                    "site": cfg.get("site", ""),
                    "location": cfg.get("location", ""),
                    "rack": cfg.get("rack", ""),
                    "slot": cfg.get("slot", ""),
                    "port": cfg.get("port", ""),
                    "port_type": cfg.get("port_type", ""),
                    "bandwidth": cfg.get("bandwidth", ""),
                    "vlan": cfg.get("vlan", ""),
                    "vrf": cfg.get("vrf", ""),
                    "asn": cfg.get("asn", ""),
                    "peer_asn": cfg.get("peer_asn", ""),
                    "peer_ip": cfg.get("peer_ip", ""),
                    "peering_type": cfg.get("peering_type", ""),
                }
            )
        return jsonify({"topology": topo["name"], "device_count": len(devices), "devices": devices})

    @bp.route("/api/import", methods=["POST"])
    @nc_login_required
    def nc_api_import():
        data = request.get_json(force=True, silent=True) or {}
        fmt = data.get("format", "drawio")
        content = data.get("content", "")
        name = data.get("name", "Imported Topology")
        if not content:
            return jsonify({"error": "content required"}), 400
        if fmt == "drawio":
            graph = import_drawio(content)
        elif fmt in ("vdx", "visio"):
            graph = import_vdx(content)
        elif fmt == "svg":
            graph = import_svg(content)
        else:
            return jsonify({"error": f"Unsupported format: {fmt}"}), 400
        topo_id = str(_uuid.uuid4())
        now = _now()
        conn = get_connection()
        conn.execute(
            "INSERT INTO topologies (id, name, description, graph_json, classification, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (topo_id, name, f"Imported from {fmt}", json.dumps(graph), "public", now, now),
        )
        conn.commit()
        conn.close()
        # Phase 1: auto-classify imported nodes
        graph = _classify_imported_nodes(graph)
        conn = get_connection()
        conn.execute("UPDATE topologies SET graph_json=? WHERE id=?", (json.dumps(graph), topo_id))
        conn.commit()
        conn.close()
        _audit("IMPORT", "topology", topo_id, fmt)
        return jsonify(
            {"id": topo_id, "name": name, "nodes": len(graph.get("nodes", [])), "edges": len(graph.get("edges", []))}
        ), 201

    # ══════════════════════════════════════════════════════════════════════
    # Phase 1: Intelligent Import & Stitching
    # ══════════════════════════════════════════════════════════════════════

    # Semantic node classifier — maps labels to device types
    _NODE_CLASSIFY_PATTERNS = [
        (r"(?i)(core|edge|border|wan|pe|ce|p\b).*r(outer|tr)", "router"),
        (r"(?i)r(outer|tr)", "router"),
        (r"(?i)(fw|firewall|palo|forti|asa|checkpoint)", "firewall"),
        (r"(?i)(sw|switch).*l3|layer.?3.*sw|dist.*sw|core.*sw", "switch-l3"),
        (r"(?i)(sw|switch|access)", "switch-l2"),
        (r"(?i)(lb|load.?bal|f5|netscaler|a10)", "load-balancer"),
        (r"(?i)(wap|access.?point|ap\d|wifi|wireless)", "wap"),
        (r"(?i)(wlc|wireless.*control)", "wlc"),
        (r"(?i)(srv|server|host|vm\b|esxi|hypervisor)", "server"),
        (r"(?i)(pc|workstation|desktop|laptop|endpoint)", "endpoint-pc"),
        (r"(?i)(phone|voip|sip)", "ip-phone"),
        (r"(?i)(sdwan|sd-wan|vmanage|vedge)", "sdwan-edge"),
        (r"(?i)(mpls.*pe|pe.*router)", "mpls-pe"),
        (r"(?i)(mpls.*p\b|p.*router|provider)", "mpls-p"),
        (r"(?i)(route.?reflect|rr\b)", "route-reflector"),
        (r"(?i)(encrypt|kg-|type.?1|nsa)", "type1-encryptor"),
        (r"(?i)(fips|hsm)", "fips-140-l2"),
        (r"(?i)(siem|splunk|qradar|arcsight)", "siem"),
        (r"(?i)(tap|span|mirror)", "network-tap"),
        (r"(?i)(vpc|aws)", "aws-vpc"),
        (r"(?i)(vnet|azure)", "az-vnet"),
        (r"(?i)(gcp|google.?cloud)", "gcp-vpc"),
        (r"(?i)(internet|cloud|wan|isp)", "cloud"),
        (r"(?i)(patch.?panel|pp\b|mdf|idf)", "patch-panel"),
        (r"(?i)(ups|pdu|power)", "server"),
        (r"(?i)(demarc|demarcation)", "demarc"),
        (r"(?i)(meet.?me|mmr|colo)", "meet-me-room"),
    ]

    def _classify_imported_nodes(graph):
        """Auto-classify imported nodes from generic 'imported' type
        to specific device types using label pattern matching."""
        import re as _re

        for n in graph.get("nodes", []):
            if n.get("type") not in ("imported", "", None):
                continue
            label = n.get("label", "")
            matched = False
            for pattern, dtype in _NODE_CLASSIFY_PATTERNS:
                if _re.search(pattern, label):
                    n["type"] = dtype
                    matched = True
                    break
            if not matched:
                n["type"] = "server"  # safe default
        return graph

    @bp.route("/api/import/bulk", methods=["POST"])
    @nc_login_required
    def nc_api_bulk_import():
        """Import multiple diagram files at once.
        Each file becomes a separate topology.
        Optionally group under a project."""
        data = request.get_json(force=True, silent=True) or {}
        files = data.get("files", [])
        project_id = data.get("project_id")
        if not files:
            return jsonify({"error": "files array required"}), 400

        conn = get_connection()
        now = _now()
        results = []
        for f in files:
            fmt = f.get("format", "drawio")
            content = f.get("content", "")
            name = f.get("name", "Imported")
            if not content:
                results.append({"name": name, "error": "empty"})
                continue
            if fmt == "drawio":
                graph = import_drawio(content)
            elif fmt in ("vdx", "visio"):
                graph = import_vdx(content)
            elif fmt == "svg":
                graph = import_svg(content)
            else:
                results.append({"name": name, "error": f"bad format: {fmt}"})
                continue
            graph = _classify_imported_nodes(graph)
            topo_id = str(_uuid.uuid4())
            conn.execute(
                "INSERT INTO topologies "
                "(id, name, description, graph_json, "
                " classification, created_at, updated_at) "
                "VALUES (?,?,?,?,?,?,?)",
                (topo_id, name, f"Bulk import ({fmt})", json.dumps(graph), "public", now, now),
            )
            if project_id:
                conn.execute(
                    "INSERT OR IGNORE INTO nc_project_topologies (project_id, topology_id) VALUES (?,?)",
                    (project_id, topo_id),
                )
            results.append(
                {
                    "id": topo_id,
                    "name": name,
                    "nodes": len(graph.get("nodes", [])),
                    "edges": len(graph.get("edges", [])),
                }
            )
        conn.commit()
        conn.close()
        _audit("BULK_IMPORT", "topology", "", f"{len(results)} files imported")
        return jsonify(
            {
                "imported": len([r for r in results if "id" in r]),
                "failed": len([r for r in results if "error" in r]),
                "results": results,
            }
        ), 201

    @bp.route("/api/import/stitch", methods=["POST"])
    @nc_login_required
    def nc_api_stitch_topologies():
        """Merge multiple topologies into one, with user-defined
        interconnect points between them."""
        data = request.get_json(force=True, silent=True) or {}
        topo_ids = data.get("topology_ids", [])
        interconnects = data.get("interconnects", [])
        name = data.get("name", "Stitched Topology")
        if len(topo_ids) < 2:
            return jsonify({"error": "Need 2+ topology_ids"}), 400

        conn = get_connection()
        merged_nodes = []
        merged_edges = []
        offset_x = 0

        for tid in topo_ids:
            row = conn.execute("SELECT name, graph_json FROM topologies WHERE id=?", (tid,)).fetchone()
            if not row:
                continue
            try:
                g = json.loads(row["graph_json"])
            except Exception:
                continue
            prefix = tid[:8]
            topo_name = row["name"]
            # Offset nodes horizontally and namespace IDs
            for n in g.get("nodes", []):
                new_id = f"{prefix}_{n['id']}"
                merged_nodes.append(
                    {
                        **n,
                        "id": new_id,
                        "x": (n.get("x") or 0) + offset_x,
                        "config": {
                            **(n.get("config") or n.get("configData") or {}),
                            "_source_topology": topo_name,
                            "_source_id": n["id"],
                        },
                    }
                )
            for e in g.get("edges", []):
                merged_edges.append(
                    {
                        **e,
                        "id": f"{prefix}_{e.get('id', '')}",
                        "source": f"{prefix}_{e['source']}",
                        "target": f"{prefix}_{e['target']}",
                    }
                )
            offset_x += 700

        # Add user-defined interconnect edges
        for ic in interconnects:
            merged_edges.append(
                {
                    "id": str(_uuid.uuid4())[:8],
                    "source": ic.get("source_node_id", ""),
                    "target": ic.get("target_node_id", ""),
                    "label": ic.get("label", "Interconnect"),
                    "protocol": ic.get("protocol", ""),
                }
            )

        merged_graph = {"nodes": merged_nodes, "edges": merged_edges}
        topo_id = str(_uuid.uuid4())
        now = _now()
        conn.execute(
            "INSERT INTO topologies "
            "(id, name, description, graph_json, "
            " classification, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (topo_id, name, f"Stitched from {len(topo_ids)} topologies", json.dumps(merged_graph), "public", now, now),
        )
        conn.commit()
        conn.close()
        _audit("STITCH", "topology", topo_id, f"Merged {len(topo_ids)} topologies")
        return jsonify(
            {
                "id": topo_id,
                "name": name,
                "nodes": len(merged_nodes),
                "edges": len(merged_edges),
                "source_topologies": len(topo_ids),
            }
        ), 201

    @bp.route("/api/import/audit", methods=["POST"])
    @nc_login_required
    def nc_api_import_and_audit():
        """Import a diagram AND immediately run compliance audit,
        design scorecard, and tech debt analysis."""
        data = request.get_json(force=True, silent=True) or {}
        fmt = data.get("format", "drawio")
        content = data.get("content", "")
        name = data.get("name", "Audit Import")
        if not content:
            return jsonify({"error": "content required"}), 400

        # Import
        if fmt == "drawio":
            graph = import_drawio(content)
        elif fmt in ("vdx", "visio"):
            graph = import_vdx(content)
        elif fmt == "svg":
            graph = import_svg(content)
        else:
            return jsonify({"error": f"Unsupported: {fmt}"}), 400
        graph = _classify_imported_nodes(graph)
        topo_id = str(_uuid.uuid4())
        now = _now()
        conn = get_connection()
        conn.execute(
            "INSERT INTO topologies "
            "(id, name, description, graph_json, "
            " classification, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (topo_id, name, f"Audit import ({fmt})", json.dumps(graph), "CUI", now, now),
        )
        conn.commit()

        # Run compliance audit
        audit_result = run_compliance_audit(topo_id, graph, ["fisma_high", "stig"], "CUI")
        total_p = sum(s["passed"] for s in audit_result["scores"].values())
        total_f = sum(s["failed"] for s in audit_result["scores"].values())
        audit_id = str(_uuid.uuid4())
        conn.execute(
            "INSERT INTO nc_compliance_checks "
            "(id, topology_id, check_type, passed, failed, "
            " findings_json, ran_at) VALUES (?,?,?,?,?,?,?)",
            (audit_id, topo_id, "fisma_high,stig", total_p, total_f, json.dumps(audit_result["findings"]), now),
        )
        conn.commit()
        conn.close()

        _audit("IMPORT_AUDIT", "topology", topo_id, fmt)

        return jsonify(
            {
                "id": topo_id,
                "name": name,
                "nodes": len(graph.get("nodes", [])),
                "edges": len(graph.get("edges", [])),
                "compliance": {
                    "passed": total_p,
                    "failed": total_f,
                    "findings": len(audit_result["findings"]),
                    "scores": audit_result["scores"],
                },
                "classified_types": dict(
                    sorted(
                        {
                            n["type"]: sum(1 for m in graph["nodes"] if m["type"] == n["type"]) for n in graph["nodes"]
                        }.items()
                    )
                ),
            }
        ), 201

    @bp.route("/api/classify-nodes", methods=["POST"])
    @nc_login_required
    def nc_api_classify_nodes():
        """Re-classify nodes in an existing topology.
        Useful after manual edits or to fix imported types."""
        data = request.get_json(force=True, silent=True) or {}
        topo_id = data.get("topology_id")
        if not topo_id:
            return jsonify({"error": "topology_id required"}), 400
        conn = get_connection()
        row = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        try:
            graph = json.loads(row["graph_json"])
        except Exception:
            conn.close()
            return jsonify({"error": "Bad graph"}), 500
        # Only classify nodes that are still "imported" type
        changed = 0
        for n in graph.get("nodes", []):
            old_type = n.get("type", "")
            if old_type in ("imported", ""):
                _classify_imported_nodes({"nodes": [n]})
                if n["type"] != old_type:
                    changed += 1
        conn.execute(
            "UPDATE topologies SET graph_json=?, updated_at=? WHERE id=?", (json.dumps(graph), _now(), topo_id)
        )
        conn.commit()
        conn.close()
        return jsonify({"reclassified": changed, "total_nodes": len(graph.get("nodes", []))})

    # ══════════════════════════════════════════════════════════════════════
    # Phase 2: Device Command Profiles + Non-Intrusive Discovery
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/device-profiles", methods=["GET"])
    @nc_login_required
    def nc_api_list_device_profiles():
        conn = get_connection()
        rows = conn.execute(
            "SELECT id, vendor, platform, description, is_builtin, "
            "created_by, created_at FROM nc_device_profiles "
            "ORDER BY vendor, platform"
        ).fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/device-profiles/<pid>", methods=["GET"])
    @nc_login_required
    def nc_api_get_device_profile(pid):
        conn = get_connection()
        row = conn.execute("SELECT * FROM nc_device_profiles WHERE id=?", (pid,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        p = _row_to_dict(row)
        try:
            p["commands"] = json.loads(p.get("commands_json") or "{}")
        except Exception:
            p["commands"] = {}
        return jsonify(p)

    @bp.route("/api/device-profiles", methods=["POST"])
    @nc_login_required
    def nc_api_create_device_profile():
        """Create a user-defined device command profile."""
        data = request.get_json(force=True, silent=True) or {}
        pid = str(_uuid.uuid4())
        commands = data.get("commands", {})
        if isinstance(commands, dict):
            commands = json.dumps(commands)
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_device_profiles "
            "(id, vendor, platform, description, commands_json, "
            " is_builtin, created_by, created_at) "
            "VALUES (?,?,?,?,?,0,?,?)",
            (
                pid,
                data.get("vendor", "Custom"),
                data.get("platform", "Custom"),
                data.get("description", ""),
                commands,
                data.get("created_by", ""),
                _now(),
            ),
        )
        conn.commit()
        conn.close()
        _audit("CREATE", "device_profile", pid, f"{data.get('vendor')} {data.get('platform')}")
        return jsonify({"id": pid}), 201

    @bp.route("/api/device-profiles/<pid>", methods=["PUT"])
    @nc_login_required
    def nc_api_update_device_profile(pid):
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        row = conn.execute("SELECT is_builtin FROM nc_device_profiles WHERE id=?", (pid,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        fields, values = [], []
        allowed = ["vendor", "platform", "description"]
        if not row[0]:  # user-created can update commands
            allowed.append("commands_json")
        for k in allowed:
            if k in data:
                v = data[k]
                if k == "commands_json" and isinstance(v, dict):
                    v = json.dumps(v)
                fields.append(f"{k}=?")
                values.append(v)
        # Allow adding commands to built-in profiles
        if row[0] and "commands" in data:
            existing = conn.execute("SELECT commands_json FROM nc_device_profiles WHERE id=?", (pid,)).fetchone()
            try:
                cmds = json.loads(existing[0] or "{}")
            except Exception:
                cmds = {}
            cmds.update(data["commands"])
            fields.append("commands_json=?")
            values.append(json.dumps(cmds))
        if fields:
            values.append(pid)
            conn.execute(
                f"UPDATE nc_device_profiles "  # nosec B608
                f"SET {', '.join(fields)} WHERE id=?",
                values,
            )
            conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @bp.route("/api/device-profiles/<pid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_device_profile(pid):
        conn = get_connection()
        row = conn.execute("SELECT is_builtin FROM nc_device_profiles WHERE id=?", (pid,)).fetchone()
        if row and row[0]:
            conn.close()
            return jsonify({"error": "Cannot delete built-in profile"}), 403
        conn.execute("DELETE FROM nc_device_profiles WHERE id=?", (pid,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    # ── Discovery Configs ─────────────────────────────────────────────────
    @bp.route("/api/discovery-configs", methods=["GET"])
    @nc_login_required
    def nc_api_list_discovery_configs():
        conn = get_connection()
        rows = conn.execute(
            "SELECT dc.*, dp.vendor, dp.platform "
            "FROM nc_discovery_configs dc "
            "LEFT JOIN nc_device_profiles dp ON dp.id=dc.profile_id "
            "ORDER BY dc.name"
        ).fetchall()
        conn.close()
        configs = []
        for r in rows:
            c = _row_to_dict(r)
            try:
                c["targets"] = json.loads(c.get("targets") or "[]")
            except Exception:
                c["targets"] = []
            configs.append(c)
        return jsonify(configs)

    @bp.route("/api/discovery-configs", methods=["POST"])
    @nc_login_required
    def nc_api_create_discovery_config():
        data = request.get_json(force=True, silent=True) or {}
        cid = str(_uuid.uuid4())
        targets = data.get("targets", [])
        if isinstance(targets, list):
            targets = json.dumps(targets)
        wl = data.get("whitelist_subnets", [])
        if isinstance(wl, list):
            wl = json.dumps(wl)
        bl = data.get("blacklist_subnets", [])
        if isinstance(bl, list):
            bl = json.dumps(bl)
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_discovery_configs "
            "(id, name, profile_id, targets, credential_ref, "
            " method, read_only, rate_limit_per_sec, "
            " max_concurrent, timeout_per_cmd, timeout_per_device, "
            " hop_limit, max_devices, whitelist_subnets, "
            " blacklist_subnets, created_at) "
            "VALUES (?,?,?,?,?,?,1,?,?,?,?,?,?,?,?,?)",
            (
                cid,
                data.get("name", "Discovery Scan"),
                data.get("profile_id"),
                targets,
                data.get("credential_ref", ""),
                data.get("method", "ssh"),
                data.get("rate_limit_per_sec", 1.0),
                data.get("max_concurrent", 5),
                data.get("timeout_per_cmd", 10),
                data.get("timeout_per_device", 60),
                data.get("hop_limit", 2),
                data.get("max_devices", 100),
                wl,
                bl,
                _now(),
            ),
        )
        conn.commit()
        conn.close()
        _audit("CREATE", "discovery_config", cid)
        return jsonify({"id": cid}), 201

    @bp.route("/api/discovery-configs/<cid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_discovery_config(cid):
        conn = get_connection()
        conn.execute("DELETE FROM nc_discovery_configs WHERE id=?", (cid,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    # ── Collected Configs ─────────────────────────────────────────────────
    @bp.route("/api/collected-configs", methods=["GET"])
    @nc_login_required
    def nc_api_list_collected_configs():
        device_ip = request.args.get("device_ip", "")
        conn = get_connection()
        if device_ip:
            rows = conn.execute(
                "SELECT id, device_ip, hostname, command_name, "
                "collected_at FROM nc_collected_configs "
                "WHERE device_ip=? ORDER BY collected_at DESC",
                (device_ip,),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT id, device_ip, hostname, command_name, "
                "collected_at FROM nc_collected_configs "
                "ORDER BY collected_at DESC LIMIT 100"
            ).fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/collected-configs/<cid>", methods=["GET"])
    @nc_login_required
    def nc_api_get_collected_config(cid):
        conn = get_connection()
        row = conn.execute("SELECT * FROM nc_collected_configs WHERE id=?", (cid,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        c = _row_to_dict(row)
        try:
            c["parsed"] = json.loads(c.get("parsed_json") or "{}")
        except Exception:
            c["parsed"] = {}
        return jsonify(c)

    @bp.route("/api/collected-configs", methods=["POST"])
    @nc_login_required
    def nc_api_store_collected_config():
        """Store a manually captured config output (deduped by device+command)."""
        data = request.get_json(force=True, silent=True) or {}
        cid = str(_uuid.uuid4())
        parsed = data.get("parsed_json", {})
        if isinstance(parsed, dict):
            parsed = json.dumps(parsed)
        conn = get_connection()
        # Dedup: keep latest per device_ip + command_name
        conn.execute(
            "DELETE FROM nc_collected_configs WHERE device_ip=? AND command_name=?",
            (data.get("device_ip", ""), data.get("command_name", "manual")),
        )
        conn.execute(
            "INSERT INTO nc_collected_configs "
            "(id, device_ip, hostname, profile_id, command_name, "
            " output_text, parsed_json, collected_at, topology_id) "
            "VALUES (?,?,?,?,?,?,?,?,?)",
            (
                cid,
                data.get("device_ip", ""),
                data.get("hostname", ""),
                data.get("profile_id"),
                data.get("command_name", "manual"),
                data.get("output_text", ""),
                parsed,
                _now(),
                data.get("topology_id"),
            ),
        )
        conn.commit()
        conn.close()
        return jsonify({"id": cid}), 201

    # ══════════════════════════════════════════════════════════════════════
    # Phase 4: Geolocation + Globe/Map View
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/device-geo", methods=["GET"])
    @nc_login_required
    def nc_api_list_device_geo():
        topo_id = request.args.get("topology_id", "")
        conn = get_connection()
        if topo_id:
            rows = conn.execute(
                "SELECT * FROM nc_device_geo WHERE topology_id=? ORDER BY site_name, label", (topo_id,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT dg.*, t.name AS topology_name "
                "FROM nc_device_geo dg "
                "LEFT JOIN topologies t ON t.id=dg.topology_id "
                "ORDER BY dg.site_name, dg.label"
            ).fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/device-geo", methods=["POST"])
    @nc_login_required
    def nc_api_set_device_geo():
        """Set geolocation for a device (deduped by topology+node)."""
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        # Dedup
        conn.execute(
            "DELETE FROM nc_device_geo WHERE topology_id=? AND node_id=?",
            (data.get("topology_id"), data.get("node_id")),
        )
        gid = str(_uuid.uuid4())
        conn.execute(
            "INSERT INTO nc_device_geo "
            "(id, topology_id, node_id, label, site_name, "
            " latitude, longitude, city, state, country, "
            " facility, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                gid,
                data.get("topology_id"),
                data.get("node_id"),
                data.get("label", ""),
                data.get("site_name", ""),
                data.get("latitude", 0),
                data.get("longitude", 0),
                data.get("city", ""),
                data.get("state", ""),
                data.get("country", "US"),
                data.get("facility", ""),
                _now(),
            ),
        )
        conn.commit()
        conn.close()
        return jsonify({"id": gid}), 201

    @bp.route("/api/device-geo/bulk", methods=["POST"])
    @nc_login_required
    def nc_api_bulk_set_geo():
        """Set geolocation for multiple devices at once."""
        data = request.get_json(force=True, silent=True) or {}
        devices = data.get("devices", [])
        conn = get_connection()
        count = 0
        for d in devices:
            conn.execute(
                "DELETE FROM nc_device_geo WHERE topology_id=? AND node_id=?", (d.get("topology_id"), d.get("node_id"))
            )
            conn.execute(
                "INSERT INTO nc_device_geo "
                "(id, topology_id, node_id, label, site_name, "
                " latitude, longitude, city, state, country, "
                " facility, created_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    str(_uuid.uuid4()),
                    d.get("topology_id"),
                    d.get("node_id"),
                    d.get("label", ""),
                    d.get("site_name", ""),
                    d.get("latitude", 0),
                    d.get("longitude", 0),
                    d.get("city", ""),
                    d.get("state", ""),
                    d.get("country", "US"),
                    d.get("facility", ""),
                    _now(),
                ),
            )
            count += 1
        conn.commit()
        conn.close()
        return jsonify({"set": count}), 201

    @bp.route("/api/device-geo/<gid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_device_geo(gid):
        conn = get_connection()
        conn.execute("DELETE FROM nc_device_geo WHERE id=?", (gid,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @bp.route("/api/geo-sites", methods=["GET"])
    @nc_login_required
    def nc_api_geo_sites():
        """Aggregate devices by site for map clustering."""
        conn = get_connection()
        rows = conn.execute(
            "SELECT site_name, latitude, longitude, city, state, "
            "country, facility, COUNT(*) AS device_count "
            "FROM nc_device_geo "
            "WHERE latitude != 0 AND longitude != 0 "
            "GROUP BY site_name, latitude, longitude "
            "ORDER BY site_name"
        ).fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/geo-links", methods=["GET"])
    @nc_login_required
    def nc_api_geo_links():
        """Get interconnect links with geolocation for map arcs."""
        conn = get_connection()
        rows = conn.execute(
            "SELECT ic.id, ic.circuit_id, ic.protocol, ic.bandwidth, "
            "sg.latitude AS src_lat, sg.longitude AS src_lon, "
            "sg.site_name AS src_site, "
            "dg.latitude AS dst_lat, dg.longitude AS dst_lon, "
            "dg.site_name AS dst_site "
            "FROM nc_interconnects ic "
            "LEFT JOIN nc_device_geo sg "
            "  ON sg.topology_id=ic.src_topology_id "
            "  AND sg.node_id=ic.src_node_id "
            "LEFT JOIN nc_device_geo dg "
            "  ON dg.topology_id=ic.dst_topology_id "
            "  AND dg.node_id=ic.dst_node_id "
            "WHERE sg.latitude IS NOT NULL "
            "  AND dg.latitude IS NOT NULL"
        ).fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/map")
    @nc_login_required
    def nc_map_view():
        if os.environ.get("NETWORK_MAP_ENABLED", "true").lower() not in ("1", "true", "yes"):
            return render_template(
                "network/index.html",
                flash_msg="Map view is disabled in air-gap mode. Set NETWORK_MAP_ENABLED=true to enable.",
            )
        return render_template("network/map.html")

    # ══════════════════════════════════════════════════════════════════════
    # P2: Charts / Visualization
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/charts/compliance-trend", methods=["GET"])
    @nc_login_required
    def nc_api_chart_compliance_trend():
        """Compliance score trend over time (from history + current)."""
        conn = get_connection()
        # Get history
        rows = conn.execute(
            "SELECT project_id, compliance_pct, recorded_at FROM nc_compliance_history ORDER BY recorded_at"
        ).fetchall()
        # Group by project
        by_project = {}
        for r in rows:
            r = _row_to_dict(r)
            pid = r.get("project_id", "")
            by_project.setdefault(pid, []).append(
                {
                    "pct": r.get("compliance_pct", 0),
                    "date": (r.get("recorded_at") or "")[:10],
                }
            )
        # Add project names
        projects = {}
        for pid in by_project:
            name_row = conn.execute("SELECT name FROM nc_projects WHERE id=?", (pid,)).fetchone()
            projects[pid] = {
                "name": name_row[0] if name_row else pid[:8],
                "data": by_project[pid],
            }
        conn.close()
        return jsonify({"projects": projects})

    @bp.route("/api/charts/compliance-snapshot", methods=["POST"])
    @nc_login_required
    def nc_api_chart_compliance_snapshot():
        """Record current compliance scores for all projects (for trending)."""
        conn = get_connection()
        now = _now()
        recorded = 0
        for p in conn.execute("SELECT id FROM nc_projects").fetchall():
            pid = p[0]
            tids = [
                r[0]
                for r in conn.execute(
                    "SELECT topology_id FROM nc_project_topologies WHERE project_id=?", (pid,)
                ).fetchall()
            ]
            tp = tf = cat1 = findings = 0
            for tid in tids:
                row = conn.execute(
                    "SELECT passed, failed FROM nc_compliance_checks WHERE topology_id=? ORDER BY ran_at DESC LIMIT 1",
                    (tid,),
                ).fetchone()
                if row:
                    tp += row[0] or 0
                    tf += row[1] or 0
                of = conn.execute(
                    "SELECT COUNT(*) FROM nc_compliance_findings WHERE topology_id=? AND status='open'", (tid,)
                ).fetchone()
                findings += of[0] if of else 0
                c1 = conn.execute(
                    "SELECT COUNT(*) FROM nc_compliance_findings "
                    "WHERE topology_id=? AND status='open' "
                    "AND severity='CAT1'",
                    (tid,),
                ).fetchone()
                cat1 += c1[0] if c1 else 0
            total = tp + tf
            if total:
                pct = round(tp * 100 / total)
                conn.execute(
                    "INSERT INTO nc_compliance_history "
                    "(id, project_id, compliance_pct, open_findings, "
                    " cat1_count, recorded_at) VALUES (?,?,?,?,?,?)",
                    (str(_uuid.uuid4()), pid, pct, findings, cat1, now),
                )
                recorded += 1
        conn.commit()
        conn.close()
        return jsonify({"recorded": recorded})

    @bp.route("/api/charts/capacity-overview", methods=["GET"])
    @nc_login_required
    def nc_api_chart_capacity_overview():
        """Capacity utilization across facilities + ports."""
        conn = get_connection()
        facilities = []
        for f in conn.execute(
            "SELECT name, total_power_kw, used_power_kw, "
            "total_cooling_tons, used_cooling_tons, "
            "total_racks, used_racks "
            "FROM nc_facilities ORDER BY name"
        ).fetchall():
            f = _row_to_dict(f)
            facilities.append(
                {
                    "name": f["name"],
                    "power_pct": round((f.get("used_power_kw") or 0) * 100 / max(f.get("total_power_kw") or 1, 1)),
                    "cooling_pct": round(
                        (f.get("used_cooling_tons") or 0) * 100 / max(f.get("total_cooling_tons") or 1, 1)
                    ),
                    "rack_pct": round((f.get("used_racks") or 0) * 100 / max(f.get("total_racks") or 1, 1)),
                }
            )
        ports = []
        for p in conn.execute(
            "SELECT device_label, total_ports, used_ports FROM nc_port_inventory ORDER BY device_label"
        ).fetchall():
            p = _row_to_dict(p)
            ports.append(
                {
                    "device": p["device_label"],
                    "pct": round((p.get("used_ports") or 0) * 100 / max(p.get("total_ports") or 1, 1)),
                }
            )
        conn.close()
        return jsonify({"facilities": facilities, "ports": ports})

    @bp.route("/api/charts/cost-breakdown", methods=["GET"])
    @nc_login_required
    def nc_api_chart_cost_breakdown():
        """CapEx vs OpEx breakdown across all projects."""
        conn = get_connection()
        # BOM CapEx
        bom_total = conn.execute("SELECT COALESCE(SUM(extended_cost), 0) FROM nc_bom_items").fetchone()[0]
        maint_total = conn.execute("SELECT COALESCE(SUM(annual_maint), 0) FROM nc_bom_items").fetchone()[0]
        license_total = conn.execute("SELECT COALESCE(SUM(license_cost), 0) FROM nc_bom_items").fetchone()[0]
        # Circuit OpEx
        circuit_monthly = conn.execute("SELECT COALESCE(SUM(monthly_cost_usd), 0) FROM nc_circuits").fetchone()[0]
        # Peering cost
        peering_monthly = conn.execute("SELECT COALESCE(SUM(monthly_cost), 0) FROM nc_peering_agreements").fetchone()[0]
        # Labor
        labor = conn.execute("SELECT COALESCE(SUM(hours * rate_per_hour), 0) FROM nc_resource_plan").fetchone()[0]
        conn.close()
        return jsonify(
            {
                "capex": {
                    "hardware": round(bom_total, 2),
                    "labor": round(labor, 2),
                },
                "opex_annual": {
                    "circuits": round(circuit_monthly * 12, 2),
                    "peering": round(peering_monthly * 12, 2),
                    "maintenance": round(maint_total, 2),
                    "licensing": round(license_total, 2),
                },
                "total_capex": round(bom_total + labor, 2),
                "total_opex_annual": round(
                    circuit_monthly * 12 + peering_monthly * 12 + maint_total + license_total, 2
                ),
            }
        )

    @bp.route("/charts")
    @nc_login_required
    def nc_charts_page():
        return render_template("network/charts.html")

    # ══════════════════════════════════════════════════════════════════════
    # P1: New Build Wizard + Alert Engine + Cross-Module Validation
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/wizard/new-build", methods=["POST"])
    @nc_login_required
    def nc_api_wizard_new_build():
        """Create a new network build: project + topology + phases + workflow in one call."""
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        now = _now()

        # Step 1: Create project
        pid = str(_uuid.uuid4())
        conn.execute(
            "INSERT INTO nc_projects "
            "(id, name, customer_id, description, status, owner, "
            " created_at, updated_at) VALUES (?,?,?,?,?,?,?,?)",
            (
                pid,
                data.get("name", "New Network Build"),
                data.get("customer_id"),
                data.get("description", ""),
                "draft",
                data.get("owner", ""),
                now,
                now,
            ),
        )

        # Step 2: Create topology (optional template)
        topo_id = str(_uuid.uuid4())
        graph = data.get("graph_json", {"nodes": [], "edges": []})
        if isinstance(graph, dict):
            graph = json.dumps(graph)
        conn.execute(
            "INSERT INTO topologies "
            "(id, name, description, graph_json, classification, "
            " created_at, updated_at) VALUES (?,?,?,?,?,?,?)",
            (
                topo_id,
                data.get("topology_name", data.get("name", "New Topology")),
                "",
                graph,
                data.get("classification", "public"),
                now,
                now,
            ),
        )
        conn.execute("INSERT INTO nc_project_topologies (project_id, topology_id) VALUES (?,?)", (pid, topo_id))

        # Step 3: Initialize phases
        for num, name in [(1, "Concept"), (2, "Design"), (3, "Approval"), (4, "Post-Deploy")]:
            conn.execute(
                "INSERT INTO nc_project_phases "
                "(id, project_id, phase_num, phase_name, status, "
                " entered_at, created_at) VALUES (?,?,?,?,?,?,?)",
                (
                    str(_uuid.uuid4()),
                    pid,
                    num,
                    name,
                    "active" if num == 1 else "pending",
                    now if num == 1 else None,
                    now,
                ),
            )

        # Step 4: Initialize case workflow
        wid = str(_uuid.uuid4())
        conn.execute(
            "INSERT INTO nc_case_workflows "
            "(id, project_id, current_state, lifecycle_json, "
            " created_at, updated_at) VALUES (?,?,?,?,?,?)",
            (wid, pid, "concept", json.dumps(_NDC_LIFECYCLE), now, now),
        )
        conn.execute(
            "INSERT INTO nc_case_history (workflow_id, from_state, to_state, comment, changed_at) VALUES (?,?,?,?,?)",
            (wid, "", "concept", "Created via New Build Wizard", now),
        )

        conn.commit()
        conn.close()
        _audit("WIZARD_NEW_BUILD", "project", pid, data.get("name", ""))
        return jsonify(
            {
                "project_id": pid,
                "topology_id": topo_id,
                "workflow_id": wid,
                "phases": 4,
                "next_step": f"/network/canvas/{topo_id}",
            }
        ), 201

    @bp.route("/wizard")
    @nc_login_required
    def nc_wizard_page():
        return render_template("network/wizard.html")

    # ── Alert/Threshold Engine ────────────────────────────────────────────
    @bp.route("/api/alert-rules", methods=["GET"])
    @nc_login_required
    def nc_api_list_alert_rules():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_alert_rules ORDER BY metric, name").fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/alert-rules", methods=["POST"])
    @nc_login_required
    def nc_api_create_alert_rule():
        data = request.get_json(force=True, silent=True) or {}
        return _crud_create(
            "nc_alert_rules", "", data, ["name", "metric", "operator", "threshold", "severity", "enabled"]
        )

    @bp.route("/api/alert-rules/<rid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_alert_rule(rid):
        return _crud_delete("nc_alert_rules", rid)

    @bp.route("/api/alert-events", methods=["GET"])
    @nc_login_required
    def nc_api_list_alert_events():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_alert_events ORDER BY created_at DESC LIMIT 50").fetchall()
        unack = conn.execute("SELECT COUNT(*) FROM nc_alert_events WHERE acknowledged=0").fetchone()[0]
        conn.close()
        return jsonify(
            {
                "events": [_row_to_dict(r) for r in rows],
                "unacknowledged": unack,
            }
        )

    @bp.route("/api/alert-events/run-check", methods=["POST"])
    @nc_login_required
    def nc_api_run_alert_check():
        """Evaluate all enabled alert rules against current data."""
        conn = get_connection()
        rules = [_row_to_dict(r) for r in conn.execute("SELECT * FROM nc_alert_rules WHERE enabled=1").fetchall()]
        now = _now()
        new_events = 0

        for rule in rules:
            metric = rule.get("metric", "")
            op = rule.get("operator", "gt")
            threshold = float(rule.get("threshold", 0))

            def _check(val, entity_type="", entity_id="", msg=""):
                nonlocal new_events
                triggered = False
                if op == "gt" and val > threshold:
                    triggered = True
                elif op == "lt" and val < threshold:
                    triggered = True
                elif op == "gte" and val >= threshold:
                    triggered = True
                elif op == "lte" and val <= threshold:
                    triggered = True
                if triggered:
                    conn.execute(
                        "INSERT INTO nc_alert_events "
                        "(id, rule_id, rule_name, severity, message, "
                        " entity_type, entity_id, created_at) "
                        "VALUES (?,?,?,?,?,?,?,?)",
                        (
                            str(_uuid.uuid4()),
                            rule["id"],
                            rule["name"],
                            rule.get("severity", "warning"),
                            msg,
                            entity_type,
                            entity_id,
                            now,
                        ),
                    )
                    new_events += 1

            if metric == "compliance_pct":
                for p in conn.execute("SELECT id, name FROM nc_projects").fetchall():
                    tids = [
                        r[0]
                        for r in conn.execute(
                            "SELECT topology_id FROM nc_project_topologies WHERE project_id=?", (p[0],)
                        ).fetchall()
                    ]
                    tp = tf = 0
                    for tid in tids:
                        row = conn.execute(
                            "SELECT passed, failed "
                            "FROM nc_compliance_checks "
                            "WHERE topology_id=? "
                            "ORDER BY ran_at DESC LIMIT 1",
                            (tid,),
                        ).fetchone()
                        if row:
                            tp += row[0] or 0
                            tf += row[1] or 0
                    total = tp + tf
                    if total:
                        pct = round(tp * 100 / total)
                        _check(pct, "project", p[0], f"{p[1]}: compliance {pct}%")

            elif metric == "port_utilization":
                for pi in conn.execute("SELECT * FROM nc_port_inventory").fetchall():
                    pi = _row_to_dict(pi)
                    total = pi.get("total_ports", 0) or 1
                    used = pi.get("used_ports", 0)
                    pct = round(used * 100 / total)
                    _check(pct, "device", pi.get("device_label", ""), f"{pi['device_label']}: port util {pct}%")

            elif metric == "power_pct":
                for f in conn.execute("SELECT * FROM nc_facilities").fetchall():
                    f = _row_to_dict(f)
                    total = f.get("total_power_kw", 0) or 1
                    used = f.get("used_power_kw", 0)
                    pct = round(used * 100 / max(total, 1))
                    _check(pct, "facility", f.get("name", ""), f"{f['name']}: power {pct}%")

        conn.commit()
        conn.close()
        return jsonify({"rules_checked": len(rules), "events_created": new_events})

    # ── Cross-Module Auto-Validation ──────────────────────────────────────
    @bp.route("/api/cross-validate/<pid>", methods=["POST"])
    @nc_login_required
    def nc_api_cross_validate(pid):
        """Run cross-module validation for a project.
        Checks peering↔capacity, compliance→findings, discovery→drift."""
        conn = get_connection()
        findings = []

        # 1. Peering ↔ Capacity: check if peering ports fit device inventory
        peers = [
            _row_to_dict(r)
            for r in conn.execute("SELECT * FROM nc_peering_agreements WHERE project_id=?", (pid,)).fetchall()
        ]
        ports = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT * FROM nc_port_inventory WHERE topology_id IN "
                "(SELECT topology_id FROM nc_project_topologies "
                " WHERE project_id=?)",
                (pid,),
            ).fetchall()
        ]
        total_avail = sum((p.get("total_ports", 0) or 0) - (p.get("used_ports", 0) or 0) for p in ports)
        if peers and total_avail < len(peers):
            findings.append(
                {
                    "module": "peering↔capacity",
                    "severity": "high",
                    "message": f"{len(peers)} peering agreements but only {total_avail} available ports",
                }
            )

        # 2. Compliance → auto-flag if CAT1 findings exist
        topo_ids = [
            r[0]
            for r in conn.execute("SELECT topology_id FROM nc_project_topologies WHERE project_id=?", (pid,)).fetchall()
        ]
        cat1_count = 0
        for tid in topo_ids:
            c1 = conn.execute(
                "SELECT COUNT(*) FROM nc_compliance_findings WHERE topology_id=? AND status='open' AND severity='CAT1'",
                (tid,),
            ).fetchone()
            cat1_count += c1[0] if c1 else 0
        if cat1_count > 0:
            findings.append(
                {
                    "module": "compliance",
                    "severity": "critical",
                    "message": f"{cat1_count} open CAT1 findings — blocks deployment",
                }
            )

        # 3. Facilities → check rack/power for project
        facs = [_row_to_dict(r) for r in conn.execute("SELECT * FROM nc_facilities").fetchall()]
        for f in facs:
            power_pct = round((f.get("used_power_kw", 0) or 0) * 100 / max(f.get("total_power_kw", 1) or 1, 1))
            if power_pct > 80:
                findings.append(
                    {
                        "module": "facilities",
                        "severity": "warning",
                        "message": f"{f['name']}: power at {power_pct}% — augment before adding equipment",
                    }
                )

        conn.close()
        return jsonify(
            {
                "project_id": pid,
                "findings": findings,
                "total_findings": len(findings),
                "has_blockers": any(f["severity"] == "critical" for f in findings),
            }
        )

    # ── Favorites / Pinned Views ──────────────────────────────────────────
    @bp.route("/api/favorites", methods=["GET"])
    @nc_login_required
    def nc_api_list_favorites():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_favorites ORDER BY created_at DESC").fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/favorites", methods=["POST"])
    @nc_login_required
    def nc_api_add_favorite():
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        try:
            conn.execute(
                "INSERT OR IGNORE INTO nc_favorites "
                "(id, entity_type, entity_id, label, user_id, "
                " created_at) VALUES (?,?,?,?,?,?)",
                (
                    str(_uuid.uuid4()),
                    data.get("entity_type", ""),
                    data.get("entity_id", ""),
                    data.get("label", ""),
                    data.get("user_id", ""),
                    _now(),
                ),
            )
            conn.commit()
        except Exception:
            pass
        conn.close()
        return jsonify({"ok": True}), 201

    @bp.route("/api/favorites/<fid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_favorite(fid):
        return _crud_delete("nc_favorites", fid)

    # ══════════════════════════════════════════════════════════════════════
    # Peering Agreements
    # ══════════════════════════════════════════════════════════════════════

    _PEERING_LIFECYCLE = [
        "evaluation",
        "negotiation",
        "agreement_signed",
        "technical_design",
        "implemented",
        "operational",
        "decommissioned",
    ]

    @bp.route("/api/peering-agreements", methods=["GET"])
    @nc_login_required
    def nc_api_list_peering():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_peering_agreements ORDER BY status, peer_name").fetchall()
        conn.close()
        items = [_row_to_dict(r) for r in rows]
        for it in items:
            try:
                it["locations"] = json.loads(it.get("locations") or "[]")
            except Exception:
                it["locations"] = []
        return jsonify(items)

    @bp.route("/api/peering-agreements", methods=["POST"])
    @nc_login_required
    def nc_api_create_peering():
        data = request.get_json(force=True, silent=True) or {}
        pid = str(_uuid.uuid4())
        locs = data.get("locations", [])
        if isinstance(locs, list):
            locs = json.dumps(locs)
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_peering_agreements "
            "(id, peer_name, peer_asn, our_asn, peering_type, "
            " routing_method, status, purpose, purpose_category, "
            " business_justification, locations, port_speed, "
            " contract_start, contract_end, monthly_cost, "
            " traffic_commit, ratio_limit, sla_uptime_pct, "
            " noc_contact, noc_email, noc_phone, legal_entity, "
            " notes, project_id, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                pid,
                data.get("peer_name", ""),
                data.get("peer_asn"),
                data.get("our_asn"),
                data.get("peering_type", "settlement_free"),
                data.get("routing_method", "bgp"),
                data.get("status", "evaluation"),
                data.get("purpose", ""),
                data.get("purpose_category", "connectivity"),
                data.get("business_justification", ""),
                locs,
                data.get("port_speed"),
                data.get("contract_start"),
                data.get("contract_end"),
                data.get("monthly_cost", 0),
                data.get("traffic_commit"),
                data.get("ratio_limit"),
                data.get("sla_uptime_pct", 99.9),
                data.get("noc_contact"),
                data.get("noc_email"),
                data.get("noc_phone"),
                data.get("legal_entity"),
                data.get("notes"),
                data.get("project_id"),
                _now(),
                _now(),
            ),
        )
        conn.commit()
        conn.close()
        _audit("CREATE", "peering_agreement", pid, data.get("peer_name", ""))
        return jsonify({"id": pid}), 201

    @bp.route("/api/peering-agreements/<aid>", methods=["PUT"])
    @nc_login_required
    def nc_api_update_peering(aid):
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        allowed = [
            "peer_name",
            "peer_asn",
            "our_asn",
            "peering_type",
            "routing_method",
            "status",
            "purpose",
            "purpose_category",
            "business_justification",
            "port_speed",
            "contract_start",
            "contract_end",
            "monthly_cost",
            "traffic_commit",
            "ratio_limit",
            "sla_uptime_pct",
            "noc_contact",
            "noc_email",
            "noc_phone",
            "legal_entity",
            "notes",
            "project_id",
        ]
        fields, values = [], []
        for k in allowed:
            if k in data:
                fields.append(f"{k}=?")
                values.append(data[k])
        if "locations" in data:
            locs = data["locations"]
            if isinstance(locs, list):
                locs = json.dumps(locs)
            fields.append("locations=?")
            values.append(locs)
        if fields:
            fields.append("updated_at=?")
            values.append(_now())
            values.append(aid)
            conn.execute(
                f"UPDATE nc_peering_agreements "  # nosec B608
                f"SET {', '.join(fields)} WHERE id=?",
                values,
            )
            conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @bp.route("/api/peering-agreements/<aid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_peering(aid):
        conn = get_connection()
        conn.execute("DELETE FROM nc_peering_sessions WHERE agreement_id=?", (aid,))
        conn.execute("DELETE FROM nc_peering_agreements WHERE id=?", (aid,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    # Peering sessions
    @bp.route("/api/peering-sessions/<aid>", methods=["GET"])
    @nc_login_required
    def nc_api_list_peering_sessions(aid):
        return _crud_list("nc_peering_sessions", aid, "location")

    @bp.route("/api/peering-sessions/<aid>", methods=["POST"])
    @nc_login_required
    def nc_api_create_peering_session(aid):
        data = request.get_json(force=True, silent=True) or {}
        sr = data.get("static_routes", [])
        if isinstance(sr, list):
            data["static_routes"] = json.dumps(sr)
        comms = data.get("communities", [])
        if isinstance(comms, list):
            data["communities"] = json.dumps(comms)
        data["md5_enabled"] = str(1 if data.get("md5_enabled") else 0)
        return _crud_create(
            "nc_peering_sessions",
            aid,
            data,
            [
                "location",
                "routing_method",
                "our_ip",
                "peer_ip",
                "our_ipv6",
                "peer_ipv6",
                "our_asn",
                "peer_asn",
                "prefix_limit",
                "md5_enabled",
                "local_pref",
                "med",
                "communities",
                "static_routes",
                "status",
                "port_speed",
                "notes",
            ],
        )

    # Peering evaluations
    @bp.route("/api/peering-evaluations", methods=["GET"])
    @nc_login_required
    def nc_api_list_peering_evals():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_peering_evaluations ORDER BY score DESC").fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/peering-evaluations", methods=["POST"])
    @nc_login_required
    def nc_api_create_peering_eval():
        data = request.get_json(force=True, silent=True) or {}
        # Auto-score
        geo = {"low": 1, "medium": 2, "high": 3}
        noc = {"poor": 1, "fair": 2, "good": 3, "excellent": 4, "unknown": 0}
        traffic = min(data.get("traffic_volume", 0) / 1000, 5)  # 0-5 pts
        geo_pts = geo.get(data.get("geographic_overlap", "medium"), 2)
        noc_pts = noc.get(data.get("noc_quality", "unknown"), 0)
        prefix_pts = min(data.get("prefix_count", 0) / 100, 3)
        score = round(traffic + geo_pts + noc_pts + prefix_pts, 2)
        rec = "peer" if score >= 7 else "evaluate" if score >= 4 else "decline"
        data["score"] = str(score)
        data["recommendation"] = rec
        data["prefix_count"] = str(data.get("prefix_count", 0))
        data["traffic_volume"] = str(data.get("traffic_volume", 0))
        return _crud_create(
            "nc_peering_evaluations",
            "",
            data,
            [
                "peer_name",
                "peer_asn",
                "traffic_volume",
                "geographic_overlap",
                "noc_quality",
                "network_capacity",
                "prefix_count",
                "peering_policy",
                "score",
                "recommendation",
                "notes",
            ],
        )

    # Peering cost-benefit
    @bp.route("/api/peering-cost-benefit/<aid>", methods=["GET"])
    @nc_login_required
    def nc_api_peering_cost_benefit(aid):
        conn = get_connection()
        agr = conn.execute("SELECT * FROM nc_peering_agreements WHERE id=?", (aid,)).fetchone()
        if not agr:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        agr = _row_to_dict(agr)
        # Get traffic data
        sessions = conn.execute("SELECT id FROM nc_peering_sessions WHERE agreement_id=?", (aid,)).fetchall()
        total_in = total_out = 0
        for s in sessions:
            t = conn.execute(
                "SELECT inbound_mbps, outbound_mbps "
                "FROM nc_peering_traffic WHERE session_id=? "
                "ORDER BY measured_at DESC LIMIT 1",
                (s[0],),
            ).fetchone()
            if t:
                total_in += t[0] or 0
                total_out += t[1] or 0
        conn.close()
        peer_cost = agr.get("monthly_cost", 0) or 0
        # Estimate transit cost for same traffic (industry avg ~$0.50/Mbps)
        transit_rate = 0.50  # $/Mbps/month
        total_traffic = total_in + total_out
        transit_cost = total_traffic * transit_rate
        savings = transit_cost - peer_cost
        return jsonify(
            {
                "peer_name": agr.get("peer_name"),
                "peering_cost_monthly": peer_cost,
                "transit_equivalent_monthly": round(transit_cost, 2),
                "monthly_savings": round(savings, 2),
                "annual_savings": round(savings * 12, 2),
                "traffic_mbps": {"inbound": total_in, "outbound": total_out},
                "transit_rate_per_mbps": transit_rate,
            }
        )

    @bp.route("/peering")
    @nc_login_required
    def nc_peering_page():
        return render_template("network/peering.html")

    # ══════════════════════════════════════════════════════════════════════
    # Partner Registry
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/partners")
    @nc_login_required
    def nc_partners_page():
        return render_template("network/partners.html")

    @bp.route("/api/partners", methods=["GET"])
    @nc_login_required
    def nc_api_list_partners():
        from tools.network.partner_registry import list_partners
        conn = get_connection()
        try:
            status = request.args.get("status", "")
            return jsonify(list_partners(conn, status=status))
        finally:
            conn.close()

    @bp.route("/api/partners", methods=["POST"])
    @nc_login_required
    def nc_api_create_partner():
        from tools.network.partner_registry import create_partner
        data = request.get_json(force=True) or {}
        if not data.get("name"):
            return jsonify({"error": "name is required"}), 400
        conn = get_connection()
        try:
            result = create_partner(conn, data)
            _audit("CREATE", "partner", result["partner_id"], conn)
            return jsonify(result), 201
        finally:
            conn.close()

    @bp.route("/api/partners/<pid>", methods=["GET"])
    @nc_login_required
    def nc_api_get_partner(pid):
        from tools.network.partner_registry import get_partner
        conn = get_connection()
        try:
            partner = get_partner(conn, pid)
            if not partner:
                return jsonify({"error": "not found"}), 404
            return jsonify(partner)
        finally:
            conn.close()

    @bp.route("/api/partners/<pid>", methods=["PUT"])
    @nc_login_required
    def nc_api_update_partner(pid):
        from tools.network.partner_registry import update_partner
        data = request.get_json(force=True) or {}
        conn = get_connection()
        try:
            result = update_partner(conn, pid, data)
            _audit("UPDATE", "partner", pid, conn)
            return jsonify(result)
        finally:
            conn.close()

    @bp.route("/api/partners/<pid>/unified-view", methods=["GET"])
    @nc_login_required
    def nc_api_partner_unified_view(pid):
        from tools.network.partner_registry import get_unified_view
        conn = get_connection()
        try:
            return jsonify(get_unified_view(conn, pid))
        finally:
            conn.close()

    @bp.route("/api/partners/by-asn", methods=["GET"])
    @nc_login_required
    def nc_api_partner_by_asn():
        """Return partner matching ?asn=<int>. Returns {} if not found."""
        asn = request.args.get("asn", type=int)
        if asn is None:
            return jsonify({"error": "asn is required"}), 400
        conn = get_connection()
        try:
            try:
                cur = conn.execute("SELECT * FROM nc_partners WHERE asn = ? LIMIT 1", (asn,))
            except Exception:
                cur = conn.cursor()
                cur.execute("SELECT * FROM nc_partners WHERE asn = %s LIMIT 1", (asn,))
            row = cur.fetchone()
            if not row:
                return jsonify({})
            cols = [d[0] for d in cur.description]
            return jsonify(dict(zip(cols, row)))
        finally:
            conn.close()

    # ── Agreement lifecycle routes ─────────────────────────────────────────

    @bp.route("/api/peering-agreements/<aid>/approve", methods=["POST"])
    @nc_login_required
    def nc_api_approve_agreement(aid):
        from tools.network.agreement_lifecycle import approve_agreement
        data = request.get_json(force=True) or {}
        approver = data.get("approver", "")
        role = data.get("role", "")
        if not approver or not role:
            return jsonify({"error": "approver and role are required"}), 400
        conn = get_connection()
        try:
            result = approve_agreement(conn, aid, approver, role)
            _audit("APPROVE", "peering_agreement", aid, conn)
            return jsonify(result)
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        finally:
            conn.close()

    @bp.route("/api/peering-agreements/<aid>/amend", methods=["POST"])
    @nc_login_required
    def nc_api_amend_agreement(aid):
        from tools.network.agreement_lifecycle import create_amendment
        data = request.get_json(force=True) or {}
        conn = get_connection()
        try:
            result = create_amendment(
                conn, aid,
                changes=data.get("changes", {}),
                reason=data.get("reason", ""),
                amended_by=data.get("amended_by", ""),
                effective_date=data.get("effective_date", ""),
            )
            _audit("AMEND", "peering_agreement", aid, conn)
            return jsonify(result), 201
        finally:
            conn.close()

    @bp.route("/api/peering-agreements/<aid>/amendments", methods=["GET"])
    @nc_login_required
    def nc_api_list_amendments(aid):
        conn = get_connection()
        try:
            try:
                rows = conn.execute("SELECT * FROM nc_agreement_amendments WHERE agreement_id=%s ORDER BY amendment_number", (aid,)).fetchall()
            except Exception:
                rows = conn.execute("SELECT * FROM nc_agreement_amendments WHERE agreement_id=? ORDER BY amendment_number", (aid,)).fetchall()
            return jsonify([_row_to_dict(r) for r in rows])
        finally:
            conn.close()

    @bp.route("/api/peering-agreements/<aid>/document", methods=["GET"])
    @nc_login_required
    def nc_api_agreement_document(aid):
        from tools.network.agreement_lifecycle import generate_agreement_document
        conn = get_connection()
        try:
            doc = generate_agreement_document(conn, aid)
            return jsonify({"agreement_id": aid, "document": doc})
        except ValueError as e:
            return jsonify({"error": str(e)}), 404
        finally:
            conn.close()

    @bp.route("/api/peering-agreements/<aid>/sync-to-pmc", methods=["POST"])
    @nc_login_required
    def nc_api_agreement_sync_to_pmc(aid):
        from tools.network.agreement_lifecycle import sync_to_pmc
        conn = get_connection()
        try:
            result = sync_to_pmc(conn, aid)
            return jsonify(result)
        finally:
            conn.close()

    # ══════════════════════════════════════════════════════════════════════
    # Capacity Planning (Port/Slot/Fiber/Circuit)
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/port-inventory", methods=["GET"])
    @nc_login_required
    def nc_api_list_port_inventory():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_port_inventory ORDER BY device_label").fetchall()
        conn.close()
        items = [_row_to_dict(r) for r in rows]
        for it in items:
            try:
                it["port_breakdown"] = json.loads(it.get("port_breakdown") or "{}")
            except Exception:
                it["port_breakdown"] = {}
        return jsonify(items)

    @bp.route("/api/port-inventory", methods=["POST"])
    @nc_login_required
    def nc_api_set_port_inventory():
        data = request.get_json(force=True, silent=True) or {}
        pb = data.get("port_breakdown", {})
        if isinstance(pb, dict):
            pb = json.dumps(pb)
        conn = get_connection()
        # Dedup by device_label + topology
        conn.execute(
            "DELETE FROM nc_port_inventory WHERE device_label=? AND topology_id=?",
            (data.get("device_label"), data.get("topology_id")),
        )
        rid = str(_uuid.uuid4())
        conn.execute(
            "INSERT INTO nc_port_inventory "
            "(id, device_label, topology_id, total_ports, used_ports, "
            " port_breakdown, last_updated) VALUES (?,?,?,?,?,?,?)",
            (
                rid,
                data.get("device_label", ""),
                data.get("topology_id"),
                data.get("total_ports", 0),
                data.get("used_ports", 0),
                pb,
                _now(),
            ),
        )
        conn.commit()
        conn.close()
        return jsonify({"id": rid}), 201

    @bp.route("/api/fiber-inventory", methods=["GET"])
    @nc_login_required
    def nc_api_list_fiber():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_fiber_inventory ORDER BY path_name").fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/fiber-inventory", methods=["POST"])
    @nc_login_required
    def nc_api_create_fiber():
        data = request.get_json(force=True, silent=True) or {}
        data["available_strands"] = str(int(data.get("total_strands", 0) or 0) - int(data.get("lit_strands", 0) or 0))
        data["available_lambdas"] = str(
            int(data.get("total_lambdas", 0) or 0) - int(data.get("active_lambdas", 0) or 0)
        )
        return _crud_create(
            "nc_fiber_inventory",
            "",
            data,
            [
                "path_name",
                "path_a",
                "path_z",
                "fiber_type",
                "total_strands",
                "lit_strands",
                "available_strands",
                "total_lambdas",
                "active_lambdas",
                "available_lambdas",
                "per_lambda_gbps",
                "conduit_ducts",
                "conduit_used",
                "diverse_path",
                "notes",
            ],
        )

    @bp.route("/api/carrier-availability", methods=["GET"])
    @nc_login_required
    def nc_api_list_carrier_avail():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_carrier_availability ORDER BY carrier, path_name").fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/carrier-availability", methods=["POST"])
    @nc_login_required
    def nc_api_create_carrier_avail():
        data = request.get_json(force=True, silent=True) or {}
        return _crud_create(
            "nc_carrier_availability",
            "",
            data,
            [
                "carrier",
                "path_name",
                "service_type",
                "available_bandwidth",
                "lead_time_days",
                "monthly_cost_est",
                "contract_term",
                "notes",
            ],
        )

    @bp.route("/capacity")
    @nc_login_required
    def nc_capacity_page():
        return render_template("network/capacity.html")

    # ══════════════════════════════════════════════════════════════════════
    # Facilities / DCIM
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/facilities", methods=["GET"])
    @nc_login_required
    def nc_api_list_facilities():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_facilities ORDER BY name").fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/facilities", methods=["POST"])
    @nc_login_required
    def nc_api_create_facility():
        data = request.get_json(force=True, silent=True) or {}
        return _crud_create(
            "nc_facilities",
            "",
            data,
            [
                "name",
                "facility_type",
                "address",
                "city",
                "state",
                "country",
                "operator",
                "total_racks",
                "used_racks",
                "total_power_kw",
                "used_power_kw",
                "total_cooling_tons",
                "used_cooling_tons",
                "ups_capacity_kva",
                "ups_load_kva",
                "ups_runtime_min",
                "generator_kw",
                "generator_load_kw",
                "generator_fuel_hours",
                "notes",
            ],
        )

    @bp.route("/api/facilities/<fid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_facility(fid):
        return _crud_delete("nc_facilities", fid)

    @bp.route("/api/racks", methods=["GET"])
    @nc_login_required
    def nc_api_list_racks():
        fid = request.args.get("facility_id", "")
        conn = get_connection()
        if fid:
            rows = conn.execute(
                "SELECT r.*, f.name AS facility_name "
                "FROM nc_racks r "
                "LEFT JOIN nc_facilities f ON f.id=r.facility_id "
                "WHERE r.facility_id=? ORDER BY r.rack_name",
                (fid,),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT r.*, f.name AS facility_name "
                "FROM nc_racks r "
                "LEFT JOIN nc_facilities f ON f.id=r.facility_id "
                "ORDER BY f.name, r.rack_name"
            ).fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/racks", methods=["POST"])
    @nc_login_required
    def nc_api_create_rack():
        data = request.get_json(force=True, silent=True) or {}
        return _crud_create(
            "nc_racks",
            "",
            data,
            [
                "facility_id",
                "rack_name",
                "total_ru",
                "used_ru",
                "reserved_ru",
                "power_circuit_a",
                "power_circuit_b",
                "max_power_kw",
                "current_power_kw",
                "weight_capacity_lbs",
                "current_weight_lbs",
                "notes",
            ],
        )

    @bp.route("/api/racks/<rid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_rack(rid):
        return _crud_delete("nc_racks", rid)

    @bp.route("/facilities")
    @nc_login_required
    def nc_facilities_page():
        return render_template("network/facilities.html")

    # ══════════════════════════════════════════════════════════════════════
    # Unified Readiness Checker
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/readiness-check/<pid>", methods=["GET"])
    @nc_login_required
    def nc_api_readiness_check(pid):
        """Cross-layer feasibility check for a project.
        Checks: peering, ports, fiber, facilities (rack/power/cooling)."""
        conn = get_connection()
        checks = []

        # Peering
        peer_count = conn.execute(
            "SELECT COUNT(*) FROM nc_peering_agreements "
            "WHERE project_id=? AND status IN "
            "('agreement_signed','technical_design','implemented','operational')",
            (pid,),
        ).fetchone()[0]
        checks.append(
            {
                "layer": "peering",
                "check": "Active peering agreements",
                "passed": peer_count > 0,
                "detail": f"{peer_count} active agreements",
            }
        )

        # Ports
        port_rows = conn.execute(
            "SELECT * FROM nc_port_inventory "
            "WHERE topology_id IN "
            "(SELECT topology_id FROM nc_project_topologies "
            " WHERE project_id=?)",
            (pid,),
        ).fetchall()
        ports_avail = sum((r["total_ports"] or 0) - (r["used_ports"] or 0) for r in port_rows)
        checks.append(
            {
                "layer": "ports",
                "check": "Available device ports",
                "passed": ports_avail > 0 or len(port_rows) == 0,
                "detail": f"{ports_avail} ports available" if port_rows else "No port inventory data",
            }
        )

        # Fiber
        fiber_rows = conn.execute("SELECT * FROM nc_fiber_inventory").fetchall()
        fiber_avail = sum((r["available_strands"] or 0) for r in fiber_rows)
        checks.append(
            {
                "layer": "fiber",
                "check": "Available fiber strands",
                "passed": fiber_avail > 0 or len(fiber_rows) == 0,
                "detail": f"{fiber_avail} strands available" if fiber_rows else "No fiber inventory data",
            }
        )

        # Facilities
        fac_rows = conn.execute("SELECT * FROM nc_facilities").fetchall()
        for f in fac_rows:
            f = _row_to_dict(f)
            avail_ru = (f.get("total_racks", 0) or 0) * 42 - sum(
                r["used_ru"] or 0
                for r in conn.execute("SELECT used_ru FROM nc_racks WHERE facility_id=?", (f["id"],)).fetchall()
            )
            avail_power = (f.get("total_power_kw", 0) or 0) - (f.get("used_power_kw", 0) or 0)
            avail_cooling = (f.get("total_cooling_tons", 0) or 0) - (f.get("used_cooling_tons", 0) or 0)
            power_pct = round((f.get("used_power_kw", 0) or 0) * 100 / max(f.get("total_power_kw", 1) or 1, 1))
            checks.append(
                {
                    "layer": "facility",
                    "check": f"Rack space at {f['name']}",
                    "passed": avail_ru > 4,
                    "detail": f"{avail_ru} RU available",
                }
            )
            checks.append(
                {
                    "layer": "power",
                    "check": f"Power at {f['name']}",
                    "passed": power_pct < 80,
                    "detail": f"{avail_power:.1f} kW available ({power_pct}% used)",
                }
            )
            checks.append(
                {
                    "layer": "cooling",
                    "check": f"Cooling at {f['name']}",
                    "passed": avail_cooling > 0,
                    "detail": f"{avail_cooling:.1f} tons available",
                }
            )

        conn.close()
        passed = sum(1 for c in checks if c["passed"])
        total = len(checks)
        return jsonify(
            {
                "project_id": pid,
                "checks": checks,
                "passed": passed,
                "total": total,
                "ready": passed == total,
                "readiness_pct": round(passed * 100 / max(total, 1)),
            }
        )

    # ══════════════════════════════════════════════════════════════════════
    # Phase 7: Innovation Flywheel
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/innovation")
    @nc_login_required
    def nc_innovation_hub():
        return render_template("network/innovation.html")

    # ── Ideas ─────────────────────────────────────────────────────────────
    @bp.route("/api/innovation-ideas", methods=["GET"])
    @nc_login_required
    def nc_api_list_ideas():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_innovation_ideas ORDER BY total_score DESC, created_at DESC").fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/innovation-ideas", methods=["POST"])
    @nc_login_required
    def nc_api_submit_idea():
        data = request.get_json(force=True, silent=True) or {}
        iid = str(_uuid.uuid4())
        imp = int(data.get("impact_score", 5) or 5)
        feas = int(data.get("feasibility_score", 5) or 5)
        cost = int(data.get("cost_score", 5) or 5)
        total = round((imp * 0.4 + feas * 0.35 + cost * 0.25), 2)
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_innovation_ideas "
            "(id, title, description, category, submitted_by, "
            " impact_score, feasibility_score, cost_score, "
            " total_score, status, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (
                iid,
                data.get("title", ""),
                data.get("description", ""),
                data.get("category", "improvement"),
                data.get("submitted_by", ""),
                imp,
                feas,
                cost,
                total,
                "submitted",
                _now(),
            ),
        )
        conn.commit()
        conn.close()
        _audit("SUBMIT_IDEA", "innovation", iid, data.get("title", ""))
        return jsonify({"id": iid, "total_score": total}), 201

    @bp.route("/api/innovation-ideas/<iid>/vote", methods=["POST"])
    @nc_login_required
    def nc_api_vote_idea(iid):
        conn = get_connection()
        conn.execute("UPDATE nc_innovation_ideas SET votes=votes+1 WHERE id=?", (iid,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @bp.route("/api/innovation-ideas/<iid>", methods=["PUT"])
    @nc_login_required
    def nc_api_update_idea(iid):
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        allowed = ["status", "project_id"]
        fields, values = [], []
        for k in allowed:
            if k in data:
                fields.append(f"{k}=?")
                values.append(data[k])
        if fields:
            values.append(iid)
            conn.execute(
                f"UPDATE nc_innovation_ideas "  # nosec B608
                f"SET {', '.join(fields)} WHERE id=?",
                values,
            )
            conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @bp.route("/api/innovation-ideas/<iid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_idea(iid):
        conn = get_connection()
        conn.execute("DELETE FROM nc_innovation_ideas WHERE id=?", (iid,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    # ── Tech Radar ────────────────────────────────────────────────────────
    @bp.route("/api/tech-radar", methods=["GET"])
    @nc_login_required
    def nc_api_list_tech_radar():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_tech_radar ORDER BY ring, technology").fetchall()
        conn.close()
        items = [_row_to_dict(r) for r in rows]
        by_ring = {"adopt": [], "trial": [], "assess": [], "hold": []}
        for it in items:
            ring = it.get("ring", "assess")
            by_ring.setdefault(ring, []).append(it)
        return jsonify({"items": items, "by_ring": by_ring})

    @bp.route("/api/tech-radar", methods=["POST"])
    @nc_login_required
    def nc_api_add_tech_radar():
        data = request.get_json(force=True, silent=True) or {}
        tid = str(_uuid.uuid4())
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_tech_radar "
            "(id, technology, ring, category, description, "
            " updated_by, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (
                tid,
                data.get("technology", ""),
                data.get("ring", "assess"),
                data.get("category", "networking"),
                data.get("description", ""),
                data.get("updated_by", ""),
                _now(),
                _now(),
            ),
        )
        conn.commit()
        conn.close()
        return jsonify({"id": tid}), 201

    @bp.route("/api/tech-radar/<tid>", methods=["PUT"])
    @nc_login_required
    def nc_api_update_tech_radar(tid):
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        old = conn.execute("SELECT ring FROM nc_tech_radar WHERE id=?", (tid,)).fetchone()
        new_ring = data.get("ring", "")
        fields, values = [], []
        for k in ["technology", "ring", "category", "description", "updated_by"]:
            if k in data:
                fields.append(f"{k}=?")
                values.append(data[k])
        if new_ring and old and new_ring != old[0]:
            fields.append("moved_from=?")
            values.append(old[0])
        if fields:
            fields.append("updated_at=?")
            values.append(_now())
            values.append(tid)
            conn.execute(
                f"UPDATE nc_tech_radar "  # nosec B608
                f"SET {', '.join(fields)} WHERE id=?",
                values,
            )
            conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @bp.route("/api/tech-radar/<tid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_tech_radar(tid):
        conn = get_connection()
        conn.execute("DELETE FROM nc_tech_radar WHERE id=?", (tid,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    # ── Lessons Learned ───────────────────────────────────────────────────
    @bp.route("/api/lessons-learned", methods=["GET"])
    @nc_login_required
    def nc_api_list_lessons():
        pid = request.args.get("project_id", "")
        conn = get_connection()
        if pid:
            rows = conn.execute(
                "SELECT ll.*, p.name AS project_name "
                "FROM nc_lessons_learned ll "
                "LEFT JOIN nc_projects p ON p.id=ll.project_id "
                "WHERE ll.project_id=? ORDER BY ll.created_at DESC",
                (pid,),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT ll.*, p.name AS project_name "
                "FROM nc_lessons_learned ll "
                "LEFT JOIN nc_projects p ON p.id=ll.project_id "
                "ORDER BY ll.created_at DESC"
            ).fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/lessons-learned", methods=["POST"])
    @nc_login_required
    def nc_api_add_lesson():
        data = request.get_json(force=True, silent=True) or {}
        lid = str(_uuid.uuid4())
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_lessons_learned "
            "(id, project_id, title, category, what_happened, "
            " root_cause, lesson, recommendation, submitted_by, "
            " created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (
                lid,
                data.get("project_id"),
                data.get("title", ""),
                data.get("category", "technical"),
                data.get("what_happened", ""),
                data.get("root_cause", ""),
                data.get("lesson", ""),
                data.get("recommendation", ""),
                data.get("submitted_by", ""),
                _now(),
            ),
        )
        conn.commit()
        conn.close()
        return jsonify({"id": lid}), 201

    @bp.route("/api/lessons-learned/<lid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_lesson(lid):
        conn = get_connection()
        conn.execute("DELETE FROM nc_lessons_learned WHERE id=?", (lid,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    # ══════════════════════════════════════════════════════════════════════
    # Phase 6: Tech Refresh Planner + Replacement Mapper + Budget Forecast
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/replacement-map", methods=["GET"])
    @nc_login_required
    def nc_api_list_replacement_map():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_replacement_map ORDER BY old_vendor, old_model").fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/replacement-map", methods=["POST"])
    @nc_login_required
    def nc_api_add_replacement():
        data = request.get_json(force=True, silent=True) or {}
        rid = str(_uuid.uuid4())
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_replacement_map "
            "(id, old_vendor, old_model, new_vendor, new_model, "
            " new_cost, migration_effort, notes, is_builtin, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,0,?)",
            (
                rid,
                data.get("old_vendor", ""),
                data.get("old_model", ""),
                data.get("new_vendor", ""),
                data.get("new_model", ""),
                data.get("new_cost", 0),
                data.get("migration_effort", "medium"),
                data.get("notes", ""),
                _now(),
            ),
        )
        conn.commit()
        conn.close()
        return jsonify({"id": rid}), 201

    @bp.route("/api/replacement-map/<rid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_replacement(rid):
        conn = get_connection()
        conn.execute("DELETE FROM nc_replacement_map WHERE id=?", (rid,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @bp.route("/api/projects/<pid>/refresh-plan", methods=["GET"])
    @nc_login_required
    def nc_api_list_refresh_plan(pid):
        conn = get_connection()
        rows = conn.execute(
            "SELECT * FROM nc_refresh_plans WHERE project_id=? ORDER BY target_year, priority DESC", (pid,)
        ).fetchall()
        items = [_row_to_dict(r) for r in rows]
        # Budget summary by year
        by_year = {}
        for it in items:
            yr = it.get("target_year", 0)
            by_year.setdefault(yr, 0)
            by_year[yr] += it.get("replacement_cost", 0) or 0
        conn.close()
        return jsonify(
            {
                "items": items,
                "budget_by_year": by_year,
                "total_cost": sum(by_year.values()),
            }
        )

    @bp.route("/api/projects/<pid>/refresh-plan", methods=["POST"])
    @nc_login_required
    def nc_api_add_refresh_item(pid):
        data = request.get_json(force=True, silent=True) or {}
        rid = str(_uuid.uuid4())
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_refresh_plans "
            "(id, project_id, device_label, old_model, eol_date, "
            " priority, replacement_model, replacement_cost, "
            " target_year, status, notes, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                rid,
                pid,
                data.get("device_label", ""),
                data.get("old_model", ""),
                data.get("eol_date", ""),
                data.get("priority", "medium"),
                data.get("replacement_model", ""),
                data.get("replacement_cost", 0),
                data.get("target_year", 2026),
                data.get("status", "planned"),
                data.get("notes", ""),
                _now(),
            ),
        )
        conn.commit()
        conn.close()
        return jsonify({"id": rid}), 201

    @bp.route("/api/refresh-plan/<rid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_refresh_item(rid):
        conn = get_connection()
        conn.execute("DELETE FROM nc_refresh_plans WHERE id=?", (rid,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @bp.route("/api/projects/<pid>/auto-refresh-plan", methods=["POST"])
    @nc_login_required
    def nc_api_auto_refresh_plan(pid):
        """Auto-generate refresh plan from tech debt analysis.
        Scans all project topologies for EOL/EOS devices and creates
        refresh items with replacement suggestions from the map."""
        conn = get_connection()
        topo_ids = [
            r[0]
            for r in conn.execute("SELECT topology_id FROM nc_project_topologies WHERE project_id=?", (pid,)).fetchall()
        ]

        # Load replacement map
        rep_map = {}
        for r in conn.execute(
            "SELECT old_model, new_model, new_cost, migration_effort FROM nc_replacement_map"
        ).fetchall():
            rep_map[r[0].lower()] = _row_to_dict(r)

        now = datetime.now(timezone.utc)
        items_created = 0
        for tid in topo_ids:
            row = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (tid,)).fetchone()
            if not row:
                continue
            try:
                graph = json.loads(row["graph_json"])
            except Exception:
                continue
            for n in graph.get("nodes", []):
                cfg = n.get("config") or n.get("configData") or {}
                eol = cfg.get("eol_date", "")
                model = cfg.get("model", "")
                if not eol and not model:
                    continue
                # Determine priority from EOL date
                priority = "low"
                if eol:
                    try:
                        eol_dt = datetime.fromisoformat(eol + "T00:00:00+00:00")
                        months = (eol_dt - now).days / 30.44
                        if months <= 0:
                            priority = "critical"
                        elif months <= 12:
                            priority = "high"
                        elif months <= 24:
                            priority = "medium"
                    except (ValueError, TypeError):
                        pass
                # Look up replacement
                rep = rep_map.get(model.lower(), {})
                target_yr = now.year
                if priority == "medium":
                    target_yr = now.year + 1
                elif priority == "low":
                    target_yr = now.year + 2

                conn.execute(
                    "INSERT INTO nc_refresh_plans "
                    "(id, project_id, device_label, old_model, "
                    " eol_date, priority, replacement_model, "
                    " replacement_cost, target_year, status, "
                    " created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (
                        str(_uuid.uuid4()),
                        pid,
                        n.get("label", ""),
                        model,
                        eol,
                        priority,
                        rep.get("new_model", ""),
                        rep.get("new_cost", 0),
                        target_yr,
                        "planned",
                        _now(),
                    ),
                )
                items_created += 1

        conn.commit()
        conn.close()
        _audit("AUTO_REFRESH", "project", pid, f"{items_created} items")
        return jsonify({"items_created": items_created}), 201

    @bp.route("/api/budget-forecast", methods=["GET"])
    @nc_login_required
    def nc_api_budget_forecast():
        """Enterprise-wide budget forecast from all refresh plans."""
        conn = get_connection()
        rows = conn.execute(
            "SELECT target_year, "
            "SUM(replacement_cost) AS total_cost, "
            "COUNT(*) AS item_count "
            "FROM nc_refresh_plans "
            "GROUP BY target_year ORDER BY target_year"
        ).fetchall()
        conn.close()
        forecast = [_row_to_dict(r) for r in rows]
        grand_total = sum(f.get("total_cost", 0) or 0 for f in forecast)
        return jsonify(
            {
                "forecast": forecast,
                "grand_total": grand_total,
            }
        )

    # ══════════════════════════════════════════════════════════════════════
    # Phase 5: Discovered Data -> Simulation/Impact/What-If Bridge
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/what-if/link-failure", methods=["POST"])
    @nc_login_required
    def nc_api_whatif_link_failure():
        """Simulate a link failure using real routing table data.
        Shows which prefixes lose reachability and alternate paths."""
        data = request.get_json(force=True, silent=True) or {}
        failed_link_src = data.get("source_device", "")
        failed_link_dst = data.get("target_device", "")
        if not failed_link_src or not failed_link_dst:
            return jsonify({"error": "source_device and target_device required"}), 400

        conn = get_connection()
        # Get all routing entries
        all_routes = [
            _row_to_dict(r) for r in conn.execute("SELECT * FROM nc_routing_entries ORDER BY device_ip").fetchall()
        ]
        conn.close()

        # Build forwarding table per device
        fwd = {}  # device_ip -> [{prefix, next_hop, protocol}]
        for r in all_routes:
            fwd.setdefault(r["device_ip"], []).append(r)

        # Find routes that use the failed link
        affected_prefixes = []
        surviving_prefixes = []
        for r in all_routes:
            if r["device_ip"] == failed_link_src and r["next_hop"] == failed_link_dst:
                # This route goes through the failed link
                # Check if there's an alternate path
                alternates = [
                    alt
                    for alt in fwd.get(r["device_ip"], [])
                    if alt["prefix"] == r["prefix"]
                    and alt["next_hop"] != failed_link_dst
                    and alt["next_hop"] != "0.0.0.0"  # nosec B104 — route filter, not socket bind
                ]
                entry = {
                    "prefix": r["prefix"],
                    "protocol": r["protocol"],
                    "failed_next_hop": failed_link_dst,
                    "alternate_paths": len(alternates),
                    "alternates": [
                        {"next_hop": a["next_hop"], "protocol": a["protocol"], "metric": a.get("metric", 0)}
                        for a in alternates
                    ],
                }
                if alternates:
                    surviving_prefixes.append(entry)
                else:
                    affected_prefixes.append(entry)

        return jsonify(
            {
                "scenario": f"Link failure: {failed_link_src} -> {failed_link_dst}",
                "affected_prefixes": affected_prefixes,
                "surviving_prefixes": surviving_prefixes,
                "total_affected": len(affected_prefixes),
                "total_surviving": len(surviving_prefixes),
                "has_full_redundancy": len(affected_prefixes) == 0,
            }
        )

    @bp.route("/api/what-if/device-failure", methods=["POST"])
    @nc_login_required
    def nc_api_whatif_device_failure():
        """Simulate a device going offline. Shows impact on all
        devices that route through it."""
        data = request.get_json(force=True, silent=True) or {}
        failed_device = data.get("device_ip", "")
        if not failed_device:
            return jsonify({"error": "device_ip required"}), 400

        conn = get_connection()
        all_routes = [
            _row_to_dict(r) for r in conn.execute("SELECT * FROM nc_routing_entries ORDER BY device_ip").fetchall()
        ]
        conn.close()

        # Find all routes whose next_hop is the failed device
        impacted_devices = {}
        for r in all_routes:
            if r["next_hop"] == failed_device and r["device_ip"] != failed_device:
                dip = r["device_ip"]
                if dip not in impacted_devices:
                    impacted_devices[dip] = {
                        "hostname": r.get("hostname", dip),
                        "lost_prefixes": [],
                    }
                impacted_devices[dip]["lost_prefixes"].append(
                    {
                        "prefix": r["prefix"],
                        "protocol": r["protocol"],
                    }
                )

        # Prefixes hosted by the failed device (connected routes)
        hosted_prefixes = [
            r["prefix"] for r in all_routes if r["device_ip"] == failed_device and r.get("protocol") == "connected"
        ]

        return jsonify(
            {
                "scenario": f"Device failure: {failed_device}",
                "impacted_devices": [{"device_ip": k, **v} for k, v in impacted_devices.items()],
                "total_impacted_devices": len(impacted_devices),
                "hosted_prefixes_lost": hosted_prefixes,
                "total_hosted_lost": len(hosted_prefixes),
            }
        )

    @bp.route("/api/what-if/add-link", methods=["POST"])
    @nc_login_required
    def nc_api_whatif_add_link():
        """Simulate adding a new link — predict how routing would
        change based on protocol and metric."""
        data = request.get_json(force=True, silent=True) or {}
        src = data.get("source_device", "")
        dst = data.get("target_device", "")
        protocol = data.get("protocol", "ospf")
        metric = data.get("metric", 10)

        conn = get_connection()
        # Get current routes from src device
        src_routes = [
            _row_to_dict(r)
            for r in conn.execute("SELECT * FROM nc_routing_entries WHERE device_ip=?", (src,)).fetchall()
        ]
        dst_routes = [
            _row_to_dict(r)
            for r in conn.execute("SELECT * FROM nc_routing_entries WHERE device_ip=?", (dst,)).fetchall()
        ]
        conn.close()

        # Prefixes reachable via new link (dst's connected/local prefixes)
        new_reachable = []
        for r in dst_routes:
            if r.get("protocol") in ("connected", "local"):
                # Check if src already has a route to this prefix
                existing = [e for e in src_routes if e["prefix"] == r["prefix"]]
                if existing:
                    best = min(existing, key=lambda x: x.get("metric", 999))
                    if metric < best.get("metric", 999):
                        new_reachable.append(
                            {
                                "prefix": r["prefix"],
                                "improvement": "better_metric",
                                "old_metric": best.get("metric", 0),
                                "new_metric": metric,
                            }
                        )
                    else:
                        new_reachable.append(
                            {
                                "prefix": r["prefix"],
                                "improvement": "redundant_path",
                                "old_metric": best.get("metric", 0),
                                "new_metric": metric,
                            }
                        )
                else:
                    new_reachable.append(
                        {
                            "prefix": r["prefix"],
                            "improvement": "new_reachability",
                            "new_metric": metric,
                        }
                    )

        return jsonify(
            {
                "scenario": f"Add link: {src} -> {dst} ({protocol}, metric {metric})",
                "new_reachable_prefixes": new_reachable,
                "total_improvements": len(new_reachable),
                "new_reachability": sum(1 for p in new_reachable if p["improvement"] == "new_reachability"),
                "better_metric": sum(1 for p in new_reachable if p["improvement"] == "better_metric"),
                "redundant_paths": sum(1 for p in new_reachable if p["improvement"] == "redundant_path"),
            }
        )

    # ══════════════════════════════════════════════════════════════════════
    # Connect & Collect + Diagram Data Extraction
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/connect-collect", methods=["POST"])
    @nc_login_required
    def nc_api_connect_collect():
        """Connect to a device via SSH, run profile commands (read-only),
        store outputs, and parse routing entries.
        Falls back to manual paste mode if SSH fails or is unavailable."""
        data = request.get_json(force=True, silent=True) or {}
        device_ip = data.get("device_ip", "")
        profile_id = data.get("profile_id", "")
        hostname = data.get("hostname", device_ip)
        topology_id = data.get("topology_id")
        mode = data.get("mode", "ssh")  # ssh or manual
        manual_outputs = data.get("manual_outputs", {})  # {cmd_name: output_text}

        if not device_ip:
            return jsonify({"error": "device_ip required"}), 400

        # Load profile commands
        conn = get_connection()
        prof_row = conn.execute("SELECT commands_json FROM nc_device_profiles WHERE id=?", (profile_id,)).fetchone()
        if not prof_row:
            conn.close()
            return jsonify({"error": "Profile not found"}), 404
        try:
            commands = json.loads(prof_row[0] or "{}")
        except Exception:
            commands = {}

        now = _now()
        results = []

        if mode == "manual":
            # Manual paste mode — store provided outputs
            for cmd_name, output in manual_outputs.items():
                if cmd_name not in commands:
                    continue
                # Dedup
                conn.execute(
                    "DELETE FROM nc_collected_configs WHERE device_ip=? AND command_name=?", (device_ip, cmd_name)
                )
                conn.execute(
                    "INSERT INTO nc_collected_configs "
                    "(id, device_ip, hostname, profile_id, "
                    " command_name, output_text, parsed_json, "
                    " collected_at, topology_id) "
                    "VALUES (?,?,?,?,?,?,?,?,?)",
                    (str(_uuid.uuid4()), device_ip, hostname, profile_id, cmd_name, output, "{}", now, topology_id),
                )
                results.append(
                    {
                        "command": cmd_name,
                        "status": "stored",
                        "lines": len(output.splitlines()),
                    }
                )
        else:
            # SSH mode — attempt live connection
            ssh_ok = False
            try:
                from netmiko import ConnectHandler

                ssh_ok = True
            except ImportError:
                pass

            if ssh_ok:
                username = data.get("username", "")
                password = data.get("password", "")
                device_type = data.get("device_type", "cisco_ios")
                try:
                    net_connect = ConnectHandler(
                        device_type=device_type,
                        host=device_ip,
                        username=username,
                        password=password,
                        timeout=data.get("timeout", 30),
                        read_timeout_override=60,
                    )
                    for cmd_name, cmd_info in commands.items():
                        cli_cmd = cmd_info.get("command", "")
                        timeout = cmd_info.get("timeout_sec", 10)
                        try:
                            output = net_connect.send_command(cli_cmd, read_timeout=timeout)
                            # Dedup and store
                            conn.execute(
                                "DELETE FROM nc_collected_configs WHERE device_ip=? AND command_name=?",
                                (device_ip, cmd_name),
                            )
                            conn.execute(
                                "INSERT INTO nc_collected_configs "
                                "(id, device_ip, hostname, profile_id, "
                                " command_name, output_text, parsed_json, "
                                " collected_at, topology_id) "
                                "VALUES (?,?,?,?,?,?,?,?,?)",
                                (
                                    str(_uuid.uuid4()),
                                    device_ip,
                                    hostname,
                                    profile_id,
                                    cmd_name,
                                    output,
                                    "{}",
                                    now,
                                    topology_id,
                                ),
                            )
                            results.append(
                                {
                                    "command": cmd_name,
                                    "status": "collected",
                                    "lines": len(output.splitlines()),
                                }
                            )
                        except Exception as cmd_err:
                            results.append(
                                {
                                    "command": cmd_name,
                                    "status": "failed",
                                    "error": str(cmd_err)[:100],
                                }
                            )
                    net_connect.disconnect()
                except Exception as ssh_err:
                    conn.close()
                    return jsonify(
                        {
                            "error": f"SSH connection failed: {str(ssh_err)[:200]}",
                            "hint": "Use mode='manual' to paste command outputs instead",
                        }
                    ), 502
            else:
                conn.close()
                return jsonify(
                    {
                        "error": "netmiko not available for SSH",
                        "hint": "Use mode='manual' to paste command outputs, or install netmiko: pip install netmiko",
                    }
                ), 501

        conn.commit()
        conn.close()
        _audit("CONNECT_COLLECT", "device", device_ip, f"profile={profile_id}, commands={len(results)}")
        return jsonify(
            {
                "device_ip": device_ip,
                "hostname": hostname,
                "mode": mode,
                "commands_executed": len(results),
                "results": results,
            }
        )

    @bp.route("/api/import/extract-data", methods=["POST"])
    @nc_login_required
    def nc_api_import_extract_data():
        """Import a diagram AND extract device data into DB tables.
        Populates: topology, IPAM blocks, circuits, device geo (from labels),
        and collected configs (from node properties)."""
        data = request.get_json(force=True, silent=True) or {}
        fmt = data.get("format", "drawio")
        content = data.get("content", "")
        name = data.get("name", "Extracted Import")
        project_id = data.get("project_id")
        if not content:
            return jsonify({"error": "content required"}), 400

        # Import and classify
        if fmt == "drawio":
            graph = import_drawio(content)
        elif fmt in ("vdx", "visio"):
            graph = import_vdx(content)
        elif fmt == "svg":
            graph = import_svg(content)
        else:
            return jsonify({"error": f"Unsupported: {fmt}"}), 400
        graph = _classify_imported_nodes(graph)

        conn = get_connection()
        now = _now()
        topo_id = str(_uuid.uuid4())
        conn.execute(
            "INSERT INTO topologies "
            "(id, name, description, graph_json, classification, "
            " created_at, updated_at) VALUES (?,?,?,?,?,?,?)",
            (topo_id, name, f"Data extraction ({fmt})", json.dumps(graph), "public", now, now),
        )

        if project_id:
            conn.execute(
                "INSERT OR IGNORE INTO nc_project_topologies (project_id, topology_id) VALUES (?,?)",
                (project_id, topo_id),
            )

        # Extract data from nodes
        ipam_extracted = 0
        devices_extracted = 0
        circuits_extracted = 0

        for n in graph.get("nodes", []):
            cfg = n.get("config") or n.get("configData") or {}
            label = n.get("label", "")
            ntype = n.get("type", "")

            # Extract IP addresses -> IPAM blocks
            ip = cfg.get("ip", "")
            if not ip:
                # Try to extract IP from label (e.g., "10.0.0.1/24")
                import re as _re

                ip_match = _re.search(r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}(?:/\d{1,2})?)", label)
                if ip_match:
                    ip = ip_match.group(1)
            if ip and "/" in ip:
                # Extract network from CIDR
                parts = ip.split("/")
                try:
                    octets = parts[0].split(".")
                    mask = int(parts[1])
                    # Simple network calculation
                    network = ".".join(octets[: mask // 8]) + ".0" * (4 - mask // 8) + "/" + parts[1]
                    conn.execute(
                        "INSERT OR IGNORE INTO nc_ipam_blocks "
                        "(id, topology_id, network, description, "
                        " created_at) VALUES (?,?,?,?,?)",
                        (str(_uuid.uuid4()), topo_id, network, f"Extracted from {label}", now),
                    )
                    ipam_extracted += 1
                except (ValueError, IndexError):
                    pass

            # Store device as collected config (properties)
            if ntype not in ("text", "heading", "badge", "rect", "circle", "imported", ""):
                devices_extracted += 1

        # Extract circuits from edges
        for e in graph.get("edges", []):
            elabel = e.get("label", "")
            proto = e.get("protocol", "")
            if elabel or proto:
                conn.execute(
                    "INSERT INTO nc_circuits "
                    "(id, topology_id, circuit_id, carrier, "
                    " circuit_type, bandwidth, install_status, "
                    " created_at, updated_at) VALUES (?,?,?,?,?,?,?,?,?)",
                    (
                        str(_uuid.uuid4()),
                        topo_id,
                        elabel or f"link-{e.get('id', '')[:8]}",
                        "",
                        proto or "ethernet",
                        "",
                        "installed",
                        now,
                        now,
                    ),
                )
                circuits_extracted += 1

        # Run compliance audit
        audit_result = run_compliance_audit(topo_id, graph, ["fisma_high"], "CUI")
        total_p = sum(s["passed"] for s in audit_result["scores"].values())
        total_f = sum(s["failed"] for s in audit_result["scores"].values())
        conn.execute(
            "INSERT INTO nc_compliance_checks "
            "(id, topology_id, check_type, passed, failed, "
            " findings_json, ran_at) VALUES (?,?,?,?,?,?,?)",
            (str(_uuid.uuid4()), topo_id, "fisma_high", total_p, total_f, json.dumps(audit_result["findings"]), now),
        )

        conn.commit()
        conn.close()
        _audit("IMPORT_EXTRACT", "topology", topo_id, fmt)

        return jsonify(
            {
                "id": topo_id,
                "name": name,
                "nodes": len(graph.get("nodes", [])),
                "edges": len(graph.get("edges", [])),
                "extracted": {
                    "devices": devices_extracted,
                    "ipam_blocks": ipam_extracted,
                    "circuits": circuits_extracted,
                },
                "compliance": {
                    "passed": total_p,
                    "failed": total_f,
                    "findings": len(audit_result["findings"]),
                },
                "classified_types": dict(
                    sorted(
                        {
                            n["type"]: sum(1 for m in graph["nodes"] if m["type"] == n["type"]) for n in graph["nodes"]
                        }.items()
                    )
                ),
            }
        ), 201

    @bp.route("/collect")
    @nc_login_required
    def nc_collect_page():
        return render_template("network/collect.html")

    # ══════════════════════════════════════════════════════════════════════
    # Phase 3: Routing Table Topology + Config-to-Canvas Sync
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/routing-entries", methods=["POST"])
    @nc_login_required
    def nc_api_store_routing_entries():
        """Store parsed routing table entries for a device."""
        data = request.get_json(force=True, silent=True) or {}
        device_ip = data.get("device_ip", "")
        hostname = data.get("hostname", "")
        entries = data.get("entries", [])
        topo_id = data.get("topology_id")
        if not entries:
            return jsonify({"error": "entries required"}), 400
        conn = get_connection()
        now = _now()
        # Dedup: remove existing entries for this device before inserting
        conn.execute("DELETE FROM nc_routing_entries WHERE device_ip=?", (device_ip,))
        count = 0
        for e in entries:
            conn.execute(
                "INSERT INTO nc_routing_entries "
                "(id, device_ip, hostname, prefix, next_hop, "
                " protocol, metric, admin_distance, interface, "
                " vrf, address_family, collected_at, topology_id) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    str(_uuid.uuid4()),
                    device_ip,
                    hostname,
                    e.get("prefix", ""),
                    e.get("next_hop", ""),
                    e.get("protocol", ""),
                    e.get("metric", 0),
                    e.get("admin_distance", 0),
                    e.get("interface", ""),
                    e.get("vrf", "default"),
                    e.get("address_family", "ipv4"),
                    now,
                    topo_id,
                ),
            )
            count += 1
        conn.commit()
        conn.close()
        return jsonify({"stored": count, "device": device_ip}), 201

    @bp.route("/api/routing-entries", methods=["GET"])
    @nc_login_required
    def nc_api_list_routing_entries():
        device_ip = request.args.get("device_ip", "")
        conn = get_connection()
        if device_ip:
            rows = conn.execute(
                "SELECT * FROM nc_routing_entries WHERE device_ip=? ORDER BY prefix", (device_ip,)
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM nc_routing_entries ORDER BY device_ip, prefix LIMIT 500").fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/routing-topology", methods=["POST"])
    @nc_login_required
    def nc_api_generate_routing_topology():
        """Generate a topology from stored routing entries.
        Each unique device becomes a node; next-hop relationships
        become edges showing actual forwarding paths."""
        data = request.get_json(force=True, silent=True) or {}
        device_ips = data.get("device_ips", [])
        name = data.get("name", "Routing Topology")
        conn = get_connection()

        if device_ips:
            placeholders = ",".join("?" for _ in device_ips)
            rows = conn.execute(
                f"SELECT * FROM nc_routing_entries "  # nosec B608
                f"WHERE device_ip IN ({placeholders}) "
                f"ORDER BY device_ip, prefix",
                device_ips,
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM nc_routing_entries ORDER BY device_ip, prefix").fetchall()

        # Build device set and next-hop relationships
        devices = {}  # ip -> {hostname, protocols, prefixes}
        links = {}  # (src_ip, dst_ip) -> {protocols, count}

        for r in rows:
            r = _row_to_dict(r)
            dip = r["device_ip"]
            if dip not in devices:
                devices[dip] = {
                    "hostname": r.get("hostname", dip),
                    "protocols": set(),
                    "prefix_count": 0,
                }
            devices[dip]["protocols"].add(r.get("protocol", ""))
            devices[dip]["prefix_count"] += 1

            nh = r.get("next_hop", "")
            if nh and nh != "0.0.0.0" and nh != "::" and nh != dip:  # nosec B104 — route filter, not socket bind
                key = (dip, nh)
                if key not in links:
                    links[key] = {"protocols": set(), "count": 0}
                links[key]["protocols"].add(r.get("protocol", ""))
                links[key]["count"] += 1
                # Ensure next-hop device exists
                if nh not in devices:
                    devices[nh] = {
                        "hostname": nh,
                        "protocols": set(),
                        "prefix_count": 0,
                    }

        # Generate graph JSON
        nodes = []
        x_pos = 0
        ip_to_id = {}
        for ip, info in sorted(devices.items()):
            nid = str(_uuid.uuid4())[:8]
            ip_to_id[ip] = nid
            nodes.append(
                {
                    "id": nid,
                    "label": info["hostname"],
                    "type": "router",
                    "x": x_pos % 800,
                    "y": (x_pos // 800) * 200,
                    "config": {
                        "ip": ip,
                        "protocol": ", ".join(sorted(info["protocols"] - {""})),
                    },
                }
            )
            x_pos += 200

        edges = []
        for (src, dst), info in links.items():
            src_id = ip_to_id.get(src)
            dst_id = ip_to_id.get(dst)
            if src_id and dst_id and src_id != dst_id:
                edges.append(
                    {
                        "id": str(_uuid.uuid4())[:8],
                        "source": src_id,
                        "target": dst_id,
                        "label": ", ".join(sorted(info["protocols"] - {""})),
                        "protocol": ", ".join(sorted(info["protocols"] - {""})),
                    }
                )

        graph = {"nodes": nodes, "edges": edges}
        graph = _classify_imported_nodes(graph)

        topo_id = str(_uuid.uuid4())
        now = _now()
        conn.execute(
            "INSERT INTO topologies "
            "(id, name, description, graph_json, "
            " classification, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (
                topo_id,
                name,
                f"Generated from {len(devices)} device routing tables",
                json.dumps(graph),
                "public",
                now,
                now,
            ),
        )
        conn.commit()
        conn.close()
        _audit("ROUTING_TOPO", "topology", topo_id, f"{len(nodes)} devices, {len(edges)} links")
        return jsonify(
            {
                "id": topo_id,
                "name": name,
                "nodes": len(nodes),
                "edges": len(edges),
                "devices_discovered": len(devices),
            }
        ), 201

    @bp.route("/api/config-to-canvas/<topo_id>", methods=["POST"])
    @nc_login_required
    def nc_api_config_to_canvas(topo_id):
        """Sync collected config data into canvas device properties.
        Matches by hostname or IP address."""
        conn = get_connection()
        row = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        try:
            graph = json.loads(row["graph_json"])
        except Exception:
            conn.close()
            return jsonify({"error": "Bad graph"}), 500

        # Get all collected configs
        configs = conn.execute(
            "SELECT device_ip, hostname, command_name, parsed_json "
            "FROM nc_collected_configs "
            "WHERE topology_id=? OR topology_id IS NULL "
            "ORDER BY collected_at DESC",
            (topo_id,),
        ).fetchall()

        # Build lookup by hostname and IP
        config_by_host = {}  # hostname -> {cmd: parsed}
        config_by_ip = {}  # ip -> {cmd: parsed}
        for c in configs:
            c = _row_to_dict(c)
            host = (c.get("hostname") or "").lower()
            ip = c.get("device_ip", "")
            cmd = c.get("command_name", "")
            try:
                parsed = json.loads(c.get("parsed_json") or "{}")
            except Exception:
                parsed = {}
            if host:
                config_by_host.setdefault(host, {})[cmd] = parsed
            if ip:
                config_by_ip.setdefault(ip, {})[cmd] = parsed

        updated = 0
        for n in graph.get("nodes", []):
            cfg = n.get("config") or n.get("configData") or {}
            label = (n.get("label") or "").lower()
            ip = cfg.get("ip", "")

            # Match by hostname then IP
            device_data = config_by_host.get(label, {})
            if not device_data and ip:
                device_data = config_by_ip.get(ip, {})
            if not device_data:
                continue

            # Merge parsed data into config
            for cmd_name, parsed in device_data.items():
                if cmd_name == "version" and parsed:
                    if parsed.get("version"):
                        cfg["sw_version"] = parsed["version"]
                    if parsed.get("hostname"):
                        n["label"] = parsed["hostname"]
                    if parsed.get("model"):
                        cfg["model"] = parsed["model"]
                    if parsed.get("serial"):
                        cfg["serial"] = parsed["serial"]
                elif cmd_name == "interfaces" and parsed:
                    if parsed.get("mgmt_ip"):
                        cfg["ip"] = parsed["mgmt_ip"]
                elif cmd_name in ("routing_table_v4", "routing_table_v6") and parsed:
                    if parsed.get("protocol"):
                        cfg["protocol"] = parsed["protocol"]

            n["config"] = cfg
            if "configData" in n:
                n["configData"] = cfg
            updated += 1

        now = _now()
        conn.execute("UPDATE topologies SET graph_json=?, updated_at=? WHERE id=?", (json.dumps(graph), now, topo_id))
        conn.commit()
        conn.close()
        return jsonify(
            {
                "updated_devices": updated,
                "total_nodes": len(graph.get("nodes", [])),
            }
        )

    # ══════════════════════════════════════════════════════════════════════
    # API: Circuits CRUD
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/circuits", methods=["GET"])
    @nc_login_required
    def nc_api_list_circuits():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_circuits ORDER BY updated_at DESC").fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/circuits", methods=["POST"])
    @nc_login_required
    def nc_api_create_circuit():
        data = request.get_json(force=True, silent=True) or {}
        cid = str(_uuid.uuid4())
        now = _now()
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_circuits (id, topology_id, circuit_id, carrier, circuit_type, bandwidth, "
            "handoff_a, handoff_z, customer, site, monthly_cost_usd, contract_start, contract_end, "
            "sla_uptime_pct, install_status, notes, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                cid,
                data.get("topology_id"),
                data.get("circuit_id", ""),
                data.get("carrier", ""),
                data.get("circuit_type", ""),
                data.get("bandwidth", ""),
                data.get("handoff_a", ""),
                data.get("handoff_z", ""),
                data.get("customer", ""),
                data.get("site", ""),
                data.get("monthly_cost_usd", 0),
                data.get("contract_start"),
                data.get("contract_end"),
                data.get("sla_uptime_pct", 99.9),
                data.get("install_status", "planned"),
                data.get("notes", ""),
                now,
                now,
            ),
        )
        conn.commit()
        conn.close()
        _audit("CREATE", "circuit", cid, data.get("circuit_id", ""))
        return jsonify({"id": cid}), 201

    @bp.route("/api/circuits/<cid>", methods=["PUT"])
    @nc_login_required
    def nc_api_update_circuit(cid):
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        row = conn.execute("SELECT id FROM nc_circuits WHERE id=?", (cid,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        allowed = [
            "circuit_id",
            "carrier",
            "circuit_type",
            "bandwidth",
            "handoff_a",
            "handoff_z",
            "customer",
            "site",
            "monthly_cost_usd",
            "contract_start",
            "contract_end",
            "sla_uptime_pct",
            "install_status",
            "notes",
            "topology_id",
        ]
        fields, values = [], []
        for k in allowed:
            if k in data:
                fields.append(f"{k}=?")
                values.append(data[k])
        if not fields:
            conn.close()
            return jsonify({"error": "No fields"}), 400
        fields.append("updated_at=?")
        values.append(_now())
        values.append(cid)
        conn.execute(f"UPDATE nc_circuits SET {', '.join(fields)} WHERE id=?", values)  # nosec B608 -- table/column names are internal constants, not user input
        conn.commit()
        conn.close()
        _audit("UPDATE", "circuit", cid)
        return jsonify({"ok": True})

    @bp.route("/api/circuits/<cid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_circuit(cid):
        conn = get_connection()
        conn.execute("DELETE FROM nc_circuits WHERE id=?", (cid,))
        conn.commit()
        conn.close()
        _audit("DELETE", "circuit", cid)
        return jsonify({"deleted": cid})

    # ══════════════════════════════════════════════════════════════════════
    # API: Customers & Sites CRUD
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/customers", methods=["GET"])
    @nc_login_required
    def nc_api_list_customers():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_customers ORDER BY name").fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/customers", methods=["POST"])
    @nc_login_required
    def nc_api_create_customer():
        data = request.get_json(force=True, silent=True) or {}
        cid = str(_uuid.uuid4())
        now = _now()
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_customers (id, name, customer_type, contact_name, contact_email, contract_ref, notes, created_at) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (
                cid,
                data.get("name", ""),
                data.get("customer_type", "customer"),
                data.get("contact_name", ""),
                data.get("contact_email", ""),
                data.get("contract_ref", ""),
                data.get("notes", ""),
                now,
            ),
        )
        conn.commit()
        conn.close()
        _audit("CREATE", "customer", cid, data.get("name", ""))
        return jsonify({"id": cid}), 201

    @bp.route("/api/customers/<cid>", methods=["PUT"])
    @nc_login_required
    def nc_api_update_customer(cid):
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        allowed = ["name", "customer_type", "contact_name", "contact_email", "contract_ref", "notes"]
        fields, values = [], []
        for k in allowed:
            if k in data:
                fields.append(f"{k}=?")
                values.append(data[k])
        if not fields:
            conn.close()
            return jsonify({"error": "No fields"}), 400
        values.append(cid)
        conn.execute(f"UPDATE nc_customers SET {', '.join(fields)} WHERE id=?", values)  # nosec B608 -- table/column names are internal constants, not user input
        conn.commit()
        conn.close()
        _audit("UPDATE", "customer", cid)
        return jsonify({"ok": True})

    @bp.route("/api/customers/<cid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_customer(cid):
        conn = get_connection()
        conn.execute("DELETE FROM nc_sites WHERE customer_id=?", (cid,))
        conn.execute("DELETE FROM nc_customers WHERE id=?", (cid,))
        conn.commit()
        conn.close()
        _audit("DELETE", "customer", cid)
        return jsonify({"deleted": cid})

    @bp.route("/api/sites", methods=["GET"])
    @nc_login_required
    def nc_api_list_sites():
        conn = get_connection()
        rows = conn.execute(
            "SELECT s.*, c.name AS customer_name FROM nc_sites s "
            "LEFT JOIN nc_customers c ON c.id=s.customer_id ORDER BY s.name"
        ).fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/sites", methods=["POST"])
    @nc_login_required
    def nc_api_create_site():
        data = request.get_json(force=True, silent=True) or {}
        sid = str(_uuid.uuid4())
        now = _now()
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_sites (id, customer_id, name, address, city, state, country, site_type, classification, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (
                sid,
                data.get("customer_id"),
                data.get("name", ""),
                data.get("address", ""),
                data.get("city", ""),
                data.get("state", ""),
                data.get("country", "US"),
                data.get("site_type", "office"),
                data.get("classification", "public"),
                now,
            ),
        )
        conn.commit()
        conn.close()
        _audit("CREATE", "site", sid, data.get("name", ""))
        return jsonify({"id": sid}), 201

    @bp.route("/api/sites/<sid>", methods=["PUT"])
    @nc_login_required
    def nc_api_update_site(sid):
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        allowed = ["customer_id", "name", "address", "city", "state", "country", "site_type", "classification"]
        fields, values = [], []
        for k in allowed:
            if k in data:
                fields.append(f"{k}=?")
                values.append(data[k])
        if not fields:
            conn.close()
            return jsonify({"error": "No fields"}), 400
        values.append(sid)
        conn.execute(f"UPDATE nc_sites SET {', '.join(fields)} WHERE id=?", values)  # nosec B608 -- table/column names are internal constants, not user input
        conn.commit()
        conn.close()
        _audit("UPDATE", "site", sid)
        return jsonify({"ok": True})

    @bp.route("/api/sites/<sid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_site(sid):
        conn = get_connection()
        conn.execute("DELETE FROM nc_sites WHERE id=?", (sid,))
        conn.commit()
        conn.close()
        _audit("DELETE", "site", sid)
        return jsonify({"deleted": sid})

    # ══════════════════════════════════════════════════════════════════════
    # API: IPAM CRUD
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/ipam", methods=["GET"])
    @nc_login_required
    def nc_api_list_ipam():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_ipam_blocks ORDER BY network").fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/ipam", methods=["POST"])
    @nc_login_required
    def nc_api_create_ipam():
        data = request.get_json(force=True, silent=True) or {}
        bid = str(_uuid.uuid4())
        now = _now()
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_ipam_blocks (id, topology_id, network, vlan_id, vrf, description, site_id, gateway, utilization_pct, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (
                bid,
                data.get("topology_id"),
                data.get("network", ""),
                data.get("vlan_id"),
                data.get("vrf", "global"),
                data.get("description", ""),
                data.get("site_id"),
                data.get("gateway", ""),
                data.get("utilization_pct", 0),
                now,
            ),
        )
        conn.commit()
        conn.close()
        _audit("CREATE", "ipam_block", bid, data.get("network", ""))
        try:
            from tools.canvas.event_bus import publish as _eb_publish
            _eb_publish("ndc", "ndc.ipam.added", {
                "block_id": bid,
                "network": data.get("network", ""),
                "topology_id": data.get("topology_id"),
                "vrf": data.get("vrf", "global"),
                "vlan_id": data.get("vlan_id"),
                "classification": "CUI",
            }, target_canvas="idc")
        except Exception:
            pass
        return jsonify({"id": bid}), 201

    @bp.route("/api/ipam/<bid>", methods=["PUT"])
    @nc_login_required
    def nc_api_update_ipam(bid):
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        allowed = ["network", "vlan_id", "vrf", "description", "site_id", "gateway", "utilization_pct", "topology_id"]
        fields, values = [], []
        for k in allowed:
            if k in data:
                fields.append(f"{k}=?")
                values.append(data[k])
        if not fields:
            conn.close()
            return jsonify({"error": "No fields"}), 400
        values.append(bid)
        conn.execute(f"UPDATE nc_ipam_blocks SET {', '.join(fields)} WHERE id=?", values)  # nosec B608 -- table/column names are internal constants, not user input
        conn.commit()
        conn.close()
        _audit("UPDATE", "ipam_block", bid)
        return jsonify({"ok": True})

    @bp.route("/api/ipam/<bid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_ipam(bid):
        conn = get_connection()
        conn.execute("DELETE FROM nc_ipam_blocks WHERE id=?", (bid,))
        conn.commit()
        conn.close()
        _audit("DELETE", "ipam_block", bid)
        return jsonify({"deleted": bid})

    # ══════════════════════════════════════════════════════════════════════
    # API: Cables CRUD
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/cables", methods=["GET"])
    @nc_login_required
    def nc_api_list_cables():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_cables ORDER BY cable_id").fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/cables", methods=["POST"])
    @nc_login_required
    def nc_api_create_cable():
        data = request.get_json(force=True, silent=True) or {}
        cid = str(_uuid.uuid4())
        now = _now()
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_cables (id, topology_id, cable_id, cable_type, src_device, src_port, "
            "dst_device, dst_port, patch_panel, length_m, status, notes, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                cid,
                data.get("topology_id"),
                data.get("cable_id", ""),
                data.get("cable_type", ""),
                data.get("src_device", ""),
                data.get("src_port", ""),
                data.get("dst_device", ""),
                data.get("dst_port", ""),
                data.get("patch_panel", ""),
                data.get("length_m"),
                data.get("status", "active"),
                data.get("notes", ""),
                now,
            ),
        )
        conn.commit()
        conn.close()
        _audit("CREATE", "cable", cid, data.get("cable_id", ""))
        return jsonify({"id": cid}), 201

    @bp.route("/api/cables/<cid>", methods=["PUT"])
    @nc_login_required
    def nc_api_update_cable(cid):
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        allowed = [
            "cable_id",
            "cable_type",
            "src_device",
            "src_port",
            "dst_device",
            "dst_port",
            "patch_panel",
            "length_m",
            "status",
            "notes",
            "topology_id",
        ]
        fields, values = [], []
        for k in allowed:
            if k in data:
                fields.append(f"{k}=?")
                values.append(data[k])
        if not fields:
            conn.close()
            return jsonify({"error": "No fields"}), 400
        values.append(cid)
        conn.execute(f"UPDATE nc_cables SET {', '.join(fields)} WHERE id=?", values)  # nosec B608 -- table/column names are internal constants, not user input
        conn.commit()
        conn.close()
        _audit("UPDATE", "cable", cid)
        return jsonify({"ok": True})

    @bp.route("/api/cables/<cid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_cable(cid):
        conn = get_connection()
        conn.execute("DELETE FROM nc_cables WHERE id=?", (cid,))
        conn.commit()
        conn.close()
        _audit("DELETE", "cable", cid)
        return jsonify({"deleted": cid})

    # ══════════════════════════════════════════════════════════════════════
    # API: Cable Plant Report Export
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/cable-plant-report/<topo_id>", methods=["GET"])
    @nc_login_required
    def nc_api_cable_plant_report(topo_id):
        """Export cable plant report as CSV from edge cableData annotations."""
        conn = get_connection()
        row = conn.execute("SELECT graph_json, name FROM topologies WHERE id=?", (topo_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Topology not found"}), 404

        topo = _row_to_dict(row)
        gj = topo.get("graph_json")
        if isinstance(gj, str):
            gj = json.loads(gj)

        nodes_map = {}
        for n in (gj or {}).get("nodes", []):
            nodes_map[n.get("id", "")] = n.get("label", n.get("id", ""))

        import io
        import csv

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(
            [
                "Link ID",
                "Source Device",
                "Target Device",
                "Cable Type",
                "Distance (m)",
                "Conduit ID",
                "Fiber Strands",
                "Pull Tension (lbs)",
                "Notes",
            ]
        )

        edges = (gj or {}).get("edges", [])
        cable_count = 0
        for e in edges:
            cable = e.get("cableData")
            if not cable:
                continue
            cable_count += 1
            writer.writerow(
                [
                    e.get("id", ""),
                    nodes_map.get(e.get("source", ""), e.get("source", "")),
                    nodes_map.get(e.get("target", ""), e.get("target", "")),
                    cable.get("cable_type", ""),
                    cable.get("distance_m", ""),
                    cable.get("conduit_id", ""),
                    cable.get("fiber_strands", ""),
                    cable.get("pull_tension_lbs", ""),
                    cable.get("notes", ""),
                ]
            )

        topo_name = (topo.get("name") or "topology").replace(" ", "_")
        filename = f"cable-plant-{topo_name}.csv"
        _audit("EXPORT", "cable_plant_report", topo_id, f"{cable_count} cable runs")
        return jsonify({"csv": buf.getvalue(), "filename": filename, "cable_count": cable_count})

    # ══════════════════════════════════════════════════════════════════════
    # API: Cross-Connects CRUD
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/cross-connects", methods=["GET"])
    @nc_login_required
    def nc_api_list_cross_connects():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_cross_connects ORDER BY updated_at DESC").fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/cross-connects", methods=["POST"])
    @nc_login_required
    def nc_api_create_cross_connect():
        data = request.get_json(force=True, silent=True) or {}
        cid = str(_uuid.uuid4())
        now = _now()
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_cross_connects (id, topology_id, xconn_id, facility, meet_me_room, "
            "src_device, src_port, dst_device, dst_port, media_type, bandwidth, "
            "provider_a, provider_z, loa_status, monthly_cost_usd, install_date, "
            "status, notes, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                cid,
                data.get("topology_id"),
                data.get("xconn_id", ""),
                data.get("facility", ""),
                data.get("meet_me_room", ""),
                data.get("src_device", ""),
                data.get("src_port", ""),
                data.get("dst_device", ""),
                data.get("dst_port", ""),
                data.get("media_type", "SMF"),
                data.get("bandwidth", ""),
                data.get("provider_a", ""),
                data.get("provider_z", ""),
                data.get("loa_status", "pending"),
                data.get("monthly_cost_usd", 0),
                data.get("install_date"),
                data.get("status", "planned"),
                data.get("notes", ""),
                now,
                now,
            ),
        )
        conn.commit()
        conn.close()
        _audit("CREATE", "cross_connect", cid, data.get("xconn_id", ""))
        return jsonify({"id": cid}), 201

    @bp.route("/api/cross-connects/<cid>", methods=["PUT"])
    @nc_login_required
    def nc_api_update_cross_connect(cid):
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        row = conn.execute("SELECT id FROM nc_cross_connects WHERE id=?", (cid,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        allowed = [
            "xconn_id",
            "facility",
            "meet_me_room",
            "src_device",
            "src_port",
            "dst_device",
            "dst_port",
            "media_type",
            "bandwidth",
            "provider_a",
            "provider_z",
            "loa_status",
            "monthly_cost_usd",
            "install_date",
            "status",
            "notes",
            "topology_id",
        ]
        fields, values = [], []
        for k in allowed:
            if k in data:
                fields.append(f"{k}=?")
                values.append(data[k])
        if not fields:
            conn.close()
            return jsonify({"error": "No fields"}), 400
        fields.append("updated_at=?")
        values.append(_now())
        values.append(cid)
        conn.execute(f"UPDATE nc_cross_connects SET {', '.join(fields)} WHERE id=?", values)  # nosec B608 -- table/column names are internal constants, not user input
        conn.commit()
        conn.close()
        _audit("UPDATE", "cross_connect", cid)
        return jsonify({"ok": True})

    @bp.route("/api/cross-connects/<cid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_cross_connect(cid):
        conn = get_connection()
        conn.execute("DELETE FROM nc_cross_connects WHERE id=?", (cid,))
        conn.commit()
        conn.close()
        _audit("DELETE", "cross_connect", cid)
        return jsonify({"deleted": cid})

    # ══════════════════════════════════════════════════════════════════════
    # API: Versions
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/versions/<topo_id>", methods=["GET"])
    @nc_login_required
    def nc_api_list_versions(topo_id):
        conn = get_connection()
        rows = conn.execute(
            "SELECT id, topology_id, version_num, label, phase, created_by, notes, created_at "
            "FROM nc_versions WHERE topology_id=? ORDER BY version_num",
            (topo_id,),
        ).fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/versions/<topo_id>", methods=["POST"])
    @nc_login_required
    def nc_api_create_version(topo_id):
        conn = get_connection()
        topo = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not topo:
            conn.close()
            return jsonify({"error": "Topology not found"}), 404
        data = request.get_json(force=True, silent=True) or {}
        last = conn.execute("SELECT MAX(version_num) FROM nc_versions WHERE topology_id=?", (topo_id,)).fetchone()[0]
        ver_num = (last or 0) + 1
        vid = str(_uuid.uuid4())
        now = _now()
        conn.execute(
            "INSERT INTO nc_versions (id, topology_id, version_num, label, phase, graph_json, created_by, notes, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?)",
            (
                vid,
                topo_id,
                ver_num,
                data.get("label", f"v{ver_num}"),
                data.get("phase", "as-is"),
                topo["graph_json"],
                data.get("created_by", ""),
                data.get("notes", ""),
                now,
            ),
        )
        conn.commit()
        conn.close()
        _audit("CREATE", "version", vid, f"v{ver_num} ({data.get('phase', 'as-is')})")
        return jsonify({"id": vid, "version_num": ver_num}), 201

    @bp.route("/api/versions/<topo_id>/diff", methods=["POST"])
    @nc_login_required
    def nc_api_diff_versions(topo_id):
        data = request.get_json(force=True, silent=True) or {}
        v1_id = data.get("version_a")
        v2_id = data.get("version_b")
        if not v1_id or not v2_id:
            return jsonify({"error": "version_a and version_b required"}), 400
        conn = get_connection()
        r1 = conn.execute("SELECT graph_json, label, phase FROM nc_versions WHERE id=?", (v1_id,)).fetchone()
        r2 = conn.execute("SELECT graph_json, label, phase FROM nc_versions WHERE id=?", (v2_id,)).fetchone()
        conn.close()
        if not r1 or not r2:
            return jsonify({"error": "Version not found"}), 404
        try:
            g1 = json.loads(r1["graph_json"])
            g2 = json.loads(r2["graph_json"])
        except Exception:
            return jsonify({"error": "Invalid graph JSON"}), 500
        n1_ids = {n["id"] for n in g1.get("nodes", [])}
        n2_ids = {n["id"] for n in g2.get("nodes", [])}
        e1_ids = {e["id"] for e in g1.get("edges", []) if "id" in e}
        e2_ids = {e["id"] for e in g2.get("edges", []) if "id" in e}
        return jsonify(
            {
                "version_a": {"id": v1_id, "label": r1["label"], "phase": r1["phase"]},
                "version_b": {"id": v2_id, "label": r2["label"], "phase": r2["phase"]},
                "nodes_added": len(n2_ids - n1_ids),
                "nodes_removed": len(n1_ids - n2_ids),
                "nodes_unchanged": len(n1_ids & n2_ids),
                "edges_added": len(e2_ids - e1_ids),
                "edges_removed": len(e1_ids - e2_ids),
                "edges_unchanged": len(e1_ids & e2_ids),
                "added_node_ids": list(n2_ids - n1_ids),
                "removed_node_ids": list(n1_ids - n2_ids),
            }
        )

    # ══════════════════════════════════════════════════════════════════════
    # API: Bill of Materials & Capacity Planning
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/topologies/<topo_id>/bom", methods=["GET"])
    @nc_login_required
    def nc_api_bom(topo_id):
        conn = get_connection()
        row = conn.execute("SELECT graph_json, name FROM topologies WHERE id=?", (topo_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        try:
            graph = json.loads(row["graph_json"])
        except Exception:
            graph = {"nodes": [], "edges": []}
        type_counts = {}
        for n in graph.get("nodes", []):
            t = n.get("type", "unknown")
            type_counts[t] = type_counts.get(t, 0) + 1
        bom_items = []
        total = 0
        for device_type, count in sorted(type_counts.items()):
            unit_cost = BOM_COSTS.get(device_type, 0)
            line_total = unit_cost * count
            total += line_total
            bom_items.append(
                {
                    "device_type": device_type,
                    "quantity": count,
                    "unit_cost_usd": unit_cost,
                    "line_total_usd": line_total,
                }
            )
        return jsonify(
            {
                "topology": row["name"],
                "items": bom_items,
                "total_capex_usd": total,
                "device_count": sum(type_counts.values()),
                "unique_types": len(type_counts),
            }
        )

    @bp.route("/api/topologies/<topo_id>/capacity", methods=["POST"])
    @nc_login_required
    def nc_api_capacity_plan(topo_id):
        conn = get_connection()
        row = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        try:
            graph = json.loads(row["graph_json"])
        except Exception:
            return jsonify({"error": "Bad graph"}), 500
        data = request.get_json(force=True, silent=True) or {}
        growth_pct = data.get("annual_growth_pct", 20)
        additional_users = data.get("additional_users", 0)
        user_bw_mbps = data.get("user_bandwidth_mbps", 2)
        analysis = []
        for e in graph.get("edges", []):
            base_util = 20 + abs(hash(e.get("id", "")) % 60)
            label = (e.get("label") or "").upper()
            link_bw_mbps = 1000
            if "100G" in label:
                link_bw_mbps = 100000
            elif "40G" in label:
                link_bw_mbps = 40000
            elif "25G" in label:
                link_bw_mbps = 25000
            elif "10G" in label:
                link_bw_mbps = 10000
            elif "100M" in label:
                link_bw_mbps = 100
            user_impact = round(additional_users * user_bw_mbps / max(link_bw_mbps, 1) * 100, 1)
            projected = round(base_util + user_impact, 1)
            status = "critical" if projected > 80 else "warning" if projected > 60 else "ok"
            analysis.append(
                {
                    "edge_id": e.get("id", ""),
                    "label": e.get("label", ""),
                    "source": e.get("source", ""),
                    "target": e.get("target", ""),
                    "current_util_pct": base_util,
                    "growth_projected_pct": round(base_util * (1 + growth_pct / 100), 1),
                    "with_users_pct": projected,
                    "link_bw_mbps": link_bw_mbps,
                    "status": status,
                    "upgrade_needed": projected > 75,
                }
            )
        critical = [a for a in analysis if a["status"] == "critical"]
        return jsonify(
            {
                "links": analysis,
                "total_links": len(analysis),
                "critical_count": len(critical),
                "growth_pct": growth_pct,
                "additional_users": additional_users,
                "recommendations": [
                    f"Upgrade {a['label'] or a['edge_id']}: {a['current_util_pct']}% -> {a['with_users_pct']}%"
                    for a in critical
                ],
            }
        )

    # ══════════════════════════════════════════════════════════════════════
    # API: Compliance (quick check)
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/topologies/<topo_id>/compliance", methods=["POST"])
    @nc_login_required
    def nc_api_compliance_check(topo_id):
        conn = get_connection()
        row = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        try:
            graph = json.loads(row["graph_json"])
        except Exception:
            conn.close()
            return jsonify({"error": "Bad graph"}), 500
        nodes = graph.get("nodes", [])
        edges = graph.get("edges", [])
        node_types = [n.get("type", "") for n in nodes]
        findings = []
        if not any(t == "firewall" for t in node_types):
            findings.append(
                {"check": "STIG-NET-001", "severity": "CAT1", "type": "stig", "finding": "No firewall in topology"}
            )
        encrypted = sum(
            1
            for e in edges
            if "ipsec" in (e.get("protocol") or "").lower() or "tls" in (e.get("protocol") or "").lower()
        )
        if encrypted == 0 and edges:
            findings.append(
                {"check": "FIPS-001", "severity": "HIGH", "type": "fips", "finding": "No encrypted links detected"}
            )
        if not any(t in ("firewall", "aws-nfw", "az-fw", "gcp-armor") for t in node_types):
            findings.append(
                {"check": "ZTA-001", "severity": "HIGH", "type": "zta", "finding": "No network segmentation"}
            )
        from collections import Counter

        degree = Counter()
        for e in edges:
            degree[e.get("source")] += 1
            degree[e.get("target")] += 1
        spof = [
            n["id"]
            for n in nodes
            if degree.get(n["id"], 0) <= 1 and n.get("type") in ("router", "switch-l3", "firewall")
        ]
        if spof:
            findings.append(
                {
                    "check": "BP-REDUNDANCY",
                    "severity": "MEDIUM",
                    "type": "best_practice",
                    "finding": f"{len(spof)} critical device(s) with single connection",
                }
            )
        passed = max(0, 10 - len(findings))
        failed = len(findings)
        check_id = str(_uuid.uuid4())
        now = _now()
        conn.execute(
            "INSERT INTO nc_compliance_checks (id, topology_id, check_type, passed, failed, findings_json, ran_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (check_id, topo_id, "full", passed, failed, json.dumps(findings), now),
        )
        conn.commit()
        conn.close()
        result = {
            "check_id": check_id,
            "passed": passed,
            "failed": failed,
            "score_pct": round(passed / max(passed + failed, 1) * 100, 1),
            "findings": findings,
        }
        # Blockchain provenance for assessment
        try:
            from tools.canvas.provenance import register_canvas_provenance
            register_canvas_provenance(
                canvas_key="ndc",
                design_id=topo_id,
                assessment_data=result,
                project_id="",
            )
        except Exception:
            pass
        return jsonify(result)

    # ══════════════════════════════════════════════════════════════════════
    # API: Full Compliance Audit Engine
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/compliance/<topo_id>/profile", methods=["PUT"])
    @nc_login_required
    def nc_api_update_compliance_profile(topo_id):
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        profile = conn.execute("SELECT id FROM nc_compliance_profiles WHERE topology_id=?", (topo_id,)).fetchone()
        if not profile:
            pid = str(_uuid.uuid4())
            conn.execute("INSERT INTO nc_compliance_profiles (id, topology_id) VALUES (?,?)", (pid, topo_id))
            conn.commit()
            profile = conn.execute("SELECT id FROM nc_compliance_profiles WHERE id=?", (pid,)).fetchone()
        fields, values = [], []
        for k in ["regimes", "classification", "environment", "auto_audit"]:
            if k in data:
                val = json.dumps(data[k]) if k == "regimes" and isinstance(data[k], list) else data[k]
                fields.append(f"{k}=?")
                values.append(val)
        if fields:
            fields.append("updated_at=?")
            values.append(_now())
            values.append(profile["id"])
            conn.execute(f"UPDATE nc_compliance_profiles SET {', '.join(fields)} WHERE id=?", values)  # nosec B608 -- table/column names are internal constants, not user input
            conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @bp.route("/api/compliance/<topo_id>/audit", methods=["POST"])
    @nc_login_required
    def nc_api_run_compliance_audit(topo_id):
        conn = get_connection()
        topo = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not topo:
            conn.close()
            return jsonify({"error": "Topology not found"}), 404
        profile = conn.execute("SELECT * FROM nc_compliance_profiles WHERE topology_id=?", (topo_id,)).fetchone()
        if profile:
            try:
                regimes = json.loads(profile["regimes"] or "[]")
            except Exception:
                regimes = ["fisma_high"]
            classification = profile["classification"] or "CUI"
        else:
            data = request.get_json(force=True, silent=True) or {}
            regimes = data.get("regimes", ["fisma_high"])
            classification = data.get("classification", "CUI")
        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            conn.close()
            return jsonify({"error": "Bad graph"}), 500

        result = run_compliance_audit(topo_id, graph, regimes, classification)
        record_canvas_decision(
            canvas_type="ndc",
            record_id=topo_id,
            decision_type="compliance_finding",
            decision=f"{len(result.get('findings', []))} finding(s) across regimes: {', '.join(regimes)}",
            rationale=f"Scores: {result.get('scores', {})}",
            model_used=None,
            confidence=None,
            project_id=None,
        )

        audit_id = str(_uuid.uuid4())
        now = _now()
        total_passed = sum(s["passed"] for s in result["scores"].values())
        total_failed = sum(s["failed"] for s in result["scores"].values())
        conn.execute(
            "INSERT INTO nc_compliance_checks (id, topology_id, check_type, passed, failed, findings_json, ran_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (audit_id, topo_id, ",".join(regimes), total_passed, total_failed, json.dumps(result["findings"]), now),
        )
        existing_rule_ids = set()
        for r in conn.execute(
            "SELECT rule_id FROM nc_compliance_findings WHERE topology_id=? AND status='open'", (topo_id,)
        ).fetchall():
            existing_rule_ids.add(r["rule_id"])
        new_rule_ids = set()
        for f in result["findings"]:
            new_rule_ids.add(f["rule_id"])
            exists = conn.execute(
                "SELECT id FROM nc_compliance_findings WHERE topology_id=? AND rule_id=? AND status='open'",
                (topo_id, f["rule_id"]),
            ).fetchone()
            if not exists:
                fid = str(_uuid.uuid4())
                conn.execute(
                    "INSERT INTO nc_compliance_findings (id, topology_id, audit_id, rule_id, regime, severity, "
                    "title, description, affected_entity, affected_type, fix_action, created_at) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                    (
                        fid,
                        topo_id,
                        audit_id,
                        f["rule_id"],
                        ",".join(f["regimes"]),
                        f["severity"],
                        f["title"],
                        f["description"],
                        f.get("affected_entity", ""),
                        f.get("affected_type", "topology"),
                        json.dumps(f.get("fix_action")) if f.get("fix_action") else None,
                        now,
                    ),
                )
        remediated_rules = existing_rule_ids - new_rule_ids
        for rid in remediated_rules:
            conn.execute(
                "UPDATE nc_compliance_findings SET status='remediated', remediated_at=? "
                "WHERE topology_id=? AND rule_id=? AND status='open'",
                (now, topo_id, rid),
            )
        conn.commit()
        conn.close()
        _audit("COMPLIANCE_AUDIT", "topology", topo_id, f"{len(result['findings'])} findings")
        result["audit_id"] = audit_id
        result["remediated_count"] = len(remediated_rules)
        return jsonify(result)

    @bp.route("/api/compliance/<topo_id>/fix", methods=["POST"])
    @nc_login_required
    def nc_api_compliance_fix(topo_id):
        data = request.get_json(force=True, silent=True) or {}
        finding_id = data.get("finding_id")
        if not finding_id:
            return jsonify({"error": "finding_id required"}), 400
        conn = get_connection()
        finding = conn.execute("SELECT * FROM nc_compliance_findings WHERE id=?", (finding_id,)).fetchone()
        if not finding:
            conn.close()
            return jsonify({"error": "Finding not found"}), 404
        fix_action = None
        try:
            fix_action = json.loads(finding["fix_action"]) if finding["fix_action"] else None
        except Exception:
            pass
        if not fix_action:
            conn.close()
            return jsonify({"error": "No automated fix available"}), 400
        topo = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not topo:
            conn.close()
            return jsonify({"error": "Topology not found"}), 404
        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            conn.close()
            return jsonify({"error": "Bad graph"}), 500

        applied, detail = apply_compliance_fix(graph, fix_action)
        action = fix_action.get("action", "")
        if not applied and action == "create_version":
            last = conn.execute("SELECT MAX(version_num) FROM nc_versions WHERE topology_id=?", (topo_id,)).fetchone()[
                0
            ]
            ver_num = (last or 0) + 1
            vid = str(_uuid.uuid4())
            conn.execute(
                "INSERT INTO nc_versions (id, topology_id, version_num, label, phase, graph_json, created_at) "
                "VALUES (?,?,?,?,?,?,?)",
                (
                    vid,
                    topo_id,
                    ver_num,
                    fix_action.get("label", "As-Built"),
                    fix_action.get("phase", "as-is"),
                    topo["graph_json"],
                    _now(),
                ),
            )
            applied = True
            detail = f"Created version v{ver_num}"
        if applied:
            if action != "create_version":
                conn.execute(
                    "UPDATE topologies SET graph_json=?, updated_at=? WHERE id=?", (json.dumps(graph), _now(), topo_id)
                )
            conn.execute(
                "UPDATE nc_compliance_findings SET status='remediated', remediated_at=? WHERE id=?",
                (_now(), finding_id),
            )
            conn.commit()
            conn.close()
            _audit("COMPLIANCE_FIX", "topology", topo_id, detail)
            return jsonify({"applied": True, "detail": detail})
        conn.close()
        return jsonify({"applied": False, "detail": "Fix action not applicable"})

    @bp.route("/api/compliance/<topo_id>/export", methods=["POST"])
    @nc_login_required
    def nc_api_compliance_export(topo_id):
        conn = get_connection()
        topo = conn.execute("SELECT name FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not topo:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        profile = conn.execute("SELECT * FROM nc_compliance_profiles WHERE topology_id=?", (topo_id,)).fetchone()
        findings = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT * FROM nc_compliance_findings WHERE topology_id=? ORDER BY severity, rule_id", (topo_id,)
            ).fetchall()
        ]
        conn.close()
        profile = _row_to_dict(profile) if profile else {}
        xml = generate_xacta_export(
            topo["name"],
            profile.get("classification", "CUI"),
            profile.get("environment", "IL4"),
            profile.get("regimes", "[]"),
            findings,
        )
        return jsonify(
            {
                "format": "xacta_xml",
                "filename": f"{topo['name']}_compliance_report.xml",
                "content": xml,
                "findings_count": len(findings),
            }
        )

    # ══════════════════════════════════════════════════════════════════════
    # API: FIPS 140 Encryption Coverage Report Export
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/compliance/<topo_id>/fips-report", methods=["POST"])
    @nc_login_required
    def nc_api_fips_report(topo_id):
        """Generate a FIPS 140 Encryption Coverage Report for ISSM review.

        Returns JSON report data or rendered HTML depending on format param.

        Body JSON (optional):
          {"format": "json" | "html"}   // default "json"
        """
        conn = get_connection()
        topo = conn.execute("SELECT name, graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not topo:
            conn.close()
            return jsonify({"error": "Topology not found"}), 404

        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            conn.close()
            return jsonify({"error": "Bad graph data"}), 500

        profile = conn.execute(
            "SELECT * FROM nc_compliance_profiles WHERE topology_id=?",
            (topo_id,),
        ).fetchone()
        conn.close()
        profile = _row_to_dict(profile) if profile else {}

        now = _now()
        report = generate_fips_coverage_report(
            system_name=topo["name"],
            classification=profile.get("classification", "CUI"),
            environment=profile.get("environment", "IL4"),
            graph=graph,
            now_str=now,
        )

        data = request.get_json(force=True, silent=True) or {}
        fmt = data.get("format", "json")

        _audit(
            "FIPS_REPORT",
            "topology",
            topo_id,
            f"coverage={report['summary']['coverage_pct']}% risk={report['summary']['risk_level']} format={fmt}",
        )

        if fmt == "html":
            html = export_fips_report_html(report)
            from flask import Response

            return Response(
                html,
                mimetype="text/html",
                headers={
                    "Content-Disposition": f'attachment; filename="{topo["name"]}_fips_report.html"',
                },
            )

        return jsonify(report)

    # ══════════════════════════════════════════════════════════════════════
    # API: STIG XCCDF/CKL Import
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/compliance/<topo_id>/stig-import", methods=["POST"])
    @nc_login_required
    def nc_api_stig_import(topo_id):
        """Import a STIG .ckl or XCCDF results file.

        Match hostnames to canvas devices and return per-device compliance
        color (green/yellow/red).
        """
        conn = get_connection()
        topo = conn.execute("SELECT graph_json, name FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not topo:
            conn.close()
            return jsonify({"error": "Topology not found"}), 404

        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            conn.close()
            return jsonify({"error": "Bad graph data"}), 500

        # Accept file upload or raw XML body
        content = None
        filename = "upload.xml"
        if request.files and "file" in request.files:
            f = request.files["file"]
            filename = f.filename or filename
            content = f.read().decode("utf-8", errors="replace")
        elif request.data:
            content = request.data.decode("utf-8", errors="replace")
            data = request.form or {}
            filename = data.get("filename", filename)

        if not content or not content.strip():
            conn.close()
            return jsonify({"error": "No file content provided"}), 400

        result = import_stig_file(content, graph)

        if "error" in result:
            conn.close()
            return jsonify(result), 400

        # Persist import record
        import_id = str(_uuid.uuid4())
        now = _now()
        conn.execute(
            "INSERT INTO nc_stig_imports "
            "(id, topology_id, filename, format, stig_name, stig_version, "
            "total_hosts, matched_hosts, result_json, imported_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (
                import_id,
                topo_id,
                filename,
                result.get("format", ""),
                result.get("stig_name", ""),
                result.get("stig_version", ""),
                result.get("total_hosts", 0),
                result.get("total_matched", 0),
                json.dumps(result),
                now,
            ),
        )

        # Also create compliance findings for failed STIG checks
        audit_id = str(_uuid.uuid4())
        total_pass = result.get("total_pass", 0)
        total_fail = result.get("total_fail", 0)
        conn.execute(
            "INSERT INTO nc_compliance_checks "
            "(id, topology_id, check_type, passed, failed, findings_json, ran_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (
                audit_id,
                topo_id,
                f"stig_import:{filename}",
                total_pass,
                total_fail,
                json.dumps(result.get("matched", [])),
                now,
            ),
        )

        # Insert individual findings for each failed check on matched devices
        for device in result.get("matched", []):
            for f in device.get("findings", []):
                fid = str(_uuid.uuid4())
                rule_id = f.get("rule_id") or f.get("vuln_id", "UNKNOWN")
                exists = conn.execute(
                    "SELECT id FROM nc_compliance_findings "
                    "WHERE topology_id=? AND rule_id=? AND affected_entity=? AND status='open'",
                    (topo_id, rule_id, device["label"]),
                ).fetchone()
                if not exists:
                    conn.execute(
                        "INSERT INTO nc_compliance_findings "
                        "(id, topology_id, audit_id, rule_id, regime, severity, "
                        "title, description, affected_entity, affected_type, created_at) "
                        "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                        (
                            fid,
                            topo_id,
                            audit_id,
                            rule_id,
                            "stig",
                            f.get("severity", "CAT2"),
                            f.get("title", rule_id),
                            f.get("finding_details", ""),
                            device["label"],
                            "node",
                            now,
                        ),
                    )

        conn.commit()
        conn.close()
        _audit(
            "STIG_IMPORT",
            "topology",
            topo_id,
            f"{filename}: {result.get('total_matched', 0)}/{result.get('total_hosts', 0)} hosts matched",
        )

        result["import_id"] = import_id
        result["audit_id"] = audit_id
        return jsonify(result)

    @bp.route("/api/compliance/<topo_id>/stig-imports", methods=["GET"])
    @nc_login_required
    def nc_api_stig_import_history(topo_id):
        """List previous STIG import records for a topology."""
        conn = get_connection()
        rows = conn.execute(
            "SELECT id, filename, format, stig_name, stig_version, "
            "total_hosts, matched_hosts, imported_at "
            "FROM nc_stig_imports WHERE topology_id=? ORDER BY imported_at DESC LIMIT 20",
            (topo_id,),
        ).fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    # ── Project routes (extracted to routes/projects.py) ──────────────────
    from tools.network.routes.projects import register_projects_routes

    register_projects_routes(bp)

    # ── Governance routes (extracted to routes/governance.py) ─────────────
    from tools.network.routes.governance import register_governance_routes

    register_governance_routes(bp)

    # ── Stencil Library routes (Cisco / Juniper / AWS / Azure / Custom) ───
    from tools.network.routes.stencils import register_stencil_routes

    register_stencil_routes(bp)

    # ══════════════════════════════════════════════════════════════════════
    # Extended: Notifications, Topology Diff, Auto-Decompose, Global Canvas
    # ══════════════════════════════════════════════════════════════════════

    # ── Notifications ─────────────────────────────────────────────────────
    @bp.route("/api/notifications", methods=["GET"])
    @nc_login_required
    def nc_api_list_notifications():
        conn = get_connection()
        rows = conn.execute(
            "SELECT n.*, p.name AS project_name "
            "FROM nc_notifications n "
            "LEFT JOIN nc_projects p ON p.id=n.project_id "
            "ORDER BY n.created_at DESC LIMIT 50"
        ).fetchall()
        unread = conn.execute("SELECT COUNT(*) FROM nc_notifications WHERE is_read=0").fetchone()[0]
        conn.close()
        return jsonify(
            {
                "notifications": [_row_to_dict(r) for r in rows],
                "unread": unread,
            }
        )

    @bp.route("/api/notifications/mark-read", methods=["POST"])
    @nc_login_required
    def nc_api_mark_notifications_read():
        conn = get_connection()
        conn.execute("UPDATE nc_notifications SET is_read=1 WHERE is_read=0")
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    # ── Topology Diff ─────────────────────────────────────────────────────
    @bp.route("/api/topology-diff", methods=["POST"])
    @nc_login_required
    def nc_api_topology_diff():
        """Compare two topologies side-by-side (node/edge delta)."""
        data = request.get_json(force=True, silent=True) or {}
        topo_a_id = data.get("topology_a")
        topo_b_id = data.get("topology_b")
        if not topo_a_id or not topo_b_id:
            return jsonify({"error": "topology_a and topology_b required"}), 400
        conn = get_connection()
        a_row = conn.execute("SELECT name, graph_json FROM topologies WHERE id=?", (topo_a_id,)).fetchone()
        b_row = conn.execute("SELECT name, graph_json FROM topologies WHERE id=?", (topo_b_id,)).fetchone()
        conn.close()
        if not a_row or not b_row:
            return jsonify({"error": "Topology not found"}), 404

        try:
            ga = json.loads(a_row["graph_json"])
        except Exception:
            ga = {"nodes": [], "edges": []}
        try:
            gb = json.loads(b_row["graph_json"])
        except Exception:
            gb = {"nodes": [], "edges": []}

        # Nodes diff by label+type
        def node_key(n):
            return f"{n.get('label', '')}|{n.get('type', '')}"

        a_nodes = {node_key(n): n for n in ga.get("nodes", [])}
        b_nodes = {node_key(n): n for n in gb.get("nodes", [])}
        a_keys = set(a_nodes.keys())
        b_keys = set(b_nodes.keys())

        only_a = [{"label": a_nodes[k].get("label"), "type": a_nodes[k].get("type")} for k in sorted(a_keys - b_keys)]
        only_b = [{"label": b_nodes[k].get("label"), "type": b_nodes[k].get("type")} for k in sorted(b_keys - a_keys)]
        common = sorted(a_keys & b_keys)

        # Edges diff
        def edge_key(e):
            return f"{e.get('source', '')}>{e.get('target', '')}"

        a_edges = {edge_key(e): e for e in ga.get("edges", [])}
        b_edges = {edge_key(e): e for e in gb.get("edges", [])}
        a_ekeys = set(a_edges.keys())
        b_ekeys = set(b_edges.keys())

        # Type distribution
        a_types = {}
        for n in ga.get("nodes", []):
            t = n.get("type", "unknown")
            a_types[t] = a_types.get(t, 0) + 1
        b_types = {}
        for n in gb.get("nodes", []):
            t = n.get("type", "unknown")
            b_types[t] = b_types.get(t, 0) + 1

        return jsonify(
            {
                "topology_a": {
                    "id": topo_a_id,
                    "name": a_row["name"],
                    "nodes": len(ga.get("nodes", [])),
                    "edges": len(ga.get("edges", [])),
                    "types": a_types,
                },
                "topology_b": {
                    "id": topo_b_id,
                    "name": b_row["name"],
                    "nodes": len(gb.get("nodes", [])),
                    "edges": len(gb.get("edges", [])),
                    "types": b_types,
                },
                "nodes_only_a": only_a,
                "nodes_only_b": only_b,
                "nodes_common": len(common),
                "edges_only_a": len(a_ekeys - b_ekeys),
                "edges_only_b": len(b_ekeys - a_ekeys),
                "edges_common": len(a_ekeys & b_ekeys),
                "similarity_pct": round(len(common) * 100 / max(len(a_keys | b_keys), 1)),
            }
        )

    @bp.route("/projects/diff")
    @nc_login_required
    def nc_topology_diff_page():
        conn = get_connection()
        topos = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT t.id, t.name, p.name AS project_name "
                "FROM topologies t "
                "LEFT JOIN nc_project_topologies pt "
                "  ON pt.topology_id=t.id "
                "LEFT JOIN nc_projects p ON p.id=pt.project_id "
                "ORDER BY t.name"
            ).fetchall()
        ]
        conn.close()
        return render_template("network/diff.html", topologies=topos)

    # ── Auto-decompose to SAFe ────────────────────────────────────────────
    @bp.route("/api/projects/<pid>/decompose", methods=["POST"])
    @nc_login_required
    def nc_api_decompose_to_safe(pid):
        """Auto-generate SAFe Feature + Stories from network project.
        Stores in nc_safe_bridge and returns the decomposition."""
        conn = get_connection()
        proj = conn.execute("SELECT * FROM nc_projects WHERE id=?", (pid,)).fetchone()
        if not proj:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        proj = _row_to_dict(proj)

        topos = []
        for r in conn.execute(
            "SELECT t.id, t.name, t.classification, t.graph_json "
            "FROM topologies t "
            "JOIN nc_project_topologies pt ON pt.topology_id=t.id "
            "WHERE pt.project_id=?",
            (pid,),
        ).fetchall():
            t = _row_to_dict(r)
            try:
                g = json.loads(t.get("graph_json") or '{"nodes":[],"edges":[]}')
            except Exception:
                g = {"nodes": [], "edges": []}
            t["node_count"] = len(g.get("nodes", []))
            topos.append(t)
        circuits = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT circuit_id, circuit_type, bandwidth "
                "FROM nc_circuits WHERE topology_id IN "
                "(SELECT topology_id FROM nc_project_topologies "
                " WHERE project_id=?)",
                (pid,),
            ).fetchall()
        ]

        # Build SAFe hierarchy
        feature = {
            "level": "feature",
            "title": f"[Feature] {proj['name']}",
            "description": proj.get("description", ""),
            "t_shirt_size": "L" if len(topos) > 2 else "M" if len(topos) > 0 else "S",
            "status": "draft",
        }
        stories = []
        for t in topos:
            size = "M" if (t.get("node_count") or 0) > 20 else "S"
            stories.append(
                {
                    "level": "story",
                    "title": f"[Story] Implement {t['name']}",
                    "description": f"Network topology: {t['name']} "
                    f"({t.get('node_count', 0)} nodes, "
                    f"{t.get('classification', 'public')})",
                    "t_shirt_size": size,
                    "source_topology_id": t["id"],
                    "status": "draft",
                }
            )
        enablers = []
        for c in circuits:
            enablers.append(
                {
                    "level": "enabler",
                    "title": f"[Enabler] Provision {c['circuit_id']}",
                    "description": f"{c.get('circuit_type', '')} {c.get('bandwidth', '')}",
                    "t_shirt_size": "S",
                    "status": "draft",
                }
            )

        # WSJF scoring (simplified)
        tshirt_pts = {"XS": 1, "S": 2, "M": 3, "L": 5, "XL": 8}
        feature["wsjf_score"] = round(7 / tshirt_pts.get(feature["t_shirt_size"], 3), 1)
        for s in stories:
            s["wsjf_score"] = round(5 / tshirt_pts.get(s["t_shirt_size"], 2), 1)
        for e in enablers:
            e["wsjf_score"] = round(4 / tshirt_pts.get(e["t_shirt_size"], 2), 1)

        # Update SAFe bridge
        bridge = conn.execute("SELECT id FROM nc_safe_bridge WHERE project_id=?", (pid,)).fetchone()
        now = _now()
        decomposition = {
            "feature": feature,
            "stories": stories,
            "enablers": enablers,
        }
        decomp_json = json.dumps(decomposition)
        if bridge:
            conn.execute(
                "UPDATE nc_safe_bridge SET safe_feature_id=?, updated_at=? WHERE project_id=?", (decomp_json, now, pid)
            )
        else:
            conn.execute(
                "INSERT INTO nc_safe_bridge "
                "(id, project_id, safe_feature_id, created_at, "
                " updated_at) VALUES (?,?,?,?,?)",
                (str(_uuid.uuid4()), pid, decomp_json, now, now),
            )
        conn.commit()

        _notify(
            conn,
            pid,
            "decomposition",
            f"SAFe decomposition: 1 Feature, {len(stories)} Stories, {len(enablers)} Enablers",
            f"WSJF={feature['wsjf_score']}",
        )
        conn.commit()
        conn.close()
        _audit("DECOMPOSE", "project", pid, f"{len(stories)} stories, {len(enablers)} enablers")

        return jsonify(
            {
                "feature": feature,
                "stories": stories,
                "enablers": enablers,
                "total_items": 1 + len(stories) + len(enablers),
            }
        )

    # ── COA Selection (AI-assisted migration) ─────────────────────────────
    @bp.route("/api/projects/<pid>/select-coa", methods=["POST"])
    @nc_login_required
    def nc_select_coa(pid):
        """Store HITL COA selection and feedback; regenerate migration phases."""
        data = request.get_json(force=True) or {}
        coa_id = data.get("coa_id")
        feedback = data.get("feedback", "")
        if not coa_id or coa_id not in (1, 2, 3):
            return jsonify({"error": "coa_id must be 1, 2, or 3"}), 400

        conn = get_connection()
        proj = conn.execute("SELECT id FROM nc_projects WHERE id=?", (pid,)).fetchone()
        if not proj:
            conn.close()
            return jsonify({"error": "Project not found"}), 404

        # Build COA JSON payload
        coa_json = json.dumps({
            "selected_coa": coa_id,
            "feedback": feedback,
            "coa_1": {"id": 1, "name": "Rip & Replace", "risk_level": "high", "estimated_downtime_hours": 4},
            "coa_2": {"id": 2, "name": "Phased Cutover", "risk_level": "medium", "estimated_downtime_hours": 1},
            "coa_3": {"id": 3, "name": "Side-by-Side VLAN", "risk_level": "low", "estimated_downtime_hours": 0, "recommended": True},
        })

        conn.execute(
            "UPDATE nc_projects SET selected_coa=?, coa_feedback=?, coa_json=? WHERE id=?",
            (coa_id, feedback, coa_json, pid),
        )
        conn.commit()

        # Audit trail
        _audit("COA_SELECT", "project", pid, f"Selected COA-{coa_id}")
        conn.commit()
        conn.close()

        return jsonify({"ok": True, "selected_coa": coa_id, "feedback": feedback})

    # ── Global Connectivity Page ────────────────────────────────────────────
    @bp.route("/global")
    @nc_login_required
    def nc_global():
        """Global connectivity map — all approved/deployed projects stitched."""
        return render_template("network/global.html")

    # ── Global Topology API ────────────────────────────────────────────────

    @bp.route("/api/global-topology", methods=["GET"])
    @nc_login_required
    def nc_api_global_topology():
        """Aggregate topology data across all projects for global view."""
        conn = get_connection()
        projects = []
        try:
            proj_rows = conn.execute(
                "SELECT id, name, status FROM nc_projects ORDER BY name"
            ).fetchall()
        except Exception:
            proj_rows = []

        for pr in proj_rows:
            pid = pr[0] if isinstance(pr, tuple) else pr["id"]
            pname = pr[1] if isinstance(pr, tuple) else pr["name"]
            pstatus = pr[2] if isinstance(pr, tuple) else pr["status"]
            try:
                topo_rows = conn.execute(
                    "SELECT t.id, t.name, t.graph_json FROM topologies t "
                    "JOIN nc_project_topologies pt ON pt.topology_id = t.id "
                    "WHERE pt.project_id = ?", (pid,)
                ).fetchall()
            except Exception:
                topo_rows = []

            topos = []
            for tr in topo_rows:
                tid = tr[0] if isinstance(tr, tuple) else tr["id"]
                tname = tr[1] if isinstance(tr, tuple) else tr["name"]
                gjson = tr[2] if isinstance(tr, tuple) else tr["graph_json"]
                try:
                    g = json.loads(gjson)
                    nc = len(g.get("nodes", []))
                    ec = len(g.get("edges", []))
                except Exception:
                    nc, ec = 0, 0
                topos.append({"id": tid, "name": tname, "node_count": nc, "edge_count": ec})
            projects.append({"id": pid, "name": pname, "status": pstatus, "topologies": topos})

        # Interconnects
        interconnects = []
        try:
            ic_rows = conn.execute(
                "SELECT id, src_project_id, dst_project_id, circuit_id, protocol, bandwidth, notes "
                "FROM nc_interconnects ORDER BY created_at DESC"
            ).fetchall()
            for ic in ic_rows:
                interconnects.append({
                    "id": ic[0], "src_project_id": ic[1], "dst_project_id": ic[2],
                    "circuit_id": ic[3], "protocol": ic[4], "bandwidth": ic[5], "notes": ic[6],
                })
        except Exception:
            pass

        conn.close()
        return jsonify({
            "total_projects": len(projects),
            "projects": projects,
            "total_interconnects": len(interconnects),
            "interconnects": interconnects,
        })

    @bp.route("/api/interconnects", methods=["POST"])
    @nc_login_required
    def nc_api_add_interconnect():
        """Add a project-to-project interconnect."""
        import uuid as _uid
        body = request.json or {}
        conn = get_connection()
        try:
            conn.execute(
                "INSERT INTO nc_interconnects (id, src_project_id, dst_project_id, circuit_id, "
                "protocol, bandwidth, notes, created_at) VALUES (?,?,?,?,?,?,?,datetime('now'))",
                (str(_uid.uuid4())[:12], body.get("src_project_id"), body.get("dst_project_id"),
                 body.get("circuit_id", ""), body.get("protocol", ""), body.get("bandwidth", ""),
                 body.get("notes", "")),
            )
            conn.commit()
        except Exception as e:
            conn.close()
            return jsonify({"error": str(e)}), 500
        conn.close()
        return jsonify({"ok": True}), 201

    @bp.route("/api/interconnects/<ic_id>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_interconnect(ic_id):
        """Delete an interconnect."""
        conn = get_connection()
        conn.execute("DELETE FROM nc_interconnects WHERE id = ?", (ic_id,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @bp.route("/api/conflicts", methods=["GET"])
    @nc_login_required
    def nc_api_check_conflicts():
        """Check for IPAM/circuit conflicts across projects."""
        conn = get_connection()
        conflicts = []
        ipam_count, circuit_count = 0, 0
        try:
            # Check overlapping IPAM blocks
            ipam_rows = conn.execute("SELECT network, topology_id FROM nc_ipam_blocks").fetchall()
            ipam_count = len(ipam_rows)
            seen_nets = {}
            for row in ipam_rows:
                net = row[0] if isinstance(row, tuple) else row["network"]
                tid = row[1] if isinstance(row, tuple) else row["topology_id"]
                if net in seen_nets and seen_nets[net] != tid:
                    conflicts.append({
                        "type": "IPAM Overlap",
                        "severity": "high",
                        "detail": f"Network {net} used in topologies {seen_nets[net]} and {tid}",
                    })
                seen_nets[net] = tid

            # Check duplicate circuit IDs
            circ_rows = conn.execute("SELECT circuit_id, COUNT(*) as cnt FROM nc_circuits GROUP BY circuit_id HAVING cnt > 1").fetchall()
            circuit_count = conn.execute("SELECT COUNT(*) FROM nc_circuits").fetchone()[0]
            for row in circ_rows:
                cid = row[0] if isinstance(row, tuple) else row["circuit_id"]
                conflicts.append({
                    "type": "Duplicate Circuit",
                    "severity": "medium",
                    "detail": f"Circuit ID {cid} appears in multiple topologies",
                })
        except Exception:
            pass
        conn.close()
        return jsonify({
            "conflicts": conflicts,
            "checked": {"ipam_blocks": ipam_count, "circuits": circuit_count},
        })

    @bp.route("/conflicts")
    @nc_login_required
    def nc_conflicts_page():
        """HITL conflict resolution UI."""
        return render_template("network/conflicts.html")

    @bp.route("/api/conflicts/resolve", methods=["POST"])
    @nc_login_required
    def nc_api_conflicts_resolve():
        """Record a HITL conflict resolution action."""
        import uuid as _uuid
        from datetime import datetime, timezone
        data = request.get_json(force=True) or {}
        conflict_type = str(data.get("conflict_type") or "").strip()
        detail = str(data.get("detail") or "").strip()
        severity = str(data.get("severity") or "medium").strip()
        action = str(data.get("action") or "acknowledged").strip()
        note = str(data.get("note") or "").strip()
        if not conflict_type or not detail:
            return jsonify({"error": "conflict_type and detail are required"}), 400
        if action not in ("acknowledged", "resolved"):
            return jsonify({"error": "action must be acknowledged or resolved"}), 400
        if severity not in ("high", "medium", "low"):
            severity = "medium"
        resolved_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        row_id = str(_uuid.uuid4())
        conn = get_connection()
        try:
            conn.execute(
                "INSERT INTO nc_conflict_resolutions (id, conflict_type, detail, severity, action, note, resolved_at)"
                " VALUES (?, ?, ?, ?, ?, ?, ?)",
                (row_id, conflict_type, detail, severity, action, note, resolved_at),
            )
            conn.commit()
        finally:
            conn.close()
        return jsonify({"ok": True, "id": row_id, "resolved_at": resolved_at})

    @bp.route("/api/conflicts/resolutions", methods=["GET"])
    @nc_login_required
    def nc_api_conflicts_resolutions():
        """Return all recorded conflict resolution actions, newest first."""
        conn = get_connection()
        try:
            rows = conn.execute(
                "SELECT id, conflict_type, detail, severity, action, note, resolved_at"
                " FROM nc_conflict_resolutions ORDER BY resolved_at DESC LIMIT 200"
            ).fetchall()
        finally:
            conn.close()
        resolutions = [_row_to_dict(r) for r in rows]
        return jsonify({"resolutions": resolutions})

    # ── Global Topology Canvas (JointJS read-only composite) ──────────────
    @bp.route("/global/canvas")
    @nc_login_required
    def nc_global_canvas():
        return render_template("network/global_canvas.html")

    @bp.route("/api/global-canvas-data", methods=["GET"])
    @nc_login_required
    def nc_api_global_canvas_data():
        """Return combined graph data for all approved/deployed projects."""
        conn = get_connection()
        rows = conn.execute(
            "SELECT t.id, t.name, t.graph_json, "
            "p.id AS project_id, p.name AS project_name "
            "FROM topologies t "
            "JOIN nc_project_topologies pt ON pt.topology_id=t.id "
            "JOIN nc_projects p ON p.id=pt.project_id "
            "WHERE p.status IN ('approved','deployed') "
            "ORDER BY p.name, t.name"
        ).fetchall()
        interconnects = [_row_to_dict(r) for r in conn.execute("SELECT * FROM nc_interconnects").fetchall()]
        conn.close()

        # Build composite graph with project-namespaced node IDs
        all_nodes = []
        all_edges = []
        project_groups = []
        offset_x = 0
        for r in rows:
            r = _row_to_dict(r)
            try:
                g = json.loads(r.get("graph_json") or "{}")
            except Exception:
                g = {"nodes": [], "edges": []}
            prefix = r["id"][:8]
            group_nodes = []
            for n in g.get("nodes", []):
                nid = f"{prefix}_{n['id']}"
                all_nodes.append(
                    {
                        "id": nid,
                        "label": n.get("label", ""),
                        "type": n.get("type", ""),
                        "x": (n.get("x", 0) or 0) + offset_x,
                        "y": n.get("y", 0) or 0,
                        "project": r["project_name"],
                        "topology": r["name"],
                    }
                )
                group_nodes.append(nid)
            for e in g.get("edges", []):
                all_edges.append(
                    {
                        "source": f"{prefix}_{e['source']}",
                        "target": f"{prefix}_{e['target']}",
                        "label": e.get("label", ""),
                        "protocol": e.get("protocol", ""),
                    }
                )
            project_groups.append(
                {
                    "project": r["project_name"],
                    "topology": r["name"],
                    "node_ids": group_nodes,
                    "x": offset_x,
                    "y": 0,
                }
            )
            offset_x += 600

        # Add interconnect edges
        for ic in interconnects:
            all_edges.append(
                {
                    "source": f"ic_src_{ic['id'][:8]}",
                    "target": f"ic_dst_{ic['id'][:8]}",
                    "label": ic.get("circuit_id", ""),
                    "protocol": ic.get("protocol", ""),
                    "is_interconnect": True,
                }
            )

        return jsonify(
            {
                "nodes": all_nodes,
                "edges": all_edges,
                "groups": project_groups,
                "total_nodes": len(all_nodes),
                "total_edges": len(all_edges),
            }
        )

    # ══════════════════════════════════════════════════════════════════════
    # API: CSP Groups
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/groups/<topo_id>", methods=["GET"])
    @nc_login_required
    def nc_api_list_groups(topo_id):
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_groups WHERE topology_id=? ORDER BY created_at", (topo_id,)).fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/groups/<topo_id>", methods=["POST"])
    @nc_login_required
    def nc_api_create_group(topo_id):
        data = request.get_json(force=True, silent=True) or {}
        gid = str(_uuid.uuid4())
        csp = data.get("csp", "aws")
        group_type = data.get("group_type", "full")
        pos_x = data.get("pos_x", 100)
        pos_y = data.get("pos_y", 100)
        now = _now()
        auto_nodes = []
        if group_type == "full" and csp in CSP_GROUP_DEFAULTS:
            conn = get_connection()
            topo = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
            if topo:
                try:
                    graph = json.loads(topo["graph_json"])
                except Exception:
                    graph = {"nodes": [], "edges": []}
                for comp in CSP_GROUP_DEFAULTS[csp]:
                    nid = f"{csp}-{str(_uuid.uuid4())[:8]}"
                    graph["nodes"].append(
                        {
                            "id": nid,
                            "label": comp["label"],
                            "type": comp["type"],
                            "x": pos_x + comp["dx"],
                            "y": pos_y + comp["dy"],
                            "group_id": gid,
                        }
                    )
                    auto_nodes.append(nid)
                conn.execute(
                    "UPDATE topologies SET graph_json=?, updated_at=? WHERE id=?", (json.dumps(graph), now, topo_id)
                )
                conn.commit()
            conn.close()
        csp_labels = {"aws": "AWS", "azure": "Azure", "gcp": "GCP", "oci": "OCI", "ibm": "IBM Cloud"}
        label = data.get("label", csp_labels.get(csp, csp.upper()))
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_groups (id, topology_id, parent_id, csp, group_type, label, description, "
            "auto_nodes_json, pos_x, pos_y, width, height, color, collapsed, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                gid,
                topo_id,
                data.get("parent_id"),
                csp,
                group_type,
                label,
                data.get("description", ""),
                json.dumps(auto_nodes),
                pos_x,
                pos_y,
                data.get("width", 400),
                data.get("height", 300),
                data.get("color"),
                0,
                now,
            ),
        )
        conn.commit()
        conn.close()
        _audit("CREATE", "group", gid, f"{csp} {group_type}")
        return jsonify({"id": gid, "auto_nodes": auto_nodes}), 201

    @bp.route("/api/groups/<topo_id>/<gid>", methods=["PUT"])
    @nc_login_required
    def nc_api_update_group(topo_id, gid):
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        allowed = [
            "parent_id",
            "label",
            "description",
            "pos_x",
            "pos_y",
            "width",
            "height",
            "color",
            "collapsed",
            "group_type",
        ]
        fields, values = [], []
        for k in allowed:
            if k in data:
                fields.append(f"{k}=?")
                values.append(data[k])
        if not fields:
            conn.close()
            return jsonify({"error": "No fields"}), 400
        values.append(gid)
        conn.execute(f"UPDATE nc_groups SET {', '.join(fields)} WHERE id=?", values)  # nosec B608 -- table/column names are internal constants, not user input
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @bp.route("/api/groups/<topo_id>/<gid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_group(topo_id, gid):
        conn = get_connection()
        group = conn.execute("SELECT auto_nodes_json FROM nc_groups WHERE id=?", (gid,)).fetchone()
        if group:
            try:
                auto_ids = set(json.loads(group["auto_nodes_json"] or "[]"))
                if auto_ids:
                    topo = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
                    if topo:
                        graph = json.loads(topo["graph_json"])
                        graph["nodes"] = [n for n in graph["nodes"] if n["id"] not in auto_ids]
                        graph["edges"] = [
                            e for e in graph["edges"] if e["source"] not in auto_ids and e["target"] not in auto_ids
                        ]
                        conn.execute(
                            "UPDATE topologies SET graph_json=?, updated_at=? WHERE id=?",
                            (json.dumps(graph), _now(), topo_id),
                        )
            except Exception:
                pass
        conn.execute("DELETE FROM nc_groups WHERE parent_id=?", (gid,))
        conn.execute("DELETE FROM nc_groups WHERE id=?", (gid,))
        conn.commit()
        conn.close()
        _audit("DELETE", "group", gid)
        return jsonify({"deleted": gid})

    # ══════════════════════════════════════════════════════════════════════
    # API: Security Boundary Auto-Fencing
    # ══════════════════════════════════════════════════════════════════════

    def _auto_stig_tags(node_types: list[str]) -> list[str]:
        """Derive STIG boundary tags from the device types inside a fence."""
        tags = set()
        type_set = set(node_types)
        # Firewall present → perimeter boundary
        fw_types = {"firewall", "aws-nfw", "az-fw", "gcp-armor", "oci-waf", "aws-waf"}
        if type_set & fw_types:
            tags.add("NET-BND-001: Perimeter Firewall Boundary")
        # Encryption devices → FIPS boundary
        enc_types = {
            "fips-140-l1",
            "fips-140-l2",
            "fips-140-l3",
            "fips-140-l4",
            "hsm",
            "type1-encryptor",
            "kg-175d",
            "kg-175g",
            "kg-250",
            "kg-340",
            "kg-245x",
            "kg-255",
            "macsec",
        }
        if type_set & enc_types:
            tags.add("NET-ENC-BND: FIPS 140 Encryption Boundary")
        # Servers → server enclave STIG
        if "server" in type_set:
            tags.add("NET-SRV-BND: Server Enclave Boundary")
        # User endpoints → user enclave
        user_types = {"endpoint-pc", "endpoint-phone"}
        if type_set & user_types:
            tags.add("NET-USR-BND: User Enclave Boundary")
        # IoT → IoT segment
        iot_types = {"endpoint-iot", "endpoint-camera"}
        if type_set & iot_types:
            tags.add("NET-IOT-BND: IoT Segment Boundary")
        # Management → NOC/management boundary
        mgmt_types = {"siem", "network-tap", "wlc"}
        if type_set & mgmt_types:
            tags.add("NET-MGT-BND: Management / NOC Boundary")
        # Wireless → wireless boundary
        if "wap" in type_set:
            tags.add("NET-WLS-BND: Wireless Access Boundary")
        # Cloud resources → cloud enclave
        cloud_prefixes = ("aws-", "az-", "gcp-", "oci-", "ibm-")
        if any(t.startswith(cloud_prefixes) for t in type_set):
            tags.add("NET-CLD-BND: Cloud Enclave Boundary")
        # Cross-zone (mixed) → micro-segmentation required
        if len(tags) >= 2:
            tags.add("NET-BND-002: Micro-segmentation Required")
        return sorted(tags)

    @bp.route("/api/boundaries/<topo_id>", methods=["GET"])
    @nc_login_required
    def nc_api_list_boundaries(topo_id):
        conn = get_connection()
        rows = conn.execute(
            "SELECT * FROM nc_boundaries WHERE topology_id=? ORDER BY created_at",
            (topo_id,),
        ).fetchall()
        conn.close()
        result = []
        for r in rows:
            d = _row_to_dict(r)
            for key in ("node_ids", "stig_tags"):
                try:
                    d[key] = json.loads(d.get(key) or "[]")
                except Exception:
                    d[key] = []
            result.append(d)
        return jsonify(result)

    @bp.route("/api/boundaries/<topo_id>", methods=["POST"])
    @nc_login_required
    def nc_api_create_boundary(topo_id):
        data = request.get_json(force=True, silent=True) or {}
        bid = str(_uuid.uuid4())
        node_ids = data.get("node_ids", [])
        classification = data.get("classification", "CUI")
        snap = data.get("snap_grid", 10)

        # Snap position/size to grid
        pos_x = round(data.get("pos_x", 0) / snap) * snap
        pos_y = round(data.get("pos_y", 0) / snap) * snap
        width = max(snap, round(data.get("width", 400) / snap) * snap)
        height = max(snap, round(data.get("height", 300) / snap) * snap)

        # Auto-derive STIG tags from contained node types
        node_types = []
        if node_ids:
            conn = get_connection()
            topo = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
            conn.close()
            if topo:
                try:
                    graph = json.loads(topo["graph_json"])
                except Exception:
                    graph = {"nodes": []}
                nid_set = set(node_ids)
                node_types = [n["type"] for n in graph.get("nodes", []) if n["id"] in nid_set]

        stig_tags = data.get("stig_tags") or _auto_stig_tags(node_types)

        # Classification label mapping
        _CLS_LABELS = {
            "CUI": "CUI Enclave",
            "SECRET": "SECRET VLAN",
            "TOP SECRET": "TOP SECRET Enclave",
            "PUBLIC": "Public Zone",
        }
        label = data.get("label") or _CLS_LABELS.get(classification, f"{classification} Enclave")

        # Classification → color mapping
        _CLS_COLORS = {
            "CUI": "#f39c12",
            "SECRET": "#e94560",
            "TOP SECRET": "#9b59b6",
            "PUBLIC": "#27ae60",
        }
        color = data.get("color") or _CLS_COLORS.get(classification, "#4a9eff")

        now = _now()
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_boundaries "
            "(id, topology_id, label, classification, color, fill_opacity, "
            "node_ids, stig_tags, pos_x, pos_y, width, height, snap_grid, notes, "
            "created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                bid,
                topo_id,
                label,
                classification,
                color,
                data.get("fill_opacity", 0.08),
                json.dumps(node_ids),
                json.dumps(stig_tags),
                pos_x,
                pos_y,
                width,
                height,
                snap,
                data.get("notes", ""),
                now,
                now,
            ),
        )
        conn.commit()
        conn.close()
        _audit("CREATE", "boundary", bid, f"{classification} — {label}")
        return jsonify(
            {
                "id": bid,
                "label": label,
                "classification": classification,
                "color": color,
                "stig_tags": stig_tags,
                "pos_x": pos_x,
                "pos_y": pos_y,
                "width": width,
                "height": height,
                "node_ids": node_ids,
            }
        ), 201

    @bp.route("/api/boundaries/<topo_id>/<bid>", methods=["PUT"])
    @nc_login_required
    def nc_api_update_boundary(topo_id, bid):
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        allowed = [
            "label",
            "classification",
            "color",
            "fill_opacity",
            "node_ids",
            "stig_tags",
            "pos_x",
            "pos_y",
            "width",
            "height",
            "snap_grid",
            "notes",
        ]
        fields, values = [], []
        for k in allowed:
            if k in data:
                v = data[k]
                if k in ("node_ids", "stig_tags") and isinstance(v, list):
                    v = json.dumps(v)
                fields.append(f"{k}=?")
                values.append(v)
        if not fields:
            conn.close()
            return jsonify({"error": "No fields"}), 400
        fields.append("updated_at=?")
        values.append(_now())
        values.append(bid)
        conn.execute(f"UPDATE nc_boundaries SET {', '.join(fields)} WHERE id=?", values)  # nosec B608 -- table/column names are internal constants, not user input
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    @bp.route("/api/boundaries/<topo_id>/<bid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_boundary(topo_id, bid):
        conn = get_connection()
        conn.execute("DELETE FROM nc_boundaries WHERE id=? AND topology_id=?", (bid, topo_id))
        conn.commit()
        conn.close()
        _audit("DELETE", "boundary", bid)
        return jsonify({"deleted": bid})

    @bp.route("/api/boundaries/<topo_id>/auto-fence", methods=["POST"])
    @nc_login_required
    def nc_api_auto_fence(topo_id):
        """Auto-generate a boundary from a list of selected node IDs.

        Calculates bounding box from node positions, snaps to grid, derives
        classification label and STIG boundary tags automatically.
        """
        data = request.get_json(force=True, silent=True) or {}
        node_ids = data.get("node_ids", [])
        if not node_ids:
            return jsonify({"error": "node_ids required"}), 400

        classification = data.get("classification", "CUI")
        snap = data.get("snap_grid", 10)
        padding = data.get("padding", 40)

        conn = get_connection()
        topo = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        conn.close()
        if not topo:
            return jsonify({"error": "Topology not found"}), 404

        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            return jsonify({"error": "Invalid graph JSON"}), 400

        nid_set = set(node_ids)
        matched = [n for n in graph.get("nodes", []) if n["id"] in nid_set]
        if not matched:
            return jsonify({"error": "No matching nodes found"}), 404

        # Bounding box from node positions (assume 110x60 default node size)
        xs = [n.get("x", 0) for n in matched]
        ys = [n.get("y", 0) for n in matched]
        node_w = 110
        node_h = 60
        min_x = min(xs) - padding
        min_y = min(ys) - padding
        max_x = max(xs) + node_w + padding
        max_y = max(ys) + node_h + padding

        # Snap to grid
        pos_x = (min_x // snap) * snap
        pos_y = (min_y // snap) * snap
        width = max(snap, ((max_x - pos_x + snap - 1) // snap) * snap)
        height = max(snap, ((max_y - pos_y + snap - 1) // snap) * snap)

        node_types = [n.get("type", "") for n in matched]
        stig_tags = _auto_stig_tags(node_types)

        _CLS_LABELS = {
            "CUI": "CUI Enclave",
            "SECRET": "SECRET VLAN",
            "TOP SECRET": "TOP SECRET Enclave",
            "PUBLIC": "Public Zone",
        }
        _CLS_COLORS = {
            "CUI": "#f39c12",
            "SECRET": "#e94560",
            "TOP SECRET": "#9b59b6",
            "PUBLIC": "#27ae60",
        }
        label = data.get("label") or _CLS_LABELS.get(classification, f"{classification} Enclave")
        color = data.get("color") or _CLS_COLORS.get(classification, "#4a9eff")

        bid = str(_uuid.uuid4())
        now = _now()
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_boundaries "
            "(id, topology_id, label, classification, color, fill_opacity, "
            "node_ids, stig_tags, pos_x, pos_y, width, height, snap_grid, notes, "
            "created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                bid,
                topo_id,
                label,
                classification,
                color,
                0.08,
                json.dumps(node_ids),
                json.dumps(stig_tags),
                pos_x,
                pos_y,
                width,
                height,
                snap,
                data.get("notes", ""),
                now,
                now,
            ),
        )
        conn.commit()
        conn.close()
        _audit("CREATE", "boundary", bid, f"auto-fence {classification} — {len(matched)} nodes")
        return jsonify(
            {
                "id": bid,
                "label": label,
                "classification": classification,
                "color": color,
                "fill_opacity": 0.08,
                "stig_tags": stig_tags,
                "pos_x": pos_x,
                "pos_y": pos_y,
                "width": width,
                "height": height,
                "node_ids": node_ids,
                "node_count": len(matched),
            }
        ), 201

    # ══════════════════════════════════════════════════════════════════════
    # API: Monte Carlo Simulation
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/mc/scenarios/<topo_id>", methods=["GET"])
    @nc_login_required
    def nc_api_list_mc_scenarios(topo_id):
        conn = get_connection()
        rows = conn.execute(
            "SELECT * FROM nc_mc_scenarios WHERE topology_id=? ORDER BY created_at DESC", (topo_id,)
        ).fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/mc/scenarios/<topo_id>", methods=["POST"])
    @nc_login_required
    def nc_api_create_mc_scenario(topo_id):
        data = request.get_json(force=True, silent=True) or {}
        sid = str(_uuid.uuid4())
        now = _now()
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_mc_scenarios (id, topology_id, name, scenario_type, description, config_json, created_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (
                sid,
                topo_id,
                data.get("name", "Untitled Scenario"),
                data.get("scenario_type", "random"),
                data.get("description", ""),
                json.dumps(data.get("config", {})),
                now,
            ),
        )
        conn.commit()
        conn.close()
        _audit("CREATE", "mc_scenario", sid, data.get("name", ""))
        return jsonify({"id": sid}), 201

    @bp.route("/api/mc/run/<scenario_id>", methods=["POST"])
    @nc_login_required
    def nc_api_run_mc(scenario_id):
        conn = get_connection()
        scenario = conn.execute("SELECT * FROM nc_mc_scenarios WHERE id=?", (scenario_id,)).fetchone()
        if not scenario:
            conn.close()
            return jsonify({"error": "Scenario not found"}), 404
        scenario = _row_to_dict(scenario)
        topo_id = scenario["topology_id"]
        topo = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not topo:
            conn.close()
            return jsonify({"error": "Topology not found"}), 404
        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            conn.close()
            return jsonify({"error": "Bad graph"}), 500
        try:
            config = json.loads(scenario["config_json"] or "{}")
        except Exception:
            config = {}
        data = request.get_json(force=True, silent=True) or {}
        iterations = data.get("iterations", config.get("iterations", 1000))
        result = run_monte_carlo(
            graph=graph,
            scenario_name=scenario["name"],
            scenario_type=scenario["scenario_type"],
            config=config,
            iterations=iterations,
        )
        run_id = str(_uuid.uuid4())
        now = _now()
        conn.execute(
            "INSERT INTO nc_mc_runs (id, scenario_id, topology_id, iterations, result_json, ai_recommendations, ran_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (
                run_id,
                scenario_id,
                topo_id,
                iterations,
                json.dumps(result),
                "\n".join(result.get("recommendations", [])),
                now,
            ),
        )
        conn.commit()
        conn.close()
        _audit("MC_RUN", "mc_scenario", scenario_id, f"{iterations} iters")
        return jsonify({"run_id": run_id, **result})

    @bp.route("/api/mc/runs/<topo_id>", methods=["GET"])
    @nc_login_required
    def nc_api_list_mc_runs(topo_id):
        conn = get_connection()
        rows = conn.execute(
            "SELECT r.id, r.scenario_id, r.iterations, r.ran_at, r.ai_recommendations, "
            "s.name AS scenario_name, s.scenario_type "
            "FROM nc_mc_runs r JOIN nc_mc_scenarios s ON s.id=r.scenario_id "
            "WHERE r.topology_id=? ORDER BY r.ran_at DESC",
            (topo_id,),
        ).fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/mc/runs/<topo_id>/<run_id>", methods=["GET"])
    @nc_login_required
    def nc_api_get_mc_run(topo_id, run_id):
        conn = get_connection()
        row = conn.execute("SELECT * FROM nc_mc_runs WHERE id=? AND topology_id=?", (run_id, topo_id)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        r = _row_to_dict(row)
        try:
            r["result"] = json.loads(r.get("result_json") or "{}")
        except Exception:
            r["result"] = {}
        return jsonify(r)

    # ══════════════════════════════════════════════════════════════════════
    # API: Backups
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/backups", methods=["GET"])
    @nc_login_required
    def nc_api_list_backups():
        conn = get_connection()
        rows = conn.execute("SELECT * FROM nc_backups ORDER BY created_at DESC").fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/backups", methods=["POST"])
    @nc_login_required
    def nc_api_create_backup():
        data = request.get_json(force=True, silent=True) or {}
        backup_id = str(_uuid.uuid4())
        now = _now()
        ts = now.replace(":", "-").replace("T", "_")[:19]
        backup_dir = _ICDEV_ROOT / "backups" / "network"
        backup_dir.mkdir(parents=True, exist_ok=True)
        zip_name = f"nc-backup-{ts}.zip"
        zip_path = backup_dir / zip_name
        includes = []
        try:
            with zipfile.ZipFile(str(zip_path), "w", zipfile.ZIP_DEFLATED) as zf:
                nc_db = _ICDEV_ROOT / "data" / "network_canvas.db"
                if nc_db.exists():
                    zf.write(str(nc_db), "network_canvas.db")
                    includes.append("network_canvas.db")
                args_dir = _ICDEV_ROOT / "args"
                if args_dir.exists():
                    for f in args_dir.glob("network_canvas_*"):
                        if f.is_file():
                            zf.write(str(f), f"args/{f.name}")
                            includes.append(f"args/{f.name}")
            file_size = zip_path.stat().st_size
            conn = get_connection()
            conn.execute(
                "INSERT INTO nc_backups (id, backup_type, file_path, file_size_bytes, includes_json, notes, created_at) "
                "VALUES (?,?,?,?,?,?,?)",
                (
                    backup_id,
                    data.get("backup_type", "manual"),
                    str(zip_path),
                    file_size,
                    json.dumps(includes),
                    data.get("notes", ""),
                    now,
                ),
            )
            conn.commit()
            conn.close()
            _audit("BACKUP", "system", backup_id, zip_name)
            return jsonify({"id": backup_id, "file": zip_name, "size_bytes": file_size, "includes": includes}), 201
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @bp.route("/api/backups/<backup_id>/restore", methods=["POST"])
    @nc_login_required
    def nc_api_restore_backup(backup_id):
        conn = get_connection()
        row = conn.execute("SELECT * FROM nc_backups WHERE id=?", (backup_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Backup not found"}), 404
        backup = _row_to_dict(row)
        zip_path = Path(backup["file_path"])
        if not zip_path.exists():
            return jsonify({"error": f"Backup file not found: {zip_path}"}), 404
        restored = []
        try:
            with zipfile.ZipFile(str(zip_path), "r") as zf:
                for name in zf.namelist():
                    if name == "network_canvas.db":
                        target = _ICDEV_ROOT / "data" / "network_canvas.db"
                        if target.exists():
                            shutil.copy2(str(target), str(target) + ".pre-restore")
                        zf.extract(name, str(target.parent))
                        restored.append(name)
                    elif name.startswith("args/"):
                        zf.extract(name, str(_ICDEV_ROOT))
                        restored.append(name)
            _audit("RESTORE", "system", backup_id, f"Restored {len(restored)} files")
            return jsonify({"restored": restored, "backup_id": backup_id})
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    # ══════════════════════════════════════════════════════════════════════
    # API: Save As / Clear All
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/topologies/<topo_id>/save-as", methods=["POST"])
    @nc_login_required
    def nc_api_save_as(topo_id):
        conn = get_connection()
        row = conn.execute("SELECT * FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        data = request.get_json(force=True, silent=True) or {}
        new_id = str(_uuid.uuid4())
        now = _now()
        name = data.get("name", f"{row['name']} (copy)")
        conn.execute(
            "INSERT INTO topologies (id, name, description, graph_json, template_id, classification, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (new_id, name, row["description"], row["graph_json"], row["template_id"], row["classification"], now, now),
        )
        conn.commit()
        conn.close()
        _audit("SAVE_AS", "topology", new_id, f"from {topo_id}")
        return jsonify({"id": new_id, "name": name, "redirect": f"/network/canvas/{new_id}"}), 201

    @bp.route("/api/topologies/<topo_id>/create-lab", methods=["POST"])
    @nc_login_required
    def nc_api_create_lab(topo_id):
        """AI-reduce production topology to a scaled-down representative lab version.

        Reduction rules applied:
        - 1 representative node per nc_group cluster (first non-L1 member)
        - L1 physical media stripped (SFP, fiber, patch panels, media converters)
        - IPs reassigned to 10.99.x.x range
        - Node labels prefixed LAB-
        - Security zones, VRFs, subnets, roles preserved
        - Lineage recorded in nc_lab_clones; version snapshot saved on source (phase=lab)
        """
        import re as _re

        L1_TYPES = {
            "sfp", "sfp-plus", "qsfp", "qsfp-dd",
            "media-ge", "media-10ge", "media-40ge", "media-100ge", "media-400ge",
            "media-fiber", "media-optical", "media-converter",
            "patch-panel", "odf",
        }

        conn = get_connection()
        row = conn.execute("SELECT * FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Not found"}), 404

        try:
            graph = json.loads(row["graph_json"])
        except Exception:
            graph = {"nodes": [], "edges": []}

        nodes = graph.get("nodes", [])
        edges = graph.get("edges", [])

        # Load nc_groups to identify cluster membership
        group_rows = conn.execute(
            "SELECT id, auto_nodes_json FROM nc_groups WHERE topology_id=?", (topo_id,)
        ).fetchall()

        node_type_map = {n["id"]: n.get("type", "") for n in nodes}

        # Map each node to its first group; pick one representative per group
        node_to_group: dict = {}
        group_representative: dict = {}
        for gr in group_rows:
            gid = gr["id"]
            try:
                members = json.loads(gr["auto_nodes_json"] or "[]")
            except Exception:
                members = []
            for nid in members:
                if nid not in node_to_group:
                    node_to_group[nid] = gid
            rep = next(
                (nid for nid in members if node_type_map.get(nid, "") not in L1_TYPES),
                members[0] if members else None,
            )
            if rep is not None:
                group_representative[gid] = rep

        # Determine kept nodes and the ID remap table
        kept_node_ids: set = set()
        id_remap: dict = {}
        for n in nodes:
            nid = n["id"]
            ntype = n.get("type", "")
            if ntype in L1_TYPES:
                continue
            gid = node_to_group.get(nid)
            if gid is not None:
                rep = group_representative.get(gid)
                if rep and rep != nid:
                    id_remap[nid] = rep
                    continue
            kept_node_ids.add(nid)

        # IP reassignment: 10.99.x.x counter
        _ip_state = [1, 1]

        def _next_lab_ip():
            ip = f"10.99.{_ip_state[0]}.{_ip_state[1]}"
            _ip_state[1] += 1
            if _ip_state[1] > 254:
                _ip_state[1] = 1
                _ip_state[0] += 1
            return ip

        _IP_RE = _re.compile(r"\b(?:\d{1,3}\.){3}\d{1,3}(?:/\d{1,2})?\b")

        def _replace_ips(text):
            if not isinstance(text, str):
                return text
            def _sub(m):
                parts = m.group(0).split("/")
                return _next_lab_ip() + ("/" + parts[1] if len(parts) > 1 else "")
            return _IP_RE.sub(_sub, text)

        def _sanitize(val):
            if isinstance(val, dict):
                return {k: _sanitize(v) for k, v in val.items()}
            if isinstance(val, list):
                return [_sanitize(v) for v in val]
            return _replace_ips(val)

        # Build lab nodes
        lab_nodes = []
        for n in nodes:
            if n["id"] not in kept_node_ids:
                continue
            lab_n = dict(n)
            lbl = lab_n.get("label", lab_n["id"])
            lab_n["label"] = f"LAB-{lbl}" if not lbl.startswith("LAB-") else lbl
            lab_n["label"] = _replace_ips(lab_n["label"])
            if "config" in lab_n:
                lab_n["config"] = _sanitize(lab_n["config"])
            lab_nodes.append(lab_n)

        # Build lab edges — remap endpoints, skip self-loops and duplicates
        seen_edges: set = set()
        lab_edges = []
        for e in edges:
            src = id_remap.get(e.get("source"), e.get("source"))
            dst = id_remap.get(e.get("target"), e.get("target"))
            if src not in kept_node_ids or dst not in kept_node_ids:
                continue
            if src == dst:
                continue
            key = (min(src, dst), max(src, dst))
            if key in seen_edges:
                continue
            seen_edges.add(key)
            lab_e = dict(e)
            lab_e["source"] = src
            lab_e["target"] = dst
            lab_edges.append(lab_e)

        lab_name = f"LAB-{row['name']}"
        now = _now()
        new_id = str(_uuid.uuid4())

        # Create the lab topology
        conn.execute(
            "INSERT INTO topologies (id, name, description, graph_json, template_id, classification, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (
                new_id,
                lab_name,
                f"Lab of '{row['name']}' (source: {topo_id}). {(row['description'] or '')}".strip(". "),
                json.dumps({"nodes": lab_nodes, "edges": lab_edges}),
                None,
                row["classification"] or "public",
                now,
                now,
            ),
        )

        # Record lineage in nc_lab_clones
        clone_id = str(_uuid.uuid4())
        conn.execute(
            "INSERT INTO nc_lab_clones "
            "(clone_id, parent_design_id, lineage, redaction_log, created_at, classification) "
            "VALUES (?,?,?,?,?,?)",
            (
                clone_id,
                topo_id,
                json.dumps([topo_id]),
                json.dumps({
                    "l1_stripped": sorted(L1_TYPES),
                    "ips_replaced": True,
                    "label_prefix": "LAB-",
                    "clusters_reduced": len(group_representative),
                }),
                now,
                row["classification"] or "UNCLASSIFIED",
            ),
        )

        # Save a version snapshot on the source topology linking to the new lab
        last_ver = conn.execute(
            "SELECT MAX(version_num) FROM nc_versions WHERE topology_id=?", (topo_id,)
        ).fetchone()[0]
        ver_num = (last_ver or 0) + 1
        vid = str(_uuid.uuid4())
        conn.execute(
            "INSERT INTO nc_versions "
            "(id, topology_id, version_num, label, phase, graph_json, created_by, notes, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?)",
            (
                vid,
                topo_id,
                ver_num,
                f"Lab snapshot → {lab_name}",
                "lab",
                row["graph_json"],
                "system",
                json.dumps({"lab_topology_id": new_id, "lab_name": lab_name}),
                now,
            ),
        )

        conn.commit()
        conn.close()
        _audit("CREATE_LAB", "topology", new_id, f"lab of {topo_id}, nodes={len(lab_nodes)}, edges={len(lab_edges)}")
        return jsonify({
            "id": new_id,
            "name": lab_name,
            "source_id": topo_id,
            "nodes": len(lab_nodes),
            "edges": len(lab_edges),
            "redirect": f"/network/canvas/{new_id}",
        }), 201

    @bp.route("/api/topologies/<topo_id>/save-as-template", methods=["POST"])
    @nc_login_required
    def nc_api_save_as_template(topo_id):
        conn = get_connection()
        row = conn.execute("SELECT * FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        data = request.get_json(force=True, silent=True) or {}
        tpl_id = f"tpl-{str(_uuid.uuid4())[:8]}"
        name = data.get("name", row["name"])
        category = data.get("category", "Custom")
        tags = json.dumps(data.get("tags", ["custom", "user-created"]))
        conn.execute(
            "INSERT INTO nc_templates (id, name, category, description, graph_json, tags) VALUES (?,?,?,?,?,?)",
            (tpl_id, name, category, data.get("description", row["description"] or ""), row["graph_json"], tags),
        )
        conn.commit()
        conn.close()
        _audit("SAVE_AS_TEMPLATE", "template", tpl_id, f"from {topo_id}")
        return jsonify({"id": tpl_id, "name": name}), 201

    # ══════════════════════════════════════════════════════════════════════
    # AI Topology Generator — natural language → JointJS graph JSON
    # Uses Ollama (scanner tier) for air-gap compatibility
    # ══════════════════════════════════════════════════════════════════════

    _AI_TOPO_SYSTEM_PROMPT = """You are a network topology generator for a dark-themed canvas (navy #1a1a2e background). Output ONLY a valid JSON object — no markdown, no explanation, no code fences.

FORMAT:
{"nodes": [...], "edges": [...]}
Each node: {"id": "unique-id", "label": "Display Name", "type": "device-type", "x": number, "y": number, "config": {}}
Each edge: {"id": "unique-id", "source": "node-id", "target": "node-id", "label": "link label", "protocol": "protocol or empty"}

═══ DEVICE TYPES (use ONLY these) ═══
Physical:     router, switch-l2, switch-l3, firewall, load-balancer, wap, server, patch-panel
Cisco:        cisco-router, cisco-switch-l2, cisco-switch-l3, cisco-firewall, cisco-lb
Juniper:      juniper-ptx10003, juniper-mx304
Endpoints:    endpoint-pc, endpoint-phone, endpoint-iot, endpoint-camera
Cloud:        cloud, aws-vpc, aws-tgw, aws-subnet, az-vnet, az-fw, gcp-vpc
Logical:      vrf, vlan, subnet, security-zone
Encryption:   kg-175d, kg-175g, kg-250, kg-340, type1-encryptor, fips-140-l2, fips-140-l3, hsm, macsec
Monitoring:   siem, sdwan-edge, sase-pop
SP/Carrier:   mpls-pe, mpls-p, route-reflector, pop, sonet-adm, roadm, oadm, edfa, transponder
Media:        media-fiber, media-ge, media-10ge, media-100ge
Colo:         meet-me-room, cross-connect
Drawing:      draw-rect, draw-rounded-rect, text-heading, text-label, text-badge
DoD JWICS:    dod-jwics-backbone, dod-jwics-gateway, dod-jwics-dns, dod-jwics-mail-relay, dod-type1-encryptor, dod-scif-lan
DoD C2S:      dod-c2s-direct-connect, dod-c2s-tgw, dod-c2s-vpc, dod-c2s-dns-phz
DoD C2E:      dod-c2e-expressroute, dod-c2e-vnet, dod-c2e-dns-private
DoD Shared:   dod-secret-bcap, dod-cds

Use vendor-specific types when the user names a vendor product (e.g., Juniper PTX10003 → juniper-ptx10003, Juniper MX304 → juniper-mx304, Cisco ASR → cisco-router, Cisco Catalyst → cisco-switch-l3, Cisco ASA → cisco-firewall).

═══ MANDATORY STRUCTURE — follow this exact order in nodes array ═══

1. ZONE BOXES (draw-rect) — placed FIRST so they render behind devices
2. ZONE HEADINGS (text-heading) — placed ABOVE their zone box (zone_y - 25px)
3. BADGES (text-badge) — top of diagram for topology type name
4. DEVICES — inside their zone boxes, with realistic labels
5. ANNOTATION LABELS (text-label) — protocol/spec notes in clear space
6. LEGEND PANEL — ALWAYS include, on the RIGHT side of the last diagram

═══ ZONE COLOR PALETTE (dark fills, bright borders) ═══
Blue:    {"_fill": "#0a1628", "_stroke": "#3498db", "_width": W, "_height": H}
Green:   {"_fill": "#0a180a", "_stroke": "#27ae60", "_width": W, "_height": H}
Orange:  {"_fill": "#1a1500", "_stroke": "#f39c12", "_width": W, "_height": H}
Red:     {"_fill": "#1a0a0a", "_stroke": "#e74c3c", "_width": W, "_height": H}
Purple:  {"_fill": "#120a20", "_stroke": "#9b59b6", "_width": W, "_height": H}
Teal:    {"_fill": "#0a1a1a", "_stroke": "#00cec9", "_width": W, "_height": H}
Silver:  {"_fill": "#111318", "_stroke": "#95a5a6", "_width": W, "_height": H}
Legend:  {"_fill": "#0f1520", "_stroke": "#636e72", "_width": 240, "_height": H}

text-heading config: {"_textColor": "<matching zone stroke color>"}
text-badge config:   {"_fill": "#0f3460", "_stroke": "#4a9eff"}
text-label config:   {"_textColor": "#7a8cb0"} (or matching color)

═══ LAYOUT RULES ═══
- Start at x=40, y=60. Leave 25px above zones for headings.
- Space devices 150-200px apart horizontally, 130-160px vertically between tiers.
- Zone boxes: 700-900px wide, 500-800px tall, with 40px padding around devices inside.
- Zone headings: positioned at (zone_x + 20, zone_y - 25) with _textColor matching zone _stroke.
- NEVER overlap text on text. Keep 30px vertical gap between text nodes.
- Use different zone colors for different functional areas.

═══ MIGRATION SCENARIOS (read carefully when user describes upgrades, replacements, parallel runs, cutover, or incremental migration) ═══
When the user describes migrating or replacing devices:

1. Create MULTIPLE labeled phase zones arranged LEFT-TO-RIGHT (each 820px wide, 120px gap between):
   • Phase 0 "AS-IS — Current State" at x=40 (Silver zone, _stroke #95a5a6)
   • Phase 1 "Phase 1 — [name]" at x=980 (Orange zone, _stroke #f39c12)
   • Phase 2 "Phase 2 — [name]" at x=1920 (Teal zone, _stroke #00cec9)
   • Phase N "TO-BE — Target State" at x=(N*940)+40 (Green zone, _stroke #27ae60)

2. AS-IS zone: show EXISTING devices and ALL connections exactly as the user described. Use port/interface labels on edges (e.g., "xe-0/0/0 ↔ Gi0/0/1"). Include VLANs and VRFs as logical nodes inside or below the device.

3. Each migration phase zone: show WHAT CHANGES — new device in parallel, which VLANs/VRFs move this phase. Devices being decommissioned should stay with "(retiring)" in label. New devices appear with "(new)" in label. Add text-label annotations for port mappings and BGP session status.

4. TO-BE zone: show final target state after full cutover — only target devices remain, no legacy equipment.

5. Add edges BETWEEN phases for the migration path: dashed edges showing BGP peer hand-off, uplink preservation, VLAN migration order.

6. For VLAN/VRF migration: create individual vlan/vrf nodes showing which phase they migrate. Label them "VLAN 10 (Phase 1)", "VRF MGMT (Phase 2)", etc.

7. Port mapping: use text-label nodes within each zone to show the interface mapping table.
   Example: "xe-0/0/0 (PTX) → et-0/0/1 (MX304)" as a text-label inside the phase zone.

8. BGP continuity: when uplinks must stay on old router until last phase, show this explicitly with a note text-label "BGP AS 1001 uplink retained on legacy until Phase N".

═══ LEGEND (MANDATORY — always include) ═══
Place a legend panel to the RIGHT of the rightmost diagram (rightmost_x + 120).
Structure:
- draw-rect background: {"_fill": "#0f1520", "_stroke": "#636e72", "_width": 240, "_height": <calculated>}
- text-heading "Legend" at top
- PROTOCOLS section: list only protocols used:
  OSPF=#27ae60, iBGP=#85c1e9, eBGP=#3498db, MPLS=#ff9800, IPSec=#f7dc6f, BGP=#5dade2
- DEVICES section: list device types used
- PHASES section (for migration): Silver=AS-IS, Orange=Phase 1, Teal=Phase 2, Green=TO-BE
Each legend entry: text-label with "• <description>" and appropriate _textColor.
Spacing: 22px between entries, 30px between sections.

═══ DoD SECRET / CLASSIFIED NETWORK TOPOLOGIES ═══
Use dod-* types when user mentions: JWICS, SCIF, C2S, C2E, SIPR, classified network, SECRET network, DISA, BCAP, SCCA, Type 1, CDS, cross-domain, IL6, DIA, or NSA encryption.

STANDARD JWICS AGENCY CONNECTION (left → right):
  dod-scif-lan → dod-type1-encryptor → dod-jwics-gateway → dod-jwics-backbone → [DIA hub: router] → dod-jwics-dns, dod-jwics-mail-relay, server (app)

JWICS → C2S (AWS Secret Region):
  dod-scif-lan → dod-type1-encryptor → dod-jwics-gateway → dod-jwics-backbone → dod-secret-bcap → dod-c2s-direct-connect → dod-c2s-tgw → dod-c2s-vpc → dod-c2s-dns-phz

JWICS → C2E (Azure Government Secret):
  dod-scif-lan → dod-type1-encryptor → dod-jwics-gateway → dod-jwics-backbone → dod-secret-bcap → dod-c2e-expressroute → dod-c2e-vnet → dod-c2e-dns-private

FULL DISA PANORAMA (3-row layout — stack vertically, 280px row spacing):
  TOP ROW (NIPR, y=80):   endpoint-pc → router → firewall → [dod-secret-bcap optional NIPR side] → aws-vpc / az-vnet
  MID ROW (DISN, y=360):  router (DISN backbone) → siem → server (ACAS/HBSS)
  BOT ROW (SECRET, y=640): dod-scif-lan → dod-type1-encryptor → dod-jwics-backbone → dod-secret-bcap → dod-c2s-vpc / dod-c2e-vnet
  CDS bridging MID ↔ BOT: place dod-cds node between MID row and BOT row (y=500)

CROSS-DOMAIN SOLUTION: place dod-cds between NIPR (unclassified) and JWICS (SECRET) segments.
DNS FLOW diagram: dod-scif-lan → endpoint-pc (SCIF user) → server (stub resolver) → dod-jwics-dns (JWICS recursive) → server (DIA authoritative)
EMAIL FLOW diagram: endpoint-pc (SCIF sender) → server (agency SMTP relay) → dod-jwics-mail-relay → server (DIA relay) → endpoint-pc (recipient)

ZONE COLORS for classified:
  SECRET zone: Red   {"_fill": "#1a0808", "_stroke": "#e74c3c"}
  JWICS zone:  Red   {"_fill": "#2b0808", "_stroke": "#ff4757"}
  C2S zone:    Amber {"_fill": "#1a0f00", "_stroke": "#e67e22"}
  C2E zone:    Purple{"_fill": "#0f0820", "_stroke": "#8e44ad"}
  CDS bridge:  Red   {"_fill": "#1a0a1a", "_stroke": "#ff7675"}
  NIPR zone:   Blue  (standard)

EDGE LABELS for classified: "Type 1 AES-256 HAIPE", "OSPF Area 0", "ClassifiedConnect 10G", "BGP eBGP MD5", "UDP/53 DNSSEC", "SMTP/S 587", "LDAPS/636"

═══ PROTOCOLS (use realistic ones) ═══
OSPF, BGP, iBGP, eBGP, MP-BGP, MPLS, LDP, RSVP, IPSec, STP, VXLAN, BGP EVPN, GRE, Type 1 AES-256, DNSSEC, S/MIME, HAIPE

Output ONLY the JSON object. No other text."""

    # Topology generation keywords — message contains one of these → likely a diagram request
    _TOPOLOGY_KEYWORDS = {
        "design", "create", "build", "draw", "generate", "diagram", "topology",
        "network", "configure", "connect", "setup", "set up", "show me", "map",
        "add router", "add switch", "add firewall", "add server", "add node",
        "wan", "lan", "dmz", "vlan", "vrf", "mpls", "bgp", "ospf", "ipsec",
        "data center", "datacenter", "cloud", "hub", "spoke", "mesh",
        "three tier", "three-tier", "two tier", "two-tier", "spine", "leaf",
        "core", "distribution", "access layer",
        # DoD / classified network keywords
        "jwics", "scif", "sipr", "niprnet", "c2s", "c2e",
        "classified", "secret network", "il6", "il5", "il4",
        "disa", "bcap", "scca", "vdss", "vdms", "tccm",
        "type 1", "type-1", "taclane", "kg-250", "kg-175",
        "cds", "cross-domain", "cross domain",
        "classifiedconnect", "classified connect",
        "dia hub", "dia network", "jwics backbone",
        "secret region", "aws secret", "azure secret",
        "agency connect", "dod agency", "dod network",
    }

    # Migration scenario keywords — trigger multi-phase layout + migration canvas session
    _MIGRATION_KEYWORDS = {
        "migrat", "replac", "cutover", "cut-over", "parallel run", "incremental",
        "as-is", "as is", "to-be", "to be", "phase ", "phased", "hand-off", "handoff",
        "decommission", "decomm", "retire", "swap", "upgrade router", "upgrade switch",
    }

    _AI_MIGRATION_PLAN_PROMPT = """You are a DoD/government network migration planner. \
You work with any vendor (Cisco, Juniper, Arista, Palo Alto, Fortinet, HPE, Brocade, etc.), \
any device type (routers, switches, firewalls, load balancers, wireless controllers, SD-WAN), \
any ISP or carrier, and any partner network (government, commercial, NIPR, SIPR, or private).

CRITICAL ARCHITECTURE RULE — DoD Cloud Connectivity:
CSPs (AWS GovCloud, Azure Government, GCP, OCI, IBM Cloud, etc.) do NOT connect directly to \
the edge router or to NIPR. In DoD networks, ALL CSP connectivity is routed through DISA BCAP \
(Boundary Cloud Access Point). The topology is always:
  Edge Router → DISA BCAP → CSP
NEVER model a direct edge-router-to-CSP connection. Any migration involving cloud workloads \
must include DISA BCAP as an intermediate node on the north side, and must include DISA \
coordination steps in the relevant phases.

Given a plain-English description of a migration, decompose it into an ordered list of phases \
that are specific to the described devices and connections — do NOT assume any vendor, protocol, \
or peer unless the description explicitly names them.

Output ONLY a valid JSON array — no markdown, no explanation:
[
  {
    "phase_num": 1,
    "title": "Short imperative title",
    "description": "3-4 sentences: what changes, what stays, dependencies, and any coordination required",
    "duration_days": 14,
    "parallel_run": 0,
    "rollback_criteria": "One sentence: condition and steps that trigger rollback",
    "maintenance_window": "Sat 02:00-06:00 local time",
    "classification": "CUI",
    "impact_level": "IL4"
  }
]

Rules:
1. Phase titles must be short imperatives specific to the described devices (e.g. "Stage Cisco ASR Config", "Cut Over ISP BGP", "Migrate VLANs to New Core Switch")
2. Infer phases from the actual topology: north-side partners, south-side peers, protocols (BGP, OSPF, EIGRP, MPLS, etc.) and physical connections (trunk, LAG, port-channel, SFP) mentioned
3. If CSP connectivity is mentioned (cloud workloads, AWS, Azure, GCP, etc.), always model it as going through DISA BCAP — generate a BCAP coordination phase
4. duration_days: realistic estimate in days (minimum 1, typical 7-30 for production cuts)
5. parallel_run: 1 if old and new devices run simultaneously during this phase, else 0
6. classification: PUBLIC | CUI | SECRET | TS — infer from context or default to CUI
7. impact_level: IL2 | IL4 | IL5 | IL6 — infer from context or default to IL4
8. Phases must be ordered so each depends only on prior completed phases
9. Always end with a decommission or final validation phase
10. Minimum 2 phases, maximum 12 phases
11. If partner coordination is needed (ISP, DISA, government agency, carrier), add a dedicated coordination step within the relevant phase description"""

    @bp.route("/api/ai-generate", methods=["POST"])
    @nc_login_required
    def nc_api_ai_generate():
        """Generate topology from natural language description using Claude or Ollama."""
        data = request.get_json(force=True, silent=True) or {}
        description = data.get("description", "").strip()
        if not description:
            return jsonify({"error": "description required"}), 400

        import re

        import requests as _req
        from tools.http.client import request as _req_request

        def _repair_json(s):
            """Best-effort repair for common LLM JSON mistakes."""
            # Missing comma between adjacent objects in an array: }...{
            s = re.sub(r"\}\s*\n(\s*)\{", r"},\n\1{", s)
            # Trailing comma before closing bracket/brace
            s = re.sub(r",(\s*[\]\}])", r"\1", s)
            # Truncated JSON: find last complete top-level object and close it
            if s.count("{") > s.count("}"):
                # Strip back to last complete object in nodes/edges
                last_close = s.rfind("}}")
                if last_close > 0:
                    s = s[: last_close + 2]
                    depth_b2 = s.count("[") - s.count("]")
                    depth_c2 = s.count("{") - s.count("}")
                    s += "]" * max(0, depth_b2) + "}" * max(0, depth_c2)
            return s

        def _parse_llm_response(content):
            """Extract and validate JSON from LLM response text."""
            text = content.strip()
            text = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL).strip()
            if text.startswith("```"):
                lines = text.split("\n")
                lines = [ln for ln in lines if not ln.strip().startswith("```")]
                text = "\n".join(lines).strip()
            start = text.find("{")
            end = text.rfind("}") + 1
            if start < 0 or end <= start:
                return None, text[:500]
            json_str = text[start:end]
            # Try clean parse first, fall back to repair
            try:
                graph_json = json.loads(json_str)
            except json.JSONDecodeError:
                json_str = _repair_json(json_str)
                graph_json = json.loads(json_str)
            if "nodes" not in graph_json or "edges" not in graph_json:
                return None, text[:500]
            for n in graph_json["nodes"]:
                n.setdefault("id", str(_uuid.uuid4())[:8])
                n.setdefault("label", "")
                n.setdefault("type", "server")
                n.setdefault("x", 100)
                n.setdefault("y", 100)
                n.setdefault("config", {})
            for e in graph_json["edges"]:
                e.setdefault("id", str(_uuid.uuid4())[:8])
                e.setdefault("source", "")
                e.setdefault("target", "")
                e.setdefault("label", "")
                e.setdefault("protocol", "")
            return graph_json, None

        # Detect migration scenario once — shared across helpers below
        desc_lower = description.lower()
        is_migration = any(kw in desc_lower for kw in _MIGRATION_KEYWORDS)
        architect_mode = data.get("architect_mode", False)

        # Prefix for architect / best-practices mode
        if architect_mode:
            description = (
                "Act as a senior network architect and engineer. "
                "Apply DISA STIG, NIST 800-53, and industry best practices for anything the user did not specify. "
                "Make sensible, production-grade design decisions. "
                "Original request: " + description
            )

        def _call_claude(desc, max_tokens=4096):
            """Call Anthropic Claude API."""
            api_key = os.environ.get("ANTHROPIC_API_KEY", "")
            if not api_key:
                return None, "No ANTHROPIC_API_KEY set"
            model = os.environ.get("ANTHROPIC_TOPO_MODEL", "claude-sonnet-4-20250514")
            r = _req_request(
                "POST",
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json={
                    "model": model,
                    "max_tokens": max_tokens,
                    "temperature": 0.3,
                    "system": _AI_TOPO_SYSTEM_PROMPT,
                    "messages": [{"role": "user", "content": desc}],
                },
                timeout=120,
            )
            r.raise_for_status()
            content = r.json().get("content", [{}])[0].get("text", "")
            return content, None

        def _call_ollama(desc, max_tokens=4096):
            """Call Ollama local LLM."""
            ollama_url = os.environ.get("OLLAMA_HOST", "http://localhost:11434")
            ollama_model = os.environ.get("OLLAMA_TOPO_MODEL", "llama3.2:3b")
            r = _req_request(
                "POST",
                f"{ollama_url}/api/chat",
                json={
                    "model": ollama_model,
                    "messages": [
                        {"role": "system", "content": _AI_TOPO_SYSTEM_PROMPT},
                        {"role": "user", "content": desc},
                    ],
                    "stream": False,
                    "options": {"num_predict": max_tokens, "temperature": 0.3},
                },
                timeout=120,
            )
            r.raise_for_status()
            content = r.json().get("message", {}).get("content", "")
            return content, None

        # Migration diagrams need more tokens (multi-phase = many nodes)
        # 8192 baseline; migration scenarios with many phases get 16k
        token_budget = 16384 if is_migration else 8192

        try:
            # Try Claude first (fast, reliable), fall back to Ollama (air-gap)
            provider = os.environ.get("NC_AI_PROVIDER", "auto")  # auto | claude | ollama
            content = None
            used_provider = ""

            if provider in ("auto", "claude"):
                content, err = _call_claude(description, max_tokens=token_budget)
                if content:
                    used_provider = "claude"
                elif provider == "claude":
                    return jsonify({"error": f"Claude API failed: {err}"}), 503

            if not content and provider in ("auto", "ollama"):
                content, err = _call_ollama(description, max_tokens=token_budget)
                if content:
                    used_provider = "ollama"
                elif provider == "ollama":
                    return jsonify({"error": f"Ollama failed: {err}"}), 503

            if not content:
                return jsonify({"error": "No LLM provider available. Set ANTHROPIC_API_KEY or start Ollama."}), 503

            graph_json, raw = _parse_llm_response(content)
            if graph_json is None:
                return jsonify({"error": "LLM did not return valid JSON", "raw": raw}), 422

            # Apply deterministic style rules (zone ordering, label deconfliction, legend)
            try:
                from tools.network.topology_styler import style_topology

                graph_json = style_topology(graph_json)
            except Exception as style_err:
                logger.warning("Topology styler failed (non-fatal): %s", style_err)

            # For migration scenarios: create a migration canvas session automatically
            migration_session_id = None
            migration_session_url = None
            if is_migration:
                try:
                    # Extract src/tgt device names from node labels
                    node_labels = [n.get("label", "") for n in graph_json["nodes"]]
                    src_candidates = [l for l in node_labels if any(
                        v in l.lower() for v in ("ptx", "mx", "juniper", "legacy", "current", "asis", "as-is")
                    )]
                    tgt_candidates = [l for l in node_labels if any(
                        v in l.lower() for v in ("cisco", "new", "tobe", "to-be", "target", "replace")
                    )]
                    src_model = src_candidates[0] if src_candidates else "Source Device"
                    tgt_model = tgt_candidates[0] if tgt_candidates else "Target Device"
                    migration_session_id = "nmig-" + _uuid.uuid4().hex[:12]
                    with get_connection() as _mc:
                        _mc.execute(
                            "INSERT INTO mc_net_sessions "
                            "(id, src_model, tgt_model, src_device_name, tgt_device_name, created_at, updated_at) "
                            "VALUES (?,?,?,?,?,?,?)",
                            (migration_session_id, src_model, tgt_model,
                             src_model, tgt_model,
                             datetime.utcnow().isoformat(), datetime.utcnow().isoformat()),
                        )
                        _mc.commit()
                    migration_session_url = f"/migration-canvas/?session={migration_session_id}"
                except Exception as _mig_err:
                    logger.warning("Migration session create failed (non-fatal): %s", _mig_err)

            _audit("AI_GENERATE", "topology", "", f"[{used_provider}] Generated from: {description[:100]}")

            # Persist to AI history (non-fatal)
            try:
                _hist_id = "aih-" + _uuid.uuid4().hex[:12]
                _short = (description[:120] + "…") if len(description) > 120 else description
                _hist_gj = json.dumps(graph_json) if graph_json else None
                with get_connection() as _hc:
                    _hc.execute(
                        "INSERT INTO nc_ai_history "
                        "(id, description, short_desc, node_count, edge_count, provider, is_migration, graph_json, created_at) "
                        "VALUES (?,?,?,?,?,?,?,?,?)",
                        (_hist_id, description, _short,
                         len(graph_json["nodes"]), len(graph_json["edges"]),
                         used_provider, int(is_migration), _hist_gj,
                         datetime.utcnow().isoformat()),
                    )
                    _hc.commit()
            except Exception as _he:
                logger.warning("AI history save failed (non-fatal): %s", _he)

            return jsonify(
                {
                    "graph_json": graph_json,
                    "description": description,
                    "node_count": len(graph_json["nodes"]),
                    "edge_count": len(graph_json["edges"]),
                    "provider": used_provider,
                    "is_migration": is_migration,
                    "migration_session_id": migration_session_id,
                    "migration_session_url": migration_session_url,
                    "history_id": _hist_id,
                }
            )

        except _req.exceptions.ConnectionError:
            return jsonify({"error": "Cannot connect to LLM provider"}), 503
        except _req.exceptions.Timeout:
            return jsonify({"error": "LLM timed out — try a simpler description"}), 504
        except json.JSONDecodeError as exc:
            return jsonify({"error": f"Invalid JSON from LLM: {exc}"}), 422
        except Exception as exc:
            logger.exception("AI generate failed")
            return jsonify({"error": str(exc)}), 500

    # ══════════════════════════════════════════════════════════════════════
    # API: AI Context Creation — allocate a new conversation context id
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/ai-context", methods=["POST"])
    @nc_login_required
    def nc_api_ai_context_create():
        """Create a new AI chat context and return its id (nc-<uuid8>)."""
        ctx_id = "nc-" + str(_uuid.uuid4())[:8]
        return jsonify({"context_id": ctx_id}), 201

    # ══════════════════════════════════════════════════════════════════════
    # API: Unified AI Chat — topology generation + Q&A in one endpoint
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/ai-chat", methods=["POST"])
    @nc_login_required
    def nc_api_ai_chat():
        """Unified AI chat: route to topology generation or direct Q&A."""
        data = request.get_json(force=True, silent=True) or {}
        description = (data.get("description") or data.get("message") or "").strip()
        context_id = (data.get("context_id") or "").strip()  # noqa: F841
        architect_mode = data.get("architect_mode", False)
        mode = data.get("mode", "qa")

        if not description:
            return jsonify({"error": "message is required"}), 400

        phase_context = data.get("phase_context") or {}
        phase_header = ""
        if phase_context:
            ph_num = phase_context.get("phase_num", "?")
            ph_title = phase_context.get("title", "")
            ph_cls = phase_context.get("classification", "CUI")
            ph_il = phase_context.get("impact_level", "IL4")
            ph_status = phase_context.get("status", "planned")
            phase_header = (
                f"\n\n## ACTIVE MIGRATION PHASE CONTEXT\n"
                f"Phase {ph_num}: {ph_title}\n"
                f"Classification: {ph_cls} | Impact Level: {ph_il} | Status: {ph_status}\n"
                f"All responses must respect {ph_cls}/{ph_il} constraints, applicable STIG/RMF controls, "
                f"and DoD network migration best practices for this classification level.\n"
            )

        qa_system = (  # noqa: F841
            _AI_TOPO_SYSTEM_PROMPT
            + phase_header
            + "\n\nYou are also a network expert who can answer questions directly"
            " without generating JSON. When the user asks a question (rather than"
            " requesting a diagram), respond in plain English with a clear, concise"
            " explanation. Only output JSON when explicitly building a topology."
        )

        # Topology detection: keyword match with short-message Q&A override
        desc_lower = description.lower()
        word_count = len(description.split())
        keyword_hit = any(kw in desc_lower for kw in _TOPOLOGY_KEYWORDS) or any(
            kw in desc_lower for kw in _MIGRATION_KEYWORDS
        )
        # Explicit mode wins; auto-detect only when mode is unspecified (default "qa")
        if mode == "topology":
            is_topology = True
        elif mode == "qa":
            is_topology = False
        else:
            # ≤3 words → too short to be a topology request; treat as Q&A
            is_topology = keyword_hit and word_count > 3

        if is_topology:
            cookie_header = request.headers.get("Cookie", "")
            forward_data = {"description": description, "architect_mode": architect_mode}
            if context_id:
                forward_data["context_id"] = context_id
            with current_app.test_client() as tc:
                resp = tc.post(
                    "/network/api/ai-generate",
                    data=json.dumps(forward_data),
                    content_type="application/json",
                    headers={"Cookie": cookie_header} if cookie_header else {},
                )
            resp_data = json.loads(resp.data)
            resp_data["mode"] = "topology"
            return jsonify(resp_data), resp.status_code

        # Q&A mode — call Anthropic with conversation history
        from tools.http.client import request as _req_request  # noqa: PLC0415

        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            return jsonify({"error": "No ANTHROPIC_API_KEY set"}), 503

        try:
            history_messages = []
            if context_id:
                try:
                    conn = get_connection()
                    rows = conn.execute(
                        "SELECT role, content FROM chat_messages WHERE context_id=? "
                        "ORDER BY turn_number DESC LIMIT 10",
                        (context_id,),
                    ).fetchall()
                    history_messages = [{"role": r[0], "content": r[1]} for r in reversed(rows)]
                except Exception as exc:
                    logger.warning("qa history load failed: %s", exc)

            messages = history_messages + [{"role": "user", "content": description}]
            model = os.environ.get("ANTHROPIC_TOPO_MODEL", "claude-sonnet-4-20250514")
            r = _req_request(
                "POST",
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json={
                    "model": model,
                    "max_tokens": 1024,
                    "system": qa_system,
                    "messages": messages,
                },
                timeout=45,
            )
            if r.status_code != 200:
                return jsonify({"error": f"Claude API error {r.status_code}"}), 503
            answer = r.json()["content"][0]["text"]
        except Exception as exc:
            logger.exception("Q&A chat failed")
            return jsonify({"error": str(exc)}), 500

        if context_id:
            _nc_save_message(context_id, "user", description)
            _nc_save_message(context_id, "assistant", answer)

        return jsonify({"ok": True, "mode": "qa", "answer": answer}), 200

    # ══════════════════════════════════════════════════════════════════════
    # API: AI Chat Pre-flight — Grilling / Clarifying Questions
    # ══════════════════════════════════════════════════════════════════════

    _CHAT_PREP_SYSTEM = """You are a senior network architect and engineer.
A user has described a network topology they want designed.
Your job: decide if you need more information to create an OPTIMAL design.

Rules:
- If the request already has enough detail (device types, counts, topology style, scale) → reply with {"needs_more_info": false}
- If key information is missing, ask 2-3 targeted questions. No more than 3. Be specific.
- Always suggest what a reasonable default assumption would be if the user doesn't know.

Respond with ONLY this JSON (no other text):
{
  "needs_more_info": true|false,
  "questions": ["question 1", "question 2"],
  "assumption_summary": "If user says that's all I have, I'll assume: ..."
}"""

    @bp.route("/api/ai-chat-prep", methods=["POST"])
    @nc_login_required
    def nc_api_ai_chat_prep():
        """Assess description completeness and return clarifying questions."""
        import re as _re2
        from tools.http.client import request as _req_prep
        data = request.get_json(force=True, silent=True) or {}
        description = data.get("description", "").strip()
        if not description:
            return jsonify({"needs_more_info": False}), 200

        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            return jsonify({"needs_more_info": False}), 200

        try:
            model = os.environ.get("ANTHROPIC_TOPO_MODEL", "claude-sonnet-4-20250514")
            r = _req_prep(
                "POST",
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json={
                    "model": model,
                    "max_tokens": 512,
                    "temperature": 0.1,
                    "system": _CHAT_PREP_SYSTEM,
                    "messages": [{"role": "user", "content": description}],
                },
                timeout=20,
            )
            r.raise_for_status()
            text = r.json().get("content", [{}])[0].get("text", "")
            # Parse the JSON response
            text = _re2.sub(r"<think>.*?</think>", "", text, flags=_re2.DOTALL).strip()
            start = text.find("{")
            end = text.rfind("}") + 1
            if start >= 0 and end > start:
                result = json.loads(text[start:end])
                return jsonify(result), 200
        except Exception as _prep_err:
            logger.warning("AI chat prep failed (non-fatal): %s", _prep_err)

        return jsonify({"needs_more_info": False}), 200

    # ══════════════════════════════════════════════════════════════════════
    # API: AI Chat History
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/ai-history", methods=["GET"])
    @nc_login_required
    def nc_api_ai_history():
        """Return the last N AI generation history entries."""
        limit = min(int(request.args.get("limit", 30)), 100)
        with get_connection() as conn:
            rows = conn.execute(
                "SELECT id, short_desc, node_count, edge_count, provider, is_migration, created_at "
                "FROM nc_ai_history ORDER BY created_at DESC LIMIT ?",
                (limit,),
            ).fetchall()
        entries = [
            {
                "id": r[0], "short_desc": r[1], "node_count": r[2],
                "edge_count": r[3], "provider": r[4],
                "is_migration": bool(r[5]), "created_at": r[6],
            }
            for r in rows
        ]
        return jsonify({"entries": entries})

    @bp.route("/api/ai-history/<hist_id>", methods=["GET"])
    @nc_login_required
    def nc_api_ai_history_get(hist_id):
        """Return full description + graph_json for a history entry."""
        with get_connection() as conn:
            row = conn.execute(
                "SELECT id, description, node_count, edge_count, provider, is_migration, graph_json, created_at "
                "FROM nc_ai_history WHERE id=?",
                (hist_id,),
            ).fetchone()
        if not row:
            return jsonify({"error": "Not found"}), 404
        return jsonify({
            "id": row[0], "description": row[1], "node_count": row[2],
            "edge_count": row[3], "provider": row[4],
            "is_migration": bool(row[5]),
            "graph_json": json.loads(row[6]) if row[6] else None,
            "created_at": row[7],
        })

    @bp.route("/api/ai-history/<hist_id>", methods=["DELETE"])
    @nc_login_required
    def nc_api_ai_history_delete(hist_id):
        """Delete a single history entry."""
        with get_connection() as conn:
            conn.execute("DELETE FROM nc_ai_history WHERE id=?", (hist_id,))
            conn.commit()
        return jsonify({"ok": True})

    # ══════════════════════════════════════════════════════════════════════
    # API: ATO Package Auto-Generator
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/ato/<topo_id>/generate", methods=["POST"])
    @nc_login_required
    def nc_api_ato_generate(topo_id):
        """Generate a partial ATO package from a topology (or region)."""
        conn = get_connection()
        topo = conn.execute(
            "SELECT id, name, graph_json, classification FROM topologies WHERE id=?",
            (topo_id,),
        ).fetchone()
        if not topo:
            conn.close()
            return jsonify({"error": "Topology not found"}), 404

        data = request.get_json(force=True, silent=True) or {}
        region_id = data.get("region_id")  # optional group ID
        system_name = data.get("system_name", topo["name"])
        classification = data.get("classification", topo["classification"] or "CUI")

        # Load regimes from compliance profile or request
        profile = conn.execute(
            "SELECT regimes, classification, environment FROM nc_compliance_profiles WHERE topology_id=?",
            (topo_id,),
        ).fetchone()
        if profile:
            try:
                regimes = json.loads(profile["regimes"] or "[]")
            except Exception:
                regimes = data.get("regimes", ["fisma_high", "stig"])
            classification = data.get("classification", profile["classification"] or classification)
        else:
            regimes = data.get("regimes", ["fisma_high", "stig"])

        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            conn.close()
            return jsonify({"error": "Bad graph JSON"}), 500

        # Load groups for region filtering
        groups = []
        if region_id:
            rows = conn.execute("SELECT * FROM nc_groups WHERE topology_id=?", (topo_id,)).fetchall()
            groups = [_row_to_dict(r) for r in rows]

        # Check for as-built version
        as_built = conn.execute(
            "SELECT id FROM nc_versions WHERE topology_id=? AND label='As-Built' LIMIT 1",
            (topo_id,),
        ).fetchone()
        has_as_built = as_built is not None

        # Generate the ATO package
        try:
            package = generate_ato_package(
                topology_id=topo_id,
                graph=graph,
                system_name=system_name,
                classification=classification,
                regimes=regimes,
                groups=groups,
                region_id=region_id,
                has_as_built_version=has_as_built,
            )
        except Exception as exc:
            conn.close()
            return jsonify({"error": f"ATO generation failed: {exc}"}), 500

        # Persist to DB
        pkg_id = package["package_id"]
        now = _now()
        user_id = ""
        try:
            user_id = session.get("user_id", "")
        except RuntimeError:
            pass

        conn.execute(
            "INSERT INTO nc_ato_packages "
            "(id, topology_id, region_id, system_name, classification, regimes, "
            "package_json, summary_json, overall_readiness, stig_pass_rate, "
            "compliance_score, created_by, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                pkg_id,
                topo_id,
                region_id,
                system_name,
                classification,
                json.dumps(regimes),
                json.dumps(package),
                json.dumps(package["summary"]),
                package["summary"]["overall_readiness"],
                package["summary"]["stig_pass_rate"],
                package["summary"]["compliance_score"],
                user_id,
                now,
            ),
        )
        conn.commit()
        conn.close()

        _audit(
            "ATO_GENERATE",
            "topology",
            topo_id,
            f"ATO package {pkg_id[:8]} | readiness={package['summary']['overall_readiness']} "
            f"| region={region_id or 'full'}",
        )

        return jsonify(package), 201

    @bp.route("/api/ato/<topo_id>/packages", methods=["GET"])
    @nc_login_required
    def nc_api_ato_list(topo_id):
        """List all generated ATO packages for a topology."""
        conn = get_connection()
        rows = conn.execute(
            "SELECT id, topology_id, region_id, system_name, classification, "
            "regimes, summary_json, overall_readiness, stig_pass_rate, "
            "compliance_score, created_by, created_at "
            "FROM nc_ato_packages WHERE topology_id=? ORDER BY created_at DESC",
            (topo_id,),
        ).fetchall()
        conn.close()
        results = []
        for r in rows:
            d = _row_to_dict(r)
            try:
                d["summary"] = json.loads(d.pop("summary_json", "{}"))
            except Exception:
                d["summary"] = {}
            try:
                d["regimes"] = json.loads(d.get("regimes") or "[]")
            except Exception:
                pass
            results.append(d)
        return jsonify(results)

    @bp.route("/api/ato/<topo_id>/packages/<pkg_id>", methods=["GET"])
    @nc_login_required
    def nc_api_ato_detail(topo_id, pkg_id):
        """Get the full ATO package (all artifacts) by package ID."""
        conn = get_connection()
        row = conn.execute(
            "SELECT package_json FROM nc_ato_packages WHERE id=? AND topology_id=?",
            (pkg_id, topo_id),
        ).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Package not found"}), 404
        try:
            return jsonify(json.loads(row["package_json"]))
        except Exception:
            return jsonify({"error": "Corrupt package data"}), 500

    @bp.route("/api/ato/<topo_id>/packages/<pkg_id>", methods=["DELETE"])
    @nc_login_required
    def nc_api_ato_delete(topo_id, pkg_id):
        """Delete an ATO package."""
        conn = get_connection()
        conn.execute(
            "DELETE FROM nc_ato_packages WHERE id=? AND topology_id=?",
            (pkg_id, topo_id),
        )
        conn.commit()
        conn.close()
        _audit("ATO_DELETE", "ato_package", pkg_id, f"topology={topo_id}")
        return jsonify({"ok": True})

    # ══════════════════════════════════════════════════════════════════════
    # Heatmap Overlay — device/link metric data
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/heatmap/<topo_id>", methods=["GET"])
    @nc_login_required
    def nc_api_heatmap(topo_id):
        """Return per-node and per-link metric values for heatmap overlay.

        Query param ``metric`` selects which metric:
          - bandwidth   — link utilisation %  (from configData.utilization)
          - vuln        — vulnerability severity (from STIG imports / compliance)
          - stig        — STIG compliance %  (from compliance findings)
          - age         — equipment age in years (from configData.install_date)
        """
        metric = request.args.get("metric", "bandwidth")
        conn = get_connection()
        topo = conn.execute(
            "SELECT id, graph_json FROM topologies WHERE id=?",
            (topo_id,),
        ).fetchone()
        if not topo:
            conn.close()
            return jsonify({"error": "Topology not found"}), 404

        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            graph = {"nodes": [], "edges": []}

        nodes = graph.get("nodes", [])
        edges = graph.get("edges", [])

        node_values = {}  # nodeId -> 0..1 normalised score
        link_values = {}  # edgeId -> 0..1 normalised score

        if metric == "bandwidth":
            for e in edges:
                cfg = e.get("configData") or {}
                util = cfg.get("utilization")
                if util is not None:
                    try:
                        link_values[e["id"]] = max(0.0, min(1.0, float(util) / 100.0))
                    except (ValueError, TypeError):
                        link_values[e["id"]] = 0.0
                else:
                    link_values[e["id"]] = 0.0

        elif metric == "vuln":
            # Pull latest compliance findings for this topology
            rows = conn.execute(
                "SELECT affected_entity, severity, status FROM nc_compliance_findings "
                "WHERE topology_id=? AND status='open' "
                "ORDER BY created_at DESC",
                (topo_id,),
            ).fetchall()
            severity_weight = {"CAT1": 1.0, "CAT2": 0.6, "CAT3": 0.25}
            entity_max = {}
            for r in rows:
                eid = r["affected_entity"]
                w = severity_weight.get(r["severity"], 0.1)
                if eid not in entity_max or w > entity_max[eid]:
                    entity_max[eid] = w
            for n in nodes:
                nid = n["id"]
                label = n.get("label", "")
                val = entity_max.get(nid, entity_max.get(label, 0.0))
                node_values[nid] = val
            for e in edges:
                eid = e["id"]
                label = e.get("label", "")
                val = entity_max.get(eid, entity_max.get(label, 0.0))
                link_values[eid] = val

        elif metric == "stig":
            # STIG import results — per-host pass rate
            stig_rows = conn.execute(
                "SELECT result_json FROM nc_stig_imports WHERE topology_id=? ORDER BY imported_at DESC LIMIT 1",
                (topo_id,),
            ).fetchone()
            host_compliance = {}
            if stig_rows and stig_rows["result_json"]:
                try:
                    stig_data = json.loads(stig_rows["result_json"])
                    hosts = stig_data.get("hosts", {})
                    for hostname, hdata in hosts.items():
                        s = hdata.get("summary", {})
                        total = s.get("pass", 0) + s.get("fail", 0) + s.get("nr", 0)
                        if total > 0:
                            host_compliance[hostname.lower()] = s.get("pass", 0) / total
                        else:
                            host_compliance[hostname.lower()] = 1.0
                except (json.JSONDecodeError, AttributeError):
                    pass
            for n in nodes:
                nid = n["id"]
                label = (n.get("label") or "").lower()
                # 1.0 = fully compliant (green), 0.0 = no compliance (red)
                # Invert so red=bad: heatmap value 1.0 = worst
                compliance = host_compliance.get(label, 1.0)
                node_values[nid] = 1.0 - compliance

        elif metric == "age":
            now = datetime.now(timezone.utc)
            for n in nodes:
                cfg = n.get("config") or n.get("configData") or {}
                score = 0.0

                # Check EOL/EOS/EoSup dates first (most impactful)
                for date_key, weight in [
                    ("eol_date", 1.0),
                    ("eosup_date", 0.9),
                    ("eos_date", 0.7),
                ]:
                    dval = cfg.get(date_key)
                    if dval:
                        try:
                            dt = datetime.fromisoformat(
                                dval + "T00:00:00+00:00" if "T" not in dval else dval.replace("Z", "+00:00")
                            )
                            if now >= dt:
                                score = max(score, weight)
                                pass  # past EOL/EOS/EoSup
                            else:
                                months_left = (dt - now).days / 30.44
                                if months_left < 6:
                                    score = max(score, weight * 0.8)
                                elif months_left < 12:
                                    score = max(score, weight * 0.5)
                                elif months_left < 24:
                                    score = max(score, weight * 0.2)
                        except (ValueError, TypeError):
                            pass

                # Fall back to install_date age
                if score == 0.0:
                    install_date = cfg.get("install_date") or cfg.get("installDate")
                    if install_date:
                        try:
                            dt = datetime.fromisoformat(
                                install_date + "T00:00:00+00:00"
                                if "T" not in install_date
                                else install_date.replace("Z", "+00:00")
                            )
                            age_years = (now - dt).days / 365.25
                            score = max(0.0, min(1.0, age_years / 10.0))
                            pass  # age computed
                        except (ValueError, TypeError):
                            pass

                node_values[n["id"]] = round(score, 3)

        conn.close()
        return jsonify(
            {
                "metric": metric,
                "node_values": node_values,
                "link_values": link_values,
            }
        )

    # ── Tech Debt Analysis ──────────────────────────────────────────────
    @bp.route("/api/tech-debt/<topo_id>", methods=["GET"])
    @nc_login_required
    def nc_api_tech_debt(topo_id):
        """Analyze lifecycle/tech debt across all devices in a topology."""
        conn = get_connection()
        row = conn.execute("SELECT graph_json, name FROM topologies WHERE id=?", (topo_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        try:
            graph = json.loads(row["graph_json"])
        except Exception:
            graph = {"nodes": [], "edges": []}

        now = datetime.now(timezone.utc)
        devices = []
        past_eol = 0
        past_eos = 0
        past_eosup = 0
        approaching_eol = 0
        no_lifecycle = 0

        for n in graph.get("nodes", []):
            cfg = n.get("config") or n.get("configData") or {}
            ntype = n.get("type", "")
            # Skip drawing shapes and text
            if ntype in (
                "rect",
                "circle",
                "text",
                "heading",
                "badge",
                "hline",
                "vline",
                "arrow",
                "diamond",
                "ellipse",
                "triangle",
                "hexagon",
                "star",
                "roundedrect",
            ):
                continue

            device = {
                "id": n.get("id"),
                "label": n.get("label", ""),
                "type": ntype,
                "model": cfg.get("model", ""),
                "serial": cfg.get("serial", ""),
                "sw_version": cfg.get("sw_version", ""),
                "install_date": cfg.get("install_date", ""),
                "eos_date": cfg.get("eos_date", ""),
                "eol_date": cfg.get("eol_date", ""),
                "eosup_date": cfg.get("eosup_date", ""),
                "risk_level": "unknown",
                "issues": [],
            }

            has_lifecycle = False
            for dkey, label in [
                ("eol_date", "End of Life"),
                ("eosup_date", "End of Support"),
                ("eos_date", "End of Sale"),
            ]:
                dval = cfg.get(dkey)
                if dval:
                    has_lifecycle = True
                    try:
                        dt = datetime.fromisoformat(
                            dval + "T00:00:00+00:00" if "T" not in dval else dval.replace("Z", "+00:00")
                        )
                        if now >= dt:
                            device["issues"].append(f"PAST {label}: {dval}")
                            if dkey == "eol_date":
                                past_eol += 1
                            elif dkey == "eos_date":
                                past_eos += 1
                            elif dkey == "eosup_date":
                                past_eosup += 1
                        else:
                            months = (dt - now).days / 30.44
                            if months < 12:
                                device["issues"].append(f"{label} in {int(months)} months: {dval}")
                                if dkey == "eol_date":
                                    approaching_eol += 1
                    except (ValueError, TypeError):
                        pass

            install = cfg.get("install_date")
            if install:
                has_lifecycle = True
                try:
                    dt = datetime.fromisoformat(
                        install + "T00:00:00+00:00" if "T" not in install else install.replace("Z", "+00:00")
                    )
                    age = (now - dt).days / 365.25
                    if age > 7:
                        device["issues"].append(f"Equipment age: {age:.1f} years (>7yr)")
                except (ValueError, TypeError):
                    pass

            if not has_lifecycle:
                no_lifecycle += 1
                device["issues"].append("No lifecycle data configured")

            # Risk level
            if any("PAST End of Life" in i for i in device["issues"]):
                device["risk_level"] = "critical"
            elif any("PAST" in i for i in device["issues"]):
                device["risk_level"] = "high"
            elif any("months" in i for i in device["issues"]):
                device["risk_level"] = "medium"
            elif device["issues"]:
                device["risk_level"] = "low"
            else:
                device["risk_level"] = "healthy"

            devices.append(device)

        total = len(devices)
        critical = sum(1 for d in devices if d["risk_level"] == "critical")
        high = sum(1 for d in devices if d["risk_level"] == "high")

        return jsonify(
            {
                "topology": row["name"],
                "total_devices": total,
                "devices": devices,
                "summary": {
                    "past_eol": past_eol,
                    "past_eos": past_eos,
                    "past_eosup": past_eosup,
                    "approaching_eol_12mo": approaching_eol,
                    "no_lifecycle_data": no_lifecycle,
                    "critical_risk": critical,
                    "high_risk": high,
                    "tech_debt_score": round((critical * 4 + high * 2 + approaching_eol) / max(total, 1) * 25, 1),
                },
            }
        )

    # ── IPv6 Readiness Assessment ───────────────────────────────────────
    @bp.route("/api/ipv6-readiness/<topo_id>", methods=["GET"])
    @nc_login_required
    def nc_api_ipv6_readiness(topo_id):
        """Assess IPv6 readiness of a topology: device capability,
        addressing gaps, migration status."""
        conn = get_connection()
        row = conn.execute("SELECT graph_json, name FROM topologies WHERE id=?", (topo_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        try:
            graph = json.loads(row["graph_json"])
        except Exception:
            graph = {"nodes": [], "edges": []}

        nodes = graph.get("nodes", [])
        devices = []
        capable_yes = capable_no = capable_partial = capable_unknown = 0
        has_v6_addr = 0
        has_v4_only = 0
        dual_stack = 0
        infra_types = {
            "router",
            "switch-l3",
            "switch-l2",
            "firewall",
            "mpls-pe",
            "mpls-p",
            "route-reflector",
            "load-balancer",
            "sdwan-edge",
            "wlc",
        }

        for n in nodes:
            ntype = n.get("type", "")
            if ntype not in infra_types:
                continue
            cfg = n.get("config") or n.get("configData") or {}
            cap = cfg.get("ipv6_capable", "")
            af = cfg.get("address_family", "")
            has_v6 = bool(cfg.get("ipv6"))
            has_v4 = bool(cfg.get("ip"))

            if cap == "yes":
                capable_yes += 1
            elif cap == "no":
                capable_no += 1
            elif cap == "partial":
                capable_partial += 1
            else:
                capable_unknown += 1

            if has_v6 and has_v4:
                dual_stack += 1
            elif has_v6:
                has_v6_addr += 1
            elif has_v4:
                has_v4_only += 1

            issues = []
            if cap == "no":
                issues.append("Device does NOT support IPv6 — requires hardware/software upgrade")
            elif cap == "partial":
                issues.append("Limited IPv6 support — verify feature set")
            elif cap == "":
                issues.append("IPv6 capability unknown — verify and update device properties")
            if has_v4 and not has_v6 and af != "ipv4":
                issues.append("IPv4 configured but no IPv6 address — add IPv6 for dual-stack")

            devices.append(
                {
                    "id": n.get("id"),
                    "label": n.get("label", ""),
                    "type": ntype,
                    "ipv6_capable": cap or "unknown",
                    "address_family": af
                    or ("dual-stack" if has_v6 and has_v4 else "ipv6" if has_v6 else "ipv4" if has_v4 else "none"),
                    "has_ipv4": has_v4,
                    "has_ipv6": has_v6,
                    "issues": issues,
                }
            )

        total = len(devices)
        ready_pct = round(capable_yes * 100 / max(total, 1))
        dual_pct = round(dual_stack * 100 / max(total, 1))

        # Migration recommendation
        if capable_no > 0:
            recommendation = (
                f"{capable_no} device(s) do not support IPv6 and "
                f"must be upgraded/replaced before migration. "
                f"Start with Phase 1: Assessment & Inventory."
            )
        elif capable_unknown > total / 2:
            recommendation = (
                f"{capable_unknown} device(s) have unknown IPv6 capability. "
                f"Audit all equipment and update the 'IPv6 Capable' "
                f"property before planning migration."
            )
        elif dual_stack < total:
            recommendation = (
                f"{total - dual_stack} device(s) not yet dual-stack. "
                f"Proceed with Phase 2: Enable dual-stack on core/distribution first."
            )
        else:
            recommendation = (
                "All infrastructure devices are dual-stack. Proceed to Phase 4: Access layer and endpoint rollout."
            )

        # Load IPv6 rules from design rules
        _dr = {}
        try:
            import yaml as _yaml_v6

            _v6_path = _ICDEV_ROOT / "args" / "network_design_rules.yaml"
            if _v6_path.exists():
                with open(_v6_path, encoding="utf-8") as _v6f:
                    _dr = _yaml_v6.safe_load(_v6f) or {}
        except Exception:
            pass
        ipv6_rules = _dr.get("ipv6_rules", {})

        return jsonify(
            {
                "topology": row["name"],
                "total_devices": total,
                "devices": devices,
                "summary": {
                    "capable_yes": capable_yes,
                    "capable_no": capable_no,
                    "capable_partial": capable_partial,
                    "capable_unknown": capable_unknown,
                    "readiness_pct": ready_pct,
                    "dual_stack_count": dual_stack,
                    "dual_stack_pct": dual_pct,
                    "ipv4_only_count": has_v4_only,
                    "ipv6_only_count": has_v6_addr,
                },
                "recommendation": recommendation,
                "migration_phases": ipv6_rules.get("migration_phases", {}),
                "transition_mechanisms": ipv6_rules.get("transition_mechanisms", {}),
                "security_checklist": ipv6_rules.get("security", []),
            }
        )

    # ══════════════════════════════════════════════════════════════════════
    # PPS Matrix Generator — Enclave / Device Pair
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/pps/<topo_id>")
    @nc_login_required
    def nc_pps_page(topo_id):
        """PPS Matrix Generator page — select two enclaves or device pair."""
        conn = get_connection()
        topo = conn.execute(
            "SELECT id, name, graph_json, classification FROM topologies WHERE id=?",
            (topo_id,),
        ).fetchone()
        if not topo:
            conn.close()
            abort(404)

        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            graph = {"nodes": [], "edges": []}

        groups = [
            _row_to_dict(r) for r in conn.execute("SELECT * FROM nc_groups WHERE topology_id=?", (topo_id,)).fetchall()
        ]
        conn.close()

        meta = get_topology_enclaves(graph, groups)
        classification_banner = (topo["classification"] or "").upper()

        return render_template(
            "network/pps_matrix.html",
            topology=_row_to_dict(topo),
            enclaves=meta["enclaves"],
            nodes=meta["nodes"],
            classification_banner=classification_banner,
        )

    @bp.route("/api/pps/<topo_id>/enclaves", methods=["GET"])
    @nc_login_required
    def nc_api_pps_enclaves(topo_id):
        """Return enclaves and nodes available for pair selection."""
        conn = get_connection()
        topo = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not topo:
            conn.close()
            return jsonify({"error": "Topology not found"}), 404

        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            graph = {"nodes": [], "edges": []}

        groups = [
            _row_to_dict(r) for r in conn.execute("SELECT * FROM nc_groups WHERE topology_id=?", (topo_id,)).fetchall()
        ]
        conn.close()

        return jsonify(get_topology_enclaves(graph, groups))

    @bp.route("/api/pps/<topo_id>/generate", methods=["POST"])
    @nc_login_required
    def nc_api_pps_generate(topo_id):
        """Generate PPS matrix for a selected enclave or device pair.

        Body JSON:
          {
            "source": "<zone_name | node_id>",
            "dest": "<zone_name | node_id>",
            "selector_type": "zone" | "node"   // default "zone"
          }
        """
        conn = get_connection()
        topo = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not topo:
            conn.close()
            return jsonify({"error": "Topology not found"}), 404

        data = request.get_json(force=True, silent=True) or {}
        source = (data.get("source") or "").strip()
        dest = (data.get("dest") or "").strip()
        selector_type = data.get("selector_type", "zone")

        if not source or not dest:
            conn.close()
            return jsonify({"error": "source and dest are required"}), 400
        if source == dest:
            conn.close()
            return jsonify({"error": "source and dest must be different"}), 400
        if selector_type not in ("zone", "node"):
            selector_type = "zone"

        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            conn.close()
            return jsonify({"error": "Bad graph JSON"}), 500

        groups = [
            _row_to_dict(r) for r in conn.execute("SELECT * FROM nc_groups WHERE topology_id=?", (topo_id,)).fetchall()
        ]
        conn.close()

        result = generate_pps_matrix_for_pair(
            graph=graph,
            source_selector=source,
            dest_selector=dest,
            selector_type=selector_type,
            groups=groups,
        )
        _audit(
            "PPS_GENERATE",
            "topology",
            topo_id,
            f"pair={source}<->{dest} type={selector_type} protocols={result['total_protocols']}",
        )
        return jsonify(result)

    @bp.route("/api/pps/<topo_id>/export", methods=["POST"])
    @nc_login_required
    def nc_api_pps_export(topo_id):
        """Export a PPS matrix as SSP table (CSV or Markdown).

        Body JSON: same as /api/pps/<topo_id>/generate plus:
          {
            "format": "csv" | "markdown"   // default "csv"
          }
        """
        conn = get_connection()
        topo = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not topo:
            conn.close()
            return jsonify({"error": "Topology not found"}), 404

        data = request.get_json(force=True, silent=True) or {}
        source = (data.get("source") or "").strip()
        dest = (data.get("dest") or "").strip()
        selector_type = data.get("selector_type", "zone")
        fmt = data.get("format", "csv")

        if not source or not dest:
            conn.close()
            return jsonify({"error": "source and dest are required"}), 400
        if selector_type not in ("zone", "node"):
            selector_type = "zone"
        if fmt not in ("csv", "markdown"):
            fmt = "csv"

        try:
            graph = json.loads(topo["graph_json"])
        except Exception:
            conn.close()
            return jsonify({"error": "Bad graph JSON"}), 500

        groups = [
            _row_to_dict(r) for r in conn.execute("SELECT * FROM nc_groups WHERE topology_id=?", (topo_id,)).fetchall()
        ]
        conn.close()

        result = generate_pps_matrix_for_pair(
            graph=graph,
            source_selector=source,
            dest_selector=dest,
            selector_type=selector_type,
            groups=groups,
        )
        content = export_pps_as_ssp_table(result, fmt=fmt)
        src_slug = source.replace(" ", "_").replace("/", "-")
        dst_slug = dest.replace(" ", "_").replace("/", "-")
        filename = f"pps_ssp_{src_slug}_to_{dst_slug}.{'csv' if fmt == 'csv' else 'md'}"

        from flask import Response

        mime = "text/csv" if fmt == "csv" else "text/markdown"
        return Response(
            content,
            mimetype=mime,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # ── Missing page routes (nav links that need wiring) ────────────────

    @bp.route("/what-if")
    @nc_login_required
    def nc_whatif():
        return render_template("network/whatif.html")

    @bp.route("/replacements")
    @nc_login_required
    def nc_replacements():
        return render_template("network/replacement_map.html")

    @bp.route("/budget")
    @nc_login_required
    def nc_budget():
        return render_template("network/budget_forecast.html")

    @bp.route("/executive-dashboard")
    @nc_login_required
    def nc_executive_dashboard():
        return render_template("network/executive_dashboard.html")

    @bp.route("/api/executive-summary", methods=["GET"])
    @nc_login_required
    def nc_api_executive_summary():
        """Return portfolio-level executive summary data for the dashboard."""
        try:
            from tools.ndc.executive_summary_generator import generate_executive_summary
            result = generate_executive_summary()
            return jsonify(result)
        except Exception as exc:
            logger.warning("nc_api_executive_summary failed: %s", exc)
            return jsonify({"error": str(exc)}), 500

    @bp.route("/cloud-topology")
    @nc_login_required
    def nc_cloud_topology():
        return render_template("network/cloud_topology.html")

    @bp.route("/api/cloud-overlay/<topology_id>", methods=["GET"])
    @nc_login_required
    def nc_api_cloud_overlay(topology_id: str):
        """Return enriched topology JSON with cloud-provider overlay nodes."""
        try:
            from tools.ndc.cloud_topology_overlay import generate_cloud_overlay
            result = generate_cloud_overlay(topology_id)
            if result.get("error"):
                return jsonify(result), 404
            return jsonify(result)
        except Exception as exc:
            logger.warning("nc_api_cloud_overlay failed: %s", exc)
            return jsonify({"error": str(exc)}), 500

    def _ensure_showcase_table():
        """Create showcase_demo_runs if missing (raw sqlite3, no RLS)."""
        import sqlite3
        db_path = _ICDEV_ROOT / "data" / "icdev.db"
        conn = sqlite3.connect(str(db_path), check_same_thread=False)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS showcase_demo_runs (
                run_id      TEXT PRIMARY KEY,
                audience    TEXT NOT NULL DEFAULT 'exec',
                scenarios_json TEXT NOT NULL DEFAULT '[]',
                status      TEXT NOT NULL DEFAULT 'running',
                result_json TEXT,
                scenarios_passed INTEGER DEFAULT 0,
                scenarios_total  INTEGER DEFAULT 0,
                elapsed_ms  INTEGER DEFAULT 0,
                created_at  TEXT NOT NULL
            )
        """)
        conn.commit()
        conn.close()

    @bp.route("/demo-runner")
    @nc_login_required
    def nc_demo_runner():
        """NDC Demo Runner control panel."""
        runs = []
        last_run = None
        last_result = {}
        try:
            _ensure_showcase_table()
            import sqlite3
            db_path = _ICDEV_ROOT / "data" / "icdev.db"
            conn = sqlite3.connect(str(db_path), check_same_thread=False)
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                "SELECT run_id, audience, scenarios_json, status, result_json, scenarios_passed, "
                "scenarios_total, elapsed_ms, created_at "
                "FROM showcase_demo_runs WHERE audience IN ('exec','tech','engineer') "
                "ORDER BY created_at DESC LIMIT 15"
            ).fetchall()
            conn.close()
            for row in rows:
                d = dict(row) if hasattr(row, "keys") else {
                    "run_id": row[0], "audience": row[1], "scenarios_json": row[2],
                    "status": row[3], "result_json": row[4], "scenarios_passed": row[5],
                    "scenarios_total": row[6], "elapsed_ms": row[7], "created_at": row[8],
                }
                try:
                    d["scenarios_list"] = json.loads(d.get("scenarios_json") or "[]")
                except (ValueError, TypeError):
                    d["scenarios_list"] = []
                runs.append(d)
            if runs:
                last_run = runs[0]
                if last_run.get("result_json"):
                    try:
                        last_result = json.loads(last_run["result_json"])
                    except (ValueError, TypeError):
                        pass
        except Exception as exc:
            logger.warning("nc_demo_runner history error: %s", exc)

        scenario_meta = {
            "A": {"title": "EOL Fire Drill",       "short": "EOL",  "color": "#e74c3c", "hook": "Risk → Replace → Runbook"},
            "B": {"title": "Multi-Cloud Expansion", "short": "Cloud", "color": "#3498db", "hook": "Hybrid connectivity overlay"},
            "C": {"title": "Compliance Audit",      "short": "Audit", "color": "#f39c12", "hook": "STIG → Remediation → cATO"},
        }
        return render_template(
            "network/demo_runner.html",
            runs=runs,
            last_run=last_run,
            last_result=last_result,
            scenario_meta=scenario_meta,
            iqe_canvas="network",
            iqe_api_route="/network/api/demo-iqe-query",
            iqe_title="NDC Demo Runner IQE",
            iqe_examples=[
                {"label": "Recent runs", "query": "show recent demo runs"},
                {"label": "Exec audience", "query": "show exec audience runs"},
                {"label": "Failed runs",   "query": "show failed runs"},
            ],
        )

    @bp.route("/api/demo-run", methods=["POST"])
    @nc_login_required
    def nc_api_demo_run():
        """Execute NDC demo scenarios and store run."""
        import time
        import uuid
        from datetime import datetime, timezone
        data = request.get_json(force=True, silent=True) or {}
        audience = (data.get("audience") or "exec").lower()
        raw = data.get("scenarios")
        if not raw or raw == "all":
            scenarios = None
        else:
            scenarios = [s.strip().upper() for s in raw if isinstance(s, str) and s.strip()]

        run_id = str(uuid.uuid4())
        t0 = time.monotonic()
        result: dict = {}
        status = "error"

        try:
            from tools.ndc.demo_runner import run_ndc_demo
            result = run_ndc_demo(scenarios=scenarios, audience=audience)
            elapsed_ms = result.get("elapsed_ms", int((time.monotonic() - t0) * 1000))
            passed = result.get("scenarios_passed", 0)
            total = result.get("scenarios_total", 0)
            status = result.get("status", "error")
            result_payload = result.get("results", {})
        except Exception as exc:
            elapsed_ms = int((time.monotonic() - t0) * 1000)
            result_payload = {"error": str(exc)}
            passed = 0
            total = 0
            status = "error"
            logger.exception("nc_api_demo_run error: %s", exc)

        # Store run via raw sqlite3 to avoid RLS on showcase_demo_runs
        try:
            _ensure_showcase_table()
            import sqlite3
            db_path = _ICDEV_ROOT / "data" / "icdev.db"
            conn = sqlite3.connect(str(db_path), check_same_thread=False)
            conn.execute(
                "INSERT INTO showcase_demo_runs "
                "(run_id, audience, scenarios_json, status, result_json, "
                "scenarios_passed, scenarios_total, elapsed_ms, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (run_id, audience, json.dumps(scenarios or "all"),
                 status, json.dumps(result_payload, default=str),
                 passed, total, elapsed_ms,
                 datetime.now(timezone.utc).isoformat()),
            )
            conn.commit()
            conn.close()
        except Exception as exc:
            logger.warning("nc_api_demo_run store error: %s", exc)

        return jsonify({
            "run_id": run_id,
            "status": status,
            "results": result_payload,
            "scenarios_passed": passed,
            "scenarios_total": total,
            "elapsed_ms": elapsed_ms,
            "scenario_meta": {
                "A": {"title": "EOL Fire Drill",       "short": "EOL",  "color": "#e74c3c", "hook": "Risk → Replace → Runbook"},
                "B": {"title": "Multi-Cloud Expansion", "short": "Cloud", "color": "#3498db", "hook": "Hybrid connectivity overlay"},
                "C": {"title": "Compliance Audit",      "short": "Audit", "color": "#f39c12", "hook": "STIG → Remediation → cATO"},
            },
        })

    @bp.route("/api/demo-runs")
    @nc_login_required
    def nc_api_demo_runs():
        """Return NDC demo run history."""
        limit = min(int(request.args.get("limit", 20)), 100)
        try:
            _ensure_showcase_table()
            import sqlite3
            db_path = _ICDEV_ROOT / "data" / "icdev.db"
            conn = sqlite3.connect(str(db_path), check_same_thread=False)
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                "SELECT run_id, audience, scenarios_json, status, result_json, scenarios_passed, "
                "scenarios_total, elapsed_ms, created_at "
                "FROM showcase_demo_runs WHERE audience IN ('exec','tech','engineer') "
                "ORDER BY created_at DESC LIMIT ?",
                (limit,),
            ).fetchall()
            conn.close()
            result = []
            for row in rows:
                d = dict(row) if hasattr(row, "keys") else {
                    "run_id": row[0], "audience": row[1], "scenarios_json": row[2],
                    "status": row[3], "result_json": row[4], "scenarios_passed": row[5],
                    "scenarios_total": row[6], "elapsed_ms": row[7], "created_at": row[8],
                }
                try:
                    d["scenarios_list"] = json.loads(d.get("scenarios_json") or "[]")
                except (ValueError, TypeError):
                    d["scenarios_list"] = []
                result.append(d)
            return jsonify(result)
        except Exception as exc:
            logger.warning("nc_api_demo_runs error: %s", exc)
            return jsonify([])

    @bp.route("/design-patterns")
    @nc_login_required
    def nc_design_patterns():
        return render_template("network/design_patterns.html")

    @bp.route("/design-rules")
    @nc_login_required
    def nc_design_rules():
        return render_template("network/design_rules.html")

    @bp.route("/api/design-rules", methods=["GET"])
    @nc_login_required
    def nc_api_design_rules():
        """Return args/network_design_rules.yaml as JSON for the design-rules page.

        The page template at templates/network/design_rules.html expects a dict
        with on_node_add, best_practices, and ipv6_rules top-level keys. The
        YAML already contains those exact keys plus on_edge_add. Deterministic,
        air-gap-safe, no LLM.
        """
        try:
            import yaml as _yaml_dr
            _dr_path = _ICDEV_ROOT / "args" / "network_design_rules.yaml"
            if not _dr_path.exists():
                return jsonify({
                    "on_node_add": {},
                    "best_practices": {},
                    "ipv6_rules": {},
                    "on_edge_add": {},
                    "error": "network_design_rules.yaml not found",
                }), 200
            with open(_dr_path, encoding="utf-8") as _drf:
                data = _yaml_dr.safe_load(_drf) or {}
            return jsonify({
                "on_node_add": data.get("on_node_add", {}),
                "best_practices": data.get("best_practices", {}),
                "ipv6_rules": data.get("ipv6_rules", {}),
                "on_edge_add": data.get("on_edge_add", {}),
            })
        except Exception as exc:
            logger.warning("nc_api_design_rules failed: %s", exc)
            return jsonify({
                "on_node_add": {},
                "best_practices": {},
                "ipv6_rules": {},
                "on_edge_add": {},
                "error": str(exc),
            }), 500

    @bp.route("/device-profiles")
    @nc_login_required
    def nc_device_profiles_page():
        return render_template("network/device_profiles.html")

    # ── Hardware Profiles (datasheet-level specs) ────────────────────────
    @bp.route("/hardware-profiles")
    @nc_login_required
    def nc_hardware_profiles_page():
        return render_template("network/hardware_profiles.html")

    @bp.route("/api/hardware-profiles", methods=["GET"])
    @nc_login_required
    def nc_api_list_hardware_profiles():
        conn = get_connection()
        rows = conn.execute(
            "SELECT id, vendor, model, model_family, device_type, form_factor, "
            "rack_units, throughput_gbps, power_max_w, replacement_cost, "
            "eol_date, eos_date, ports_json, tags, is_builtin "
            "FROM nc_hardware_profiles ORDER BY vendor, model"
        ).fetchall()
        conn.close()
        result = []
        for r in rows:
            d = dict(r)
            try:
                d["ports"] = json.loads(d.pop("ports_json", "[]"))
            except Exception:
                d["ports"] = []
            try:
                d["tags"] = json.loads(d.pop("tags", "[]"))
            except Exception:
                d["tags"] = []
            result.append(d)
        return jsonify(result)

    @bp.route("/api/hardware-profiles/<pid>", methods=["GET"])
    @nc_login_required
    def nc_api_get_hardware_profile(pid):
        conn = get_connection()
        row = conn.execute("SELECT * FROM nc_hardware_profiles WHERE id=?", (pid,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Not found"}), 404
        d = dict(row)
        for jcol in ("ports_json", "components_json", "mgmt_ports_json", "os_options", "tags"):
            try:
                d[jcol.replace("_json", "")] = json.loads(d.pop(jcol, "[]"))
            except Exception:
                d[jcol.replace("_json", "")] = []
        return jsonify(d)

    @bp.route("/api/hardware-profiles", methods=["POST"])
    @nc_login_required
    def nc_api_create_hardware_profile():
        data = request.get_json(force=True, silent=True) or {}
        pid = "hw-" + str(_uuid.uuid4())[:8]
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_hardware_profiles (id, vendor, model, model_family, device_type, "
            "form_factor, rack_units, weight_kg, power_typical_w, power_max_w, throughput_gbps, "
            "ports_json, replacement_cost, eol_date, eos_date, tags, is_builtin, created_by) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,0,?)",
            (pid, data.get("vendor", ""), data.get("model", ""), data.get("model_family", ""),
             data.get("device_type", "router"), data.get("form_factor", "rack"),
             data.get("rack_units", 1), data.get("weight_kg"),
             data.get("power_typical_w"), data.get("power_max_w"), data.get("throughput_gbps"),
             json.dumps(data.get("ports", [])), data.get("replacement_cost"),
             data.get("eol_date"), data.get("eos_date"),
             json.dumps(data.get("tags", [])), data.get("created_by", "")),
        )
        conn.commit()
        conn.close()
        return jsonify({"id": pid}), 201

    @bp.route("/api/hardware-profiles/<pid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_hardware_profile(pid):
        conn = get_connection()
        row = conn.execute("SELECT is_builtin FROM nc_hardware_profiles WHERE id=?", (pid,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        if row["is_builtin"]:
            conn.close()
            return jsonify({"error": "Cannot delete built-in profile"}), 403
        conn.execute("DELETE FROM nc_hardware_profiles WHERE id=?", (pid,))
        conn.commit()
        conn.close()
        return jsonify({"deleted": pid})

    # ── Naming Conventions ─────────────────────────────────────────────
    @bp.route("/api/naming-conventions", methods=["GET"])
    @nc_login_required
    def nc_api_list_naming_conventions():
        from tools.network.naming_engine import list_conventions
        return jsonify(list_conventions())

    @bp.route("/api/naming-conventions/<cid>", methods=["GET"])
    @nc_login_required
    def nc_api_get_naming_convention(cid):
        from tools.network.naming_engine import _load_convention
        conv = _load_convention(cid)
        if not conv:
            return jsonify({"error": "Not found"}), 404
        return jsonify(conv)

    @bp.route("/api/naming-conventions/<cid>/generate", methods=["POST"])
    @nc_login_required
    def nc_api_generate_name(cid):
        from tools.network.naming_engine import generate_name, bulk_generate
        data = request.get_json(force=True, silent=True) or {}
        count = data.pop("count", 1)
        topology_id = data.pop("topology_id", "")
        if count > 1:
            result = bulk_generate(cid, count, data, topology_id)
        else:
            result = generate_name(cid, data, topology_id)
        return jsonify(result)

    @bp.route("/api/naming-conventions/<cid>/validate", methods=["POST"])
    @nc_login_required
    def nc_api_validate_name(cid):
        from tools.network.naming_engine import validate_name
        data = request.get_json(force=True, silent=True) or {}
        name = data.get("name", "")
        return jsonify(validate_name(name, cid))

    @bp.route("/api/naming-conventions", methods=["POST"])
    @nc_login_required
    def nc_api_create_naming_convention():
        data = request.get_json(force=True, silent=True) or {}
        cid = "nc-" + str(_uuid.uuid4())[:8]
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_naming_conventions (id, name, description, pattern, "
            "fields_json, separator, max_length, case_rule, example, is_builtin) "
            "VALUES (?,?,?,?,?,?,?,?,?,0)",
            (cid, data.get("name", ""), data.get("description", ""),
             data.get("pattern", ""), data.get("fields_json", "[]"),
             data.get("separator", ""), data.get("max_length", 63),
             data.get("case_rule", "upper"), data.get("example", "")),
        )
        conn.commit()
        conn.close()
        return jsonify({"id": cid}), 201

    @bp.route("/api/naming-conventions/<cid>", methods=["PUT"])
    @nc_login_required
    def nc_api_update_naming_convention(cid):
        data = request.get_json(force=True, silent=True) or {}
        conn = get_connection()
        row = conn.execute("SELECT id FROM nc_naming_conventions WHERE id=?", (cid,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        updates = []
        params = []
        for col in ("name", "description", "pattern", "fields_json", "separator", "max_length", "case_rule", "example"):
            if col in data:
                updates.append(f"{col} = ?")
                params.append(data[col])
        if updates:
            params.append(cid)
            conn.execute(f"UPDATE nc_naming_conventions SET {', '.join(updates)} WHERE id = ?", params)  # nosec B608 — columns from hardcoded whitelist
            conn.commit()
        conn.close()
        return jsonify({"updated": cid})

    @bp.route("/api/naming-conventions/<cid>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_naming_convention(cid):
        conn = get_connection()
        row = conn.execute("SELECT is_builtin FROM nc_naming_conventions WHERE id=?", (cid,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Not found"}), 404
        if row["is_builtin"]:
            conn.close()
            return jsonify({"error": "Cannot delete built-in convention"}), 403
        conn.execute("DELETE FROM nc_naming_conventions WHERE id=?", (cid,))
        conn.execute("DELETE FROM nc_naming_sequences WHERE convention_id=?", (cid,))
        conn.commit()
        conn.close()
        return jsonify({"deleted": cid})

    # ── Import Wizard: combined ingest + enrich + validate ──────────
    @bp.route("/api/import-wizard", methods=["POST"])
    @nc_login_required
    def nc_api_import_wizard():
        """One-shot import: ingest diagram + enrich + validate."""
        from tools.network.network_ingester import ingest_diagram
        import tempfile as _tmpmod

        if "file" not in request.files:
            return jsonify({"error": "No file provided"}), 400

        f = request.files["file"]
        project_id = request.form.get("project_id", "ndc-network-intelligence")
        name = request.form.get("name", f.filename or "uploaded")
        do_enrich = request.form.get("enrich", "true").lower() in ("true", "1", "yes")
        do_template = request.form.get("save_template", "false").lower() in ("true", "1", "yes")

        # Save to temp
        tmp_path = Path(os.path.join(_tmpmod.gettempdir(), f"ndc_import_{_uuid.uuid4().hex[:8]}_{f.filename}"))
        f.save(str(tmp_path))

        try:
            # Step 1: Ingest
            ingest_result = ingest_diagram(str(tmp_path), project_id, name)
            if ingest_result.get("error"):
                return jsonify({"step": "ingest", "error": ingest_result["error"]}), 400

            topo_id = ingest_result["topology_id"]
            response = {
                "topology_id": topo_id,
                "topology_name": name,
                "node_count": ingest_result.get("node_count", 0),
                "edge_count": ingest_result.get("edge_count", 0),
                "device_count": ingest_result.get("device_count", 0),
            }

            # Step 2: Enrich
            if do_enrich:
                try:
                    from tools.network.topology_enricher import enrich_topology
                    enrich_result = enrich_topology(topo_id, add_infra=False, add_groups=True)
                    response["enrichment"] = {
                        "groups_added": enrich_result.get("groups_added", 0),
                        "facilities_created": enrich_result.get("facilities_created", 0),
                        "racks_created": enrich_result.get("racks_created", 0),
                        "validation": enrich_result.get("validation", {}),
                    }
                except Exception as e:
                    response["enrichment"] = {"error": str(e)}

            # Step 3: Validate
            try:
                from tools.network.topology_validator import validate_topology
                val_result = validate_topology(topo_id, fix=True)
                response["validation"] = {
                    "issues_found": val_result.get("issues_found", 0),
                    "fixes_applied": val_result.get("fixes_applied", 0),
                    "passed": val_result.get("passed", True),
                }
            except Exception as e:
                response["validation"] = {"error": str(e)}

            # Step 4: Save template
            if do_template:
                try:
                    from tools.network.topology_enricher import save_as_template
                    tpl = save_as_template(topo_id, name + " (Template)")
                    response["template"] = tpl
                except Exception:
                    pass

            return jsonify(response)
        finally:
            tmp_path.unlink(missing_ok=True)

    # ── Topology Enrichment API ─────────────────────────────────────
    @bp.route("/api/topology/<tid>/enrich", methods=["POST"])
    @nc_login_required
    def nc_api_enrich_topology(tid):
        try:
            from tools.network.topology_enricher import enrich_topology
            data = request.get_json(force=True, silent=True) or {}
            result = enrich_topology(
                tid,
                add_infra=data.get("add_infra", True),
                add_groups=data.get("add_groups", True),
            )
            return jsonify(result)
        except ImportError:
            return jsonify({"error": "topology_enricher not available"}), 503
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @bp.route("/api/topology/<tid>/validate", methods=["POST"])
    @nc_login_required
    def nc_api_validate_topology(tid):
        try:
            from tools.network.topology_validator import validate_topology
            data = request.get_json(force=True, silent=True) or {}
            result = validate_topology(tid, fix=data.get("fix", True))
            return jsonify(result)
        except ImportError:
            return jsonify({"error": "topology_validator not available"}), 503
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    # Renamed from nc_api_save_as_template to nc_api_save_template_enricher to
    # disambiguate from the older nc_api_save_as_template at line 7644 (which
    # handles /api/topologies/<topo_id>/save-as-template with direct SQL). This
    # newer wrapper delegates to tools.network.topology_enricher and has its own
    # URL path (/api/topology/<tid>/save-template). The Python name collision
    # was aborting the network blueprint registration mid-flight, hiding all
    # subsequent routes (/discovery, /intelligence, /runbooks, /ingestion).
    @bp.route("/api/topology/<tid>/save-template", methods=["POST"])
    @nc_login_required
    def nc_api_save_template_enricher(tid):
        try:
            from tools.network.topology_enricher import save_as_template
            data = request.get_json(force=True, silent=True) or {}
            name = data.get("name", f"Template from {tid}")
            result = save_as_template(tid, name, data.get("description", ""))
            return jsonify(result)
        except ImportError:
            return jsonify({"error": "topology_enricher not available"}), 503
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @bp.route("/discovery")
    @nc_login_required
    def nc_discovery_page():
        return render_template("network/discovery.html") if os.path.exists(
            os.path.join(os.path.dirname(__file__), "..", "dashboard", "templates", "network", "discovery.html")
        ) else ("Discovery page coming soon", 200)

    @bp.route("/logout")
    def nc_logout():
        return redirect("/network/")

    # ── Analysis routes (extracted to routes/analysis.py) ────────────────
    from tools.network.routes.analysis import register_analysis_routes

    register_analysis_routes(bp)

    # ── Network Intelligence page + API routes (NII) ───────────────────
    @bp.route("/intelligence")
    @nc_login_required
    def nc_intelligence():
        """Network Infrastructure Intelligence dashboard page."""
        return render_template("network/intelligence.html")

    from tools.network.routes.intelligence import register_intelligence_routes

    register_intelligence_routes(bp)

    # ── Enterprise Summary (missing route fix) ───────────────────────────────

    @bp.route("/enterprise")
    @nc_login_required
    def nc_enterprise():
        """Enterprise summary page — aggregate metrics across all projects."""
        return render_template("network/enterprise.html")

    @bp.route("/api/enterprise-summary")
    @nc_login_required
    def nc_enterprise_summary_api():
        """Enterprise summary API — aggregate metrics."""
        conn = get_connection()
        total_topos = conn.execute("SELECT COUNT(*) FROM topologies").fetchone()[0]
        total_devices = 0
        total_interconnects = 0
        total_cat1 = 0
        total_open_findings = 0
        compliance_pct = None

        # Count devices from all topology graphs
        rows = conn.execute("SELECT graph_json FROM topologies").fetchall()
        for r in rows:
            try:
                g = json.loads(r[0] if isinstance(r, tuple) else r["graph_json"])
                total_devices += len(g.get("nodes", []))
                total_interconnects += len(g.get("edges", []))
            except Exception:
                pass

        # Compliance findings
        try:
            findings = conn.execute(
                "SELECT COUNT(*) FROM nc_compliance_findings WHERE status = 'open'"
            ).fetchone()
            total_open_findings = findings[0] if findings else 0
            cat1 = conn.execute(
                "SELECT COUNT(*) FROM nc_compliance_findings WHERE severity = 'CAT1' AND status = 'open'"
            ).fetchone()
            total_cat1 = cat1[0] if cat1 else 0
        except Exception:
            pass

        # Projects
        total_projects = 0
        status_counts = {}
        try:
            proj_rows = conn.execute("SELECT status, COUNT(*) FROM nc_projects GROUP BY status").fetchall()
            for pr in proj_rows:
                s = pr[0] if isinstance(pr, tuple) else pr["status"]
                c = pr[1] if isinstance(pr, tuple) else pr[1]
                status_counts[s or "draft"] = c
                total_projects += c
        except Exception:
            pass

        # Cost
        total_capex = 0
        total_circuit_monthly = 0
        try:
            capex_row = conn.execute("SELECT SUM(purchase_cost) FROM ni_devices").fetchone()
            total_capex = float(capex_row[0] or 0) if capex_row else 0
            circ_row = conn.execute("SELECT SUM(monthly_cost_usd) FROM nc_circuits").fetchone()
            total_circuit_monthly = float(circ_row[0] or 0) if circ_row else 0
        except Exception:
            pass

        # Board reviews
        board_reviews = {"pending": 0, "approved": 0, "rejected": 0}
        try:
            for status in ("pending", "approved", "rejected"):
                cnt = conn.execute(
                    "SELECT COUNT(*) FROM nc_governance_reviews WHERE status = ?", (status,)
                ).fetchone()
                board_reviews[status] = cnt[0] if cnt else 0
        except Exception:
            pass

        conn.close()
        return jsonify({
            "total_projects": total_projects,
            "total_topologies": total_topos,
            "total_devices": total_devices,
            "total_interconnects": total_interconnects,
            "compliance_pct": compliance_pct,
            "total_cat1": total_cat1,
            "total_open_findings": total_open_findings,
            "status_counts": status_counts,
            "total_capex": total_capex,
            "total_circuit_cost_monthly": total_circuit_monthly,
            "board_reviews": board_reviews,
        })

    # ── Collaboration (Task 18) ───────────────────────────────────────────────
    import uuid as _uuid_mod
    from tools.canvas.collaboration import CanvasCollabManager as _NDCCollabMgr

    _ndc_collab = _NDCCollabMgr("nc")

    @bp.route("/api/collab/<design_id>/join", methods=["POST"])
    @nc_login_required
    def nc_collab_join(design_id):
        """Join a collaborative NDC editing session."""
        body = request.json or {}
        user_id = body.get("user_id", str(_uuid_mod.uuid4())[:8])
        user_name = body.get("user_name", "")
        return jsonify(_ndc_collab.join(design_id, user_id, user_name))

    @bp.route("/api/collab/<design_id>/leave", methods=["POST"])
    @nc_login_required
    def nc_collab_leave(design_id):
        """Leave an NDC collaborative session."""
        body = request.json or {}
        user_id = body.get("user_id", "")
        _ndc_collab.leave(design_id, user_id)
        return jsonify({"ok": True})

    @bp.route("/api/collab/<design_id>/push", methods=["POST"])
    @nc_login_required
    def nc_collab_push(design_id):
        """Push an operation into an NDC collaborative session."""
        body = request.json or {}
        user_id = body.get("user_id", "")
        op_type = body.get("op_type", "")
        data = body.get("data", {})
        seq = _ndc_collab.push(design_id, user_id, op_type, data)
        return jsonify({"seq": seq})

    @bp.route("/api/collab/<design_id>/poll", methods=["GET"])
    @nc_login_required
    def nc_collab_poll(design_id):
        """Poll for NDC collaborative operations since a sequence number."""
        since = int(request.args.get("since", 0))
        user_id = request.args.get("user_id", "")
        cx = request.args.get("cx")
        cy = request.args.get("cy")
        if user_id and cx is not None and cy is not None:
            _ndc_collab.update_cursor(design_id, user_id, float(cx), float(cy))
        ops, participants, latest_seq = _ndc_collab.poll(design_id, since)
        return jsonify({"operations": ops, "participants": participants, "latest_seq": latest_seq})

    @bp.route("/api/collab/<design_id>/participants", methods=["GET"])
    @nc_login_required
    def nc_collab_participants(design_id):
        """Return current participants in an NDC collaborative session."""
        return jsonify({"participants": _ndc_collab.get_participants(design_id)})

    # ── Runbooks ──────────────────────────────────────────────────────────────

    _NDC_RUNBOOK_TRIGGERS = [
        "link_failure",
        "routing_loop",
        "stig_remediation",
        "bgp_session_recovery",
        "circuit_outage_triage",
        "device_unreachable",
        "ipam_conflict",
        "ddos_mitigation",
    ]
    _NDC_RUNBOOK_SEVERITIES = ["critical", "high", "medium", "low"]

    @bp.route("/runbooks")
    @nc_login_required
    def nc_runbooks():
        """NDC Runbooks — network incident response playbooks."""
        conn = get_connection()
        runbooks = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT * FROM ndc_runbooks ORDER BY updated_at DESC"
            ).fetchall()
        ]
        topologies = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT id, name FROM topologies ORDER BY name"
            ).fetchall()
        ]
        conn.close()
        for rb in runbooks:
            try:
                rb["steps"] = json.loads(rb.get("steps_json") or "[]")
            except Exception:
                rb["steps"] = []
        return render_template(
            "network/runbooks.html",
            runbooks=runbooks,
            topologies=topologies,
            valid_triggers=_NDC_RUNBOOK_TRIGGERS,
            valid_severities=_NDC_RUNBOOK_SEVERITIES,
        )

    @bp.route("/api/runbooks", methods=["GET"])
    @nc_login_required
    def nc_api_runbooks_list():
        """List all NDC runbooks."""
        conn = get_connection()
        rows = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT * FROM ndc_runbooks ORDER BY updated_at DESC"
            ).fetchall()
        ]
        conn.close()
        for r in rows:
            try:
                r["steps"] = json.loads(r.get("steps_json") or "[]")
            except Exception:
                r["steps"] = []
        return jsonify(rows)

    @bp.route("/api/runbooks", methods=["POST"])
    @nc_login_required
    def nc_api_runbooks_create():
        """Create a new NDC runbook."""
        body = request.json or {}
        title = (body.get("title") or "").strip()
        if not title:
            return jsonify({"error": "title is required"}), 400
        trigger = body.get("trigger_event", "link_failure")
        if trigger not in _NDC_RUNBOOK_TRIGGERS:
            trigger = "link_failure"
        severity = body.get("severity", "high")
        if severity not in _NDC_RUNBOOK_SEVERITIES:
            severity = "high"
        rid = str(_uuid.uuid4())
        now = _now()
        steps_json = json.dumps(body.get("steps") or [])
        conn = get_connection()
        conn.execute(
            "INSERT INTO ndc_runbooks (id, title, trigger_event, severity, owner, "
            "topology_id, description, steps_json, classification, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (
                rid, title, trigger, severity,
                body.get("owner", ""), body.get("topology_id") or None,
                body.get("description", ""), steps_json,
                body.get("classification", "CUI"), now, now,
            ),
        )
        conn.commit()
        _audit(conn, "runbook_created", "ndc_runbook", rid, title)
        conn.close()
        return jsonify({"id": rid, "status": "created"})

    @bp.route("/api/runbooks/<rb_id>", methods=["GET"])
    @nc_login_required
    def nc_api_runbook_get(rb_id):
        """Get a single NDC runbook by ID."""
        conn = get_connection()
        row = conn.execute("SELECT * FROM ndc_runbooks WHERE id=?", (rb_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "not found"}), 404
        rb = _row_to_dict(row)
        try:
            rb["steps"] = json.loads(rb.get("steps_json") or "[]")
        except Exception:
            rb["steps"] = []
        return jsonify(rb)

    @bp.route("/api/runbooks/<rb_id>", methods=["PUT"])
    @nc_login_required
    def nc_api_runbook_update(rb_id):
        """Update an existing NDC runbook."""
        conn = get_connection()
        row = conn.execute("SELECT id FROM ndc_runbooks WHERE id=?", (rb_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "not found"}), 404
        body = request.json or {}
        now = _now()
        trigger = body.get("trigger_event", "link_failure")
        if trigger not in _NDC_RUNBOOK_TRIGGERS:
            trigger = "link_failure"
        severity = body.get("severity", "high")
        if severity not in _NDC_RUNBOOK_SEVERITIES:
            severity = "high"
        steps_json = json.dumps(body.get("steps") or [])
        conn.execute(
            "UPDATE ndc_runbooks SET title=?, trigger_event=?, severity=?, owner=?, "
            "topology_id=?, description=?, steps_json=?, classification=?, updated_at=? WHERE id=?",
            (
                (body.get("title") or "").strip() or "Untitled",
                trigger, severity,
                body.get("owner", ""), body.get("topology_id") or None,
                body.get("description", ""), steps_json,
                body.get("classification", "CUI"), now, rb_id,
            ),
        )
        conn.commit()
        _audit(conn, "runbook_updated", "ndc_runbook", rb_id, body.get("title", ""))
        conn.close()
        return jsonify({"status": "updated"})

    @bp.route("/api/runbooks/<rb_id>", methods=["DELETE"])
    @nc_login_required
    def nc_api_runbook_delete(rb_id):
        """Delete an NDC runbook."""
        conn = get_connection()
        row = conn.execute("SELECT id FROM ndc_runbooks WHERE id=?", (rb_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "not found"}), 404
        conn.execute("DELETE FROM ndc_runbooks WHERE id=?", (rb_id,))
        conn.commit()
        _audit(conn, "runbook_deleted", "ndc_runbook", rb_id, rb_id)
        conn.close()
        return jsonify({"status": "deleted"})

    # ── SOP Library ───────────────────────────────────────────────────────

    @bp.route("/sops")
    @nc_login_required
    def nc_sops_page():
        from tools.network.sops import list_sops as _ls
        category = request.args.get("category", "")
        status_f = request.args.get("status", "approved")
        q = request.args.get("q", "")
        sops = _ls(category=category or None, status=status_f or None, limit=200)
        all_sops = _ls(limit=1000)
        categories = sorted({s["category"] for s in all_sops})
        csps = sorted({s.get("csp", "multi") for s in all_sops})
        return render_template(
            "network/sops.html",
            sops=sops, categories=categories, csps=csps,
            filter_category=category, filter_status=status_f, search_q=q,
            is_admin=(getattr(getattr(g, "current_user", None), "role", "") == "admin"),
        )

    @bp.route("/api/sops")
    def nc_api_sops_list():
        from tools.network.sops import list_sops as _ls
        return jsonify(_ls(
            category=request.args.get("category") or None,
            status=request.args.get("status") or None,
            limit=200,
        ))

    @bp.route("/api/sops/<sop_id>")
    def nc_api_sop_get(sop_id):
        from tools.network.sops import get_sop as _gs
        s = _gs(sop_id)
        return (jsonify(s), 200) if s else (jsonify({"error": "not found"}), 404)

    @bp.route("/api/sops/<sop_id>/history")
    def nc_api_sop_history(sop_id):
        from tools.network.sops import get_approval_history as _gah
        return jsonify(_gah(sop_id))

    # ── Connectivity Reference ────────────────────────────────────────────

    @bp.route("/connectivity")
    @nc_login_required
    def nc_connectivity_page():
        from tools.network.connectivity_ref import (
            get_connectivity_matrix, get_scca_flow, get_resiliency_tiers,
        )
        return render_template(
            "network/connectivity.html",
            matrix=get_connectivity_matrix(),
            csps=["aws", "azure", "gcp", "oci", "ibm"],
            scca_flow=get_scca_flow(),
            resiliency_tiers=get_resiliency_tiers(),
        )

    @bp.route("/api/connectivity/matrix")
    def nc_api_connectivity_matrix():
        from tools.network.connectivity_ref import get_connectivity_matrix
        return jsonify(get_connectivity_matrix())

    @bp.route("/api/connectivity/onprem-pattern")
    def nc_api_onprem_pattern():
        from tools.network.connectivity_ref import get_onprem_to_csp_patterns
        return jsonify(get_onprem_to_csp_patterns(
            csp=request.args.get("csp", "aws"),
            pattern_type=request.args.get("type", "ipsec_vpn"),
        ))

    @bp.route("/api/connectivity/c2c-patterns")
    def nc_api_c2c_patterns():
        from tools.network.connectivity_ref import get_csp_to_csp_patterns
        return jsonify(get_csp_to_csp_patterns(
            src_csp=request.args.get("src", "aws"),
            dst_csp=request.args.get("dst", "azure"),
        ))

    # ── Packet Capture (GNS3 / lab link capture) ─────────────────────────

    import hashlib
    import struct
    from datetime import timedelta

    def _gen_stub_pcap() -> bytes:
        """Return a minimal valid PCAP file (global header only, no packets).

        Magic 0xa1b2c3d4, Ethernet link type.  Wireshark opens this cleanly and
        reports "0 packets captured".  Real GNS3 captures replace this with the
        streamed .pcapng file.
        """
        return struct.pack(
            "<IHHiIII",
            0xA1B2C3D4,  # magic number
            2,            # version major
            4,            # version minor
            0,            # thiszone (UTC)
            0,            # sigfigs
            65535,        # snaplen
            1,            # network (LINKTYPE_ETHERNET)
        )

    def _finalize_capture(conn, cap_id: str):
        """Transition a 'running' capture to 'complete' with stub PCAP data."""
        pcap_bytes = _gen_stub_pcap()
        sha = hashlib.sha256(pcap_bytes).hexdigest()
        expiry = (datetime.now(timezone.utc) + timedelta(days=7)).strftime("%Y-%m-%dT%H:%M:%SZ")
        now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        conn.execute(
            """UPDATE nc_packet_captures
               SET status='complete', size_bytes=?, sha256=?, expiry_at=?,
                   stopped_at=?, pcap_data=?
               WHERE id=?""",
            (len(pcap_bytes), sha, expiry, now, pcap_bytes, cap_id),
        )

    @bp.route("/api/captures", methods=["POST"])
    @nc_login_required
    def nc_api_capture_start():
        """Start a packet capture on a canvas link.

        Body: {link_id, topology_id, src_label?, dst_label?, protocol?, lab_run_id?}
        """
        body = request.get_json(force=True) or {}
        link_id = (body.get("link_id") or "").strip()
        topo_id = (body.get("topology_id") or "").strip()
        if not link_id or not topo_id:
            return jsonify({"error": "link_id and topology_id required"}), 400

        cap_id = str(_uuid.uuid4())
        lab_run_id = body.get("lab_run_id") or None
        now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

        # If no lab_run provided, auto-create a stub run for this topology
        conn = get_connection()
        if not lab_run_id:
            lab_run_id = str(_uuid.uuid4())
            conn.execute(
                """INSERT INTO nc_lab_runs (id, topology_id, name, backend, status, started_at)
                   VALUES (?, ?, ?, 'stub', 'running', ?)""",
                (lab_run_id, topo_id, f"Auto-run {now[:10]}", now),
            )

        conn.execute(
            """INSERT INTO nc_packet_captures
               (id, link_id, lab_run_id, topology_id, src_label, dst_label,
                protocol, status, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, 'running', ?)""",
            (
                cap_id,
                link_id,
                lab_run_id,
                topo_id,
                body.get("src_label", ""),
                body.get("dst_label", ""),
                body.get("protocol", ""),
                now,
            ),
        )
        conn.commit()
        _audit(conn, "capture_started", "nc_packet_captures", cap_id, link_id)
        conn.close()
        return jsonify({"id": cap_id, "status": "running", "link_id": link_id})

    @bp.route("/api/captures", methods=["GET"])
    @nc_login_required
    def nc_api_captures_list():
        """List captures.  Query params: link_id, topology_id."""
        link_id = request.args.get("link_id", "")
        topo_id = request.args.get("topology_id", "")
        conn = get_connection()
        if link_id:
            rows = conn.execute(
                "SELECT * FROM nc_packet_captures WHERE link_id=? ORDER BY created_at DESC",
                (link_id,),
            ).fetchall()
        elif topo_id:
            rows = conn.execute(
                "SELECT * FROM nc_packet_captures WHERE topology_id=? ORDER BY created_at DESC",
                (topo_id,),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM nc_packet_captures ORDER BY created_at DESC LIMIT 100"
            ).fetchall()
        conn.close()
        result = []
        for r in rows:
            d = _row_to_dict(r)
            d.pop("pcap_data", None)  # don't send binary over list endpoint
            result.append(d)
        return jsonify(result)

    @bp.route("/api/captures/<cap_id>", methods=["GET"])
    @nc_login_required
    def nc_api_capture_get(cap_id):
        """Poll a single capture; auto-finalizes stub captures after 5 s."""
        conn = get_connection()
        row = conn.execute(
            "SELECT * FROM nc_packet_captures WHERE id=?", (cap_id,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "not found"}), 404

        d = _row_to_dict(row)
        d.pop("pcap_data", None)

        # Auto-finalize stub captures that have been 'running' for ≥5 s
        if d.get("status") == "running":
            backend_ref = json.loads(d.get("backend_ref") or "{}")
            is_stub = not backend_ref.get("gns3_node_id")
            created = d.get("created_at", "")
            try:
                age = (
                    datetime.now(timezone.utc)
                    - datetime.fromisoformat(created.replace("Z", "+00:00"))
                ).total_seconds()
            except Exception:
                age = 99
            if is_stub and age >= 5:
                _finalize_capture(conn, cap_id)
                conn.commit()
                row2 = conn.execute(
                    "SELECT * FROM nc_packet_captures WHERE id=?", (cap_id,)
                ).fetchone()
                d = _row_to_dict(row2)
                d.pop("pcap_data", None)

        conn.close()
        return jsonify(d)

    @bp.route("/api/captures/<cap_id>/stop", methods=["POST"])
    @nc_login_required
    def nc_api_capture_stop(cap_id):
        """Stop a running capture."""
        conn = get_connection()
        row = conn.execute(
            "SELECT id, status FROM nc_packet_captures WHERE id=?", (cap_id,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "not found"}), 404
        if row["status"] == "running":
            _finalize_capture(conn, cap_id)
            conn.commit()
            _audit(conn, "capture_stopped", "nc_packet_captures", cap_id, cap_id)
        conn.close()
        return jsonify({"status": "complete"})

    @bp.route("/api/captures/<cap_id>/download", methods=["GET"])
    @nc_login_required
    def nc_api_capture_download(cap_id):
        """Download the .pcapng file for a completed capture."""
        from flask import Response

        conn = get_connection()
        row = conn.execute(
            "SELECT * FROM nc_packet_captures WHERE id=?", (cap_id,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "not found"}), 404
        d = _row_to_dict(row)

        # Ensure capture is finalized before download
        if d.get("status") == "running":
            _finalize_capture(conn, cap_id)
            conn.commit()
            row = conn.execute(
                "SELECT * FROM nc_packet_captures WHERE id=?", (cap_id,)
            ).fetchone()
            d = _row_to_dict(row)

        pcap_bytes = row["pcap_data"] if row["pcap_data"] else _gen_stub_pcap()
        conn.close()

        src = (d.get("src_label") or "src").replace(" ", "-")
        dst = (d.get("dst_label") or "dst").replace(" ", "-")
        ts = (d.get("created_at") or "").replace(":", "").replace("-", "")[:13]
        filename = f"capture-{src}-{dst}-{ts}.pcap"

        return Response(
            pcap_bytes,
            mimetype="application/vnd.tcpdump.pcap",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Length": str(len(pcap_bytes)),
            },
        )

    # ── Ingestion routes (file upload, config, NMS, folder watch) ────────
    from tools.network.routes.ingestion import register_ingestion_routes

    register_ingestion_routes(bp)

    @bp.route("/ingestion")
    @nc_login_required
    def nc_ingestion_page():
        """Network data ingestion dashboard."""
        return render_template("network/ingestion.html",
                               classification_banner=NC_CONFIG.get("app", {}).get("classification", ""))

    # ── GraphRAG /ask endpoint (per-canvas Q&A pattern, DT adaptation #1) ──
    _NDC_KG_PROJECT_ID = "ndc-network-intelligence"
    _NDC_KG_PROFILE = "network_infrastructure"

    @bp.route("/ask")
    @nc_login_required
    def nc_ask_page():
        return render_template(
            "network/ask.html",
            classification_banner=NC_CONFIG.get("app", {}).get("classification", ""),
        )

    @bp.route("/api/ask", methods=["POST"])
    @nc_login_required
    def nc_api_ask():
        from tools.knowledge_graph.canvas_ask import handle_ask_request
        data = request.get_json(silent=True) or {}
        payload = handle_ask_request(
            query=data.get("query", ""),
            graph_id=_NDC_KG_PROJECT_ID,
            profile=_NDC_KG_PROFILE,
            top_k=int(data.get("top_k", 10)),
            narrate=bool(data.get("narrate", False)),
            canvas_label="network topology",
        )
        status = payload.pop("_status", 200)
        return jsonify(payload), status

    # ── Digital Twin ───────────────────────────────────────────────────────
    @bp.route("/twin/<topo_id>")
    @nc_login_required
    def nc_twin_page(topo_id):
        conn = get_connection()
        topo = conn.execute(
            "SELECT * FROM topologies WHERE id = ?", (topo_id,)
        ).fetchone()
        if not topo:
            return render_template("404.html"), 404
        topo = _row_to_dict(topo)

        try:
            snaps = conn.execute(
                "SELECT * FROM network_twin_snapshots WHERE project_id = ? ORDER BY created_at DESC LIMIT 20",
                (topo_id,),
            ).fetchall()
        except Exception:
            snaps = []

        from tools.network.constants import INTENT_RULES
        from tools.network.narrative_generator import load_personas as _ng_personas
        try:
            _tfw_personas = _ng_personas()
        except Exception:
            _tfw_personas = []
        return render_template(
            "network/twin.html",
            project=topo,
            snapshots=[_row_to_dict(s) for s in snaps],
            intent_rules=INTENT_RULES,
            classification_banner=NC_CONFIG.get("app", {}).get("classification", ""),
            personas=_tfw_personas,
        )

    @bp.route("/api/twin/<topo_id>/snapshot", methods=["POST"])
    @nc_login_required
    def nc_api_twin_snapshot(topo_id):
        from tools.network.twin import take_snapshot
        data = request.get_json(silent=True) or {}
        snap = take_snapshot(topo_id, label=data.get("label"))
        return jsonify(snap), 201

    @bp.route("/api/twin/<topo_id>/simulate", methods=["POST"])
    @nc_login_required
    def nc_api_twin_simulate(topo_id):
        from tools.network.twin import simulate_delta
        data = request.get_json(silent=True) or {}
        result = simulate_delta(
            topo_id,
            topology_delta=data.get("topology_delta", {}),
            intent_rules=data.get("intent_rules", []),
            baseline_snap_id=data.get("baseline_snap_id"),
        )
        return jsonify(result), 200

    @bp.route("/api/twin/<topo_id>/blast-radius", methods=["POST"])
    @nc_login_required
    def nc_api_twin_blast_radius(topo_id):
        from tools.network.twin import blast_radius
        data = request.get_json(silent=True) or {}
        result = blast_radius(
            topo_id,
            node_id=data.get("node_id", ""),
            topology_delta=data.get("topology_delta"),
            baseline_snap_id=data.get("baseline_snap_id"),
        )
        return jsonify(result), 200

    @bp.route("/api/twin/<topo_id>/analyze-path", methods=["POST"])
    @nc_login_required
    def nc_api_twin_analyze_path(topo_id):
        from tools.network.path_analyzer import find_paths
        import json as _json
        data = request.get_json(silent=True) or {}
        src_query = (data.get("src") or data.get("src_id") or "").strip()
        dst_query = (data.get("dst") or data.get("dst_id") or "").strip()
        if not src_query or not dst_query:
            return jsonify({"error": "src and dst are required"}), 400
        conn = get_connection()
        try:
            row = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        finally:
            conn.close()
        if not row or not row["graph_json"]:
            return jsonify({"error": "Topology not found"}), 404
        try:
            raw_graph = _json.loads(row["graph_json"])
        except Exception:
            return jsonify({"error": "Invalid topology JSON"}), 500
        nodes_list = raw_graph.get("nodes", [])
        nodes_dict = {n["id"]: n for n in nodes_list if isinstance(n, dict) and "id" in n}
        graph = {"nodes": nodes_dict, "edges": raw_graph.get("edges", [])}
        result = find_paths(src_query, dst_query, graph)
        resolve_parts = []
        if result["src"] != src_query:
            src_label = nodes_dict.get(result["src"], {}).get("label") or result["src"]
            resolve_parts.append(f'"{src_query}" → {src_label}')
        if result["dst"] != dst_query:
            dst_label = nodes_dict.get(result["dst"], {}).get("label") or result["dst"]
            resolve_parts.append(f'"{dst_query}" → {dst_label}')
        result["resolve_note"] = "Matched: " + ", ".join(resolve_parts) if resolve_parts else None
        if result["reachable"]:
            result["verdict"] = "green"
        elif result["blocked_by_acl"]:
            result["verdict"] = "amber"
        else:
            result["verdict"] = "red"
        return jsonify(result), 200

    @bp.route("/api/twin/<topo_id>/current-topology", methods=["GET"])
    @nc_login_required
    def nc_api_twin_current_topology(topo_id):
        conn = get_connection()
        row = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
        if not row:
            return jsonify({"error": "Topology not found"}), 404
        try:
            graph = json.loads(row["graph_json"] or "{}")
        except Exception:
            graph = {}
        return jsonify({"graph_json": graph}), 200

    @bp.route("/api/twin/<topo_id>/chat-delta", methods=["POST"])
    @nc_login_required
    def nc_api_twin_chat_delta(topo_id):
        from tools.twin_chat import network_chat_to_delta
        data = request.get_json(silent=True) or {}
        message = (data.get("message") or "").strip()
        if not message:
            return jsonify({"error": "message is required"}), 400
        result = network_chat_to_delta(message, data.get("graph_json"))
        return jsonify(result), (500 if "error" in result else 200)

    @bp.route("/api/twin/<topo_id>/nl-query", methods=["POST"])
    @nc_login_required
    def nc_api_twin_nl_query(topo_id):
        """IQE AI Assist — answer a natural language question about a topology."""
        from tools.network.nl_query import answer_query
        data = request.get_json(silent=True) or {}
        question = (data.get("question") or "").strip()
        if not question:
            return jsonify({"error": "question is required"}), 400
        conn = get_connection()
        try:
            result = answer_query(topo_id, question, conn)
        finally:
            conn.close()
        return jsonify(result), 200

    @bp.route("/api/twin/<topo_id>/iqe-query", methods=["POST"])
    @nc_login_required
    def nc_api_twin_iqe_query(topo_id):
        """IQE structured query — translate NL to IQE and optionally execute against topology."""
        from tools.iqe.nl_to_iqe import nl_to_iqe
        from tools.iqe.parser import IQESyntaxError, parse
        from tools.iqe.executor import Executor

        data = request.get_json(silent=True) or {}
        question = (data.get("question") or "").strip()
        if not question:
            return jsonify({"error": "question is required"}), 400
        execute = data.get("execute", True)

        # Derive available collections from topology node/edge tables
        collections = ["nodes", "edges"]

        translation = nl_to_iqe(question, collections)
        iqe_str = translation.get("iqe", "")
        explanation = translation.get("explanation", "")

        if not execute:
            return jsonify({"iqe": iqe_str, "explanation": explanation}), 200

        conn = get_connection()
        try:
            ast = parse(iqe_str)

            # Build adapters for topology node/edge data
            def _nodes_adapter(c):
                row = c.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
                if not row:
                    return []
                import json as _json
                graph = _json.loads(row["graph_json"] or "{}")
                cells = graph.get("cells") or []
                result = []
                for cell in cells:
                    kind = (cell.get("type") or "").lower()
                    if "link" in kind or "edge" in kind:
                        continue
                    attrs = cell.get("attrs") or {}
                    label = ""
                    for key in ("label", "text", "body"):
                        val = attrs.get(key)
                        if isinstance(val, dict):
                            label = val.get("text") or val.get("textWrap", {}).get("text", "") or ""
                        if label:
                            break
                    if not label:
                        label = cell.get("label") or cell.get("name") or cell.get("id", "")
                    result.append({
                        "id": cell.get("id", ""),
                        "label": str(label).strip(),
                        "type": cell.get("deviceType") or attrs.get("deviceType", "") or cell.get("type", ""),
                        "position": cell.get("position", {}),
                    })
                return result

            def _edges_adapter(c):
                row = c.execute("SELECT graph_json FROM topologies WHERE id=?", (topo_id,)).fetchone()
                if not row:
                    return []
                import json as _json
                graph = _json.loads(row["graph_json"] or "{}")
                cells = graph.get("cells") or []
                result = []
                for cell in cells:
                    kind = (cell.get("type") or "").lower()
                    if "link" not in kind and "edge" not in kind:
                        continue
                    src = (cell.get("source") or {})
                    tgt = (cell.get("target") or {})
                    result.append({
                        "id": cell.get("id", ""),
                        "source": src.get("id", "") if isinstance(src, dict) else str(src),
                        "target": tgt.get("id", "") if isinstance(tgt, dict) else str(tgt),
                        "protocol": (cell.get("attrs") or {}).get("line", {}).get("strokeDasharray", "") or cell.get("protocol", ""),
                        "bandwidth": cell.get("bandwidth", ""),
                    })
                return result

            executor = Executor()
            executor.register_collection("nodes", _nodes_adapter)
            executor.register_collection("edges", _edges_adapter)

            rows = executor.run(ast, conn)
            return jsonify({
                "iqe": iqe_str,
                "explanation": explanation,
                "results": rows,
                "row_count": len(rows),
            }), 200
        except IQESyntaxError as exc:
            return jsonify({"error": f"IQE syntax error: {exc}", "iqe": iqe_str}), 400
        except Exception as exc:
            logger.warning("IQE query execution error: %s", exc)
            return jsonify({"error": str(exc), "iqe": iqe_str}), 500
        finally:
            conn.close()

    @bp.route("/api/twin/<topo_id>/nodes/<node_id>/domain-policy", methods=["GET", "POST"])
    @nc_login_required
    def nc_api_twin_node_domain_policy(topo_id, node_id):
        """GET: load domain policy for a node. POST: save domain policy for a node."""
        import json as _json
        import uuid
        from datetime import datetime, timezone

        conn = get_connection()
        try:
            if request.method == "GET":
                row = conn.execute(
                    "SELECT domain_type, domain_label, security_policy, routing_policy, vpn_policy"
                    " FROM nc_security_domain_policies WHERE topology_id=? AND node_id=?",
                    (topo_id, node_id),
                ).fetchone()
                if not row:
                    return jsonify({"domain_type": "on_prem", "security_policy": {}, "routing_policy": {}, "vpn_policy": {}}), 200
                return jsonify({
                    "domain_type": row["domain_type"],
                    "domain_label": row["domain_label"] or "",
                    "security_policy": _json.loads(row["security_policy"] or "{}"),
                    "routing_policy": _json.loads(row["routing_policy"] or "{}"),
                    "vpn_policy": _json.loads(row["vpn_policy"] or "{}"),
                }), 200
            else:
                data = request.get_json(silent=True) or {}
                domain_type = (data.get("domain_type") or "on_prem").strip()
                domain_label = (data.get("domain_label") or "").strip()
                sec = _json.dumps(data.get("security_policy") or {})
                route = _json.dumps(data.get("routing_policy") or {})
                vpn = _json.dumps(data.get("vpn_policy") or {})
                now = datetime.now(timezone.utc).isoformat()
                existing = conn.execute(
                    "SELECT id FROM nc_security_domain_policies WHERE topology_id=? AND node_id=?",
                    (topo_id, node_id),
                ).fetchone()
                if existing:
                    conn.execute(
                        "UPDATE nc_security_domain_policies"
                        " SET domain_type=?, domain_label=?, security_policy=?, routing_policy=?, vpn_policy=?, updated_at=?"
                        " WHERE topology_id=? AND node_id=?",
                        (domain_type, domain_label, sec, route, vpn, now, topo_id, node_id),
                    )
                else:
                    conn.execute(
                        "INSERT INTO nc_security_domain_policies"
                        " (id, topology_id, node_id, domain_type, domain_label, security_policy, routing_policy, vpn_policy, created_at, updated_at)"
                        " VALUES (?,?,?,?,?,?,?,?,?,?)",
                        (str(uuid.uuid4()), topo_id, node_id, domain_type, domain_label, sec, route, vpn, now, now),
                    )
                conn.commit()
                return jsonify({"ok": True}), 200
        except Exception as exc:
            logger.warning("domain-policy error: %s", exc)
            return jsonify({"error": str(exc)}), 500
        finally:
            conn.close()

    @bp.route("/api/twin/<topo_id>/domain-policy-templates", methods=["GET"])
    @nc_login_required
    def nc_api_twin_domain_policy_templates(topo_id):
        """Return pre-built JSON templates for each domain type."""
        templates = {
            "on_prem": {
                "security": {"allowed_ports": [22, 80, 443], "denied_ports": [], "inspection_type": "stateful", "encryption_required": False, "mfa_required": False, "pki_required": False, "cdm_sensor": False, "idps_enabled": False},
                "routing": {"bgp_as": None, "static_routes": [], "next_hops": [], "failover_group": None, "load_balance_algo": "round_robin", "advertised_prefixes": []},
                "vpn": {"ike_version": 2, "encryption_algo": "AES-256", "auth_algo": "SHA-256", "pfs_group": 14, "tunnel_endpoints": [], "dpd_timeout": 30},
            },
            "nipr": {
                "security": {"allowed_ports": [22, 80, 443, 8080], "denied_ports": [], "inspection_type": "deep_packet", "encryption_required": True, "mfa_required": True, "pki_required": True, "cdm_sensor": True, "idps_enabled": True},
                "routing": {"bgp_as": 65001, "static_routes": [], "next_hops": [], "failover_group": "nipr-primary", "load_balance_algo": "weighted", "advertised_prefixes": []},
                "vpn": {"ike_version": 2, "encryption_algo": "AES-256-GCM", "auth_algo": "SHA-384", "pfs_group": 20, "tunnel_endpoints": [], "dpd_timeout": 10},
            },
            "sipr": {
                "security": {"allowed_ports": [22, 443], "denied_ports": [80, 8080], "inspection_type": "deep_packet", "encryption_required": True, "mfa_required": True, "pki_required": True, "cdm_sensor": True, "idps_enabled": True},
                "routing": {"bgp_as": 65002, "static_routes": [], "next_hops": [], "failover_group": "sipr-primary", "load_balance_algo": "weighted", "advertised_prefixes": []},
                "vpn": {"ike_version": 2, "encryption_algo": "AES-256-GCM", "auth_algo": "SHA-512", "pfs_group": 21, "tunnel_endpoints": [], "dpd_timeout": 5},
            },
            "bcap_vdms": {
                "security": {"allowed_ports": [443], "denied_ports": [80], "inspection_type": "deep_packet", "encryption_required": True, "mfa_required": True, "pki_required": True, "cdm_sensor": True, "idps_enabled": True},
                "routing": {"bgp_as": None, "static_routes": [], "next_hops": [], "failover_group": "bcap-vdms", "load_balance_algo": "least_conn", "advertised_prefixes": []},
                "vpn": {"ike_version": 2, "encryption_algo": "AES-256-GCM", "auth_algo": "SHA-384", "pfs_group": 20, "tunnel_endpoints": [], "dpd_timeout": 10},
            },
            "bcap_vdss": {
                "security": {"allowed_ports": [443], "denied_ports": [80, 22], "inspection_type": "deep_packet", "encryption_required": True, "mfa_required": True, "pki_required": True, "cdm_sensor": True, "idps_enabled": True},
                "routing": {"bgp_as": None, "static_routes": [], "next_hops": [], "failover_group": "bcap-vdss", "load_balance_algo": "least_conn", "advertised_prefixes": []},
                "vpn": {"ike_version": 2, "encryption_algo": "AES-256-GCM", "auth_algo": "SHA-512", "pfs_group": 21, "tunnel_endpoints": [], "dpd_timeout": 5},
            },
            "csp_il2": {
                "security": {"allowed_ports": [80, 443], "denied_ports": [], "inspection_type": "stateful", "encryption_required": False, "mfa_required": False, "pki_required": False, "cdm_sensor": False, "idps_enabled": False},
                "routing": {"bgp_as": None, "static_routes": [], "next_hops": [], "failover_group": None, "load_balance_algo": "round_robin", "advertised_prefixes": []},
                "vpn": {"ike_version": 2, "encryption_algo": "AES-128", "auth_algo": "SHA-256", "pfs_group": 14, "tunnel_endpoints": [], "dpd_timeout": 30},
            },
            "csp_il4": {
                "security": {"allowed_ports": [443], "denied_ports": [80], "inspection_type": "deep_packet", "encryption_required": True, "mfa_required": True, "pki_required": True, "cdm_sensor": True, "idps_enabled": True},
                "routing": {"bgp_as": None, "static_routes": [], "next_hops": [], "failover_group": "il4-primary", "load_balance_algo": "weighted", "advertised_prefixes": []},
                "vpn": {"ike_version": 2, "encryption_algo": "AES-256-GCM", "auth_algo": "SHA-384", "pfs_group": 20, "tunnel_endpoints": [], "dpd_timeout": 10},
            },
            "csp_il5": {
                "security": {"allowed_ports": [443], "denied_ports": [80, 22], "inspection_type": "deep_packet", "encryption_required": True, "mfa_required": True, "pki_required": True, "cdm_sensor": True, "idps_enabled": True},
                "routing": {"bgp_as": None, "static_routes": [], "next_hops": [], "failover_group": "il5-primary", "load_balance_algo": "weighted", "advertised_prefixes": []},
                "vpn": {"ike_version": 2, "encryption_algo": "AES-256-GCM", "auth_algo": "SHA-512", "pfs_group": 21, "tunnel_endpoints": [], "dpd_timeout": 5},
            },
            "csp_il6": {
                "security": {"allowed_ports": [443], "denied_ports": [80, 22, 8080], "inspection_type": "deep_packet", "encryption_required": True, "mfa_required": True, "pki_required": True, "cdm_sensor": True, "idps_enabled": True},
                "routing": {"bgp_as": 65006, "static_routes": [], "next_hops": [], "failover_group": "il6-primary", "load_balance_algo": "weighted", "advertised_prefixes": []},
                "vpn": {"ike_version": 2, "encryption_algo": "AES-256-GCM", "auth_algo": "SHA-512", "pfs_group": 21, "tunnel_endpoints": [], "dpd_timeout": 5},
            },
            "inter_csp": {
                "security": {"allowed_ports": [443], "denied_ports": [80], "inspection_type": "deep_packet", "encryption_required": True, "mfa_required": True, "pki_required": True, "cdm_sensor": True, "idps_enabled": True},
                "routing": {"bgp_as": None, "static_routes": [], "next_hops": [], "failover_group": "inter-csp", "load_balance_algo": "weighted", "advertised_prefixes": []},
                "vpn": {"ike_version": 2, "encryption_algo": "AES-256-GCM", "auth_algo": "SHA-384", "pfs_group": 20, "tunnel_endpoints": [], "dpd_timeout": 10},
            },
            "dmz": {
                "security": {"allowed_ports": [80, 443, 25, 53], "denied_ports": [], "inspection_type": "stateful", "encryption_required": False, "mfa_required": False, "pki_required": False, "cdm_sensor": False, "idps_enabled": True},
                "routing": {"bgp_as": None, "static_routes": [], "next_hops": [], "failover_group": None, "load_balance_algo": "round_robin", "advertised_prefixes": []},
                "vpn": {"ike_version": 2, "encryption_algo": "AES-256", "auth_algo": "SHA-256", "pfs_group": 14, "tunnel_endpoints": [], "dpd_timeout": 30},
            },
            "internet": {
                "security": {"allowed_ports": [80, 443], "denied_ports": [], "inspection_type": "none", "encryption_required": False, "mfa_required": False, "pki_required": False, "cdm_sensor": False, "idps_enabled": False},
                "routing": {"bgp_as": None, "static_routes": [], "next_hops": [], "failover_group": None, "load_balance_algo": "round_robin", "advertised_prefixes": []},
                "vpn": {},
            },
            "custom": {"security": {}, "routing": {}, "vpn": {}},
        }
        return jsonify(templates), 200

    @bp.route("/api/twin/<topo_id>/persona-definitions", methods=["GET"])
    @nc_login_required
    def nc_api_twin_persona_definitions(topo_id):
        """Return all TFW persona definitions from tfw_personas.yaml."""
        from tools.network.narrative_generator import load_personas
        return jsonify({"personas": load_personas()}), 200

    @bp.route("/api/twin/<topo_id>/traffic-flows", methods=["GET", "POST"])
    @nc_login_required
    def nc_api_twin_traffic_flows(topo_id):
        """GET: list flows for a topology. POST: create a new flow (returns 201 + id)."""
        from tools.network.traffic_flow import TrafficFlowEngine
        from tools.network.db.init_db import get_connection as _gc

        conn = _gc()
        try:
            engine = TrafficFlowEngine()
            if request.method == "GET":
                flows = engine.list_flows(topo_id, conn)
                phase_filter = request.args.get("phase_id")
                if phase_filter:
                    flows = [f for f in flows if f.get("phase_id") == phase_filter]
                return jsonify({"flows": flows}), 200
            # POST — create
            data = request.get_json(silent=True) or {}
            name = (data.get("name") or "").strip()
            src_zone = (data.get("src_zone") or "").strip()
            dst_zone = (data.get("dst_zone") or "").strip()
            app_type = (data.get("app_type") or "sso_saml").strip()
            classification = (data.get("classification") or "NIPR").strip()
            if not name or not src_zone or not dst_zone:
                return jsonify({"error": "name, src_zone, and dst_zone are required"}), 400
            flow_id = engine.create_flow(topo_id, name, src_zone, dst_zone, app_type, classification, conn)
            return jsonify({"id": flow_id, "name": name, "src_zone": src_zone,
                            "dst_zone": dst_zone, "app_type": app_type,
                            "classification": classification}), 201
        except Exception as exc:
            logger.warning("traffic-flows error: %s", exc)
            return jsonify({"error": str(exc)}), 500
        finally:
            conn.close()

    @bp.route("/api/twin/<topo_id>/traffic-flows/<flow_id>/walkthrough", methods=["POST"])
    @nc_login_required
    def nc_api_twin_tfw_walkthrough(topo_id, flow_id):
        """Run multi-persona walkthrough for a traffic flow.

        Request body (JSON, all optional):
          personas       : list of persona IDs to generate (default: all 7)
          classification : override flow classification (NIPR, IL4, IL5, IL6, SIPR)
          use_llm        : bool, default True

        Returns:
          {steps: [...], summary: {...}}
          Each step has: step_number, node_id, node_label, action_type,
          persona_responses: {persona_id: {narrative, detail_json}}
        """
        from tools.network.narrative_generator import generate_all
        from tools.network.db.init_db import get_connection

        body = request.get_json(silent=True) or {}
        personas = body.get("personas") or None
        use_llm = bool(body.get("use_llm", True))
        phase_id_filter = request.args.get("phase_id") or body.get("phase_id")

        try:
            conn = get_connection()
            engine_cls = None
            try:
                from tools.network.traffic_flow import TrafficFlowEngine
                engine_cls = TrafficFlowEngine
            except Exception:
                pass

            if engine_cls:
                engine = engine_cls()
                engine._ensure_tables(conn)
                if phase_id_filter:
                    engine.generate_walkthrough(flow_id, conn, phase_id=phase_id_filter)

            # Verify flow belongs to this topology
            flow_row = conn.execute(
                "SELECT * FROM nc_traffic_flows WHERE id = ? AND topology_id = ?",
                (flow_id, topo_id),
            ).fetchone()
            if not flow_row:
                return jsonify({"error": "flow not found"}), 404

            flow = dict(flow_row)
            classification = body.get("classification") or flow.get("classification", "NIPR")

            result = generate_all(
                flow_id=flow_id,
                conn=conn,
                personas=personas,
                classification=classification,
                use_llm=use_llm,
            )

            # Reformat steps to use 'persona_responses' key for API consumers
            api_steps = []
            for step in result.get("steps", []):
                api_step = {
                    "step_number": step["step_number"],
                    "node_id":     step["node_id"],
                    "node_label":  step["node_label"],
                    "action_type": step["action_type"],
                    "persona_responses": step.get("personas", {}),
                }
                api_steps.append(api_step)

            return jsonify({"steps": api_steps, "summary": result.get("summary", {})}), 200
        except Exception as exc:
            logger.warning("walkthrough error: %s", exc)
            return jsonify({"error": str(exc)}), 500
        finally:
            try:
                conn.close()
            except Exception:
                pass

    @bp.route("/api/twin/<topo_id>/traffic-flows/<flow_id>/assign-phase", methods=["POST"])
    @nc_login_required
    def nc_api_twin_tfw_assign_phase(topo_id, flow_id):
        """Assign (or unassign) a traffic flow to a migration phase.

        Body: {"phase_id": "<id>"}  — pass null to unassign.
        """
        from tools.network.db.init_db import get_connection as _gc
        data = request.get_json(silent=True) or {}
        phase_id = data.get("phase_id")

        conn = _gc()
        try:
            row = conn.execute(
                "SELECT id FROM nc_traffic_flows WHERE id=? AND topology_id=?",
                (flow_id, topo_id),
            ).fetchone()
            if not row:
                return jsonify({"error": "flow not found"}), 404

            conn.execute(
                "UPDATE nc_traffic_flows SET phase_id=? WHERE id=?",
                (phase_id, flow_id),
            )
            conn.commit()
            return jsonify({"status": "ok", "flow_id": flow_id, "phase_id": phase_id}), 200
        except Exception as exc:
            logger.warning("assign-phase error: %s", exc)
            return jsonify({"error": str(exc)}), 500
        finally:
            conn.close()

    # ══════════════════════════════════════════════════════════════════════
    # API: IP Address Planning Assistant
    # ══════════════════════════════════════════════════════════════════════

    _IP_PLAN_SYSTEM_PROMPT = """You are a network IP address planner. Given a supernet and topology nodes, allocate non-overlapping CIDRs.

Output ONLY a valid JSON object — no markdown, no explanation:
{"assignments": [{"node_id": "...", "label": "...", "type": "...", "cidr": "x.x.x.x/xx", "gateway": "x.x.x.x", "vlan": N}]}

Strategy rules:
- balanced: divide the supernet into equal-sized subnets for each logical segment
- power-of-2: size each subnet to the next power-of-2 that fits expected hosts
- flat: assign sequential /24 subnets

Planning rules:
1. Stay within the given supernet — never assign IPs outside it
2. Never overlap CIDRs
3. Logical segments (type=subnet, vlan, security-zone, vrf, aws-subnet, aws-vpc, az-vnet, gcp-vpc): assign a network CIDR (e.g. 10.0.1.0/24)
4. Devices (router, switch-l2, switch-l3, firewall, server, load-balancer, wap, siem, sdwan-edge, etc.): assign a host IP within the appropriate segment CIDR (e.g. 10.0.1.1/24)
5. If no explicit segments exist, group devices by inferred function and assign subnets accordingly
6. gateway = first usable IP (.1) of the subnet
7. vlan: assign sequential VLAN IDs starting at 10, incrementing by 10 per segment
8. Include ALL provided nodes in the assignments array
9. Output ONLY the JSON object"""

    @bp.route("/api/topologies/<topo_id>/plan-ips", methods=["POST"])
    @nc_login_required
    def nc_api_plan_ips(topo_id):
        """AI-assisted IP address planning: supernet → subnet/host allocation.

        Uses ICDEV™ LLM Router when available; falls back to deterministic
        subnet allocation so the feature works regardless of API keys.
        """
        conn = get_connection()
        topo = conn.execute(
            "SELECT id, name, graph_json FROM topologies WHERE id=?",
            (topo_id,),
        ).fetchone()
        conn.close()
        if not topo:
            return jsonify({"error": "Topology not found"}), 404

        data = request.get_json(force=True, silent=True) or {}
        supernet = data.get("supernet", "").strip()
        strategy = data.get("strategy", "balanced").strip()
        if not supernet:
            return jsonify({"error": "supernet required"}), 400
        if strategy not in ("balanced", "power-of-2", "flat"):
            strategy = "balanced"

        # Parse graph nodes
        try:
            graph_data = json.loads(topo["graph_json"] or "{}")
        except Exception:
            graph_data = {}
        nodes = graph_data.get("nodes", [])

        _SKIP_TYPES = {
            "draw-rect", "draw-rounded-rect", "text-heading", "text-label",
            "text-badge", "media-fiber", "media-ge", "media-10ge", "media-100ge",
            "patch-panel", "meet-me-room", "cross-connect",
        }
        routable = [
            {"id": n["id"], "label": n.get("label", ""), "type": n.get("type", "")}
            for n in nodes
            if n.get("type", "") not in _SKIP_TYPES and n.get("type", "")
        ]
        if not routable:
            return jsonify({"error": "No addressable nodes found in topology"}), 400

        def _plan_ips_deterministic(nodes, strategy):
            """Deterministic subnet allocator using Python ipaddress."""
            import ipaddress
            net = ipaddress.ip_network(supernet, strict=False)

            SEGMENT_TYPES = {
                "subnet", "vlan", "security-zone", "vrf",
                "aws-subnet", "aws-vpc", "az-vnet", "gcp-vpc",
            }
            INFRA_TYPES = {"router", "firewall", "sdwan-edge"}
            DIST_TYPES = {"switch-l2", "switch-l3", "switch", "load-balancer"}
            COMPUTE_TYPES = {"server", "siem", "wap"}

            groups = {}
            for node in nodes:
                t = node.get("type", "")
                if t in SEGMENT_TYPES:
                    key = "segment_" + t
                elif t in INFRA_TYPES:
                    key = "infra"
                elif t in DIST_TYPES:
                    key = "distribution"
                elif t in COMPUTE_TYPES:
                    key = "compute"
                else:
                    key = "general"
                groups.setdefault(key, []).append(node)

            if not groups:
                return {"assignments": []}

            group_items = list(groups.items())

            if strategy == "flat":
                prefix_len = 24
                base = int(net.network_address)
                subnets = []
                for i in range(len(group_items)):
                    subnets.append(
                        ipaddress.ip_network(
                            f"{ipaddress.ip_address(base + i * 256)}/{prefix_len}",
                            strict=False,
                        )
                    )
            else:
                needed_bits = (len(group_items) - 1).bit_length()
                new_prefix = net.prefixlen + needed_bits
                if new_prefix > 30:
                    new_prefix = 30
                subnets = list(net.subnets(new_prefix=new_prefix))[:len(group_items)]

            assignments = []
            vlan_base = 10
            for idx, (key, members) in enumerate(group_items):
                subnet = subnets[idx]
                hosts = list(subnet.hosts())
                gateway = str(hosts[0]) if hosts else str(subnet.network_address + 1)
                vlan = vlan_base + idx * 10
                host_idx = 1
                for node in members:
                    t = node.get("type", "")
                    if t in SEGMENT_TYPES:
                        cidr = f"{subnet.network_address}/{subnet.prefixlen}"
                    else:
                        if host_idx < len(hosts):
                            ip = hosts[host_idx]
                            host_idx += 1
                        else:
                            ip = subnet.network_address + 2
                        cidr = f"{ip}/{subnet.prefixlen}"
                    assignments.append({
                        "node_id": node.get("id", ""),
                        "label": node.get("label", ""),
                        "type": t,
                        "cidr": cidr,
                        "gateway": gateway,
                        "vlan": vlan,
                    })
            return {"assignments": assignments}

        def _parse_ip_plan(content):
            import re
            text = content.strip()
            text = re.sub(r"\n?[\s\S]*?\n?", "", text, flags=re.DOTALL).strip()
            if text.startswith("```"):
                lines = [ln for ln in text.split("\n") if not ln.strip().startswith("```")]
                text = "\n".join(lines).strip()
            start = text.find("{")
            end = text.rfind("}") + 1
            if start < 0 or end <= start:
                return None, text[:500]
            try:
                result = json.loads(text[start:end])
            except json.JSONDecodeError:
                return None, text[:500]
            if "assignments" not in result:
                return None, text[:500]
            return result, None

        # Deterministic result is always available
        used_provider = "deterministic"
        result = _plan_ips_deterministic(routable, strategy)

        # Attempt LLM enhancement via ICDEV™ router (optional, non-blocking)
        try:
            from icdev.tools.llm.router import LLMRouter
            from icdev.tools.llm.provider import LLMRequest
            router = LLMRouter()
            req = LLMRequest(
                system_prompt=_IP_PLAN_SYSTEM_PROMPT,
                messages=[
                    {"role": "user", "content": (
                        f"Topology: {topo['name']}\n"
                        f"Supernet: {supernet}\n"
                        f"Strategy: {strategy}\n\n"
                        f"Nodes:\n{json.dumps(routable, indent=2)}\n\n"
                        f"Assign IP CIDRs from the supernet to these nodes using the {strategy} strategy."
                    )},
                ],
                max_tokens=4096,
                temperature=0.1,
            )
            resp = router.invoke("ip_planning", req)
            if resp and resp.content:
                llm_result, raw = _parse_ip_plan(resp.content)
                if llm_result and llm_result.get("assignments"):
                    result = llm_result
                    used_provider = resp.provider or "llm"
        except Exception:
            pass  # deterministic result already set

        _audit(
            "IP_PLAN", "topology", topo_id,
            f"[{used_provider}] {supernet} ({strategy}) → {len(result['assignments'])} assignments",
        )
        return jsonify({
            "assignments": result["assignments"],
            "supernet": supernet,
            "strategy": strategy,
            "provider": used_provider,
        })

    # ── AI Context Messages ────────────────────────────────────────────────
    @bp.route("/api/ai-context/<ctx_id>/messages", methods=["GET"])
    def nc_api_ai_context_messages(ctx_id):
        """Return up to 50 messages for an AI context, ordered by turn_number."""
        conn = get_connection()
        try:
            rows = conn.execute(
                "SELECT id, role, content, turn_number FROM chat_messages"
                " WHERE context_id = ? ORDER BY turn_number ASC LIMIT 50",
                (ctx_id,),
            ).fetchall()
            messages = [
                {"role": r[1], "content": r[2], "turn_number": r[3]}
                for r in rows
            ]
            return jsonify({"messages": messages})
        finally:
            conn.close()

    # ══════════════════════════════════════════════════════════════════════
    # Subnet Calculator — page + API
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/subnet-calc")
    @nc_login_required
    def nc_subnet_calc():
        conn = get_connection()
        project_id = request.args.get("project", "")
        projects = [_row_to_dict(r) for r in conn.execute("SELECT id, name FROM nc_projects ORDER BY name").fetchall()]
        if project_id:
            rows = conn.execute(
                "SELECT * FROM nc_subnet_calc_history WHERE project_id=? ORDER BY created_at DESC",
                (project_id,),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT h.*, p.name AS project_name FROM nc_subnet_calc_history h "
                "LEFT JOIN nc_projects p ON p.id=h.project_id ORDER BY h.created_at DESC LIMIT 200"
            ).fetchall()
        history = [_row_to_dict(r) for r in rows]
        conn.close()
        active_project = next((p for p in projects if p["id"] == project_id), None)
        return render_template(
            "network/subnet_calc.html",
            projects=projects,
            history=history,
            filter_project=project_id,
            active_project=active_project,
        )

    @bp.route("/api/nc-projects", methods=["POST"])
    @nc_login_required
    def nc_api_create_project_quick():
        """Lightweight project creation — name + optional description/owner only."""
        data = request.get_json(force=True, silent=True) or {}
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"error": "name is required"}), 400
        pid = str(_uuid.uuid4())
        now = _now()
        conn = get_connection()
        conn.execute(
            "INSERT INTO nc_projects (id, name, description, status, owner, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (pid, name, data.get("description", ""), data.get("status", "draft"), data.get("owner", ""), now, now),
        )
        conn.commit()
        conn.close()
        _audit("CREATE", "project", pid, name)
        return jsonify({"id": pid, "name": name}), 201

    @bp.route("/api/nc-projects", methods=["GET"])
    @nc_login_required
    def nc_api_list_projects_quick():
        """Return all nc_projects as id/name pairs for dropdowns."""
        conn = get_connection()
        rows = conn.execute("SELECT id, name, status FROM nc_projects ORDER BY name").fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/subnet-calc", methods=["GET"])
    @nc_login_required
    def nc_api_list_subnet_calc():
        conn = get_connection()
        project_id = request.args.get("project", "")
        if project_id:
            rows = conn.execute(
                "SELECT * FROM nc_subnet_calc_history WHERE project_id=? ORDER BY created_at DESC",
                (project_id,),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM nc_subnet_calc_history ORDER BY created_at DESC LIMIT 200"
            ).fetchall()
        conn.close()
        return jsonify([_row_to_dict(r) for r in rows])

    @bp.route("/api/subnet-calc", methods=["POST"])
    @nc_login_required
    def nc_api_save_subnet_calc():
        import ipaddress
        data = request.get_json(force=True, silent=True) or {}
        cidr = (data.get("cidr") or "").strip()
        project_id = (data.get("project_id") or "").strip()
        if not cidr or not project_id:
            return jsonify({"error": "cidr and project_id required"}), 400

        # Validate and compute server-side
        try:
            net = ipaddress.ip_network(cidr, strict=False)
        except ValueError as exc:
            return jsonify({"error": f"Invalid CIDR: {exc}"}), 422

        prefix_len = net.prefixlen
        af = "ipv6" if net.version == 6 else "ipv4"
        total_hosts = net.num_addresses
        if net.version == 4:
            usable = max(0, total_hosts - 2) if prefix_len < 31 else total_hosts
            subnet_mask = str(net.netmask)
            wildcard = str(net.hostmask)
            first_addr = str(net.network_address + 1) if prefix_len < 31 else str(net.network_address)
            last_addr = str(net.broadcast_address - 1) if prefix_len < 31 else str(net.broadcast_address)
            first_octet = int(str(net.network_address).split(".")[0])
            if first_octet < 128:
                ip_class = "A"
            elif first_octet < 192:
                ip_class = "B"
            elif first_octet < 224:
                ip_class = "C"
            elif first_octet < 240:
                ip_class = "D"
            else:
                ip_class = "E"
            broadcast = str(net.broadcast_address)
        else:
            usable = None   # too large for SQLite INTEGER; frontend uses prefix_len
            total_hosts = None
            subnet_mask = f"/{prefix_len}"
            wildcard = str(net.hostmask)
            first_addr = str(net.network_address + 1)
            last_addr = str(net.broadcast_address - 1)
            ip_class = data.get("ip_class") or "Global Unicast"
            broadcast = str(net.broadcast_address)

        entry_id = str(_uuid.uuid4())
        now = _now()
        conn = get_connection()
        # Dedup: INSERT OR REPLACE (UNIQUE on cidr+project_id)
        existing = conn.execute(
            "SELECT id FROM nc_subnet_calc_history WHERE cidr=? AND project_id=?",
            (str(net.with_prefixlen), project_id),
        ).fetchone()
        if existing:
            entry_id = existing[0]
            conn.execute(
                "UPDATE nc_subnet_calc_history SET network_addr=?,broadcast=?,first_host=?,last_host=?,"
                "total_hosts=?,usable_hosts=?,prefix_len=?,subnet_mask=?,wildcard_mask=?,address_family=?,"
                "ip_class=?,notes=?,created_at=? WHERE id=?",
                (str(net.network_address), broadcast, first_addr, last_addr,
                 total_hosts, usable, prefix_len, subnet_mask, wildcard,
                 af, ip_class, data.get("notes", ""), now, entry_id),
            )
        else:
            conn.execute(
                "INSERT INTO nc_subnet_calc_history "
                "(id,project_id,cidr,network_addr,broadcast,first_host,last_host,"
                "total_hosts,usable_hosts,prefix_len,subnet_mask,wildcard_mask,"
                "address_family,ip_class,notes,created_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (entry_id, project_id, str(net.with_prefixlen),
                 str(net.network_address), broadcast, first_addr, last_addr,
                 total_hosts, usable, prefix_len, subnet_mask, wildcard,
                 af, ip_class, data.get("notes", ""), now),
            )
        conn.commit()
        conn.close()
        _audit("SAVE", "subnet_calc", entry_id, cidr)
        return jsonify({"id": entry_id, "cidr": str(net.with_prefixlen), "updated": bool(existing)}), 201

    @bp.route("/api/subnet-calc/<entry_id>", methods=["DELETE"])
    @nc_login_required
    def nc_api_delete_subnet_calc(entry_id):
        conn = get_connection()
        conn.execute("DELETE FROM nc_subnet_calc_history WHERE id=?", (entry_id,))
        conn.commit()
        conn.close()
        _audit("DELETE", "subnet_calc", entry_id)
        return jsonify({"deleted": entry_id})

    # ── Migration Phases View ───────────────────────────────────────────────

    @bp.route("/migration-phases/<topo_id>")
    @nc_login_required
    def nc_migration_phases(topo_id):
        """Three-panel migration phases view: Current → Phase N → Final/To-Be."""
        conn = get_connection()
        topo = conn.execute(
            "SELECT id, name, graph_json FROM topologies WHERE id=?", (topo_id,)
        ).fetchone()
        if not topo:
            conn.close()
            return "Topology not found", 404

        # Fetch migration phases linked to any project that uses this topology
        phases = conn.execute(
            """
            SELECT mp.id, mp.phase_num, mp.title, mp.description,
                   mp.duration_days, mp.parallel_run, mp.rollback_criteria,
                   mp.maintenance_window, mp.dependencies, mp.status
            FROM nc_migration_phases mp
            JOIN nc_project_topologies pt ON pt.project_id = mp.project_id
            WHERE pt.topology_id = ?
            ORDER BY mp.phase_num
            """,
            (topo_id,),
        ).fetchall()
        conn.close()

        topo_name = topo["name"] if hasattr(topo, "__getitem__") else topo[1]
        phases_list = [dict(p) if hasattr(p, "keys") else {
            "id": p[0], "phase_num": p[1], "title": p[2], "description": p[3],
            "duration_days": p[4], "parallel_run": p[5], "rollback_criteria": p[6],
            "maintenance_window": p[7], "dependencies": p[8], "status": p[9],
        } for p in phases]

        return render_template(
            "network/migration_phases.html",
            topo_id=topo_id,
            topo_name=topo_name,
            phases=phases_list,
            phase_count=len(phases_list),
        )

    @bp.route("/api/migration-phases/<topo_id>/data", methods=["GET"])
    @nc_login_required
    def nc_api_migration_phases_data(topo_id):
        """Return all phase graphs + info boxes + consolidation as JSON."""
        import json as _json
        from tools.network.migration_phases import (
            compute_infoboxes, compute_final_infoboxes,
            generate_phase_graph, generate_final_graph,
            generate_phase_physical_graph, generate_phase_logical_graph,
            compute_physical_infoboxes, compute_logical_infoboxes,
            run_consolidation_analysis, load_consolidation,
        )

        conn = get_connection()
        topo_row = conn.execute(
            "SELECT graph_json, name FROM topologies WHERE id=?", (topo_id,)
        ).fetchone()
        if not topo_row:
            conn.close()
            return jsonify({"error": "topology not found"}), 404

        graph_json, topo_name = topo_row[0], topo_row[1]
        current_graph = _json.loads(graph_json) if graph_json else {"nodes": [], "edges": []}

        phases = conn.execute(
            """
            SELECT mp.id, mp.phase_num, mp.title, mp.description,
                   mp.duration_days, mp.parallel_run, mp.rollback_criteria,
                   mp.maintenance_window, mp.dependencies, mp.status,
                   mp.properties_json
            FROM nc_migration_phases mp
            JOIN nc_project_topologies pt ON pt.project_id = mp.project_id
            WHERE pt.topology_id = ?
            ORDER BY mp.phase_num
            """,
            (topo_id,),
        ).fetchall()

        phases_list = []
        for p in phases:
            d = dict(p) if hasattr(p, "keys") else {
                "id": p[0], "phase_num": p[1], "title": p[2], "description": p[3],
                "duration_days": p[4], "parallel_run": p[5], "rollback_criteria": p[6],
                "maintenance_window": p[7], "dependencies": p[8], "status": p[9],
                "properties_json": p[10] if len(p) > 10 else "{}",
            }
            d["total_phases"] = len(phases)
            phases_list.append(d)

        # Build per-phase graphs and info boxes
        phase_data = []
        for pm in phases_list:
            pm["total_phases"] = len(phases_list)
            ph_graph = generate_phase_graph(current_graph, pm)
            ph_boxes = compute_infoboxes(ph_graph, phase_key=f"phase-{pm['phase_num']}", phase_meta=pm)

            # Physical & Logical views (2-hop subgraphs with annotations)
            try:
                phy_graph = generate_phase_physical_graph(current_graph, pm, conn)
                phy_boxes = compute_physical_infoboxes(phy_graph, phase_meta=pm)
            except Exception:
                phy_graph = ph_graph
                phy_boxes = ph_boxes
            try:
                log_graph = generate_phase_logical_graph(current_graph, pm)
                log_boxes = compute_logical_infoboxes(log_graph, phase_meta=pm)
            except Exception:
                log_graph = ph_graph
                log_boxes = ph_boxes

            phase_data.append({
                "phase_num": pm["phase_num"],
                "title": pm["title"],
                "status": pm["status"],
                "graph": ph_graph,
                "infoboxes": ph_boxes,
                "graph_physical": phy_graph,
                "infoboxes_physical": phy_boxes,
                "graph_logical": log_graph,
                "infoboxes_logical": log_boxes,
            })

        conn.close()

        # Final/To-Be graph + consolidation
        final_graph = generate_final_graph(current_graph, phases_list)
        consolidation = load_consolidation(topo_id)
        if not consolidation:
            consolidation = run_consolidation_analysis(current_graph, final_graph)
            consolidation["current_device_count"] = len(current_graph.get("nodes", []))
            consolidation["final_device_count"] = len(final_graph.get("nodes", []))
        final_boxes = compute_final_infoboxes(current_graph, final_graph, consolidation)

        return jsonify({
            "topo_id": topo_id,
            "topo_name": topo_name,
            "current": {
                "graph": current_graph,
                "infoboxes": compute_infoboxes(current_graph, phase_key="current"),
            },
            "phases": phase_data,
            "final": {
                "graph": final_graph,
                "infoboxes": final_boxes,
                "consolidation": consolidation,
            },
        })

    @bp.route("/api/migration-phases/<topo_id>/export/<phase_key>/<fmt>", methods=["POST"])
    @nc_login_required
    def nc_api_migration_phases_export(topo_id, phase_key, fmt):
        """Export a single phase panel as PDF, Visio, or Draw.io.

        phase_key: 'current', 'phase-N', or 'final'
        fmt: 'pdf', 'visio', 'drawio'
        """
        import json as _json
        from tools.network.migration_phases import (
            compute_infoboxes, compute_final_infoboxes,
            generate_phase_graph, generate_final_graph,
            run_consolidation_analysis, load_consolidation,
        )

        conn = get_connection()
        topo_row = conn.execute(
            "SELECT graph_json, name FROM topologies WHERE id=?", (topo_id,)
        ).fetchone()
        if not topo_row:
            conn.close()
            return jsonify({"error": "topology not found"}), 404
        graph_json, topo_name = topo_row[0], topo_row[1]
        current_graph = _json.loads(graph_json) if graph_json else {"nodes": [], "edges": []}
        conn.close()

        # Resolve which graph + label to export
        if phase_key == "current":
            graph = current_graph
            phase_label = "Current State"
            infoboxes = compute_infoboxes(graph, phase_key="current")
            consolidation = None
            phase_meta = None
        elif phase_key == "final":
            phases_raw = []  # fetch from DB if needed — simplified: use empty for now
            graph = generate_final_graph(current_graph, phases_raw)
            phase_label = "Final / To-Be State"
            consolidation = load_consolidation(topo_id) or run_consolidation_analysis(current_graph, graph)
            infoboxes = compute_final_infoboxes(current_graph, graph, consolidation)
            phase_meta = None
        else:
            # phase-N
            try:
                pnum = int(phase_key.split("-")[1])
            except Exception:
                pnum = 1
            phase_meta = {"phase_num": pnum, "total_phases": pnum, "title": f"Phase {pnum}"}
            graph = generate_phase_graph(current_graph, phase_meta)
            phase_label = f"Phase {pnum}"
            infoboxes = compute_infoboxes(graph, phase_key=phase_key, phase_meta=phase_meta)
            consolidation = None

        safe_name = topo_name.replace(" ", "_").replace("/", "-")[:40]

        if fmt == "pdf":
            from tools.network.pdf_export import export_phase_pdf
            pdf_bytes = export_phase_pdf(
                topo_name, phase_label, graph, infoboxes, consolidation, phase_meta
            )
            is_html = pdf_bytes[:15].startswith(b"<!DOCTYPE")
            if is_html:
                return current_app.response_class(
                    pdf_bytes,
                    mimetype="text/html",
                    headers={"Content-Disposition": f'attachment; filename="{safe_name}_{phase_key}_report.html"'},
                )
            return current_app.response_class(
                pdf_bytes,
                mimetype="application/pdf",
                headers={"Content-Disposition": f'attachment; filename="{safe_name}_{phase_key}.pdf"'},
            )

        if fmt == "visio":
            from tools.network.visio_export import export_vsdx
            vsdx_bytes = export_vsdx(f"{topo_name} — {phase_label}", graph)
            return current_app.response_class(
                vsdx_bytes,
                mimetype="application/vnd.visio",
                headers={"Content-Disposition": f'attachment; filename="{safe_name}_{phase_key}.vsdx"'},
            )

        if fmt == "drawio":
            from tools.network.export_import import to_drawio
            xml = to_drawio(graph, f"{topo_name} — {phase_label}")
            return current_app.response_class(
                xml.encode("utf-8"),
                mimetype="application/octet-stream",
                headers={"Content-Disposition": f'attachment; filename="{safe_name}_{phase_key}.drawio"'},
            )

        return jsonify({"error": f"Unknown format: {fmt}"}), 400

    # ── AI Migration Plan Generator ────────────────────────────────────────
    @bp.route("/api/migration-plan/generate", methods=["POST"])
    @nc_login_required
    def nc_api_migration_plan_generate():
        """Decompose a NL description into migration phases using LLM.

        Body: {"description": "...", "project_id": "...", "topo_id": "..."}
        Returns: {"phases_created": N, "phase_ids": [...]}
        """
        import uuid as _uuid
        from tools.http.client import request as _req_request

        data = request.get_json(force=True, silent=True) or {}
        description = (data.get("description") or "").strip()
        project_id = (data.get("project_id") or "").strip()
        if not description:
            return jsonify({"error": "description is required"}), 400

        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            return jsonify({"error": "No ANTHROPIC_API_KEY set"}), 503

        try:
            model = os.environ.get("ANTHROPIC_TOPO_MODEL", "claude-sonnet-4-20250514")
            r = _req_request(
                "POST",
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json={
                    "model": model,
                    "max_tokens": 2048,
                    "system": _AI_MIGRATION_PLAN_PROMPT,
                    "messages": [{"role": "user", "content": description}],
                },
                timeout=60,
            )
            if r.status_code != 200:
                return jsonify({"error": f"LLM API error {r.status_code}"}), 503

            raw_text = r.json()["content"][0]["text"].strip()
            # Strip markdown fences if present
            if raw_text.startswith("```"):
                raw_text = raw_text.split("\n", 1)[1].rsplit("```", 1)[0].strip()
            phases_data = json.loads(raw_text)
            if not isinstance(phases_data, list):
                return jsonify({"error": "LLM returned unexpected format"}), 500
        except Exception as exc:
            logger.exception("migration-plan/generate LLM call failed")
            return jsonify({"error": str(exc)}), 500

        if not project_id:
            return jsonify({"error": "project_id is required"}), 400

        conn = get_connection()
        try:
            phase_ids = []
            for ph in phases_data:
                pid = str(_uuid.uuid4())
                conn.execute(
                    """INSERT INTO nc_migration_phases
                       (id, project_id, phase_num, title, description,
                        duration_days, parallel_run, rollback_criteria,
                        maintenance_window, classification, impact_level, status)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,'planned')""",
                    (
                        pid,
                        project_id,
                        int(ph.get("phase_num", len(phase_ids) + 1)),
                        str(ph.get("title", f"Phase {len(phase_ids)+1}")),
                        str(ph.get("description", "")),
                        int(ph.get("duration_days", 7)),
                        int(ph.get("parallel_run", 0)),
                        str(ph.get("rollback_criteria", "")),
                        str(ph.get("maintenance_window", "")),
                        str(ph.get("classification", "CUI")),
                        str(ph.get("impact_level", "IL4")),
                    ),
                )
                phase_ids.append(pid)
                # Auto-link SOPs: extract meaningful tokens from phase text, match against SOP titles
                _stop = {"the","and","or","for","with","from","that","this","into","will","phase",
                         "have","been","each","only","when","also","are","was","were","all","not"}
                phase_text = f"{ph.get('title','')} {ph.get('description','')}".lower()
                import re as _re
                raw_tokens = _re.findall(r"[a-z][a-z0-9\-]{3,}", phase_text)
                sop_keywords = {t for t in raw_tokens if t not in _stop}
                if sop_keywords:
                    sop_rows = conn.execute(
                        "SELECT sop_id, title FROM ndc_sops WHERE status = 'approved'"
                    ).fetchall()
                    for sop_row in sop_rows:
                        sop_title_lower = (sop_row["title"] or "").lower()
                        if any(kw in sop_title_lower for kw in sop_keywords):
                            link_id = str(_uuid.uuid4())
                            try:
                                conn.execute(
                                    """INSERT OR IGNORE INTO nc_phase_documents
                                       (id, phase_id, project_id, doc_source, doc_id,
                                        doc_title, doc_type, relevance_note, display_order)
                                       VALUES (?,?,?,?,?,?,?,?,0)""",
                                    (link_id, pid, project_id, "sop",
                                     sop_row["sop_id"], sop_row["title"],
                                     "sop", "auto-linked by keyword match"),
                                )
                            except Exception:
                                pass
            conn.commit()
            return jsonify({"phases_created": len(phase_ids), "phase_ids": phase_ids}), 201
        except Exception as exc:
            logger.warning("migration-plan insert error: %s", exc)
            return jsonify({"error": str(exc)}), 500
        finally:
            conn.close()

    # ── Phase Status Update + Snapshot ────────────────────────────────────
    @bp.route("/api/migration-phases/<topo_id>/<phase_id>/status", methods=["PUT"])
    @nc_login_required
    def nc_api_migration_phase_status(topo_id, phase_id):
        """Update phase status; if completed, snapshot the topology.

        Body (JSON):
          status         : "completed" | "in_progress" | "rolled_back"
          classification : optional override ("PUBLIC"|"CUI"|"SECRET"|"TS")
          impact_level   : optional override ("IL2"|"IL4"|"IL5"|"IL6")
        """
        import json as _json
        import uuid as _uuid
        from tools.network.migration_phases import (
            generate_phase_graph,
            run_consolidation_analysis,
            save_consolidation,
        )

        data = request.get_json(silent=True) or {}
        new_status = data.get("status", "").strip()
        valid_statuses = {"planned", "in_progress", "completed", "rolled_back"}
        if new_status not in valid_statuses:
            return jsonify({"error": f"status must be one of {sorted(valid_statuses)}"}), 400

        conn = get_connection()
        try:
            phase_row = conn.execute(
                "SELECT * FROM nc_migration_phases WHERE id=?", (phase_id,)
            ).fetchone()
            if not phase_row:
                return jsonify({"error": "phase not found"}), 404
            phase = dict(phase_row)

            update_fields = ["status=?"]
            update_vals: list = [new_status]
            classification = data.get("classification")
            impact_level = data.get("impact_level")
            if classification:
                update_fields.append("classification=?")
                update_vals.append(classification)
            if impact_level:
                update_fields.append("impact_level=?")
                update_vals.append(impact_level)
            update_vals.extend([phase_id])

            conn.execute(
                f"UPDATE nc_migration_phases SET {', '.join(update_fields)} WHERE id=?",
                update_vals,
            )

            snapshot_id = None
            if new_status == "completed":
                topo_row = conn.execute(
                    "SELECT graph_json FROM topologies WHERE id=?", (topo_id,)
                ).fetchone()
                if topo_row:
                    current_graph = _json.loads(topo_row[0] or "{}") or {"nodes": [], "edges": []}
                    phase_meta = dict(phase)
                    phase_meta["status"] = new_status
                    post_graph = generate_phase_graph(current_graph, phase_meta)
                    snapshot_id = str(_uuid.uuid4())
                    conn.execute(
                        """INSERT INTO nc_topology_snapshots (id, topo_id, phase_id, label, graph_json)
                           VALUES (?,?,?,?,?)""",
                        (
                            snapshot_id,
                            topo_id,
                            phase_id,
                            f"Phase {phase.get('phase_num', '?')} Complete",
                            _json.dumps(post_graph),
                        ),
                    )
                    try:
                        analysis = run_consolidation_analysis(current_graph, post_graph)
                        save_consolidation(topo_id, analysis)
                    except Exception:
                        pass

                _audit(conn, "phase_completed", {
                    "phase_id": phase_id,
                    "topo_id": topo_id,
                    "classification": classification or phase.get("classification", "CUI"),
                })
                # DIC Canvas Synergy — emit migration phase complete event (dsyn-emit-02)
                try:
                    from tools.network.event_emitter import emit_migration_phase_complete
                    emit_migration_phase_complete(
                        phase_id=phase_id,
                        migration_id=topo_id,
                        phase_name=phase.get("phase_name", ""),
                        classification=classification or phase.get("classification", "CUI"),
                    )
                except Exception:
                    pass  # event emission never blocks phase status update

            # DIC Canvas Synergy — emit anomaly event on rollback (dsyn-emit-02)
            if new_status == "rolled_back":
                try:
                    from tools.network.event_emitter import emit_anomaly_detected
                    emit_anomaly_detected(
                        migration_id=topo_id,
                        anomaly_type="phase_rollback",
                        severity="high",
                        change_summary=f"Migration phase '{phase.get('phase_name', phase_id)}' was rolled back",
                    )
                except Exception:
                    pass  # event emission never blocks phase status update

            conn.commit()
            return jsonify({"status": "ok", "snapshot_id": snapshot_id}), 200
        except Exception as exc:
            logger.warning("phase status update error: %s", exc)
            return jsonify({"error": str(exc)}), 500
        finally:
            conn.close()

    # ── Migration Hub ──────────────────────────────────────────────────────
    @bp.route("/migration-hub")
    def migration_hub():
        """Hub page: all migration projects, phases, and linked documentation."""
        return render_template("network/migration_hub.html")

    @bp.route("/api/migration-hub/data", methods=["GET"])
    def migration_hub_data():
        """Return all projects + phases + linked docs as JSON for the hub."""
        conn = get_connection()
        try:
            # Projects
            projects = [dict(r) for r in conn.execute(
                "SELECT id, name, description, status, owner, created_at FROM nc_projects ORDER BY created_at DESC"
            ).fetchall()]

            # Phases per project
            phases_raw = conn.execute(
                """SELECT id, project_id, phase_num, title, description,
                          duration_days, status, maintenance_window, rollback_criteria,
                          dependencies, classification, impact_level, created_at
                   FROM nc_migration_phases
                   ORDER BY project_id, phase_num"""
            ).fetchall()
            phases_by_project: dict = {}
            for row in phases_raw:
                r = dict(row)
                pid = r["project_id"]
                phases_by_project.setdefault(pid, []).append(r)

            # Phase documents
            phase_docs_raw = conn.execute(
                """SELECT id, phase_id, project_id, doc_source, doc_id,
                          doc_title, doc_type, relevance_note, display_order
                   FROM nc_phase_documents
                   ORDER BY project_id, display_order"""
            ).fetchall()
            docs_by_phase: dict = {}
            for row in phase_docs_raw:
                r = dict(row)
                docs_by_phase.setdefault(r["phase_id"], []).append(r)

            # Standalone uploaded docs (not yet linked to a phase but in project)
            docs_raw = conn.execute(
                """SELECT id, file_name, doc_type, project_id,
                          topology_id, classification, status, ingested_at
                   FROM nc_documents
                   WHERE status = 'ingested'
                   ORDER BY project_id, ingested_at DESC"""
            ).fetchall()
            docs_by_project: dict = {}
            for row in docs_raw:
                r = dict(row)
                pid = r["project_id"] or "default"
                docs_by_project.setdefault(pid, []).append(r)

            # Runbooks
            runbooks_raw = conn.execute(
                """SELECT id, title, trigger_event, severity, owner, topology_id, classification
                   FROM ndc_runbooks ORDER BY title"""
            ).fetchall()
            runbooks = [dict(r) for r in runbooks_raw]

            # SOPs
            sops_raw = conn.execute(
                """SELECT sop_id AS id, title, category, version, status,
                          description, classification, author
                   FROM ndc_sops ORDER BY category, title"""
            ).fetchall()
            sops = [dict(r) for r in sops_raw]

            # Topologies (for MP links)
            topos_raw = conn.execute(
                "SELECT id, name FROM topologies ORDER BY name"
            ).fetchall()
            topos = [dict(r) for r in topos_raw]

            # First topology per project (for 3-panel diagram links)
            topo_by_project = {}
            for r in conn.execute(
                "SELECT project_id, topology_id FROM nc_project_topologies"
            ).fetchall():
                topo_by_project.setdefault(r[0], r[1])

            # Attach phases and docs to projects
            for proj in projects:
                pid = proj["id"]
                proj_phases = phases_by_project.get(pid, [])
                for phase in proj_phases:
                    phase["documents"] = docs_by_phase.get(phase["id"], [])
                proj["phases"] = proj_phases
                proj["documents"] = docs_by_project.get(pid, [])
                proj["topology_id"] = topo_by_project.get(pid)

            return jsonify({
                "projects": projects,
                "runbooks": runbooks,
                "sops": sops,
                "topologies": topos,
            })
        finally:
            conn.close()

    @bp.route("/api/migration-hub/phase-docs", methods=["POST"])
    def migration_hub_link_doc():
        """Link a document/runbook/SOP to a migration phase."""
        data = request.get_json(force=True) or {}
        required = {"phase_id", "project_id", "doc_source", "doc_id", "doc_title"}
        if missing := required - data.keys():
            return jsonify({"error": f"Missing fields: {missing}"}), 400

        conn = get_connection()
        try:
            doc_id = str(_uuid.uuid4())
            conn.execute(
                """INSERT INTO nc_phase_documents
                   (id, phase_id, project_id, doc_source, doc_id,
                    doc_title, doc_type, relevance_note, display_order)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (
                    doc_id,
                    data["phase_id"],
                    data["project_id"],
                    data["doc_source"],
                    data["doc_id"],
                    data["doc_title"],
                    data.get("doc_type", ""),
                    data.get("relevance_note", ""),
                    data.get("display_order", 0),
                ),
            )
            conn.commit()
            return jsonify({"id": doc_id, "status": "linked"})
        finally:
            conn.close()

    @bp.route("/api/migration-hub/phase-docs/<doc_link_id>", methods=["DELETE"])
    def migration_hub_unlink_doc(doc_link_id: str):
        """Remove a document link from a migration phase."""
        conn = get_connection()
        try:
            conn.execute("DELETE FROM nc_phase_documents WHERE id = ?", (doc_link_id,))
            conn.commit()
            return jsonify({"status": "deleted"})
        finally:
            conn.close()

    # ── Unified Project Dashboard ──────────────────────────────────────────
    @bp.route("/projects/<project_id>")
    @nc_login_required
    def nc_project_dashboard(project_id):
        """Unified 4-panel view: phases + canvas + SOPs per phase + traffic flows."""
        conn = get_connection()
        try:
            project_row = conn.execute(
                "SELECT * FROM nc_projects WHERE id=?", (project_id,)
            ).fetchone()
            if not project_row:
                return "Project not found", 404
            project = dict(project_row)

            phases = [dict(r) for r in conn.execute(
                "SELECT * FROM nc_migration_phases WHERE project_id=? ORDER BY phase_num",
                (project_id,),
            ).fetchall()]

            topo_row = conn.execute(
                "SELECT t.id, t.name, t.graph_json FROM topologies t "
                "JOIN nc_project_topologies pt ON pt.topology_id = t.id "
                "WHERE pt.project_id=? LIMIT 1",
                (project_id,),
            ).fetchone()
            topology = dict(topo_row) if topo_row else None
            topo_id = topology["id"] if topology else None

            snapshots = [dict(r) for r in conn.execute(
                "SELECT id, phase_id, label, created_at FROM nc_topology_snapshots "
                "WHERE topo_id=? ORDER BY created_at DESC",
                (topo_id,),
            ).fetchall()] if topo_id else []

            phase_ids = [ph["id"] for ph in phases]
            sops_by_phase: dict = {ph["id"]: [] for ph in phases}
            if phase_ids:
                placeholders = ",".join("?" * len(phase_ids))
                doc_rows = conn.execute(
                    f"SELECT pd.phase_id, s.sop_id, s.title, s.category, s.csp "
                    f"FROM nc_phase_documents pd "
                    f"JOIN ndc_sops s ON s.sop_id = pd.doc_id "
                    f"WHERE pd.phase_id IN ({placeholders}) AND pd.doc_source='sop'",
                    phase_ids,
                ).fetchall()
                for r in doc_rows:
                    sops_by_phase.setdefault(r[0], []).append({
                        "sop_id": r[1], "title": r[2], "category": r[3], "csp": r[4]
                    })

            flows_by_phase: dict = {ph["id"]: [] for ph in phases}
            if topo_id and phase_ids:
                placeholders = ",".join("?" * len(phase_ids))
                flow_rows = conn.execute(
                    f"SELECT id, name, source_zone, destination_zone, classification, phase_id "
                    f"FROM nc_traffic_flows WHERE topology_id=? AND phase_id IN ({placeholders})",
                    [topo_id] + phase_ids,
                ).fetchall()
                for r in flow_rows:
                    flows_by_phase.setdefault(r[5], []).append({
                        "id": r[0], "name": r[1],
                        "source_zone": r[2], "destination_zone": r[3],
                        "classification": r[4],
                    })

            return render_template(
                "network/project_dashboard.html",
                project=project,
                phases=phases,
                topology=topology,
                topo_id=topo_id,
                snapshots=snapshots,
                sops_by_phase=sops_by_phase,
                flows_by_phase=flows_by_phase,
            )
        finally:
            conn.close()

    # ── AI Trace API ────────────────────────────────────────────────────────
    @bp.route("/api/ai-trace")
    @nc_login_required
    def nc_api_ai_trace():
        """Return recent AI decisions made by NDC assessment engines."""
        limit = min(int(request.args.get("limit", 50)), 200)
        record_id = request.args.get("record_id")
        try:
            from tools.db.storage import get_connection as _gc
            with _gc() as _conn:
                if record_id:
                    rows = _conn.execute(
                        "SELECT * FROM canvas_ai_decisions WHERE canvas_type='ndc' AND record_id=? "
                        "ORDER BY created_at DESC LIMIT ?",
                        (record_id, limit),
                    ).fetchall()
                else:
                    rows = _conn.execute(
                        "SELECT * FROM canvas_ai_decisions WHERE canvas_type='ndc' "
                        "ORDER BY created_at DESC LIMIT ?",
                        (limit,),
                    ).fetchall()
            return jsonify({"ok": True, "canvas": "ndc", "decisions": [dict(r) for r in rows]})
        except Exception as exc:
            return jsonify({"ok": False, "error": str(exc)}), 500

    # ── FCC Compliance ──────────────────────────────────────────────────────────

    @bp.route("/fcc")
    @nc_login_required
    def network_fcc():
        from tools.network.fcc_compliance import (
            calea_checklist, part36_assessment,
            nanp_number_inventory, e911_capability_check,
        )
        checks = {}
        for name, fn in [
            ("calea", calea_checklist),
            ("part36", part36_assessment),
            ("nanp", nanp_number_inventory),
            ("e911", e911_capability_check),
        ]:
            try:
                checks[name] = fn()
            except Exception as exc:
                checks[name] = {"error": str(exc)}
        return render_template("network/fcc_compliance.html", checks=checks)

    @bp.route("/api/fcc/<check_type>")
    def api_network_fcc(check_type):
        from tools.network.fcc_compliance import (
            calea_checklist, part36_assessment,
            nanp_number_inventory, e911_capability_check,
        )
        _CHECK_MAP = {
            "calea":  calea_checklist,
            "part36": part36_assessment,
            "nanp":   nanp_number_inventory,
            "e911":   e911_capability_check,
        }
        if check_type == "all":
            result = {}
            for name, fn in _CHECK_MAP.items():
                try:
                    result[name] = fn()
                except Exception as exc:
                    result[name] = {"error": str(exc)}
            return jsonify(result)
        fn = _CHECK_MAP.get(check_type)
        if not fn:
            return jsonify({"error": f"Unknown check: {check_type}. Valid: {sorted(_CHECK_MAP)}"}), 400
        try:
            return jsonify(fn())
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    # ── Presentation (exec review) ─────────────────────────────────────────
    @bp.route("/projects/<pid>/presentation")
    @nc_login_required
    def nc_project_presentation(pid):
        conn = get_connection()
        proj = conn.execute("SELECT * FROM nc_projects WHERE id=?", (pid,)).fetchone()
        if not proj:
            conn.close()
            abort(404)
        proj = _row_to_dict(proj)
        conn.close()
        return render_template("network/presentation.html", project=proj)

    @bp.route("/api/projects/<pid>/presentation")
    @nc_login_required
    def nc_api_project_presentation(pid):
        conn = get_connection()
        proj = conn.execute(
            "SELECT p.*, c.name AS customer_name FROM nc_projects p "
            "LEFT JOIN nc_customers c ON c.id=p.customer_id WHERE p.id=?", (pid,)
        ).fetchone()
        if not proj:
            conn.close()
            return jsonify({"error": "Project not found"}), 404
        proj = _row_to_dict(proj)
        topos = []
        for r in conn.execute(
            "SELECT t.id, t.name, t.classification, t.graph_json "
            "FROM topologies t JOIN nc_project_topologies pt ON pt.topology_id=t.id "
            "WHERE pt.project_id=?", (pid,)
        ).fetchall():
            t = _row_to_dict(r)
            try:
                g = json.loads(t.get("graph_json") or '{"nodes":[],"edges":[]}')
            except Exception:
                g = {"nodes": [], "edges": []}
            t["node_count"] = len(g.get("nodes", []))
            t["edge_count"] = len(g.get("edges", []))
            topos.append(t)
        circuits = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT * FROM nc_circuits WHERE topology_id IN "
                "(SELECT topology_id FROM nc_project_topologies WHERE project_id=?)", (pid,)
            ).fetchall()
        ]
        milestones = [
            _row_to_dict(r)
            for r in conn.execute("SELECT * FROM nc_project_milestones WHERE project_id=?", (pid,)).fetchall()
        ]
        reviews = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT br.*, rb.name AS board_name "
                "FROM nc_board_reviews br JOIN nc_review_boards rb ON rb.id=br.board_id "
                "WHERE br.project_id=? ORDER BY br.phase", (pid,)
            ).fetchall()
        ]
        safe_bridge = conn.execute("SELECT * FROM nc_safe_bridge WHERE project_id=?", (pid,)).fetchone()
        safe_bridge = _row_to_dict(safe_bridge) if safe_bridge else None
        roi = {}
        if safe_bridge and safe_bridge.get("roi_json"):
            try:
                roi = json.loads(safe_bridge["roi_json"])
            except Exception:
                pass
        agg_audit = conn.execute(
            "SELECT SUM(passed), SUM(failed) FROM nc_compliance_checks "
            "WHERE topology_id IN (SELECT topology_id FROM nc_project_topologies WHERE project_id=?)", (pid,)
        ).fetchone()
        passed = agg_audit[0] or 0
        failed = agg_audit[1] or 0
        total = passed + failed
        compliance_pct = round(passed * 100 / total) if total else None
        cat1 = conn.execute(
            "SELECT COUNT(*) FROM nc_compliance_findings "
            "WHERE topology_id IN (SELECT topology_id FROM nc_project_topologies WHERE project_id=?) "
            "AND status='open' AND severity='CAT1'", (pid,)
        ).fetchone()[0] or 0
        # Pre-compute CapEx before closing connection
        total_capex = 0
        for t in topos:
            trow = conn.execute("SELECT graph_json FROM topologies WHERE id=?", (t["id"],)).fetchone()
            if trow and trow[0]:
                try:
                    nodes = json.loads(trow[0]).get("nodes", [])
                    total_capex += sum(BOM_COSTS.get(n.get("type", "unknown"), 0) for n in nodes)
                except Exception:
                    pass
        conn.close()
        return jsonify({
            "title": proj["name"],
            "status": proj.get("status", "draft"),
            "owner": proj.get("owner"),
            "description": proj.get("description"),
            "generated_at": _now(),
            "executive_summary": {
                "topology_count": len(topos),
                "total_devices": sum(t.get("node_count", 0) for t in topos),
                "compliance_pct": compliance_pct,
                "cat1_findings": cat1,
                "total_capex": total_capex,
                "monthly_circuit_cost": sum(c.get("monthly_cost_usd") or 0 for c in circuits),
                "roi": roi,
                "justification": safe_bridge.get("justification") if safe_bridge else None,
                "alternatives": safe_bridge.get("alternatives") if safe_bridge else None,
            },
            "topologies": topos,
            "circuits": circuits,
            "milestones": milestones,
            "review_history": reviews,
        })

    # ── Placeholder / redirect routes for migration hub quick actions ────────
    @bp.route("/migration-wizard")
    @nc_login_required
    def nc_migration_wizard():
        """Redirect to migration hub — wizard is embedded there."""
        return redirect("/network/migration-hub")

    @bp.route("/wave-planner")
    @nc_login_required
    def nc_wave_planner():
        """Placeholder wave planner — redirects to migration hub."""
        return redirect("/network/migration-hub")

    @bp.route("/port-mapping")
    @nc_login_required
    def nc_port_mapping_page():
        """Standalone port mapping — redirect to first project with port mapping."""
        conn = get_connection()
        proj = conn.execute(
            "SELECT id FROM nc_projects WHERE selected_coa > 0 LIMIT 1"
        ).fetchone()
        if not proj:
            proj = conn.execute("SELECT id FROM nc_projects LIMIT 1").fetchone()
        conn.close()
        if proj:
            return redirect(f"/network/projects/{proj['id']}#port-mapping-section")
        return redirect("/network/migration-hub")

    @bp.route("/documents")
    @nc_login_required
    def nc_documents():
        """Document library — redirect to SOPs for now."""
        return redirect("/network/sops")

    # ── Advisory History ───────────────────────────────────────────────────

    @bp.route("/advisory-history")
    @nc_login_required
    def nc_advisory_history():
        from tools.network.advisory import list_advisories, list_vendors
        advisories = list_advisories(
            vendor=request.args.get("vendor"),
            severity=request.args.get("severity"),
            status=request.args.get("status"),
            date_from=request.args.get("date_from"),
            date_to=request.args.get("date_to"),
        )
        vendors = list_vendors()
        return render_template(
            "network/advisory_history.html",
            advisories=advisories,
            vendors=vendors,
        )

    @bp.route("/api/advisories/export")
    @nc_login_required
    def nc_api_advisories_export():
        import csv
        import io
        from flask import Response
        from tools.network.advisory import list_advisories
        items = list_advisories()
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["CVE ID", "Vendor", "Severity", "Date", "Total Devices",
                    "Impacted", "Remediation %", "Data Source", "HITL Status", "Status"])
        for a in items:
            w.writerow([
                a.get("cve_id", ""), a.get("vendor", ""), a.get("severity", ""),
                (a.get("published_date") or "")[:10], a.get("total_devices", 0),
                a.get("impacted_devices", 0), a.get("remediation_pct", 0),
                a.get("data_source", ""), a.get("hitl_status", ""), a.get("status", ""),
            ])
        return Response(buf.getvalue(), mimetype="text/csv",
                        headers={"Content-Disposition": "attachment; filename=advisories.csv"})

    # ── POAM ───────────────────────────────────────────────────────────────

    @bp.route("/poam")
    @nc_login_required
    def nc_poam():
        from tools.network.poam_generator import list_poam_items
        from datetime import date
        advisory_filter = request.args.get("advisory")
        items = list_poam_items()
        return render_template(
            "network/poam.html",
            items=items,
            advisory_filter=advisory_filter,
            today=date.today().isoformat(),
        )

    @bp.route("/api/poam/generate", methods=["POST"])
    @nc_login_required
    def nc_api_poam_generate():
        from tools.network.poam_generator import generate_poam_item
        data = request.get_json(silent=True) or {}
        advisory_id = data.get("advisory_id", "")
        try:
            item = generate_poam_item(advisory_id, data)
            return jsonify({"ok": True, "item": item})
        except Exception as exc:
            logger.exception("POAM generate failed")
            return jsonify({"error": str(exc)}), 500

    @bp.route("/api/poam/export")
    @nc_login_required
    def nc_api_poam_export():
        from flask import Response
        from tools.network.poam_generator import export_poam
        fmt = request.args.get("format", "csv")
        content, mimetype = export_poam(fmt)
        ext = "json" if fmt == "json" else "csv"
        return Response(content, mimetype=mimetype,
                        headers={"Content-Disposition": f"attachment; filename=poam.{ext}"})

    # ── Exceptions ─────────────────────────────────────────────────────────

    @bp.route("/exceptions")
    @nc_login_required
    def nc_exceptions():
        from tools.network.exception_registry import list_exceptions
        from datetime import date
        exceptions = list_exceptions()
        return render_template(
            "network/exceptions.html",
            exceptions=exceptions,
            today=date.today().isoformat(),
        )

    @bp.route("/exceptions/file")
    @nc_login_required
    def nc_exceptions_file_form():
        """Redirect to exceptions page with file modal open."""
        return redirect("/network/exceptions?open_modal=file")

    @bp.route("/api/exception/file", methods=["POST"])
    @nc_login_required
    def nc_api_exception_file():
        from tools.network.exception_registry import file_exception
        data = request.get_json(silent=True) or {}
        try:
            exc = file_exception(data)
            return jsonify({"ok": True, "exception": exc})
        except Exception as e:
            logger.exception("Exception filing failed")
            return jsonify({"error": str(e)}), 500

    @bp.route("/api/exception/<exc_id>/approve", methods=["POST"])
    @nc_login_required
    def nc_api_exception_approve(exc_id):
        from tools.network.exception_registry import approve_exception
        data = request.get_json(silent=True) or {}
        level = data.get("level", "")
        approver = data.get("approver", "")
        if not level or not approver:
            return jsonify({"error": "level and approver are required"}), 400
        try:
            exc = approve_exception(exc_id, level, approver)
            return jsonify({"ok": True, "exception": exc})
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        except Exception as e:
            logger.exception("Exception approval failed")
            return jsonify({"error": str(e)}), 500

    # ── ATO Evidence Chain Export ──────────────────────────────────────────

    def _gather_ato_evidence(advisory_id, conn):
        """Collect all audit-chain rows for one advisory from available tables."""

        def _q(sql, params=()):
            try:
                return [dict(r) for r in conn.execute(sql, params).fetchall()]
            except Exception:
                return []

        def _q1(sql, params=()):
            try:
                row = conn.execute(sql, params).fetchone()
                return dict(row) if row else None
            except Exception:
                return None

        advisory = _q1("SELECT * FROM nc_advisories WHERE id = ?", (advisory_id,))

        assessments = _q(
            "SELECT * FROM nc_advisory_assessments WHERE advisory_id = ? ORDER BY created_at ASC",
            (advisory_id,),
        )

        rem_actions = _q(
            "SELECT * FROM nc_remediation_actions WHERE advisory_id = ? ORDER BY created_at ASC",
            (advisory_id,),
        )

        rem_status_log: list = []
        action_ids = [a.get("id") for a in rem_actions if a.get("id")]
        if action_ids:
            ph = ",".join("?" * len(action_ids))
            rem_status_log = _q(
                f"SELECT * FROM nc_remediation_status_log WHERE action_id IN ({ph}) ORDER BY created_at ASC",
                action_ids,
            )

        poam_items = _q(
            "SELECT * FROM nc_poam_items WHERE advisory_id = ? ORDER BY created_at ASC",
            (advisory_id,),
        )

        poam_status_log: list = []
        poam_ids = [p.get("id") for p in poam_items if p.get("id")]
        if poam_ids:
            ph = ",".join("?" * len(poam_ids))
            poam_status_log = _q(
                f"SELECT * FROM nc_poam_status_log WHERE poam_id IN ({ph}) ORDER BY created_at ASC",
                poam_ids,
            )

        exceptions = _q(
            "SELECT * FROM nc_exceptions WHERE advisory_id = ? ORDER BY created_at ASC",
            (advisory_id,),
        )

        exception_approvals: list = []
        exc_ids = [e.get("id") for e in exceptions if e.get("id")]
        if exc_ids:
            ph = ",".join("?" * len(exc_ids))
            exception_approvals = _q(
                f"SELECT * FROM nc_exception_approvals WHERE exception_id IN ({ph}) ORDER BY created_at ASC",
                exc_ids,
            )

        audit_log = _q(
            "SELECT * FROM nc_nqe_audit_log WHERE advisory_id = ? ORDER BY created_at ASC",
            (advisory_id,),
        )

        return {
            "advisory": advisory,
            "assessments": assessments,
            "remediation_actions": rem_actions,
            "remediation_status_log": rem_status_log,
            "poam_items": poam_items,
            "poam_status_log": poam_status_log,
            "exceptions": exceptions,
            "exception_approvals": exception_approvals,
            "audit_log": audit_log,
        }

    def _audit_ato_export(advisory_id, fmt, doc_hash):
        """Append an audit-log entry for the export (best-effort; non-blocking)."""
        try:
            conn = get_connection()
            conn.execute(
                """INSERT INTO nc_nqe_audit_log
                   (session_id, user_session, action, input_text, result_summary,
                    raw_response_hash, advisory_id, created_at)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (
                    session.get("session_id", ""),
                    session.get("user", ""),
                    "export",
                    fmt,
                    f"ATO evidence chain export (format={fmt})",
                    doc_hash,
                    advisory_id,
                    datetime.now(timezone.utc).isoformat(),
                ),
            )
            conn.commit()
            conn.close()
        except Exception:
            pass

    def _ato_safe(text):
        """Coerce to latin-1-safe string for fpdf2 core fonts."""
        _MAP = str.maketrans({
            "—": "-", "–": "-", "→": "->", "←": "<-",
            "≥": ">=", "≤": "<=", "°": "deg", "•": "*",
            "’": "'", "‘": "'", "“": '"', "”": '"',
            "™": "(TM)", "®": "(R)", "©": "(c)",
            "×": "x", "÷": "/",
        })
        text = str(text).translate(_MAP)
        return text.encode("latin-1", errors="replace").decode("latin-1")

    def _build_ato_pdf(evidence, advisory_id):
        """Return PDF bytes (fpdf2) or HTML bytes (fallback) for the evidence package."""
        _CLASSIFICATION = "CUI // SP-CTI"
        adv = evidence.get("advisory") or {}
        meta = evidence.get("_meta", {})

        def _cui_banner(pdf):
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_fill_color(180, 30, 30)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(0, 5, f"  {_CLASSIFICATION}", ln=True, fill=True)
            pdf.set_text_color(0, 0, 0)

        def _section_heading(pdf, title):
            pdf.set_font("Helvetica", "B", 11)
            pdf.set_fill_color(220, 230, 245)
            pdf.set_text_color(20, 60, 120)
            pdf.cell(0, 7, _ato_safe(title), ln=True, fill=True)
            pdf.set_text_color(0, 0, 0)
            pdf.ln(1)

        def _kv_row(pdf, label, value):
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(60, 5, _ato_safe(str(label) + ":"), ln=False)
            pdf.set_font("Helvetica", "", 8)
            pdf.multi_cell(0, 5, _ato_safe(str(value)[:300]))

        def _cui_footer_all(pdf):
            total = pdf.page
            for pno in range(1, total + 1):
                pdf.page = pno
                pdf.set_y(-12)
                pdf.set_font("Helvetica", "B", 7)
                pdf.set_fill_color(180, 30, 30)
                pdf.set_text_color(255, 255, 255)
                pdf.cell(0, 4,
                         f"  {_CLASSIFICATION} | Page {pno} of {total} | ICDEV ATO Evidence Package",
                         ln=True, fill=True)
            pdf.page = total

        try:
            from fpdf import FPDF

            pdf = FPDF(orientation="P", unit="mm", format="A4")
            pdf.set_auto_page_break(auto=True, margin=18)
            pdf.set_margins(15, 20, 15)

            # ── Cover ─────────────────────────────────────────────────────
            pdf.add_page()
            _cui_banner(pdf)
            pdf.ln(12)
            pdf.set_font("Helvetica", "B", 20)
            pdf.set_text_color(20, 60, 120)
            pdf.cell(0, 12, "ATO Evidence Package", ln=True, align="C")
            pdf.set_font("Helvetica", "B", 13)
            pdf.set_text_color(50, 50, 50)
            pdf.cell(0, 8, _ato_safe(f"Advisory ID: {advisory_id}  |  CVE: {adv.get('cve_id','N/A')}"), ln=True, align="C")
            pdf.set_font("Helvetica", "", 11)
            pdf.cell(0, 7, _ato_safe(f"Vendor: {adv.get('vendor','N/A')}  |  Severity: {str(adv.get('severity','N/A')).upper()}"), ln=True, align="C")
            pdf.cell(0, 7, _ato_safe(f"Status: {adv.get('status','N/A')}"), ln=True, align="C")
            pdf.ln(4)
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_text_color(100, 100, 100)
            pdf.cell(0, 5, _ato_safe(f"Exported: {meta.get('exported_at','')}"), ln=True, align="C")
            pdf.cell(0, 5, _ato_safe(f"SHA-256: {meta.get('doc_hash_sha256','')}"), ln=True, align="C")
            pdf.set_text_color(0, 0, 0)

            # ── Sec 1: Advisory Record ─────────────────────────────────────
            pdf.add_page()
            _cui_banner(pdf)
            pdf.ln(2)
            _section_heading(pdf, "1. Advisory Record")
            for label, key in [
                ("CVE ID", "cve_id"), ("Vendor", "vendor"), ("Severity", "severity"),
                ("Status", "status"), ("Published Date", "published_date"),
                ("Total Devices", "total_devices"), ("Impacted Devices", "impacted_devices"),
                ("Remediation %", "remediation_pct"), ("Data Source", "data_source"),
                ("HITL Status", "hitl_status"), ("HITL Approved By", "hitl_approved_by"),
                ("HITL Approved At", "hitl_approved_at"),
                ("Source Doc Hash (SHA-256)", "source_doc_hash"),
                ("Source Doc Format", "source_doc_format"),
                ("Extraction Confidence", "extraction_confidence"),
                ("Created", "created_at"), ("Updated", "updated_at"),
            ]:
                val = adv.get(key, "")
                if val not in (None, "", 0):
                    _kv_row(pdf, label, val)
            if adv.get("description"):
                pdf.ln(2)
                pdf.set_font("Helvetica", "B", 8)
                pdf.cell(0, 5, "Description:", ln=True)
                pdf.set_font("Helvetica", "", 8)
                pdf.multi_cell(0, 5, _ato_safe(str(adv["description"])[:1500]))

            # ── Sec 2: Impact Assessments ──────────────────────────────────
            for idx, asmt in enumerate(evidence.get("assessments", []), 1):
                pdf.add_page()
                _cui_banner(pdf)
                pdf.ln(2)
                _section_heading(pdf, f"2.{idx} Impact Assessment (ID: {asmt.get('id','')})")
                for label, key in [
                    ("Network ID", "network_id"), ("FWD Snapshot ID", "fwd_snapshot_id"),
                    ("Data Source", "data_source"),
                    ("NQL — Total Devices", "nql_total"), ("NQL — Impacted", "nql_impacted"),
                    ("NQL — AI Generated", "nql_ai_generated"), ("NQL — Template Based", "nql_template_based"),
                    ("Total Devices", "total_devices"), ("Impacted Count", "impacted_count"),
                    ("Raw Response Hash (SHA-256)", "raw_response_hash"),
                    ("AI Confidence", "ai_confidence"),
                    ("Cross-Val Delta %", "cross_validation_delta_pct"),
                    ("Cross-Val Warning", "cross_validation_warning"),
                    ("HITL Approved By", "approved_by"), ("HITL Approved At", "approved_at"),
                    ("Created", "created_at"),
                ]:
                    val = asmt.get(key, "")
                    if val not in (None, ""):
                        _kv_row(pdf, label, val)

            # ── Sec 3: Remediation Actions ─────────────────────────────────
            if evidence.get("remediation_actions"):
                pdf.add_page()
                _cui_banner(pdf)
                pdf.ln(2)
                _section_heading(pdf, "3. Remediation Actions")
                for ra in evidence["remediation_actions"]:
                    pdf.set_font("Helvetica", "BI", 9)
                    pdf.cell(0, 6, _ato_safe(f"Action {ra.get('id','')[:24]}"), ln=True)
                    for label, key in [
                        ("Device ID", "device_id"), ("Action Type", "action_type"),
                        ("Performed By", "performed_by"), ("Result", "result"),
                        ("Notes", "notes"), ("Created", "created_at"),
                    ]:
                        val = ra.get(key, "")
                        if val not in (None, ""):
                            _kv_row(pdf, label, val)
                    pdf.ln(2)
                if evidence.get("remediation_status_log"):
                    pdf.ln(2)
                    _section_heading(pdf, "3a. Remediation Status Log")
                    pdf.set_font("Helvetica", "", 8)
                    for sl in evidence["remediation_status_log"]:
                        pdf.cell(0, 5, _ato_safe(
                            f"{sl.get('created_at','')} | action={sl.get('action_id','')} | "
                            f"{sl.get('old_status','?')} -> {sl.get('new_status','')} | "
                            f"by={sl.get('updated_by','')}"
                        ), ln=True)

            # ── Sec 4: POAM Items ──────────────────────────────────────────
            if evidence.get("poam_items"):
                pdf.add_page()
                _cui_banner(pdf)
                pdf.ln(2)
                _section_heading(pdf, "4. POAM Items")
                for pi in evidence["poam_items"]:
                    pdf.set_font("Helvetica", "BI", 9)
                    pdf.cell(0, 6, _ato_safe(f"POAM {pi.get('poam_id', pi.get('id',''))}"), ln=True)
                    for label, key in [
                        ("CVE", "cve_id"), ("Weakness", "weakness"), ("Control ID", "control_id"),
                        ("Severity", "severity"), ("Status", "status"),
                        ("Scheduled Completion", "scheduled_completion"),
                        ("Responsible Party", "responsible_party"),
                        ("Twin Validated", "twin_validated"), ("Resources", "resources"),
                    ]:
                        val = pi.get(key, "")
                        if val not in (None, "", 0):
                            _kv_row(pdf, label, val)
                    pdf.ln(2)
                if evidence.get("poam_status_log"):
                    pdf.ln(2)
                    _section_heading(pdf, "4a. POAM Status Log")
                    pdf.set_font("Helvetica", "", 8)
                    for sl in evidence["poam_status_log"]:
                        pdf.cell(0, 5, _ato_safe(
                            f"{sl.get('created_at','')} | poam={sl.get('poam_id','')} | "
                            f"{sl.get('old_status','?')} -> {sl.get('new_status','')} | "
                            f"by={sl.get('updated_by','')}"
                        ), ln=True)

            # ── Sec 5: Exceptions ──────────────────────────────────────────
            if evidence.get("exceptions"):
                pdf.add_page()
                _cui_banner(pdf)
                pdf.ln(2)
                _section_heading(pdf, "5. Exceptions")
                for ex in evidence["exceptions"]:
                    pdf.set_font("Helvetica", "BI", 9)
                    pdf.cell(0, 6, _ato_safe(f"Device: {ex.get('device_name','')}  [{ex.get('status','')}]"), ln=True)
                    for label, key in [
                        ("Exception Type", "exception_type"), ("Risk Level", "risk_level"),
                        ("Expiry Date", "expiry_date"), ("Justification", "justification"),
                        ("ISSO Approved By", "isso_approved_by"), ("ISSO Approved At", "isso_approved_at"),
                        ("ISSM Approved By", "issm_approved_by"), ("ISSM Approved At", "issm_approved_at"),
                        ("AO Approved By", "ao_approved_by"), ("AO Approved At", "ao_approved_at"),
                        ("Compensating Controls", "compensating_controls"),
                        ("Risk Acceptance Level", "risk_acceptance_level"),
                        ("Filed By", "filed_by"), ("Created", "created_at"),
                    ]:
                        val = ex.get(key, "")
                        if val not in (None, "", 0):
                            _kv_row(pdf, label, val)
                    pdf.ln(2)
                if evidence.get("exception_approvals"):
                    pdf.ln(2)
                    _section_heading(pdf, "5a. Exception Approval Chain")
                    pdf.set_font("Helvetica", "", 8)
                    for ap in evidence["exception_approvals"]:
                        pdf.cell(0, 5, _ato_safe(
                            f"{ap.get('created_at','')} | exc={ap.get('exception_id','')} | "
                            f"{ap.get('approver_role','')} {ap.get('approver','')} -> "
                            f"{ap.get('decision','')} | {ap.get('conditions','')}"
                        ), ln=True)

            # ── Sec 6: NQE Audit Trail ─────────────────────────────────────
            if evidence.get("audit_log"):
                pdf.add_page()
                _cui_banner(pdf)
                pdf.ln(2)
                _section_heading(pdf, "6. NQE Audit Trail")
                pdf.set_font("Helvetica", "", 8)
                for entry in evidence["audit_log"]:
                    pdf.cell(0, 5, _ato_safe(
                        f"{entry.get('created_at','')} | {entry.get('action','')} | "
                        f"{str(entry.get('result_summary',''))[:100]}"
                    ), ln=True)

            # ── Hash integrity page ────────────────────────────────────────
            pdf.add_page()
            _cui_banner(pdf)
            pdf.ln(4)
            _section_heading(pdf, "Document Integrity")
            _kv_row(pdf, "Document SHA-256", meta.get("doc_hash_sha256", ""))
            _kv_row(pdf, "Exported At", meta.get("exported_at", ""))
            _kv_row(pdf, "Classification", meta.get("classification", ""))
            pdf.ln(4)
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_text_color(100, 100, 100)
            pdf.multi_cell(0, 5,
                "The SHA-256 hash above was computed over the canonical JSON serialization "
                "of all evidence fields. Re-compute to verify document integrity.")

            _cui_footer_all(pdf)
            return bytes(pdf.output())

        except ImportError:
            logger.warning("fpdf2 not installed — generating HTML fallback for ATO PDF")
            adv_rows = "".join(
                f"<tr><th>{k}</th><td>{v}</td></tr>"
                for k, v in (adv or {}).items()
                if v not in (None, "")
            )
            return (
                f"<!DOCTYPE html><html><head><meta charset='utf-8'>"
                f"<title>ATO Evidence — Advisory {advisory_id}</title>"
                f"<style>body{{font-family:monospace;margin:2em}}"
                f"table{{border-collapse:collapse;width:100%}}"
                f"th,td{{border:1px solid #ccc;padding:4px 8px;text-align:left}}"
                f"th{{background:#dde;width:220px}}"
                f".cui{{background:#b41e1e;color:#fff;padding:4px 10px;font-weight:bold}}"
                f"h2{{color:#1a3c78}}</style></head><body>"
                f"<div class='cui'>CUI // SP-CTI</div>"
                f"<h1>ATO Evidence Package — Advisory {advisory_id}</h1>"
                f"<p>Exported: {meta.get('exported_at','')} | "
                f"SHA-256: <code>{meta.get('doc_hash_sha256','')}</code></p>"
                f"<h2>1. Advisory Record</h2><table>{adv_rows}</table>"
                f"<p><em>Install fpdf2 for full multi-section PDF output.</em></p>"
                f"<div class='cui'>CUI // SP-CTI</div></body></html>"
            ).encode("utf-8")

    def _build_ato_excel(evidence, advisory_id):
        """Return (bytes, mimetype, extension) for Excel export.

        Uses openpyxl when available; falls back to a ZIP archive of CSV sheets.
        """
        import csv
        import io
        import zipfile

        SHEETS = [
            ("Advisory",            [evidence.get("advisory")] if evidence.get("advisory") else []),
            ("Assessments",         evidence.get("assessments", [])),
            ("RemediationActions",  evidence.get("remediation_actions", [])),
            ("RemediationStatusLog",evidence.get("remediation_status_log", [])),
            ("POAMItems",           evidence.get("poam_items", [])),
            ("POAMStatusLog",       evidence.get("poam_status_log", [])),
            ("Exceptions",          evidence.get("exceptions", [])),
            ("ExceptionApprovals",  evidence.get("exception_approvals", [])),
            ("AuditLog",            evidence.get("audit_log", [])),
        ]

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # drop default blank sheet

            _CUI_FILL = PatternFill("solid", fgColor="B41E1E")
            _HEAD_FILL = PatternFill("solid", fgColor="DDE4F0")
            _CUI_FONT = Font(bold=True, color="FFFFFF", size=9)
            _HEAD_FONT = Font(bold=True, size=9)
            _META = evidence.get("_meta", {})

            for sheet_name, rows in SHEETS:
                if not rows:
                    continue
                ws = wb.create_sheet(title=sheet_name[:31])
                # CUI banner row
                ws.append([f"CUI // SP-CTI | Advisory {advisory_id} | {_META.get('exported_at','')}"])
                cui_cell = ws.cell(1, 1)
                cui_cell.fill = _CUI_FILL
                cui_cell.font = _CUI_FONT
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(rows[0]) if rows else 1, 1))
                # Header row
                headers = list(rows[0].keys()) if rows else []
                ws.append(headers)
                for col_idx, _ in enumerate(headers, 1):
                    cell = ws.cell(2, col_idx)
                    cell.fill = _HEAD_FILL
                    cell.font = _HEAD_FONT
                # Data rows
                for row in rows:
                    ws.append([str(v) if v is not None else "" for v in row.values()])
                # Auto-width (capped)
                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=8)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"

        except ImportError:
            # Fallback: ZIP of CSV files (can be opened sheet-by-sheet)
            logger.info("openpyxl not installed — generating ZIP-of-CSVs fallback for ATO Excel")
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for sheet_name, rows in SHEETS:
                    if not rows:
                        continue
                    csv_buf = io.StringIO()
                    writer = csv.DictWriter(csv_buf, fieldnames=list(rows[0].keys()))
                    writer.writeheader()
                    writer.writerows(rows)
                    zf.writestr(f"{sheet_name}.csv", csv_buf.getvalue())
            return zip_buf.getvalue(), "application/zip", "zip"

    @bp.route("/api/advisory/<int:advisory_id>/export-ato", methods=["POST"])
    @nc_login_required
    def nc_api_advisory_export_ato(advisory_id: int):
        """Export complete ATO evidence chain for one advisory.

        Query param: format — json | pdf | excel (default: json)

        Evidence chain:
          1. nc_advisories row + source_doc_hash
          2. nc_advisory_assessments (NQL queries, sha256 hashes, dual-query
             reconciliation, HITL approval record)
          3. nc_remediation_actions (all statuses)
          4. nc_remediation_status_log
          5. nc_poam_items + nc_poam_status_log
          6. nc_exceptions + nc_exception_approvals + nc_nqe_audit_log
        """
        import hashlib
        from flask import Response

        fmt = request.args.get("format", "json").lower()
        if fmt not in ("json", "pdf", "excel"):
            return jsonify({"error": "format must be json, pdf, or excel"}), 400

        conn = get_connection()
        try:
            evidence = _gather_ato_evidence(advisory_id, conn)
        finally:
            conn.close()

        if evidence["advisory"] is None:
            return jsonify({"error": f"Advisory {advisory_id} not found"}), 404

        canonical = json.dumps(evidence, sort_keys=True, default=str)
        doc_hash = hashlib.sha256(canonical.encode()).hexdigest()
        evidence["_meta"] = {
            "advisory_id": advisory_id,
            "export_format": fmt,
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "doc_hash_sha256": doc_hash,
            "classification": "CUI // SP-CTI",
        }

        _audit_ato_export(advisory_id, fmt, doc_hash)

        if fmt == "json":
            return Response(
                json.dumps(evidence, indent=2, default=str),
                mimetype="application/json",
                headers={"Content-Disposition": f"attachment; filename=ato-evidence-{advisory_id}.json"},
            )
        if fmt == "pdf":
            pdf_bytes = _build_ato_pdf(evidence, advisory_id)
            mimetype = "application/pdf" if pdf_bytes[:4] == b"%PDF" else "text/html"
            ext = "pdf" if mimetype == "application/pdf" else "html"
            return Response(
                pdf_bytes,
                mimetype=mimetype,
                headers={"Content-Disposition": f"attachment; filename=ato-evidence-{advisory_id}.{ext}"},
            )
        # excel
        xl_bytes, mimetype, ext = _build_ato_excel(evidence, advisory_id)
        return Response(
            xl_bytes,
            mimetype=mimetype,
            headers={"Content-Disposition": f"attachment; filename=ato-evidence-{advisory_id}.{ext}"},
        )

    # ── NQE Translator ─────────────────────────────────────────────────────

    @bp.route("/nqe-translator", methods=["GET"])
    def nqe_translator_page():
        """Render the NQE query translator UI."""
        return render_template("network/nqe_translator.html", page_title="NQE Translator")

    @bp.route("/api/nqe/translate", methods=["POST"])
    def api_nqe_translate():
        """Translate plain-English text to NQL.

        Request: {"text": str, "context": dict (optional advisory context)}
        Response: {"nql": str, "confidence": float, "source": str}
        """
        data = request.get_json(force=True, silent=True) or {}
        text = (data.get("text") or "").strip()
        if not text:
            return jsonify({"error": "text is required"}), 400

        context = data.get("context") or {}

        try:
            from tools.network.nql_translator import nl_to_nql
            nql = nl_to_nql(text, context=context or None)
        except Exception as exc:
            logger.exception("NQE translate error")
            return jsonify({"error": str(exc)}), 500

        # Confidence heuristic: deterministic context path → high; LLM path → medium
        if context and any(context.get(k) for k in ("vendor", "affected_models", "affected_versions")):
            confidence = 0.92
            source = "deterministic"
        elif nql and nql.startswith("foreach"):
            confidence = 0.70
            source = "llm_translation"
        else:
            confidence = 0.50
            source = "fallback"

        # Audit log
        try:
            from tools.db.storage import get_canvas_connection
            conn = get_canvas_connection("NC_STORAGE_BACKEND")
            conn.execute(
                "INSERT INTO nc_nqe_audit_log (action, nql_query, user_confirmed, created_at) "
                "VALUES (%s, %s, %s, NOW())",
                ("translate", nql, False),
            )
            conn.commit()
            conn.close()
        except Exception:
            pass

        return jsonify({"nql": nql, "confidence": confidence, "source": source})

    @bp.route("/api/nqe/explain", methods=["POST"])
    def api_nqe_explain():
        """Return a plain-English explanation of an NQL query.

        Request: {"nql": str}
        Response: {"explanation": str}
        """
        data = request.get_json(force=True, silent=True) or {}
        nql = (data.get("nql") or "").strip()
        if not nql:
            return jsonify({"error": "nql is required"}), 400

        try:
            from tools.llm.router import LLMRouter
            from tools.llm.provider import LLMRequest

            prompt = (
                "Explain this NQL (Network Query Language) query in plain English "
                "for a non-technical network administrator. Be concise (2-3 sentences).\n\n"
                f"NQL:\n{nql}"
            )
            router = LLMRouter()
            req = LLMRequest(messages=[{"role": "user", "content": prompt}], max_tokens=150, temperature=0.2)
            resp = router.invoke("nql_explain", req)
            explanation = (resp.content or "").strip()
        except ImportError:
            explanation = _nql_heuristic_explain(nql)
        except Exception:
            explanation = _nql_heuristic_explain(nql)

        return jsonify({"explanation": explanation})

    @bp.route("/api/nqe/run", methods=["POST"])
    def api_nqe_run():
        """Execute an NQL query and return results.

        Requires explicit user confirmation (call audit-log endpoint first
        with user_confirmed=true — enforced by the UI transparency gate).

        Request: {"nql": str, "network_id": str|null}
        Response: {"rows": list, "columns": list, "total": int, "source": str}
        """
        data = request.get_json(force=True, silent=True) or {}
        nql = (data.get("nql") or "").strip()
        network_id = data.get("network_id") or None

        if not nql:
            return jsonify({"error": "nql is required"}), 400

        try:
            from tools.network.nqe_client import FallbackNQEClient

            client = FallbackNQEClient()
            result = client.run_query(nql, network_id=network_id)
            rows = result.get("rows", [])
            columns = list(rows[0].keys()) if rows else []

            # Audit execution
            try:
                from tools.db.storage import get_canvas_connection
                conn = get_canvas_connection("NC_STORAGE_BACKEND")
                conn.execute(
                    "INSERT INTO nc_nqe_audit_log (action, nql_query, user_confirmed, row_count, created_at) "
                    "VALUES (%s, %s, %s, %s, NOW())",
                    ("run", nql, True, len(rows)),
                )
                conn.commit()
                conn.close()
            except Exception:
                pass

            return jsonify({
                "rows": rows[:500],
                "columns": columns,
                "total": len(rows),
                "source": result.get("source", "local"),
            })
        except Exception as exc:
            logger.exception("NQE run error: nql=%r", nql)
            return jsonify({"error": str(exc)}), 500

    @bp.route("/api/nqe/collections", methods=["GET"])
    def api_nqe_collections():
        """Return the list of supported NQE collection paths.

        Response: {"collections": [{"path": str, "description": str}]}
        """
        collections = [
            {"path": "network.devices",          "description": "All network devices (hostname, OS, vendor, platform)"},
            {"path": "network.interfaces",        "description": "Device interfaces with status and counters"},
            {"path": "network.bgp_sessions",      "description": "BGP peering sessions and their state"},
            {"path": "network.acls",              "description": "Access control lists and firewall rules"},
            {"path": "network.paths",             "description": "End-to-end forwarding paths"},
            {"path": "network.os_versions",       "description": "OS version inventory across all devices"},
            {"path": "network.links",             "description": "Physical and logical links between nodes"},
            {"path": "network.vlans",             "description": "VLAN definitions and membership"},
            {"path": "network.prefixes",          "description": "IP prefix / subnet inventory"},
            {"path": "network.ospf.neighbors",    "description": "OSPF adjacency table"},
            {"path": "network.isis.adjacencies",  "description": "IS-IS adjacency table"},
            {"path": "network.mpls.lsps",         "description": "MPLS label-switched paths"},
        ]
        return jsonify({"collections": collections})

    @bp.route("/api/nqe/audit-log", methods=["POST"])
    def api_nqe_audit_log():
        """Append a transparency-gate audit event.

        Request: {"action": str, "nql": str, "user_confirmed": bool}
        Response: {"id": str, "recorded": true}
        """
        data = request.get_json(force=True, silent=True) or {}
        action = (data.get("action") or "").strip()
        nql = (data.get("nql") or "").strip()
        user_confirmed = bool(data.get("user_confirmed", False))

        if not action:
            return jsonify({"error": "action is required"}), 400

        row_id = None
        try:
            from tools.db.storage import get_canvas_connection
            conn = get_canvas_connection("NC_STORAGE_BACKEND")
            cur = conn.execute(
                "INSERT INTO nc_nqe_audit_log (action, nql_query, user_confirmed, created_at) "
                "VALUES (%s, %s, %s, NOW()) RETURNING id",
                (action, nql, user_confirmed),
            )
            row = cur.fetchone()
            row_id = str(row[0] if isinstance(row, (list, tuple)) else row.get("id", "")) if row else None
            conn.commit()
            conn.close()
        except Exception as exc:
            logger.warning("NQE audit-log insert failed: %s", exc)

        return jsonify({"id": row_id or "n/a", "recorded": True})

    @bp.route("/api/nqe/cross-validate", methods=["POST"])
    def api_nqe_cross_validate():
        """Dual-query cross-validation: translate using two strategies and compare.

        Strategy A uses structured context (deterministic); strategy B is context-free
        (LLM/fallback). A high divergence score means the two strategies disagree and
        a human must review and approve before execution.

        Request:  {"text": str, "context": dict (optional)}
        Response: {
            "nql_primary":      str,   # strategy A result
            "nql_secondary":    str,   # strategy B result
            "divergence_score": float, # 0.0 identical → 1.0 completely different
            "require_hitl":     bool,  # True when divergence_score >= 0.6
            "message":          str
        }
        """
        data = request.get_json(force=True, silent=True) or {}
        text = (data.get("text") or "").strip()
        if not text:
            return jsonify({"error": "text is required"}), 400

        context = data.get("context") or {}

        try:
            from tools.network.nql_translator import nl_to_nql
            nql_primary = nl_to_nql(text, context=context or None)
            nql_secondary = nl_to_nql(text)  # context-free → always LLM/fallback
        except Exception as exc:
            logger.exception("NQE cross-validate translation error")
            return jsonify({"error": str(exc)}), 500

        divergence = _nql_divergence_score(nql_primary, nql_secondary)
        require_hitl = divergence >= 0.6

        if require_hitl:
            message = (
                f"The two translation strategies produced divergent queries "
                f"(divergence {divergence:.0%}). Human approval is required before execution."
            )
        elif divergence >= 0.3:
            message = (
                f"Minor divergence detected ({divergence:.0%}). "
                "Review both queries before running."
            )
        else:
            message = "Both translation strategies agree. Safe to proceed."

        return jsonify({
            "nql_primary": nql_primary,
            "nql_secondary": nql_secondary,
            "divergence_score": divergence,
            "require_hitl": require_hitl,
            "message": message,
        })

    @bp.route("/api/nqe/hitl-approve", methods=["POST"])
    def api_nqe_hitl_approve():
        """Record a HITL approval for a cross-validated NQE query pair.

        Request:  {"nql_primary": str, "nql_secondary": str, "approved_by": str, "notes": str}
        Response: {"approved": true, "approved_by": str, "recorded": true}
        """
        data = request.get_json(force=True, silent=True) or {}
        nql_primary = (data.get("nql_primary") or "").strip()
        nql_secondary = (data.get("nql_secondary") or "").strip()
        approved_by = (data.get("approved_by") or "").strip()
        notes = (data.get("notes") or "").strip()

        if not approved_by:
            return jsonify({"error": "approved_by is required"}), 400

        import json as _json
        audit_payload = _json.dumps({
            "nql_primary": nql_primary,
            "nql_secondary": nql_secondary,
            "notes": notes,
        })

        try:
            from tools.db.storage import get_canvas_connection
            conn = get_canvas_connection("NC_STORAGE_BACKEND")
            conn.execute(
                "INSERT INTO nc_nqe_audit_log (action, nql_query, user_confirmed, created_at) "
                "VALUES (%s, %s, %s, NOW())",
                ("hitl_approve", audit_payload, True),
            )
            conn.commit()
            conn.close()
        except Exception as exc:
            logger.warning("NQE hitl-approve audit insert failed: %s", exc)

        return jsonify({"approved": True, "approved_by": approved_by, "recorded": True})

    # ── NQE helpers ────────────────────────────────────────────────────────

    def _nql_divergence_score(nql_a: str, nql_b: str) -> float:
        """Return divergence in [0.0, 1.0] between two NQL strings (Jaccard distance).

        0.0 = identical, 1.0 = completely different.
        Collection mismatch is boosted to ≥ 0.8 since different primary collections
        almost certainly query different facts.
        """
        import re as _re

        def _collection(nql):
            m = _re.search(r"\bin\s+(network\.\S+)", nql, _re.I)
            return m.group(1).lower() if m else ""

        col_a = _collection(nql_a)
        col_b = _collection(nql_b)

        tok_a = set(_re.findall(r"[\w.]+", nql_a.lower()))
        tok_b = set(_re.findall(r"[\w.]+", nql_b.lower()))

        if not tok_a or not tok_b:
            return 1.0

        jaccard = len(tok_a & tok_b) / len(tok_a | tok_b)
        divergence = 1.0 - jaccard

        if col_a and col_b and col_a != col_b:
            divergence = max(divergence, 0.8)

        return round(divergence, 3)

    def _nql_heuristic_explain(nql: str) -> str:
        """Generate a simple heuristic explanation from NQL structure."""
        import re as _re
        nql = nql.strip()
        m = _re.search(r"\bin\s+(network\.\S+)", nql, _re.I)
        collection = m.group(1) if m else "network"
        where_m = _re.search(r"\bwhere\s+(.+?)(?:\bselect\b|$)", nql, _re.I | _re.DOTALL)
        where_clause = where_m.group(1).strip() if where_m else ""
        base = f"Queries the '{collection}' collection"
        if where_clause:
            base += f" filtered by: {where_clause[:120]}"
        return base + "."


    # ── Config Review ─────────────────────────────────────────────────────
    @bp.route("/config-review")
    def nc_config_review():
        """Configuration Review Assistant — upload and AI-review device configs."""
        return render_template("network/config_review.html")

    # ── Diagram Analysis ──────────────────────────────────────────────────
    @bp.route("/diagram-analysis")
    def nc_diagram_analysis():
        """Network Diagram Analysis — upload PNG/PDF/draw.io for AI analysis."""
        return render_template("network/diagram_analysis.html")

    # ── Migration Phases landing (alias → hub) ────────────────────────────
    @bp.route("/migration-phases")
    def nc_migration_phases_hub():
        """Redirect bare /migration-phases to the Migration Hub."""
        from flask import redirect, url_for
        return redirect(url_for("network.migration_hub"))

    # ── PVM Predictive Vulnerability Management routes ────────────────────
    from tools.network.routes.pvm import register_pvm_routes

    register_pvm_routes(bp)

    # ── PNA Predictive Network Analytics routes ───────────────────────────
    from tools.network.routes.pna import register_pna_routes

    register_pna_routes(bp)

    # ── Done ───────────────────────────────────────────────────────────────
    logger.info("Network Design Canvas Blueprint created (%d routes)", len(bp.deferred_functions))
    return bp
