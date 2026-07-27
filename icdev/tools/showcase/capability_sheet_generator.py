#!/usr/bin/env python3
"""Generate the ICDEV Capabilities & Solutions Excel from args/capability_sheet.yaml."""
import argparse
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]

# ── colour palette ──────────────────────────────────────────────────────────
NAVY        = "1F3864"
NAVY2       = "2E4A7A"
ODD_ROW     = "F2F7FF"
EVEN_ROW    = "FFFFFF"
SW_FILL     = "E8F4FD"
SW_FONT_CLR = "1565C0"
FOREST      = "155724"
BROWN       = "7D5A3C"
GRAY        = "666666"
# New-row highlight (amber/gold)
NEW_ROW     = "FFF8E1"
NEW_BORDER  = "F9A825"
SW_NEW_FILL = "FFF3E0"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, italic=False, size=11, color="000000", name="Calibri"):
    return Font(bold=bold, italic=italic, size=size, color=color, name=name)

def _align(wrap=True, h="left", v="top"):
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)

def apply_cell(ws, row, col, value, fill=None, font=None, alignment=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:
        c.fill = fill
    if font:
        c.font = font
    if alignment:
        c.alignment = alignment
    return c

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_title_rows(ws, n_cols, title, subtitle):
    last_col = get_column_letter(n_cols)
    # Row 1 — title
    ws.merge_cells(f"A1:{last_col}1")
    apply_cell(ws, 1, 1, title,
               fill=_fill(NAVY),
               font=_font(bold=True, size=14, color="FFFFFF"),
               alignment=_align(h="center", v="center"))
    ws.row_dimensions[1].height = 28

    # Row 2 — subtitle
    ws.merge_cells(f"A2:{last_col}2")
    apply_cell(ws, 2, 1, subtitle,
               fill=_fill(NAVY2),
               font=_font(italic=True, size=11, color="FFFFFF"),
               alignment=_align(h="left", v="center"))
    ws.row_dimensions[2].height = 22

def write_header_row(ws, col_labels):
    for col, label in enumerate(col_labels, 1):
        apply_cell(ws, 3, col, label,
                   fill=_fill(NAVY),
                   font=_font(bold=True, size=11, color="FFFFFF"),
                   alignment=_align(h="center", v="center", wrap=True))
    ws.row_dimensions[3].height = 20

def row_fill(idx, is_new=False):
    if is_new:
        return _fill(NEW_ROW)
    return _fill(ODD_ROW) if idx % 2 == 0 else _fill(EVEN_ROW)

def _new_border():
    gold = Side(style="medium", color=NEW_BORDER)
    thin = Side(style="thin", color="DDDDDD")
    return Border(left=gold, right=thin, top=thin, bottom=thin)

def _std_border():
    thin = Side(style="thin", color="DDDDDD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _apply_row_borders(ws, row, n_cols, is_new):
    border = _new_border() if is_new else _std_border()
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).border = border

def build_sheet1(ws, challenges, last_updated):
    col_labels = ["#", "Domain", "Challenge", "Business Impact",
                  "ICDEV Solution", "Key ICDEV Modules / Tools",
                  "Compliance Frameworks", "Quantified Outcome / ROI*", "So What"]
    widths = [6, 20, 40, 36, 50, 40, 30, 34, 44]
    set_col_widths(ws, widths)

    write_title_rows(ws, 9,
        "ICDEV™ Platform  —  Capability Showcase: Challenges & Solutions  (Updated June 2026)",
        "AI-orchestrated, compliance-hardened software delivery for DoD & Government. "
        "Each solution combines multiple ICDEV capabilities. "
        f"Updated: {last_updated}")
    write_header_row(ws, col_labels)
    ws.freeze_panes = "A4"

    prev_domain = None
    for idx, ch in enumerate(challenges):
        r = idx + 4
        is_new = bool(ch.get("is_new"))
        rf = row_fill(idx, is_new)
        domain_start = ch.get("domain") != prev_domain
        prev_domain = ch.get("domain")
        row_bold = domain_start and not is_new  # bold only on first row of each existing domain group

        data = [
            ch.get("id"), ch.get("domain"), ch.get("challenge"),
            ch.get("business_impact"), ch.get("solution"), ch.get("modules"),
            ch.get("compliance"), ch.get("outcome"),
        ]
        for col, val in enumerate(data, 1):
            apply_cell(ws, r, col, val, fill=rf,
                       font=_font(bold=row_bold, size=11), alignment=_align())
        # So What — styled column (deeper amber on new rows)
        sw_fill = _fill(SW_NEW_FILL) if is_new else _fill(SW_FILL)
        apply_cell(ws, r, 9, ch.get("so_what"),
                   fill=sw_fill,
                   font=_font(italic=True, size=11, color=SW_FONT_CLR),
                   alignment=_align())
        _apply_row_borders(ws, r, 9, is_new)
        ws.row_dimensions[r].height = 80

    ws.sheet_properties.tabColor = NAVY

def build_sheet2(ws, capabilities, last_updated):
    col_labels = ["#", "Capability", "What It Does",
                  "Key ICDEV Modules / Tools", "Compliance Frameworks",
                  "Solves Challenge(s)", "So What"]
    widths = [6, 28, 54, 40, 30, 16, 44]
    set_col_widths(ws, widths)

    write_title_rows(ws, 7,
        "ICDEV™ Platform  —  Capability Catalog  (Updated June 2026)",
        "What each ICDEV capability is and what it does. "
        "'Solves Challenge(s)' cross-references rows on the Challenges & Solutions sheet. "
        "33 capability clusters from 600+ registered tool domains.")
    write_header_row(ws, col_labels)
    ws.freeze_panes = "A4"

    for idx, cap in enumerate(capabilities):
        r = idx + 4
        is_new = bool(cap.get("is_new"))
        rf = row_fill(idx, is_new)
        data = [
            cap.get("id"), cap.get("capability"), cap.get("what_it_does"),
            cap.get("modules"), cap.get("compliance"), cap.get("solves"),
        ]
        for col, val in enumerate(data, 1):
            apply_cell(ws, r, col, val, fill=rf,
                       font=_font(size=11), alignment=_align())
        sw_fill = _fill(SW_NEW_FILL) if is_new else _fill(SW_FILL)
        apply_cell(ws, r, 7, cap.get("so_what"),
                   fill=sw_fill,
                   font=_font(italic=True, size=11, color=SW_FONT_CLR),
                   alignment=_align())
        _apply_row_borders(ws, r, 7, is_new)
        ws.row_dimensions[r].height = 80

    # Footer
    footer_row = len(capabilities) + 4
    ws.merge_cells(f"A{footer_row}:G{footer_row}")
    footer_text = (
        f"ICDEV™ orchestrates 16 AI agents and 600+ deterministic tools via the FORGE framework. "
        f"Updated: {last_updated}"
    )
    apply_cell(ws, footer_row, 1, footer_text,
               font=_font(italic=True, size=10, color=GRAY),
               alignment=_align(h="center", v="center"))
    ws.row_dimensions[footer_row].height = 18

    ws.sheet_properties.tabColor = FOREST

def build_sheet3(ws, canvases):
    col_labels = ["#", "Canvas", "Route", "Use Case (Who / When)",
                  "Challenge It Solves", "How the Canvas Solves It"]
    widths = [6, 26, 18, 34, 38, 54]
    set_col_widths(ws, widths)

    write_title_rows(ws, 6,
        "ICDEV™ Platform  —  Design Canvas Breakdown  (Updated June 2026)",
        "21 registered canvases. Each shares a DB, Flask blueprint, YAML config, "
        "feature flag, NIST 800-53 assessment, and cross-canvas integration.")
    write_header_row(ws, col_labels)
    ws.freeze_panes = "A4"

    for idx, cv in enumerate(canvases):
        r = idx + 4
        is_new = bool(cv.get("is_new"))
        rf = row_fill(idx, is_new)
        data = [
            cv.get("id"), cv.get("canvas"), cv.get("route"),
            cv.get("use_case"), cv.get("challenge"), cv.get("how_it_solves"),
        ]
        for col, val in enumerate(data, 1):
            apply_cell(ws, r, col, val, fill=rf,
                       font=_font(size=11), alignment=_align())
        _apply_row_borders(ws, r, 6, is_new)
        ws.row_dimensions[r].height = 75

    ws.sheet_properties.tabColor = BROWN

def generate(yaml_path: Path, output_path: Path):
    with open(yaml_path, encoding="utf-8") as f:
        data = yaml.safe_load(f)

    meta = data.get("meta", {})
    last_updated = meta.get("last_updated", "2026-06-27")

    challenges   = [c for c in data.get("challenges",   []) if c.get("status") == "approved"]
    capabilities = [c for c in data.get("capabilities", []) if c.get("status") == "approved"]
    canvases     = [c for c in data.get("canvases",     []) if c.get("status") == "approved"]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Challenges & Solutions"
    build_sheet1(ws1, challenges, last_updated)

    ws2 = wb.create_sheet("Capability Catalog")
    build_sheet2(ws2, capabilities, last_updated)

    ws3 = wb.create_sheet("Design Canvases")
    build_sheet3(ws3, canvases)

    wb.save(str(output_path))
    print(f"[capability-sheet] Generated: {output_path} "
          f"({len(challenges)} challenges, {len(capabilities)} capabilities, "
          f"{len(canvases)} canvases)")

def main():
    parser = argparse.ArgumentParser(description="Generate ICDEV capability sheet Excel")
    parser.add_argument("--yaml",   default=str(ROOT / "args" / "capability_sheet.yaml"))
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    yaml_path = Path(args.yaml)
    if args.output:
        output_path = Path(args.output)
    else:
        with open(yaml_path, encoding="utf-8") as f:
            meta_data = yaml.safe_load(f)
        output_path = Path(meta_data.get("meta", {}).get(
            "output_path",
            str(ROOT / "data" / "ICDEV_Capabilities_Challenges_Solutions.xlsx")))

    generate(yaml_path, output_path)

if __name__ == "__main__":
    main()
