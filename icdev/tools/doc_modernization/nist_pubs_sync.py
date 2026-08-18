# CUI // SP-CTI
"""NIST publication revision cache — policy_refs evidence with air-gap fallback.

Structurally cloned from tools/doc_modernization/eol_products_sync.py (the
endoflife.date pattern): a mutable cache table (docmod_nist_pubs), a static YAML
seed for air-gapped sites, a best-effort live pull, and an import_dataset() path
for bundled exports.

The policy_refs pack flags any document that cites an OLDER revision than the one
recorded here (deterministic numeric comparison — TRUST rule 1, no LLM).

LIVE SOURCE (cef-fnd-02). The original live pull read the CSRC RSS feed at
``/CSRC/media/feeds/rss/publications.xml``. **That feed is retired** — measured
2026-08-17 it returns HTTP 404 while csrc.nist.gov itself answers 200, so the pull
had never landed a single row (every row in the live cache was ``source='seed'``).
CSRC today advertises exactly one publications feed, ``drafts-open-for-comment.xml``,
and that one is deliberately NOT used: a DRAFT does not supersede a final
publication, so treating a draft Rev 6 as the current revision would manufacture
false "superseded" findings — the opposite of this cache's job.

The authoritative source is the spreadsheet CSRC publishes of current draft AND
final publications, ``NIST-Cybersecurity-Publications.xlsx``. It carries a ``Stage``
column, so the sync can keep **Final rows only** and read the revision straight out
of ``Publication Number`` ('800-53 Rev. 5'). openpyxl is already a declared
dependency. parse_feed() is retained for operators who configure a working
RSS/Atom feed of their own.

A publication whose number carries NO revision (e.g. SP 800-207) is deliberately
NOT cached: inventing "Rev 1" for it would let a bogus citation read as current,
and omitting it yields verdict 'unknown', which produces no finding. Absent
evidence must not present as evidence.

Air-gap discipline (mirrors eol_products_sync): honors the docmod_config
``offline`` flag, is https-only with TLS verification and a tight timeout, and
swallows every network/parse error — the cache keeps its seed/import rows and the
nightly sweep never fails because egress is unavailable. True push/webhooks are
out of scope: this is scheduled pull only.

**A dead URL and a dead network are different failures** and this module no longer
merges them. A 4xx means the server answered and the resource is gone — a
MISCONFIGURATION, logged at warning and reported as ``url_dead_http_404``. A
connection/TLS/timeout error means no egress — a normal air-gap state, logged at
info and reported as ``unreachable``. Reporting the first as "offline?" is exactly
how a retired feed hid for as long as it did.

CLI:
    python -m tools.doc_modernization.nist_pubs_sync                 # refresh
    python -m tools.doc_modernization.nist_pubs_sync --seed --json
    python -m tools.doc_modernization.nist_pubs_sync --sync --json
    python -m tools.doc_modernization.nist_pubs_sync --sync --force --json
    python -m tools.doc_modernization.nist_pubs_sync --import <path> --json
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path

from tools.logging.icdev_logger import get_logger

logger = get_logger(__name__)

_REPO_ROOT = Path(__file__).resolve().parents[2]
SEED_PATH = _REPO_ROOT / "args" / "docmod" / "nist_pubs.yaml"

# CSRC's spreadsheet of current draft + final publications — the live source of
# record. Operators may override in args/docmod/docmod_config.yaml
# (nist_pubs_catalog_url); the static seed is the air-gap source of truth when no
# egress is available.
DEFAULT_CATALOG_URL = (
    "https://csrc.nist.gov/files/pubs/shared/docs/NIST-Cybersecurity-Publications.xlsx"
)

# No default RSS/Atom feed: the CSRC publications feed this module shipped against
# is retired (HTTP 404). Operators with their own mirror set nist_pubs_feed_url.
DEFAULT_FEED_URL = ""

# Cap the catalog download so a redirect to something enormous cannot exhaust
# memory on a scheduled sweep. The real file is ~450 KB.
_MAX_CATALOG_BYTES = 32 * 1024 * 1024

# "NIST SP 800-53 Rev. 5" / "SP 800-171r3" -> pub_id 'SP 800-53', revision 5.
_PUB_RE = re.compile(
    r"(?i)\b(?:NIST\s+)?(SP\s?800-\d+[A-Za-z]?)\s?"
    r"(?:Rev(?:ision|\.)?\s?|r)(\d+)\b"
)


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _connect():
    from tools.db.storage import get_connection
    return get_connection()


def _config() -> dict:
    from .pack_loader import load_config
    return load_config()


def _normalize_pub_id(raw: str) -> str:
    """'SP 800-53' / 'sp800-53' -> canonical 'SP 800-53'."""
    m = re.match(r"(?i)\s*SP\s?(800-\d+[A-Za-z]?)\s*", raw or "")
    return f"SP {m.group(1)}" if m else (raw or "").strip()


def _upsert(conn, pub_id: str, row: dict, source: str) -> None:
    """Upsert by pub_id. source ∈ the DDL CHECK: 'nist.gov'|'seed'|'manual'."""
    conn.execute("DELETE FROM docmod_nist_pubs WHERE pub_id = %s", (pub_id,))
    conn.execute(
        """INSERT INTO docmod_nist_pubs
           (id, pub_id, latest_revision, revision_num, title, url,
            published_date, source, synced_at)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
        (
            f"nistpub-{uuid.uuid4().hex[:10]}", pub_id,
            row.get("latest_revision"),
            int(row["revision_num"]) if row.get("revision_num") is not None else None,
            row.get("title"), row.get("url"), row.get("published_date"),
            source, _now(),
        ),
    )


def _keep_highest(best: dict[str, dict], row: dict) -> None:
    """Keep only the highest revision seen for each pub_id."""
    cur = best.get(row["pub_id"])
    if cur is None or row["revision_num"] > cur["revision_num"]:
        best[row["pub_id"]] = row


def parse_feed(xml_text: str) -> list[dict]:
    """Parse RSS or Atom feed XML into per-publication latest-revision rows.

    Pure function (no DB, no network) — the unit-testable core. For each pub_id
    only the HIGHEST revision seen is kept. Returns a list of dicts:
    {pub_id, latest_revision, revision_num, title, url, published_date}.
    """
    # Feed XML is untrusted network content — parse with defusedxml (blocks
    # entity-expansion / external-entity attacks). Any parse failure (malformed
    # or hostile) degrades to an empty result so the sweep never crashes.
    import defusedxml.ElementTree as ET

    try:
        root = ET.fromstring(xml_text)
    except Exception as exc:
        logger.warning("nist_pubs_sync: feed parse error: %s", exc)
        return []

    def _local(tag: str) -> str:
        return tag.rsplit("}", 1)[-1] if "}" in tag else tag

    # Collect (title, link, date) triples from RSS <item> or Atom <entry>.
    items: list[tuple[str, str, str]] = []
    for el in root.iter():
        if _local(el.tag) not in ("item", "entry"):
            continue
        title = link = date = ""
        for child in el:
            name = _local(child.tag)
            if name == "title":
                title = (child.text or "").strip()
            elif name == "link":
                # RSS: text; Atom: href attribute.
                link = (child.text or "").strip() or child.get("href", "")
            elif name in ("pubDate", "updated", "published", "date"):
                date = date or (child.text or "").strip()
        if title:
            items.append((title, link, date))

    best: dict[str, dict] = {}
    for title, link, date in items:
        m = _PUB_RE.search(title)
        if not m:
            continue
        rev = int(m.group(2))
        _keep_highest(best, {
            "pub_id": _normalize_pub_id(m.group(1)),
            "latest_revision": f"Rev {rev}",
            "revision_num": rev,
            "title": title,
            "url": link,
            "published_date": date,
        })
    return list(best.values())


def parse_catalog(xlsx_bytes: bytes) -> list[dict]:
    """Parse the CSRC publications spreadsheet into latest-revision rows.

    Pure function (no DB, no network) — the unit-testable core, same output shape
    as parse_feed(). **Final rows only**: a draft does not supersede anything, and
    a draft revision in this cache would manufacture false findings.

    Any structural surprise (missing sheet, renamed column, hostile file) degrades
    to an empty list rather than raising — a scheduled sweep must not die because
    NIST reshaped a spreadsheet.
    """
    try:
        import io

        import openpyxl
    except Exception as exc:  # openpyxl is declared, but never hard-fail the sweep
        logger.warning("nist_pubs_sync: openpyxl unavailable (%s)", exc)
        return []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("nist_pubs_sync: catalog parse error: %s", exc)
        return []

    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        try:
            header = [str(c or "").strip() for c in next(rows)]
        except StopIteration:
            return []
        idx = {name: i for i, name in enumerate(header)}
        # Required columns. A rename upstream means we cannot trust the parse at
        # all — report nothing rather than a partially-understood catalog.
        needed = ("Stage", "Series", "Publication Number")
        if any(c not in idx for c in needed):
            logger.warning(
                "nist_pubs_sync: catalog schema changed, missing %s",
                [c for c in needed if c not in idx],
            )
            return []

        def cell(row, name: str) -> str:
            i = idx.get(name)
            if i is None or i >= len(row) or row[i] is None:
                return ""
            return str(row[i]).strip()

        best: dict[str, dict] = {}
        for row in rows:
            if not row:
                continue
            if cell(row, "Stage").lower() != "final":
                continue  # drafts never supersede
            number = cell(row, "Publication Number")
            if not number:
                continue
            label = f"{cell(row, 'Series')} {number}".strip()
            m = _PUB_RE.search(label)
            if not m:
                continue  # no explicit revision (e.g. SP 800-207) — do not invent one
            rev = int(m.group(2))
            _keep_highest(best, {
                "pub_id": _normalize_pub_id(m.group(1)),
                "latest_revision": f"Rev {rev}",
                "revision_num": rev,
                "title": cell(row, "Title") or label,
                "url": cell(row, "CurrentURL") or cell(row, "URL"),
                "published_date": cell(row, "Release Date"),
            })
        return list(best.values())
    except Exception as exc:
        logger.warning("nist_pubs_sync: catalog read error: %s", exc)
        return []
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _fetch(url: str, timeout: int) -> tuple[bytes | None, str]:
    """Best-effort https-only GET. Returns (body, status).

    The status token distinguishes failures that a zero row count would otherwise
    merge into one indistinguishable "offline?":

      * ``ok``                  — body returned
      * ``not_configured``      — no URL set for this source
      * ``refused_non_https``   — plaintext URL, refused before any request
      * ``url_dead_http_<code>``— server ANSWERED 4xx: the resource is gone, a
                                  misconfiguration this deployment must fix
      * ``server_error_http_<code>`` — 5xx, transient upstream fault
      * ``unreachable``         — connection/TLS/timeout: no egress, a normal
                                  air-gap state and not a defect
      * ``oversized``           — body exceeded the download cap
    """
    if not url:
        return None, "not_configured"
    if not url.lower().startswith("https://"):
        logger.warning("nist_pubs_sync: refusing non-https url: %s", url)
        return None, "refused_non_https"

    import urllib.error
    import urllib.request

    req = urllib.request.Request(url, headers={"User-Agent": "ICDEV-docmod/1.0"})
    try:
        # https-only enforced above; default SSL context verifies the cert chain.
        with urllib.request.urlopen(req, timeout=timeout) as resp:  # nosec B310 -- https scheme validated above; TLS verified
            body = resp.read(_MAX_CATALOG_BYTES + 1)
        if len(body) > _MAX_CATALOG_BYTES:
            logger.warning("nist_pubs_sync: response exceeded %d bytes: %s",
                           _MAX_CATALOG_BYTES, url)
            return None, "oversized"
        return body, "ok"
    except urllib.error.HTTPError as exc:
        # The server answered and refused. A 404 here is a RETIRED URL, not an
        # air-gap — it must not read as "offline", which is how the retired CSRC
        # RSS feed stayed broken while reporting a benign-looking skip.
        code = int(getattr(exc, "code", 0) or 0)
        if 400 <= code < 500:
            logger.warning(
                "nist_pubs_sync: feed url is dead (HTTP %s) — fix the configured "
                "URL, this is NOT an air-gap condition: %s", code, url,
            )
            return None, f"url_dead_http_{code}"
        logger.info("nist_pubs_sync: upstream error HTTP %s for %s", code, url)
        return None, f"server_error_http_{code}"
    except Exception as exc:  # network/TLS/timeout — offline is a normal state
        logger.info("nist_pubs_sync: url unreachable (%s)", exc)
        return None, "unreachable"


def _fetch_feed(url: str, timeout: int) -> str | None:
    """Back-compat shim: body text, or None on any failure."""
    body, status = _fetch(url, timeout)
    if status != "ok" or body is None:
        return None
    return body.decode("utf-8", errors="replace")


def _within_cadence(conn, cadence_hours: float) -> bool:
    """True when a live sync ran within the cadence window (skip this run)."""
    if cadence_hours <= 0:
        return False
    try:
        row = conn.execute(
            "SELECT MAX(synced_at) AS last FROM docmod_nist_pubs WHERE source = 'nist.gov'"
        ).fetchone()
    except Exception:
        return False
    last = dict(row).get("last") if row else None
    if not last:
        return False
    try:
        ts = datetime.fromisoformat(str(last))
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
    except ValueError:
        return False
    return datetime.now(timezone.utc) - ts < timedelta(hours=cadence_hours)


def load_seed(path: Path | None = None) -> dict:
    """Load args/docmod/nist_pubs.yaml into the cache (source='seed')."""
    import yaml

    path = path or SEED_PATH
    if not path.exists():
        return {"loaded": 0, "error": f"seed not found: {path}"}
    raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    conn = _connect()
    loaded = 0
    try:
        for pub_id, row in (raw.get("publications") or {}).items():
            _upsert(conn, _normalize_pub_id(pub_id), row or {}, "seed")
            loaded += 1
        conn.commit()
    finally:
        conn.close()
    return {"loaded": loaded, "source": "seed"}


def sync(force: bool = False) -> dict:
    """Best-effort live pull of the NIST publication revisions.

    Tries the CSRC catalog spreadsheet first (Final rows only — the authority for
    supersession), then any operator-configured RSS/Atom feed. Offline-safe:
    honors the docmod_config ``offline`` flag, gates on ``nist_pubs_cadence_hours``
    (skip if a live sync ran inside the window unless ``force``), and swallows
    network errors so the sweep never fails.

    ``sources`` in the result records the per-source status token, so a retired
    URL (``url_dead_http_404``) is never reported as an air-gap.
    """
    cfg = _config()
    if cfg.get("offline"):
        return {"synced": 0, "skipped": "offline flag set"}

    cadence = float(cfg.get("nist_pubs_cadence_hours", 24) or 0)
    conn = _connect()
    try:
        if not force and _within_cadence(conn, cadence):
            return {"synced": 0, "skipped": "within cadence window"}

        timeout = int(cfg.get("nist_pubs_timeout_seconds", 15) or 15)
        catalog_url = cfg.get("nist_pubs_catalog_url", DEFAULT_CATALOG_URL)
        if catalog_url is None:
            catalog_url = DEFAULT_CATALOG_URL
        feed_url = cfg.get("nist_pubs_feed_url") or DEFAULT_FEED_URL

        attempts: list[tuple[str, str, callable]] = [
            ("catalog", catalog_url, parse_catalog),
            ("feed", feed_url, lambda b: parse_feed(b.decode("utf-8", errors="replace"))),
        ]

        statuses: dict[str, str] = {}
        rows: list[dict] = []
        landed_via = None
        for name, url, parse in attempts:
            body, status = _fetch(url, timeout)
            statuses[name] = status
            if status != "ok" or body is None:
                continue
            parsed = parse(body)
            if parsed:
                rows, landed_via = parsed, name
                break
            statuses[name] = "parsed_empty"

        if not rows:
            return {
                "synced": 0,
                "skipped": "no live source yielded rows",
                "sources": statuses,
            }

        synced = 0
        for row in rows:
            _upsert(conn, row["pub_id"], row, "nist.gov")
            synced += 1
        conn.commit()
        return {
            "synced": synced,
            "source": landed_via,
            "sources": statuses,
            "publications": sorted(r["pub_id"] for r in rows),
        }
    finally:
        conn.close()


def _row_count(conn) -> int:
    try:
        row = conn.execute("SELECT COUNT(*) AS n FROM docmod_nist_pubs").fetchone()
        return int(dict(row).get("n") or 0)
    except Exception:
        try:
            conn.rollback()  # PG: failed statement poisons the transaction
        except Exception:
            pass
        return 0


def refresh(force: bool = False) -> dict:
    """Ensure the cache has a substrate: live sync, seeding if it is still empty.

    This is what a bare ``python -m tools.doc_modernization.nist_pubs_sync`` runs.
    An empty cache makes policy_refs answer 'unknown' forever, so a run that lands
    nothing live falls back to the static seed rather than leaving the pack with
    no substrate at all. The result reports which path supplied the rows — the
    seed fallback is never presented as a successful live pull.
    """
    result: dict = {"sync": sync(force=force)}
    conn = _connect()
    try:
        before = _row_count(conn)
    finally:
        conn.close()
    if before == 0:
        result["seed"] = load_seed()
    conn = _connect()
    try:
        result["rows"] = _row_count(conn)
    finally:
        conn.close()
    return result


def import_dataset(path: str | Path) -> dict:
    """Air-gap import of a JSON/YAML bundle: {publications: {pub_id: {...}}}.
    Rows land with source='manual' (the DDL vocabulary for operator loads)."""
    import yaml

    p = Path(path)
    if not p.exists():
        return {"imported": 0, "error": f"not found: {p}"}
    text = p.read_text(encoding="utf-8")
    raw = json.loads(text) if p.suffix == ".json" else yaml.safe_load(text)
    conn = _connect()
    imported = 0
    try:
        for pub_id, row in ((raw or {}).get("publications") or {}).items():
            _upsert(conn, _normalize_pub_id(pub_id), row or {}, "manual")
            imported += 1
        conn.commit()
    finally:
        conn.close()
    return {"imported": imported, "source": "manual"}


def get_latest_revision(pub_id: str, conn=None) -> dict | None:
    """Latest known revision for a publication, or None if uncached."""
    slug = _normalize_pub_id(pub_id)
    own = conn is None
    if own:
        conn = _connect()
    try:
        row = conn.execute(
            "SELECT * FROM docmod_nist_pubs WHERE pub_id = %s", (slug,)
        ).fetchone()
        return dict(row) if row else None
    finally:
        if own:
            conn.close()


def evidence_hash(conn=None) -> str:
    """sha256 over the cache contents — a refresh re-triggers policy_refs scans."""
    import hashlib

    own = conn is None
    if own:
        conn = _connect()
    try:
        rows = conn.execute(
            "SELECT pub_id, revision_num FROM docmod_nist_pubs ORDER BY pub_id"
        ).fetchall()
        payload = "|".join(f"{r['pub_id']}:{r['revision_num']}" for r in (dict(x) for x in rows))
    except Exception:
        try:
            conn.rollback()  # PG: failed statement poisons the transaction
        except Exception:
            pass
        payload = "no-nist-pubs"
    finally:
        if own:
            conn.close()
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description="docmod NIST publication revision cache")
    parser.add_argument("--seed", action="store_true", help="load static seed")
    parser.add_argument("--sync", action="store_true", help="live pull of NIST publications")
    parser.add_argument("--force", action="store_true", help="ignore cadence gate on --sync")
    parser.add_argument("--import", dest="import_path", metavar="PATH", help="air-gap bundle import")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)

    result: dict = {}
    if args.seed:
        result["seed"] = load_seed()
    if args.sync:
        result["sync"] = sync(force=args.force)
    if args.import_path:
        result["import"] = import_dataset(args.import_path)
    if not result:
        # Bare invocation: refresh the substrate rather than printing help. An
        # empty docmod_nist_pubs makes policy_refs answer 'unknown' forever.
        result["refresh"] = refresh(force=args.force)
    print(json.dumps(result, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
