# CUI // SP-CTI
"""
Curated deck: Innovation Lab — Executive Investment Brief.

The company name is NOT hardcoded: it is resolved from the `own_company` profile
in args/govcon_company_profiles.yaml (ICDEV is open source — no employer identity
ships in the repo). Slide copy carries a `{company}` token, substituted at build.

Generates a 12-slide, narrative/strategic deck inside the ICDEV Slides canvas.
Sources differentiators and quantified outcomes from:
  ICDEV_Capabilities_Challenges_Solutions.xlsx
and builds editable ROI placeholder models.

Usage:
    python tools/slides/curated_decks/innovation_lab_business_case.py

Output:
    Creates a slides_decks record + slides_slides rows, then builds a .pptx.
    Prints deck_id, pptx_path, and a one-line summary.
"""
from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

import sys

# kax-conflict-05: run by path, sys.path[0] is this file's own directory — never
# the import root. Bootstrap it before the first first-party import below.
# parents[N] is whatever holds this file's `tools` package: the repo root in
# tools/, and <repo>/icdev in the icdev/ mirror (which is what a wheel ships).
_REPO_ROOT = Path(__file__).resolve().parents[3]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from tools.logging.icdev_logger import get_logger
# Use the same namespace as the rest of tools/slides/ to avoid shim/object mismatch.
from tools.slides.db.init_db import get_connection, init_db
from tools.slides.pptx_builder import build

logger = get_logger(__name__)

ROOT = Path(__file__).resolve().parents[3]
EXCEL_PATH = ROOT / "ICDEV_Capabilities_Challenges_Solutions.xlsx"

DECK_TITLE = "{company} Innovation Lab: Infrastructure for the Next Generation of Government Technology"
THEME = "midnight_executive"
DECK_TYPE = "executive_overview"


# ── Excel parsing helpers ─────────────────────────────────────────────────────

@dataclass
class ExcelHighlights:
    challenges: list[dict]
    capabilities: list[dict]
    canvases: list[dict]


def _fallback_highlights() -> ExcelHighlights:
    """Inline highlights used if openpyxl is unavailable or the workbook is missing."""
    return ExcelHighlights(
        challenges=[
            {
                "domain": "Compliance & ATO",
                "challenge": "Authority to Operate (ATO) takes 12-18 months of manual paperwork",
                "solution": "Treat compliance as code: auto-generate ATO artifacts directly from the build with continuous (cATO) monitoring on an append-only audit trail.",
                "outcome": "18-month ATO cycle → 6-8 weeks target; evidence auto-generated",
            },
            {
                "domain": "Risk & Simulation",
                "challenge": "Program risk is assessed after the fact, with no real what-if capability",
                "solution": "Digital Program Twin runs Monte Carlo simulation across schedule, cost, risk, compliance, architecture, and staffing, generating three Courses of Action.",
                "outcome": "Spreadsheet guesswork → 6-dimension simulation + 3 scored COAs",
            },
            {
                "domain": "Security & ZTA",
                "challenge": "Zero Trust is a slogan; segmentation is static and unmeasured",
                "solution": "Score NIST 800-207 ZTA maturity, harden the DevSecOps pipeline, and model network/security topology on interactive canvases.",
                "outcome": "Static segmentation → scored ZTA maturity + modeled controls",
            },
            {
                "domain": "Supply Chain",
                "challenge": "Vendor and dependency risk enters the system unmanaged",
                "solution": "Supply-chain intelligence aggregates the dependency graph, triages CVEs with SLA tracking, warns on ISA expiry, and screens for NDAA Section 889 with SBOMs on every build.",
                "outcome": "Reactive incidents → proactive, SLA-tracked supply-chain risk",
            },
        ],
        capabilities=[
            {"name": "FORGE Framework", "what": "Separates probabilistic AI reasoning from deterministic tool execution across six layers so multi-step workflows stay reliable."},
            {"name": "ANVIL TDD Build", "what": "5-phase true test-driven workflow that writes failing tests first, generates implementation, then runs adversarial critique and security scans."},
            {"name": "Compliance & ATO Automation", "what": "Generates SSP, POAM, STIG checklist, and SBOM directly from the build with CUI markings and cATO monitoring."},
            {"name": "42-Framework Crosswalk", "what": "Maps one implemented control to 42+ compliance frameworks automatically and flags coverage gaps."},
            {"name": "Requirements Intake (RICOAS)", "what": "Detects ambiguity and missing security/compliance requirements and decomposes them into SAFe epics/stories."},
            {"name": "AI Security (MITRE ATLAS)", "what": "Defends the AI supply chain with prompt-injection detection, prompt-chain validation, and adversarial-ML red-teaming."},
        ],
        canvases=[
            {"name": "NDC — Network Design Canvas", "use_case": "Living, validated network topology with ACAS/Nessus overlay and NetBox sync."},
            {"name": "IDC — Infrastructure Design Canvas", "use_case": "Multi-CSP design with CSP equivalence mapping and one-click IaC emit."},
            {"name": "SDC — Security Design Canvas", "use_case": "STRIDE threat modeling with NIST/FedRAMP/CMMC control mapping and attack-path digital twin."},
            {"name": "AADC — Agentic AI Design Canvas", "use_case": "40+ node types, 7 vetted solution packs, risk register, and MITRE ATLAS scenarios."},
            {"name": "DIC — Document Intelligence Canvas", "use_case": "Own RAG+KG over documents with grounded, cited, no-LLM search and AI-labeled generation."},
        ],
    )


def parse_excel(path: Path) -> ExcelHighlights:
    """Parse the Excel workbook and return sanitized highlights."""
    try:
        import openpyxl
    except Exception:
        logger.warning("openpyxl unavailable; using fallback highlights")
        return _fallback_highlights()

    if not path.exists():
        logger.warning("%s not found; using fallback highlights", path)
        return _fallback_highlights()

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        logger.warning("failed to load workbook (%s); using fallback highlights", exc)
        return _fallback_highlights()

    def _sheet_rows(name: str, max_rows: int = 12) -> list[tuple]:
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        rows: list[tuple] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append(row)
        return rows

    # ── Challenges & Solutions ──────────────────────────────────────────────
    rows = _sheet_rows("Challenges & Solutions", max_rows=15)
    challenges: list[dict] = []
    # Header is row index 3 (0-based: 3) based on prior inspection.
    for row in rows[4:]:
        if not row or not row[0]:
            continue
        try:
            num = int(row[0])
        except Exception:
            continue
        challenges.append({
            "num": num,
            "domain": str(row[1] or "").strip(),
            "challenge": str(row[2] or "").strip(),
            "impact": str(row[3] or "").strip(),
            "solution": str(row[4] or "").strip(),
            "modules": str(row[5] or "").strip(),
            "frameworks": str(row[6] or "").strip(),
            "outcome": str(row[7] or "").strip(),
        })

    # ── Capability Catalog ────────────────────────────────────────────────────
    rows = _sheet_rows("Capability Catalog", max_rows=15)
    capabilities: list[dict] = []
    for row in rows[4:]:
        if not row or not row[0]:
            continue
        try:
            num = int(row[0])
        except Exception:
            continue
        capabilities.append({
            "num": num,
            "name": str(row[1] or "").strip(),
            "what": str(row[2] or "").strip(),
            "modules": str(row[3] or "").strip(),
            "frameworks": str(row[4] or "").strip(),
            "solves": str(row[5] or "").strip(),
        })

    # ── Design Canvases ─────────────────────────────────────────────────────
    rows = _sheet_rows("Design Canvases", max_rows=15)
    canvases: list[dict] = []
    for row in rows[4:]:
        if not row or not row[0]:
            continue
        try:
            num = int(row[0])
        except Exception:
            continue
        canvases.append({
            "num": num,
            "name": str(row[1] or "").strip().replace("\n", " "),
            "route": str(row[2] or "").strip(),
            "use_case": str(row[3] or "").strip(),
            "challenge": str(row[4] or "").strip(),
            "how": str(row[5] or "").strip(),
        })

    if not challenges:
        return _fallback_highlights()

    return ExcelHighlights(
        challenges=challenges[:4],
        capabilities=capabilities[:6],
        canvases=canvases[:5],
    )


# ── Deck content ──────────────────────────────────────────────────────────────

def company_name() -> str:
    """The company this deck is for, from the `own_company` GovCon profile.

    ICDEV is open source: no employer's identity is hardcoded anywhere in it. An
    unfilled profile yields a neutral placeholder rather than a wrong name — a
    deck that says "Our Company" is obviously a template; one that says the wrong
    company is a mistake you hand to an executive.
    """
    try:
        import yaml

        path = ROOT / "args" / "govcon_company_profiles.yaml"
        profiles = (yaml.safe_load(path.read_text(encoding="utf-8")) or {}).get("profiles", {})
        name = (profiles.get("own_company", {}) or {}).get("entity_name", "").strip()
    except Exception:  # noqa: BLE001 — a missing profile must never break the build
        name = ""
    # The shipped template literally reads "[YOUR COMPANY NAME]".
    if not name or name.startswith("["):
        return "Our Company"
    return name


def _resolve(value, company: str):
    """Substitute the {company} token through nested slide structures."""
    if isinstance(value, str):
        return value.replace("{company}", company)
    if isinstance(value, list):
        return [_resolve(v, company) for v in value]
    if isinstance(value, dict):
        return {k: _resolve(v, company) for k, v in value.items()}
    return value


def build_slides(highlights: ExcelHighlights) -> list[dict]:
    ch = highlights.challenges
    cap = highlights.capabilities
    cvs = highlights.canvases

    # ROI placeholder assumptions — clearly labeled as editable.
    roi_assumptions = {
        "active_pipeline_m": 500,
        "win_rate_lift_pct": 15,
        "reshaped_opps_per_year": 4,
        "avg_rfp_size_m": 50,
        "retained_top_talent": 10,
        "loaded_cost_per_hire_k": 200,
        "partner_hosts": 5,
        "hosting_fee_per_partner_k": 100,
        "infrastructure_investment_m": 5,
    }
    pipeline_lift = roi_assumptions["active_pipeline_m"] * (roi_assumptions["win_rate_lift_pct"] / 100)
    reshaped_value = roi_assumptions["reshaped_opps_per_year"] * roi_assumptions["avg_rfp_size_m"]
    recruitment_value = roi_assumptions["retained_top_talent"] * (roi_assumptions["loaded_cost_per_hire_k"] / 1000)
    partner_value = roi_assumptions["partner_hosts"] * (roi_assumptions["hosting_fee_per_partner_k"] / 1000)
    total_annual_value = pipeline_lift + reshaped_value + recruitment_value + partner_value

    first_challenge = ch[0] if ch else _fallback_highlights().challenges[0]
    second_challenge = ch[1] if len(ch) > 1 else _fallback_highlights().challenges[1]

    company = company_name()
    return _resolve([
        {
            "position": 1,
            "slide_type": "title",
            "title": "{company} Innovation Lab",
            "bullets": [
                "Infrastructure → Application/AI/Agentic Development → Hosting",
                "Not another lab. A revenue-generating, talent-magnet, thought-leadership engine.",
            ],
            "speaker_notes": (
                "Open with the framing: this lab is not a showcase room with whiteboards and VR headsets. "
                "It is an operating capability — infrastructure as the foundation, AI/ML and agentic development as the engine, "
                "and hosting as the outward-facing proof-of-value for customers, partners, and future hires."
            ),
        },
        {
            "position": 2,
            "slide_type": "content",
            "title": "The Ask",
            "bullets": [
                "Build a secure, scalable innovation infrastructure stack: compute, network, GPU, cloud, and K8s.",
                "Stand up an AI/ML + Agentic AI development factory on top of that foundation.",
                "Operate a hosting layer where partners and vendors demonstrate SOTA products and digital-twin customer environments for POCs.",
                "Target decision: approve investment to build Phase 1 (infrastructure + factory) and pilot hosting with 2–3 partners.",
            ],
            "speaker_notes": (
                "Be explicit about the investment layers. Infrastructure is not optional — it is the prerequisite for everything else. "
                "Hosting is the revenue and recruiting accelerator. Ask for a Phase 1 go/no-go, not a full multi-year commitment."
            ),
        },
        {
            "position": 3,
            "slide_type": "content",
            "title": "Why Now?",
            "bullets": [
                "DoD/IC customers are drowning in RFI/RFP volume and expect vendor-led innovation, not slide decks.",
                "AI/ML and Agentic AI are moving from experiment to mission-critical procurement evaluation criteria.",
                "Digital Twin and interoperability testing are now table stakes for large systems integrators.",
                "Competitors are standing up labs; the window to be perceived as a thought leader is closing.",
            ],
            "speaker_notes": (
                "The market window is the urgency. Customers are asking 'show me, don't tell me' in RFPs. "
                "If {company} does not have a live environment where AI, digital twin, and interoperability can be demonstrated, "
                "we will be relegated to responding to other people's innovation in proposals."
            ),
        },
        {
            "position": 4,
            "slide_type": "content",
            "title": "So What? — Four Executive Outcomes",
            "bullets": [
                "Opportunity pipeline: accelerate RFI/RFP response and reshape future pursuits with live proof points.",
                "Recruitment pipeline: become the 'I want to work for {company} because they...' destination for top AI/ML and systems engineers.",
                "Customer stickiness: move from vendor to co-innovation partner through hosted POCs and digital twins.",
                "Thought leadership: demonstrate SOTA capabilities to customers, employees, partners, and future hires — not as slides, as working systems.",
            ],
            "speaker_notes": (
                "These are the only outcomes executives care about. Every capability we fund must map to one of these four. "
                "The lab is not a cost center if it produces pipeline, people, stickiness, and perception."
            ),
        },
        {
            "position": 5,
            "slide_type": "content",
            "title": "Why {company}? — Our Differentiators",
            "bullets": [
                f"Compliance-as-generator, not compliance-as-tax: {first_challenge['outcome']}.",
                f"Risk simulation before commitment: {second_challenge['outcome']}.",
                "Air-gap and classified-ready by design — not a cloud-only prototype.",
                "Private, security-hardened lab: DoD/IC interoperability testing without exposing customer data or relying on public sandboxes.",
                f"One integrated stack: {cap[0]['name'] if cap else 'FORGE'} separates AI reasoning from deterministic execution so missions stay reliable.",
            ],
            "speaker_notes": (
                "This is the answer to 'why not another lab?' {company} already has the compliance, security, and systems-integration DNA. "
                "The differentiators are not the hardware — they are the ability to generate ATO evidence, simulate risk, operate in classified environments, "
                "and do it on one deterministic platform instead of 12 disconnected tools."
            ),
        },
        {
            "position": 6,
            "slide_type": "content",
            "title": "The Three Layers",
            "bullets": [
                "1. Infrastructure — servers, networking, CPU/GPU, cloud fabric, K8s, storage, and security enclaves.",
                "2. Application, AI/ML, and Agentic AI Development — ANVIL TDD factory, agent orchestration, model training/inference, and governance.",
                "3. Hosting — partner/vendor showcase, customer digital-twin environments, and POC interoperability testbeds.",
                "Layer 1 enables Layer 2; Layer 2 produces the capabilities Layer 3 demonstrates.",
            ],
            "speaker_notes": (
                "This slide is the architecture. Do not let the conversation jump to hosting before infrastructure is funded. "
                "Hosting without a development layer is just a demo room. Hosting with a development layer is a perpetual innovation engine."
            ),
        },
        {
            "position": 7,
            "slide_type": "content",
            "title": "Use Case 1 — AI/ML & Agentic AI Factory",
            "bullets": [
                f"Develop and harden AI/ML pipelines using {cap[1]['name'] if len(cap) > 1 else 'ANVIL TDD'}: tests first, implementation second, adversarial critique third.",
                f"Design multi-agent systems on the {cvs[3]['name'] if len(cvs) > 3 else 'Agentic AI Design Canvas'} with built-in risk registers and MITRE ATLAS scenarios.",
                "Govern AI with model cards, transparency, and OMB M-25-21 / NIST AI 600-1 controls baked in.",
                "Outcome: deployable, governed agents — not experimental notebooks.",
            ],
            "speaker_notes": (
                "The AI factory use case answers the RFP evaluator who asks 'can {company} actually build and secure an agentic system?' "
                "ANVIL and the Agentic AI Design Canvas are the proof. They turn AI from a prototype into a product."
            ),
        },
        {
            "position": 8,
            "slide_type": "content",
            "title": "Use Case 2 — Digital Twin & Interoperability Testbed",
            "bullets": [
                "Mirror customer environments in a private, controlled sandbox for requirements validation and integration rehearsal.",
                "Run boundary impact analysis (BDC) so scope changes that would trigger re-authorization are caught before code is written.",
                f"Score NIST 800-207 ZTA maturity and model network/security topology with {cvs[0]['name'] if cvs else 'NDC'} + {cvs[2]['name'] if len(cvs) > 2 else 'SDC'}.",
                "Outcome: de-risk interoperability, accelerate ATO, and win the 'show me' moment in customer engagements.",
            ],
            "speaker_notes": (
                "Digital twin is the use case that changes customer conversations. Instead of promising interoperability in a proposal, "
                "we prove it in their mirror environment. The Boundary and Security Design Canvases keep the ATO boundary intact while we iterate."
            ),
        },
        {
            "position": 9,
            "slide_type": "content",
            "title": "Use Case 3 — Partner & Vendor Showcase Hosting",
            "bullets": [
                "Host partners and vendors who want to demonstrate SOTA products with {company} customers.",
                "Provide digital-twin customer environments so partners can build POCs against realistic, representative topologies.",
                "Generate co-branded thought leadership, capture RFI/RFP intel, and identify resale/OEM opportunities.",
                "Outcome: revenue from hosting fees, partner-funded capability expansion, and a continuous pipeline of shaped opportunities.",
            ],
            "speaker_notes": (
                "Hosting is the business model layer. Partners pay to access our environment and our customer relationships. "
                "Customers see live solutions. {company} captures the intellectual property of how those solutions map to mission problems. "
                "This is how the lab pays for itself."
            ),
        },
        {
            "position": 10,
            "slide_type": "content",
            "title": "ROI Model — Editable Placeholders",
            "bullets": [
                f"Active opportunity pipeline uplift: ${pipeline_lift:.0f}M/year ({roi_assumptions['win_rate_lift_pct']}% lift on ${roi_assumptions['active_pipeline_m']}M pipeline).",
                f"Reshaped opportunities: ${reshaped_value:.0f}M/year ({roi_assumptions['reshaped_opps_per_year']} opps × ${roi_assumptions['avg_rfp_size_m']}M average RFP).",
                f"Recruitment/retention value: ${recruitment_value:.1f}M/year ({roi_assumptions['retained_top_talent']} hires × ${roi_assumptions['loaded_cost_per_hire_k']}K loaded cost).",
                f"Partner hosting revenue: ${partner_value:.1f}M/year ({roi_assumptions['partner_hosts']} partners × ${roi_assumptions['hosting_fee_per_partner_k']}K/year).",
                f"Indicative annual value: ${total_annual_value:.1f}M vs. Phase 1 investment of ${roi_assumptions['infrastructure_investment_m']}M.",
            ],
            "speaker_notes": (
                "These numbers are placeholders. The value statement is in the structure, not the exact dollars. "
                f"Assumptions: ${roi_assumptions['active_pipeline_m']}M active pipeline, {roi_assumptions['win_rate_lift_pct']}% win-rate lift, "
                f"{roi_assumptions['reshaped_opps_per_year']} reshaped RFPs at ${roi_assumptions['avg_rfp_size_m']}M each, "
                f"{roi_assumptions['retained_top_talent']} retained/attracted hires, and {roi_assumptions['partner_hosts']} hosted partners. "
                "Replace each assumption with {company}-specific data before the final readout."
            ),
        },
        {
            "position": 11,
            "slide_type": "content",
            "title": "Risk Mitigation & Path Forward",
            "bullets": [
                "Phase 1 (0–6 months): secure infrastructure, AI/ML factory, and 1–2 internal pilot use cases.",
                "Phase 2 (6–12 months): partner onboarding, digital-twin hosting, and first customer-facing demonstrations.",
                "Phase 3 (12–24 months): scaled hosting revenue stream, recurring cATO monitoring, and interoperability certification offerings.",
                "Governance: executive steering, security/ATO gate, and quarterly ROI reconciliation using the same audit trail that powers compliance.",
            ],
            "speaker_notes": (
                "De-risk the investment by phasing it. Phase 1 proves the stack on internal problems. "
                "Phase 2 brings in partners. Phase 3 turns hosting into a recurring revenue and thought-leadership flywheel. "
                "Governance is built on ICDEV's own append-only audit trail — the same mechanism we sell to customers."
            ),
        },
        {
            "position": 12,
            "slide_type": "outro",
            "title": "The Decision",
            "bullets": [
                "Approve Phase 1 investment in infrastructure + AI/ML/Agentic factory.",
                "Authorize pilot hosting agreements with 2–3 strategic partners.",
                "Position {company} as the thought-leading, co-innovation partner for DoD/IC transformation.",
            ],
            "speaker_notes": (
                "Close with the decision. The deck is not asking for a lab. It is asking for a capability that produces pipeline, people, "
                "and perception. Every month of delay is a month competitors spend answering 'show me' while {company} is still telling."
            ),
        },
    ], company)


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    init_db()
    highlights = parse_excel(EXCEL_PATH)
    slides = build_slides(highlights)
    deck_title = DECK_TITLE.replace("{company}", company_name())

    conn = get_connection()
    try:
        cur = conn.cursor()

        # Insert deck record
        cur.execute(
            "INSERT INTO slides_decks (title, deck_type, theme, status, source_types, slide_count) "
            "VALUES (%s, %s, %s, 'pending', %s, 0) RETURNING deck_id",
            (deck_title, DECK_TYPE, THEME, json.dumps(["curated", "excel_capabilities"])),
        )
        row = cur.fetchone()
        deck_id = row["deck_id"] if isinstance(row, dict) else row[0]

        # Insert slides
        for s in slides:
            cur.execute(
                "INSERT INTO slides_slides (deck_id, position, slide_type, title, bullets, speaker_notes, image_path) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s)",
                (
                    deck_id,
                    s["position"],
                    s["slide_type"],
                    s["title"],
                    json.dumps(s["bullets"]),
                    s.get("speaker_notes", ""),
                    s.get("image_path", ""),
                ),
            )
        conn.commit()

        # Build PPTX
        slide_dicts = [
            {
                "title": s["title"],
                "bullets": s["bullets"],
                "speaker_notes": s.get("speaker_notes", ""),
                "slide_type": s["slide_type"],
                "image_path": s.get("image_path", ""),
            }
            for s in slides
        ]
        pptx_path = build(slide_dicts, theme=THEME, title=deck_title)

        # Finalize deck record
        cur.execute(
            "UPDATE slides_decks SET status='completed', pptx_path=%s, slide_count=%s, completed_at=CURRENT_TIMESTAMP WHERE deck_id=%s",
            (pptx_path, len(slide_dicts), deck_id),
        )
        conn.commit()

        print(f"DECK_ID={deck_id}")
        print(f"PPTX_PATH={pptx_path}")
        print(f"SLIDE_COUNT={len(slide_dicts)}")
        print(f"TITLE={deck_title}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
