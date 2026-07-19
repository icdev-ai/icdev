# CUI // SP-CTI
"""ICDEV Network Design Canvas -- misc route group.

Extracted verbatim from tools/network/blueprint.py (cvx-net-01 monolith split).
Registered on the shared NDC blueprint via register_misc_routes(bp).
"""
from __future__ import annotations

import json
import uuid as _uuid
from datetime import datetime, timezone
from flask import abort, jsonify, redirect, render_template, request, session
from tools.network.routes._common import logger
from tools.db.storage import sql_placeholder
from tools.network.blueprint_helpers import _audit, _now, _row_to_dict, nc_login_required
from tools.network.constants import BOM_COSTS
from tools.network.db.init_db import get_connection


def register_misc_routes(bp):
    """Register misc routes on the NDC blueprint."""

    @bp.route("/projects/<project_id>")
    @nc_login_required
    def nc_project_dashboard(project_id):
        """Unified 4-panel view: phases + canvas + SOPs per phase + traffic flows."""
        conn = get_connection()
        _ph = sql_placeholder(conn)
        try:
            project_row = conn.execute(
                f"SELECT * FROM nc_projects WHERE id={_ph}", (project_id,)
            ).fetchone()
            if not project_row:
                return "Project not found", 404
            project = dict(project_row)

            phases = [dict(r) for r in conn.execute(
                f"SELECT * FROM nc_migration_phases WHERE project_id={_ph} ORDER BY phase_num",
                (project_id,),
            ).fetchall()]

            topo_row = conn.execute(
                "SELECT t.id, t.name, t.graph_json FROM topologies t "
                "JOIN nc_project_topologies pt ON pt.topology_id = t.id "
                f"WHERE pt.project_id={_ph} LIMIT 1",
                (project_id,),
            ).fetchone()
            topology = dict(topo_row) if topo_row else None
            topo_id = topology["id"] if topology else None

            snapshots = [dict(r) for r in conn.execute(
                "SELECT id, phase_id, label, created_at FROM nc_topology_snapshots "
                f"WHERE topo_id={_ph} ORDER BY created_at DESC",
                (topo_id,),
            ).fetchall()] if topo_id else []

            phase_ids = [ph["id"] for ph in phases]
            sops_by_phase: dict = {ph["id"]: [] for ph in phases}
            if phase_ids:
                placeholders = ",".join([_ph] * len(phase_ids))
                doc_rows = conn.execute(
                    f"SELECT pd.phase_id, s.sop_id, s.title, s.category, s.csp "
                    f"FROM nc_phase_documents pd "
                    f"JOIN ndc_sops s ON s.sop_id = pd.doc_id "
                    f"WHERE pd.phase_id IN ({placeholders}) AND pd.doc_source='sop'",
                    phase_ids,
                ).fetchall()
                for r in doc_rows:
                    sops_by_phase.setdefault(r[0], []).append({
                        "sop_id": r[1], "title": r[2], "category": r[3], "csp": r[4]
                    })

            flows_by_phase: dict = {ph["id"]: [] for ph in phases}
            if topo_id and phase_ids:
                placeholders = ",".join([_ph] * len(phase_ids))
                flow_rows = conn.execute(
                    f"SELECT id, name, src_zone, dst_zone, classification, phase_id "
                    f"FROM nc_traffic_flows WHERE topology_id={_ph} AND phase_id IN ({placeholders})",
                    [topo_id] + phase_ids,
                ).fetchall()
                for r in flow_rows:
                    flows_by_phase.setdefault(r[5], []).append({
                        "id": r[0], "name": r[1],
                        # response keys kept as source_zone/destination_zone for the
                        # project_dashboard.html template contract; DB columns are src_zone/dst_zone.
                        "source_zone": r[2], "destination_zone": r[3],
                        "classification": r[4],
                    })

            return render_template(
                "network/project_dashboard.html",
                project=project,
                phases=phases,
                topology=topology,
                topo_id=topo_id,
                snapshots=snapshots,
                sops_by_phase=sops_by_phase,
                flows_by_phase=flows_by_phase,
            )
        finally:
            conn.close()

    # ── AI Trace API ────────────────────────────────────────────────────────
    @bp.route("/api/ai-trace")
    @nc_login_required
    def nc_api_ai_trace():
        """Return recent AI decisions made by NDC assessment engines."""
        limit = min(int(request.args.get("limit", 50)), 200)
        record_id = request.args.get("record_id")
        try:
            from tools.db.storage import get_connection as _gc
            with _gc() as _conn:
                _ph = sql_placeholder(_conn)
                if record_id:
                    rows = _conn.execute(
                        f"SELECT * FROM canvas_ai_decisions WHERE canvas_type='ndc' AND record_id={_ph} "
                        f"ORDER BY created_at DESC LIMIT {_ph}",
                        (record_id, limit),
                    ).fetchall()
                else:
                    rows = _conn.execute(
                        "SELECT * FROM canvas_ai_decisions WHERE canvas_type='ndc' "
                        f"ORDER BY created_at DESC LIMIT {_ph}",
                        (limit,),
                    ).fetchall()
            return jsonify({"ok": True, "canvas": "ndc", "decisions": [dict(r) for r in rows]})
        except Exception as exc:
            return jsonify({"ok": False, "error": str(exc)}), 500

    # ── FCC Compliance ──────────────────────────────────────────────────────────

    @bp.route("/fcc")
    @nc_login_required
    def network_fcc():
        from tools.network.fcc_compliance import (
            calea_checklist, part36_assessment,
            nanp_number_inventory, e911_capability_check,
        )
        checks = {}
        for name, fn in [
            ("calea", calea_checklist),
            ("part36", part36_assessment),
            ("nanp", nanp_number_inventory),
            ("e911", e911_capability_check),
        ]:
            try:
                checks[name] = fn()
            except Exception as exc:
                checks[name] = {"error": str(exc)}
        return render_template("network/fcc_compliance.html", checks=checks)

    @bp.route("/api/fcc/<check_type>")
    def api_network_fcc(check_type):
        from tools.network.fcc_compliance import (
            calea_checklist, part36_assessment,
            nanp_number_inventory, e911_capability_check,
        )
        _CHECK_MAP = {
            "calea":  calea_checklist,
            "part36": part36_assessment,
            "nanp":   nanp_number_inventory,
            "e911":   e911_capability_check,
        }
        if check_type == "all":
            result = {}
            for name, fn in _CHECK_MAP.items():
                try:
                    result[name] = fn()
                except Exception as exc:
                    result[name] = {"error": str(exc)}
            return jsonify(result)
        fn = _CHECK_MAP.get(check_type)
        if not fn:
            return jsonify({"error": f"Unknown check: {check_type}. Valid: {sorted(_CHECK_MAP)}"}), 400
        try:
            return jsonify(fn())
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    # ── Presentation (exec review) ─────────────────────────────────────────
    @bp.route("/projects/<pid>/presentation")
    @nc_login_required
    def nc_project_presentation(pid):
        conn = get_connection()
        _ph = sql_placeholder(conn)
        proj = conn.execute(f"SELECT * FROM nc_projects WHERE id={_ph}", (pid,)).fetchone()
        if not proj:
            conn.close()
            abort(404)
        proj = _row_to_dict(proj)
        conn.close()
        return render_template("network/presentation.html", project=proj)

    @bp.route("/api/projects/<pid>/presentation")
    @nc_login_required
    def nc_api_project_presentation(pid):
        conn = get_connection()
        _ph = sql_placeholder(conn)
        proj = conn.execute(
            "SELECT p.*, c.name AS customer_name FROM nc_projects p "
            f"LEFT JOIN nc_customers c ON c.id=p.customer_id WHERE p.id={_ph}", (pid,)
        ).fetchone()
        if not proj:
            conn.close()
            return jsonify({"error": "Project not found"}), 404
        proj = _row_to_dict(proj)
        topos = []
        for r in conn.execute(
            "SELECT t.id, t.name, t.classification, t.graph_json "
            "FROM topologies t JOIN nc_project_topologies pt ON pt.topology_id=t.id "
            f"WHERE pt.project_id={_ph}", (pid,)
        ).fetchall():
            t = _row_to_dict(r)
            try:
                g = json.loads(t.get("graph_json") or '{"nodes":[],"edges":[]}')
            except Exception:
                g = {"nodes": [], "edges": []}
            t["node_count"] = len(g.get("nodes", []))
            t["edge_count"] = len(g.get("edges", []))
            topos.append(t)
        circuits = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT * FROM nc_circuits WHERE topology_id IN "
                f"(SELECT topology_id FROM nc_project_topologies WHERE project_id={_ph})", (pid,)
            ).fetchall()
        ]
        milestones = [
            _row_to_dict(r)
            for r in conn.execute(f"SELECT * FROM nc_project_milestones WHERE project_id={_ph}", (pid,)).fetchall()
        ]
        reviews = [
            _row_to_dict(r)
            for r in conn.execute(
                "SELECT br.*, rb.name AS board_name "
                "FROM nc_board_reviews br JOIN nc_review_boards rb ON rb.id=br.board_id "
                f"WHERE br.project_id={_ph} ORDER BY br.phase", (pid,)
            ).fetchall()
        ]
        safe_bridge = conn.execute(f"SELECT * FROM nc_safe_bridge WHERE project_id={_ph}", (pid,)).fetchone()
        safe_bridge = _row_to_dict(safe_bridge) if safe_bridge else None
        roi = {}
        if safe_bridge and safe_bridge.get("roi_json"):
            try:
                roi = json.loads(safe_bridge["roi_json"])
            except Exception:
                pass
        agg_audit = conn.execute(
            "SELECT SUM(passed), SUM(failed) FROM nc_compliance_checks "
            f"WHERE topology_id IN (SELECT topology_id FROM nc_project_topologies WHERE project_id={_ph})", (pid,)
        ).fetchone()
        passed = agg_audit[0] or 0
        failed = agg_audit[1] or 0
        total = passed + failed
        compliance_pct = round(passed * 100 / total) if total else None
        cat1 = conn.execute(
            "SELECT COUNT(*) FROM nc_compliance_findings "
            f"WHERE topology_id IN (SELECT topology_id FROM nc_project_topologies WHERE project_id={_ph}) "
            "AND status='open' AND severity='CAT1'", (pid,)
        ).fetchone()[0] or 0
        # Pre-compute CapEx before closing connection
        total_capex = 0
        for t in topos:
            trow = conn.execute(f"SELECT graph_json FROM topologies WHERE id={_ph}", (t["id"],)).fetchone()
            if trow and trow[0]:
                try:
                    nodes = json.loads(trow[0]).get("nodes", [])
                    total_capex += sum(BOM_COSTS.get(n.get("type", "unknown"), 0) for n in nodes)
                except Exception:
                    pass
        conn.close()
        return jsonify({
            "title": proj["name"],
            "status": proj.get("status", "draft"),
            "owner": proj.get("owner"),
            "description": proj.get("description"),
            "generated_at": _now(),
            "executive_summary": {
                "topology_count": len(topos),
                "total_devices": sum(t.get("node_count", 0) for t in topos),
                "compliance_pct": compliance_pct,
                "cat1_findings": cat1,
                "total_capex": total_capex,
                "monthly_circuit_cost": sum(c.get("monthly_cost_usd") or 0 for c in circuits),
                "roi": roi,
                "justification": safe_bridge.get("justification") if safe_bridge else None,
                "alternatives": safe_bridge.get("alternatives") if safe_bridge else None,
            },
            "topologies": topos,
            "circuits": circuits,
            "milestones": milestones,
            "review_history": reviews,
        })

    # ── Placeholder / redirect routes for migration hub quick actions ────────
    @bp.route("/migration-wizard")
    @nc_login_required
    def nc_migration_wizard():
        """Redirect to migration hub — wizard is embedded there."""
        return redirect("/network/migration-hub")

    @bp.route("/wave-planner")
    @nc_login_required
    def nc_wave_planner():
        """Placeholder wave planner — redirects to migration hub."""
        return redirect("/network/migration-hub")

    @bp.route("/port-mapping")
    @nc_login_required
    def nc_port_mapping_page():
        """Standalone port mapping — redirect to first project with port mapping."""
        conn = get_connection()
        proj = conn.execute(
            "SELECT id FROM nc_projects WHERE selected_coa > 0 LIMIT 1"
        ).fetchone()
        if not proj:
            proj = conn.execute("SELECT id FROM nc_projects LIMIT 1").fetchone()
        conn.close()
        if proj:
            return redirect(f"/network/projects/{proj['id']}#port-mapping-section")
        return redirect("/network/migration-hub")

    @bp.route("/documents")
    @nc_login_required
    def nc_documents():
        """Document library — redirect to SOPs for now."""
        return redirect("/network/sops")

    # ── Advisory History ───────────────────────────────────────────────────

    @bp.route("/advisory-history")
    @nc_login_required
    def nc_advisory_history():
        from tools.network.advisory import list_advisories, list_vendors
        advisories = list_advisories(
            vendor=request.args.get("vendor"),
            severity=request.args.get("severity"),
            status=request.args.get("status"),
            date_from=request.args.get("date_from"),
            date_to=request.args.get("date_to"),
        )
        vendors = list_vendors()
        return render_template(
            "network/advisory_history.html",
            advisories=advisories,
            vendors=vendors,
        )

    @bp.route("/api/advisories/export")
    @nc_login_required
    def nc_api_advisories_export():
        import csv
        import io
        from flask import Response
        from tools.network.advisory import list_advisories
        items = list_advisories()
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["CVE ID", "Vendor", "Severity", "Date", "Total Devices",
                    "Impacted", "Remediation %", "Data Source", "HITL Status", "Status"])
        for a in items:
            w.writerow([
                a.get("cve_id", ""), a.get("vendor", ""), a.get("severity", ""),
                (a.get("published_date") or "")[:10], a.get("total_devices", 0),
                a.get("impacted_devices", 0), a.get("remediation_pct", 0),
                a.get("data_source", ""), a.get("hitl_status", ""), a.get("status", ""),
            ])
        return Response(buf.getvalue(), mimetype="text/csv",
                        headers={"Content-Disposition": "attachment; filename=advisories.csv"})

    # ── POAM ───────────────────────────────────────────────────────────────

    @bp.route("/poam")
    @nc_login_required
    def nc_poam():
        from tools.network.poam_generator import list_poam_items
        from datetime import date
        advisory_filter = request.args.get("advisory")
        items = list_poam_items()
        return render_template(
            "network/poam.html",
            items=items,
            advisory_filter=advisory_filter,
            today=date.today().isoformat(),
        )

    @bp.route("/api/poam/generate", methods=["POST"])
    @nc_login_required
    def nc_api_poam_generate():
        from tools.network.poam_generator import generate_poam_item
        data = request.get_json(silent=True) or {}
        advisory_id = data.get("advisory_id", "")
        try:
            item = generate_poam_item(advisory_id, data)
            return jsonify({"ok": True, "item": item})
        except Exception as exc:
            logger.exception("POAM generate failed")
            return jsonify({"error": str(exc)}), 500

    @bp.route("/api/poam/export")
    @nc_login_required
    def nc_api_poam_export():
        from flask import Response
        from tools.network.poam_generator import export_poam
        fmt = request.args.get("format", "csv")
        content, mimetype = export_poam(fmt)
        ext = "json" if fmt == "json" else "csv"
        return Response(content, mimetype=mimetype,
                        headers={"Content-Disposition": f"attachment; filename=poam.{ext}"})

    # ── Exceptions ─────────────────────────────────────────────────────────

    @bp.route("/exceptions")
    @nc_login_required
    def nc_exceptions():
        from tools.network.exception_registry import list_exceptions
        from datetime import date
        exceptions = list_exceptions()
        return render_template(
            "network/exceptions.html",
            exceptions=exceptions,
            today=date.today().isoformat(),
        )

    @bp.route("/exceptions/file")
    @nc_login_required
    def nc_exceptions_file_form():
        """Redirect to exceptions page with file modal open."""
        return redirect("/network/exceptions?open_modal=file")

    @bp.route("/api/exception/file", methods=["POST"])
    @nc_login_required
    def nc_api_exception_file():
        from tools.network.exception_registry import file_exception
        data = request.get_json(silent=True) or {}
        try:
            exc = file_exception(data)
            return jsonify({"ok": True, "exception": exc})
        except Exception as e:
            logger.exception("Exception filing failed")
            return jsonify({"error": str(e)}), 500

    @bp.route("/api/exception/<exc_id>/approve", methods=["POST"])
    @nc_login_required
    def nc_api_exception_approve(exc_id):
        from tools.network.exception_registry import approve_exception
        data = request.get_json(silent=True) or {}
        level = data.get("level", "")
        approver = data.get("approver", "")
        if not level or not approver:
            return jsonify({"error": "level and approver are required"}), 400
        try:
            exc = approve_exception(exc_id, level, approver)
            return jsonify({"ok": True, "exception": exc})
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        except Exception as e:
            logger.exception("Exception approval failed")
            return jsonify({"error": str(e)}), 500

    # ── ATO Evidence Chain Export ──────────────────────────────────────────

    def _gather_ato_evidence(advisory_id, conn):
        """Collect all audit-chain rows for one advisory from available tables."""
        _ph = sql_placeholder(conn)

        def _q(sql, params=()):
            try:
                return [dict(r) for r in conn.execute(sql, params).fetchall()]
            except Exception:
                return []

        def _q1(sql, params=()):
            try:
                row = conn.execute(sql, params).fetchone()
                return dict(row) if row else None
            except Exception:
                return None

        advisory = _q1(f"SELECT * FROM nc_advisories WHERE id = {_ph}", (advisory_id,))

        assessments = _q(
            f"SELECT * FROM nc_advisory_assessments WHERE advisory_id = {_ph} ORDER BY created_at ASC",
            (advisory_id,),
        )

        rem_actions = _q(
            f"SELECT * FROM nc_remediation_actions WHERE advisory_id = {_ph} ORDER BY created_at ASC",
            (advisory_id,),
        )

        rem_status_log: list = []
        action_ids = [a.get("id") for a in rem_actions if a.get("id")]
        if action_ids:
            ph = ",".join([_ph] * len(action_ids))
            rem_status_log = _q(
                f"SELECT * FROM nc_remediation_status_log WHERE action_id IN ({ph}) ORDER BY created_at ASC",
                action_ids,
            )

        poam_items = _q(
            f"SELECT * FROM nc_poam_items WHERE advisory_id = {_ph} ORDER BY created_at ASC",
            (advisory_id,),
        )

        poam_status_log: list = []
        poam_ids = [p.get("id") for p in poam_items if p.get("id")]
        if poam_ids:
            ph = ",".join([_ph] * len(poam_ids))
            poam_status_log = _q(
                f"SELECT * FROM nc_poam_status_log WHERE poam_id IN ({ph}) ORDER BY created_at ASC",
                poam_ids,
            )

        exceptions = _q(
            f"SELECT * FROM nc_exceptions WHERE advisory_id = {_ph} ORDER BY created_at ASC",
            (advisory_id,),
        )

        exception_approvals: list = []
        exc_ids = [e.get("id") for e in exceptions if e.get("id")]
        if exc_ids:
            ph = ",".join([_ph] * len(exc_ids))
            exception_approvals = _q(
                f"SELECT * FROM nc_exception_approvals WHERE exception_id IN ({ph}) ORDER BY created_at ASC",
                exc_ids,
            )

        audit_log = _q(
            f"SELECT * FROM nc_nqe_audit_log WHERE advisory_id = {_ph} ORDER BY created_at ASC",
            (advisory_id,),
        )

        return {
            "advisory": advisory,
            "assessments": assessments,
            "remediation_actions": rem_actions,
            "remediation_status_log": rem_status_log,
            "poam_items": poam_items,
            "poam_status_log": poam_status_log,
            "exceptions": exceptions,
            "exception_approvals": exception_approvals,
            "audit_log": audit_log,
        }

    def _audit_ato_export(advisory_id, fmt, doc_hash):
        """Append an audit-log entry for the export (best-effort; non-blocking)."""
        try:
            conn = get_connection()
            _ph = sql_placeholder(conn)
            conn.execute(
                f"""INSERT INTO nc_nqe_audit_log
                   (session_id, user_session, action, input_text, result_summary,
                    raw_response_hash, advisory_id, created_at)
                   VALUES ({_ph},{_ph},{_ph},{_ph},{_ph},{_ph},{_ph},{_ph})""",

                (
                    session.get("session_id", ""),
                    session.get("user", ""),
                    "export",
                    fmt,
                    f"ATO evidence chain export (format={fmt})",
                    doc_hash,
                    advisory_id,
                    datetime.now(timezone.utc).isoformat(),
                ),
            )
            conn.commit()
            conn.close()
        except Exception:
            pass

    def _ato_safe(text):
        """Coerce to latin-1-safe string for fpdf2 core fonts."""
        _MAP = str.maketrans({
            "—": "-", "–": "-", "→": "->", "←": "<-",
            "≥": ">=", "≤": "<=", "°": "deg", "•": "*",
            "’": "'", "‘": "'", "“": '"', "”": '"',
            "™": "(TM)", "®": "(R)", "©": "(c)",
            "×": "x", "÷": "/",
        })
        text = str(text).translate(_MAP)
        return text.encode("latin-1", errors="replace").decode("latin-1")

    def _build_ato_pdf(evidence, advisory_id):
        """Return PDF bytes (fpdf2) or HTML bytes (fallback) for the evidence package."""
        _CLASSIFICATION = "CUI // SP-CTI"
        adv = evidence.get("advisory") or {}
        meta = evidence.get("_meta", {})

        def _cui_banner(pdf):
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_fill_color(180, 30, 30)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(0, 5, f"  {_CLASSIFICATION}", ln=True, fill=True)
            pdf.set_text_color(0, 0, 0)

        def _section_heading(pdf, title):
            pdf.set_font("Helvetica", "B", 11)
            pdf.set_fill_color(220, 230, 245)
            pdf.set_text_color(20, 60, 120)
            pdf.cell(0, 7, _ato_safe(title), ln=True, fill=True)
            pdf.set_text_color(0, 0, 0)
            pdf.ln(1)

        def _kv_row(pdf, label, value):
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(60, 5, _ato_safe(str(label) + ":"), ln=False)
            pdf.set_font("Helvetica", "", 8)
            pdf.multi_cell(0, 5, _ato_safe(str(value)[:300]))

        def _cui_footer_all(pdf):
            total = pdf.page
            for pno in range(1, total + 1):
                pdf.page = pno
                pdf.set_y(-12)
                pdf.set_font("Helvetica", "B", 7)
                pdf.set_fill_color(180, 30, 30)
                pdf.set_text_color(255, 255, 255)
                pdf.cell(0, 4,
                         f"  {_CLASSIFICATION} | Page {pno} of {total} | ICDEV ATO Evidence Package",
                         ln=True, fill=True)
            pdf.page = total

        try:
            from fpdf import FPDF

            pdf = FPDF(orientation="P", unit="mm", format="A4")
            pdf.set_auto_page_break(auto=True, margin=18)
            pdf.set_margins(15, 20, 15)

            # ── Cover ─────────────────────────────────────────────────────
            pdf.add_page()
            _cui_banner(pdf)
            pdf.ln(12)
            pdf.set_font("Helvetica", "B", 20)
            pdf.set_text_color(20, 60, 120)
            pdf.cell(0, 12, "ATO Evidence Package", ln=True, align="C")
            pdf.set_font("Helvetica", "B", 13)
            pdf.set_text_color(50, 50, 50)
            pdf.cell(0, 8, _ato_safe(f"Advisory ID: {advisory_id}  |  CVE: {adv.get('cve_id','N/A')}"), ln=True, align="C")
            pdf.set_font("Helvetica", "", 11)
            pdf.cell(0, 7, _ato_safe(f"Vendor: {adv.get('vendor','N/A')}  |  Severity: {str(adv.get('severity','N/A')).upper()}"), ln=True, align="C")
            pdf.cell(0, 7, _ato_safe(f"Status: {adv.get('status','N/A')}"), ln=True, align="C")
            pdf.ln(4)
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_text_color(100, 100, 100)
            pdf.cell(0, 5, _ato_safe(f"Exported: {meta.get('exported_at','')}"), ln=True, align="C")
            pdf.cell(0, 5, _ato_safe(f"SHA-256: {meta.get('doc_hash_sha256','')}"), ln=True, align="C")
            pdf.set_text_color(0, 0, 0)

            # ── Sec 1: Advisory Record ─────────────────────────────────────
            pdf.add_page()
            _cui_banner(pdf)
            pdf.ln(2)
            _section_heading(pdf, "1. Advisory Record")
            for label, key in [
                ("CVE ID", "cve_id"), ("Vendor", "vendor"), ("Severity", "severity"),
                ("Status", "status"), ("Published Date", "published_date"),
                ("Total Devices", "total_devices"), ("Impacted Devices", "impacted_devices"),
                ("Remediation %", "remediation_pct"), ("Data Source", "data_source"),
                ("HITL Status", "hitl_status"), ("HITL Approved By", "hitl_approved_by"),
                ("HITL Approved At", "hitl_approved_at"),
                ("Source Doc Hash (SHA-256)", "source_doc_hash"),
                ("Source Doc Format", "source_doc_format"),
                ("Extraction Confidence", "extraction_confidence"),
                ("Created", "created_at"), ("Updated", "updated_at"),
            ]:
                val = adv.get(key, "")
                if val not in (None, "", 0):
                    _kv_row(pdf, label, val)
            if adv.get("description"):
                pdf.ln(2)
                pdf.set_font("Helvetica", "B", 8)
                pdf.cell(0, 5, "Description:", ln=True)
                pdf.set_font("Helvetica", "", 8)
                pdf.multi_cell(0, 5, _ato_safe(str(adv["description"])[:1500]))

            # ── Sec 2: Impact Assessments ──────────────────────────────────
            for idx, asmt in enumerate(evidence.get("assessments", []), 1):
                pdf.add_page()
                _cui_banner(pdf)
                pdf.ln(2)
                _section_heading(pdf, f"2.{idx} Impact Assessment (ID: {asmt.get('id','')})")
                for label, key in [
                    ("Network ID", "network_id"), ("FWD Snapshot ID", "fwd_snapshot_id"),
                    ("Data Source", "data_source"),
                    ("NQL — Total Devices", "nql_total"), ("NQL — Impacted", "nql_impacted"),
                    ("NQL — AI Generated", "nql_ai_generated"), ("NQL — Template Based", "nql_template_based"),
                    ("Total Devices", "total_devices"), ("Impacted Count", "impacted_count"),
                    ("Raw Response Hash (SHA-256)", "raw_response_hash"),
                    ("AI Confidence", "ai_confidence"),
                    ("Cross-Val Delta %", "cross_validation_delta_pct"),
                    ("Cross-Val Warning", "cross_validation_warning"),
                    ("HITL Approved By", "approved_by"), ("HITL Approved At", "approved_at"),
                    ("Created", "created_at"),
                ]:
                    val = asmt.get(key, "")
                    if val not in (None, ""):
                        _kv_row(pdf, label, val)

            # ── Sec 3: Remediation Actions ─────────────────────────────────
            if evidence.get("remediation_actions"):
                pdf.add_page()
                _cui_banner(pdf)
                pdf.ln(2)
                _section_heading(pdf, "3. Remediation Actions")
                for ra in evidence["remediation_actions"]:
                    pdf.set_font("Helvetica", "BI", 9)
                    pdf.cell(0, 6, _ato_safe(f"Action {ra.get('id','')[:24]}"), ln=True)
                    for label, key in [
                        ("Device ID", "device_id"), ("Action Type", "action_type"),
                        ("Performed By", "performed_by"), ("Result", "result"),
                        ("Notes", "notes"), ("Created", "created_at"),
                    ]:
                        val = ra.get(key, "")
                        if val not in (None, ""):
                            _kv_row(pdf, label, val)
                    pdf.ln(2)
                if evidence.get("remediation_status_log"):
                    pdf.ln(2)
                    _section_heading(pdf, "3a. Remediation Status Log")
                    pdf.set_font("Helvetica", "", 8)
                    for sl in evidence["remediation_status_log"]:
                        pdf.cell(0, 5, _ato_safe(
                            f"{sl.get('created_at','')} | action={sl.get('action_id','')} | "
                            f"{sl.get('old_status','?')} -> {sl.get('new_status','')} | "
                            f"by={sl.get('updated_by','')}"
                        ), ln=True)

            # ── Sec 4: POAM Items ──────────────────────────────────────────
            if evidence.get("poam_items"):
                pdf.add_page()
                _cui_banner(pdf)
                pdf.ln(2)
                _section_heading(pdf, "4. POAM Items")
                for pi in evidence["poam_items"]:
                    pdf.set_font("Helvetica", "BI", 9)
                    pdf.cell(0, 6, _ato_safe(f"POAM {pi.get('poam_id', pi.get('id',''))}"), ln=True)
                    for label, key in [
                        ("CVE", "cve_id"), ("Weakness", "weakness"), ("Control ID", "control_id"),
                        ("Severity", "severity"), ("Status", "status"),
                        ("Scheduled Completion", "scheduled_completion"),
                        ("Responsible Party", "responsible_party"),
                        ("Twin Validated", "twin_validated"), ("Resources", "resources"),
                    ]:
                        val = pi.get(key, "")
                        if val not in (None, "", 0):
                            _kv_row(pdf, label, val)
                    pdf.ln(2)
                if evidence.get("poam_status_log"):
                    pdf.ln(2)
                    _section_heading(pdf, "4a. POAM Status Log")
                    pdf.set_font("Helvetica", "", 8)
                    for sl in evidence["poam_status_log"]:
                        pdf.cell(0, 5, _ato_safe(
                            f"{sl.get('created_at','')} | poam={sl.get('poam_id','')} | "
                            f"{sl.get('old_status','?')} -> {sl.get('new_status','')} | "
                            f"by={sl.get('updated_by','')}"
                        ), ln=True)

            # ── Sec 5: Exceptions ──────────────────────────────────────────
            if evidence.get("exceptions"):
                pdf.add_page()
                _cui_banner(pdf)
                pdf.ln(2)
                _section_heading(pdf, "5. Exceptions")
                for ex in evidence["exceptions"]:
                    pdf.set_font("Helvetica", "BI", 9)
                    pdf.cell(0, 6, _ato_safe(f"Device: {ex.get('device_name','')}  [{ex.get('status','')}]"), ln=True)
                    for label, key in [
                        ("Exception Type", "exception_type"), ("Risk Level", "risk_level"),
                        ("Expiry Date", "expiry_date"), ("Justification", "justification"),
                        ("ISSO Approved By", "isso_approved_by"), ("ISSO Approved At", "isso_approved_at"),
                        ("ISSM Approved By", "issm_approved_by"), ("ISSM Approved At", "issm_approved_at"),
                        ("AO Approved By", "ao_approved_by"), ("AO Approved At", "ao_approved_at"),
                        ("Compensating Controls", "compensating_controls"),
                        ("Risk Acceptance Level", "risk_acceptance_level"),
                        ("Filed By", "filed_by"), ("Created", "created_at"),
                    ]:
                        val = ex.get(key, "")
                        if val not in (None, "", 0):
                            _kv_row(pdf, label, val)
                    pdf.ln(2)
                if evidence.get("exception_approvals"):
                    pdf.ln(2)
                    _section_heading(pdf, "5a. Exception Approval Chain")
                    pdf.set_font("Helvetica", "", 8)
                    for ap in evidence["exception_approvals"]:
                        pdf.cell(0, 5, _ato_safe(
                            f"{ap.get('created_at','')} | exc={ap.get('exception_id','')} | "
                            f"{ap.get('approver_role','')} {ap.get('approver','')} -> "
                            f"{ap.get('decision','')} | {ap.get('conditions','')}"
                        ), ln=True)

            # ── Sec 6: NQE Audit Trail ─────────────────────────────────────
            if evidence.get("audit_log"):
                pdf.add_page()
                _cui_banner(pdf)
                pdf.ln(2)
                _section_heading(pdf, "6. NQE Audit Trail")
                pdf.set_font("Helvetica", "", 8)
                for entry in evidence["audit_log"]:
                    pdf.cell(0, 5, _ato_safe(
                        f"{entry.get('created_at','')} | {entry.get('action','')} | "
                        f"{str(entry.get('result_summary',''))[:100]}"
                    ), ln=True)

            # ── Hash integrity page ────────────────────────────────────────
            pdf.add_page()
            _cui_banner(pdf)
            pdf.ln(4)
            _section_heading(pdf, "Document Integrity")
            _kv_row(pdf, "Document SHA-256", meta.get("doc_hash_sha256", ""))
            _kv_row(pdf, "Exported At", meta.get("exported_at", ""))
            _kv_row(pdf, "Classification", meta.get("classification", ""))
            pdf.ln(4)
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_text_color(100, 100, 100)
            pdf.multi_cell(0, 5,
                "The SHA-256 hash above was computed over the canonical JSON serialization "
                "of all evidence fields. Re-compute to verify document integrity.")

            _cui_footer_all(pdf)
            return bytes(pdf.output())

        except ImportError:
            logger.warning("fpdf2 not installed — generating HTML fallback for ATO PDF")
            adv_rows = "".join(
                f"<tr><th>{k}</th><td>{v}</td></tr>"
                for k, v in (adv or {}).items()
                if v not in (None, "")
            )
            return (
                f"<!DOCTYPE html><html><head><meta charset='utf-8'>"
                f"<title>ATO Evidence — Advisory {advisory_id}</title>"
                f"<style>body{{font-family:monospace;margin:2em}}"
                f"table{{border-collapse:collapse;width:100%}}"
                f"th,td{{border:1px solid #ccc;padding:4px 8px;text-align:left}}"
                f"th{{background:#dde;width:220px}}"
                f".cui{{background:#b41e1e;color:#fff;padding:4px 10px;font-weight:bold}}"
                f"h2{{color:#1a3c78}}</style></head><body>"
                f"<div class='cui'>CUI // SP-CTI</div>"
                f"<h1>ATO Evidence Package — Advisory {advisory_id}</h1>"
                f"<p>Exported: {meta.get('exported_at','')} | "
                f"SHA-256: <code>{meta.get('doc_hash_sha256','')}</code></p>"
                f"<h2>1. Advisory Record</h2><table>{adv_rows}</table>"
                f"<p><em>Install fpdf2 for full multi-section PDF output.</em></p>"
                f"<div class='cui'>CUI // SP-CTI</div></body></html>"
            ).encode("utf-8")

    def _build_ato_excel(evidence, advisory_id):
        """Return (bytes, mimetype, extension) for Excel export.

        Uses openpyxl when available; falls back to a ZIP archive of CSV sheets.
        """
        import csv
        import io
        import zipfile

        SHEETS = [
            ("Advisory",            [evidence.get("advisory")] if evidence.get("advisory") else []),
            ("Assessments",         evidence.get("assessments", [])),
            ("RemediationActions",  evidence.get("remediation_actions", [])),
            ("RemediationStatusLog",evidence.get("remediation_status_log", [])),
            ("POAMItems",           evidence.get("poam_items", [])),
            ("POAMStatusLog",       evidence.get("poam_status_log", [])),
            ("Exceptions",          evidence.get("exceptions", [])),
            ("ExceptionApprovals",  evidence.get("exception_approvals", [])),
            ("AuditLog",            evidence.get("audit_log", [])),
        ]

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # drop default blank sheet

            _CUI_FILL = PatternFill("solid", fgColor="B41E1E")
            _HEAD_FILL = PatternFill("solid", fgColor="DDE4F0")
            _CUI_FONT = Font(bold=True, color="FFFFFF", size=9)
            _HEAD_FONT = Font(bold=True, size=9)
            _META = evidence.get("_meta", {})

            for sheet_name, rows in SHEETS:
                if not rows:
                    continue
                ws = wb.create_sheet(title=sheet_name[:31])
                # CUI banner row
                ws.append([f"CUI // SP-CTI | Advisory {advisory_id} | {_META.get('exported_at','')}"])
                cui_cell = ws.cell(1, 1)
                cui_cell.fill = _CUI_FILL
                cui_cell.font = _CUI_FONT
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(rows[0]) if rows else 1, 1))
                # Header row
                headers = list(rows[0].keys()) if rows else []
                ws.append(headers)
                for col_idx, _ in enumerate(headers, 1):
                    cell = ws.cell(2, col_idx)
                    cell.fill = _HEAD_FILL
                    cell.font = _HEAD_FONT
                # Data rows
                for row in rows:
                    ws.append([str(v) if v is not None else "" for v in row.values()])
                # Auto-width (capped)
                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=8)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"

        except ImportError:
            # Fallback: ZIP of CSV files (can be opened sheet-by-sheet)
            logger.info("openpyxl not installed — generating ZIP-of-CSVs fallback for ATO Excel")
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for sheet_name, rows in SHEETS:
                    if not rows:
                        continue
                    csv_buf = io.StringIO()
                    writer = csv.DictWriter(csv_buf, fieldnames=list(rows[0].keys()))
                    writer.writeheader()
                    writer.writerows(rows)
                    zf.writestr(f"{sheet_name}.csv", csv_buf.getvalue())
            return zip_buf.getvalue(), "application/zip", "zip"

    @bp.route("/api/advisory/<int:advisory_id>/export-ato", methods=["POST"])
    @nc_login_required
    def nc_api_advisory_export_ato(advisory_id: int):
        """Export complete ATO evidence chain for one advisory.

        Query param: format — json | pdf | excel (default: json)

        Evidence chain:
          1. nc_advisories row + source_doc_hash
          2. nc_advisory_assessments (NQL queries, sha256 hashes, dual-query
             reconciliation, HITL approval record)
          3. nc_remediation_actions (all statuses)
          4. nc_remediation_status_log
          5. nc_poam_items + nc_poam_status_log
          6. nc_exceptions + nc_exception_approvals + nc_nqe_audit_log
        """
        import hashlib
        from flask import Response

        fmt = request.args.get("format", "json").lower()
        if fmt not in ("json", "pdf", "excel"):
            return jsonify({"error": "format must be json, pdf, or excel"}), 400

        conn = get_connection()
        try:
            evidence = _gather_ato_evidence(advisory_id, conn)
        finally:
            conn.close()

        if evidence["advisory"] is None:
            return jsonify({"error": f"Advisory {advisory_id} not found"}), 404

        canonical = json.dumps(evidence, sort_keys=True, default=str)
        doc_hash = hashlib.sha256(canonical.encode()).hexdigest()
        evidence["_meta"] = {
            "advisory_id": advisory_id,
            "export_format": fmt,
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "doc_hash_sha256": doc_hash,
            "classification": "CUI // SP-CTI",
        }

        _audit_ato_export(advisory_id, fmt, doc_hash)

        if fmt == "json":
            return Response(
                json.dumps(evidence, indent=2, default=str),
                mimetype="application/json",
                headers={"Content-Disposition": f"attachment; filename=ato-evidence-{advisory_id}.json"},
            )
        if fmt == "pdf":
            pdf_bytes = _build_ato_pdf(evidence, advisory_id)
            mimetype = "application/pdf" if pdf_bytes[:4] == b"%PDF" else "text/html"
            ext = "pdf" if mimetype == "application/pdf" else "html"
            return Response(
                pdf_bytes,
                mimetype=mimetype,
                headers={"Content-Disposition": f"attachment; filename=ato-evidence-{advisory_id}.{ext}"},
            )
        # excel
        xl_bytes, mimetype, ext = _build_ato_excel(evidence, advisory_id)
        return Response(
            xl_bytes,
            mimetype=mimetype,
            headers={"Content-Disposition": f"attachment; filename=ato-evidence-{advisory_id}.{ext}"},
        )

    # ── NQE Translator ─────────────────────────────────────────────────────

    @bp.route("/nqe-translator", methods=["GET"])
    def nqe_translator_page():
        """Render the NQE query translator UI."""
        return render_template("network/nqe_translator.html", page_title="NQE Translator")

    @bp.route("/api/nqe/translate", methods=["POST"])
    def api_nqe_translate():
        """Translate plain-English text to NQL.

        Request: {"text": str, "context": dict (optional advisory context)}
        Response: {"nql": str, "confidence": float, "source": str}
        """
        data = request.get_json(force=True, silent=True) or {}
        text = (data.get("text") or "").strip()
        if not text:
            return jsonify({"error": "text is required"}), 400

        context = data.get("context") or {}

        try:
            from tools.network.nql_translator import nl_to_nql
            nql = nl_to_nql(text, context=context or None)
        except Exception as exc:
            logger.exception("NQE translate error")
            return jsonify({"error": str(exc)}), 500

        # Confidence heuristic: deterministic context path → high; LLM path → medium
        if context and any(context.get(k) for k in ("vendor", "affected_models", "affected_versions")):
            confidence = 0.92
            source = "deterministic"
        elif nql and nql.startswith("foreach"):
            confidence = 0.70
            source = "llm_translation"
        else:
            confidence = 0.50
            source = "fallback"

        # Audit log
        try:
            from tools.db.storage import get_canvas_connection
            conn = get_canvas_connection("NC_STORAGE_BACKEND")
            conn.execute(
                "INSERT INTO nc_nqe_audit_log (action, nql_query, user_confirmed, created_at) "
                "VALUES (%s, %s, %s, NOW())",
                ("translate", nql, False),
            )
            conn.commit()
            conn.close()
        except Exception:
            pass

        return jsonify({"nql": nql, "confidence": confidence, "source": source})

    @bp.route("/api/nqe/explain", methods=["POST"])
    def api_nqe_explain():
        """Return a plain-English explanation of an NQL query.

        Request: {"nql": str}
        Response: {"explanation": str}
        """
        data = request.get_json(force=True, silent=True) or {}
        nql = (data.get("nql") or "").strip()
        if not nql:
            return jsonify({"error": "nql is required"}), 400

        try:
            from tools.llm.router import LLMRouter
            from tools.llm.provider import LLMRequest

            prompt = (
                "Explain this NQL (Network Query Language) query in plain English "
                "for a non-technical network administrator. Be concise (2-3 sentences).\n\n"
                f"NQL:\n{nql}"
            )
            router = LLMRouter()
            req = LLMRequest(messages=[{"role": "user", "content": prompt}], max_tokens=150, temperature=0.2)
            resp = router.invoke("nql_explain", req)
            explanation = (resp.content or "").strip()
        except ImportError:
            explanation = _nql_heuristic_explain(nql)
        except Exception:
            explanation = _nql_heuristic_explain(nql)

        return jsonify({"explanation": explanation})

    @bp.route("/api/nqe/run", methods=["POST"])
    def api_nqe_run():
        """Execute an NQL query and return results.

        Requires explicit user confirmation (call audit-log endpoint first
        with user_confirmed=true — enforced by the UI transparency gate).

        Request: {"nql": str, "network_id": str|null}
        Response: {"rows": list, "columns": list, "total": int, "source": str}
        """
        data = request.get_json(force=True, silent=True) or {}
        nql = (data.get("nql") or "").strip()
        network_id = data.get("network_id") or None

        if not nql:
            return jsonify({"error": "nql is required"}), 400

        try:
            from tools.network.nqe_client import FallbackNQEClient

            client = FallbackNQEClient()
            result = client.run_query(nql, network_id=network_id)
            rows = result.get("rows", [])
            columns = list(rows[0].keys()) if rows else []

            # Audit execution
            try:
                from tools.db.storage import get_canvas_connection
                conn = get_canvas_connection("NC_STORAGE_BACKEND")
                conn.execute(
                    "INSERT INTO nc_nqe_audit_log (action, nql_query, user_confirmed, row_count, created_at) "
                    "VALUES (%s, %s, %s, %s, NOW())",
                    ("run", nql, True, len(rows)),
                )
                conn.commit()
                conn.close()
            except Exception:
                pass

            return jsonify({
                "rows": rows[:500],
                "columns": columns,
                "total": len(rows),
                "source": result.get("source", "local"),
            })
        except Exception as exc:
            logger.exception("NQE run error: nql=%r", nql)
            return jsonify({"error": str(exc)}), 500

    @bp.route("/api/nqe/collections", methods=["GET"])
    def api_nqe_collections():
        """Return the list of supported NQE collection paths.

        Response: {"collections": [{"path": str, "description": str}]}
        """
        collections = [
            {"path": "network.devices",          "description": "All network devices (hostname, OS, vendor, platform)"},
            {"path": "network.interfaces",        "description": "Device interfaces with status and counters"},
            {"path": "network.bgp_sessions",      "description": "BGP peering sessions and their state"},
            {"path": "network.acls",              "description": "Access control lists and firewall rules"},
            {"path": "network.paths",             "description": "End-to-end forwarding paths"},
            {"path": "network.os_versions",       "description": "OS version inventory across all devices"},
            {"path": "network.links",             "description": "Physical and logical links between nodes"},
            {"path": "network.vlans",             "description": "VLAN definitions and membership"},
            {"path": "network.prefixes",          "description": "IP prefix / subnet inventory"},
            {"path": "network.ospf.neighbors",    "description": "OSPF adjacency table"},
            {"path": "network.isis.adjacencies",  "description": "IS-IS adjacency table"},
            {"path": "network.mpls.lsps",         "description": "MPLS label-switched paths"},
        ]
        return jsonify({"collections": collections})

    @bp.route("/api/nqe/audit-log", methods=["POST"])
    def api_nqe_audit_log():
        """Append a transparency-gate audit event.

        Request: {"action": str, "nql": str, "user_confirmed": bool}
        Response: {"id": str, "recorded": true}
        """
        data = request.get_json(force=True, silent=True) or {}
        action = (data.get("action") or "").strip()
        nql = (data.get("nql") or "").strip()
        user_confirmed = bool(data.get("user_confirmed", False))

        if not action:
            return jsonify({"error": "action is required"}), 400

        row_id = None
        try:
            from tools.db.storage import get_canvas_connection
            conn = get_canvas_connection("NC_STORAGE_BACKEND")
            cur = conn.execute(
                "INSERT INTO nc_nqe_audit_log (action, nql_query, user_confirmed, created_at) "
                "VALUES (%s, %s, %s, NOW()) RETURNING id",
                (action, nql, user_confirmed),
            )
            row = cur.fetchone()
            row_id = str(row[0] if isinstance(row, (list, tuple)) else row.get("id", "")) if row else None
            conn.commit()
            conn.close()
        except Exception as exc:
            logger.warning("NQE audit-log insert failed: %s", exc)

        return jsonify({"id": row_id or "n/a", "recorded": True})

    @bp.route("/api/nqe/cross-validate", methods=["POST"])
    def api_nqe_cross_validate():
        """Dual-query cross-validation: translate using two strategies and compare.

        Strategy A uses structured context (deterministic); strategy B is context-free
        (LLM/fallback). A high divergence score means the two strategies disagree and
        a human must review and approve before execution.

        Request:  {"text": str, "context": dict (optional)}
        Response: {
            "nql_primary":      str,   # strategy A result
            "nql_secondary":    str,   # strategy B result
            "divergence_score": float, # 0.0 identical → 1.0 completely different
            "require_hitl":     bool,  # True when divergence_score >= 0.6
            "message":          str
        }
        """
        data = request.get_json(force=True, silent=True) or {}
        text = (data.get("text") or "").strip()
        if not text:
            return jsonify({"error": "text is required"}), 400

        context = data.get("context") or {}

        try:
            from tools.network.nql_translator import nl_to_nql
            nql_primary = nl_to_nql(text, context=context or None)
            nql_secondary = nl_to_nql(text)  # context-free → always LLM/fallback
        except Exception as exc:
            logger.exception("NQE cross-validate translation error")
            return jsonify({"error": str(exc)}), 500

        divergence = _nql_divergence_score(nql_primary, nql_secondary)
        require_hitl = divergence >= 0.6

        if require_hitl:
            message = (
                f"The two translation strategies produced divergent queries "
                f"(divergence {divergence:.0%}). Human approval is required before execution."
            )
        elif divergence >= 0.3:
            message = (
                f"Minor divergence detected ({divergence:.0%}). "
                "Review both queries before running."
            )
        else:
            message = "Both translation strategies agree. Safe to proceed."

        return jsonify({
            "nql_primary": nql_primary,
            "nql_secondary": nql_secondary,
            "divergence_score": divergence,
            "require_hitl": require_hitl,
            "message": message,
        })

    @bp.route("/api/nqe/hitl-approve", methods=["POST"])
    def api_nqe_hitl_approve():
        """Record a HITL approval for a cross-validated NQE query pair.

        Request:  {"nql_primary": str, "nql_secondary": str, "approved_by": str, "notes": str}
        Response: {"approved": true, "approved_by": str, "recorded": true}
        """
        data = request.get_json(force=True, silent=True) or {}
        nql_primary = (data.get("nql_primary") or "").strip()
        nql_secondary = (data.get("nql_secondary") or "").strip()
        approved_by = (data.get("approved_by") or "").strip()
        notes = (data.get("notes") or "").strip()

        if not approved_by:
            return jsonify({"error": "approved_by is required"}), 400

        import json as _json
        audit_payload = _json.dumps({
            "nql_primary": nql_primary,
            "nql_secondary": nql_secondary,
            "notes": notes,
        })

        try:
            from tools.db.storage import get_canvas_connection
            conn = get_canvas_connection("NC_STORAGE_BACKEND")
            conn.execute(
                "INSERT INTO nc_nqe_audit_log (action, nql_query, user_confirmed, created_at) "
                "VALUES (%s, %s, %s, NOW())",
                ("hitl_approve", audit_payload, True),
            )
            conn.commit()
            conn.close()
        except Exception as exc:
            logger.warning("NQE hitl-approve audit insert failed: %s", exc)

        return jsonify({"approved": True, "approved_by": approved_by, "recorded": True})

    # ── NQE helpers ────────────────────────────────────────────────────────

    def _nql_divergence_score(nql_a: str, nql_b: str) -> float:
        """Return divergence in [0.0, 1.0] between two NQL strings (Jaccard distance).

        0.0 = identical, 1.0 = completely different.
        Collection mismatch is boosted to ≥ 0.8 since different primary collections
        almost certainly query different facts.
        """
        import re as _re

        def _collection(nql):
            m = _re.search(r"\bin\s+(network\.\S+)", nql, _re.I)
            return m.group(1).lower() if m else ""

        col_a = _collection(nql_a)
        col_b = _collection(nql_b)

        tok_a = set(_re.findall(r"[\w.]+", nql_a.lower()))
        tok_b = set(_re.findall(r"[\w.]+", nql_b.lower()))

        if not tok_a or not tok_b:
            return 1.0

        jaccard = len(tok_a & tok_b) / len(tok_a | tok_b)
        divergence = 1.0 - jaccard

        if col_a and col_b and col_a != col_b:
            divergence = max(divergence, 0.8)

        return round(divergence, 3)

    def _nql_heuristic_explain(nql: str) -> str:
        """Generate a simple heuristic explanation from NQL structure."""
        import re as _re
        nql = nql.strip()
        m = _re.search(r"\bin\s+(network\.\S+)", nql, _re.I)
        collection = m.group(1) if m else "network"
        where_m = _re.search(r"\bwhere\s+(.+?)(?:\bselect\b|$)", nql, _re.I | _re.DOTALL)
        where_clause = where_m.group(1).strip() if where_m else ""
        base = f"Queries the '{collection}' collection"
        if where_clause:
            base += f" filtered by: {where_clause[:120]}"
        return base + "."


    # ── Compliance Audit landing ───────────────────────────────────────────
    @bp.route("/compliance-audit")
    def nc_compliance_audit_index():
        """List topologies so the user can pick one to audit."""
        conn = get_connection()
        topos = [
            dict(r) if hasattr(r, "keys") else {"id": r[0], "name": r[1]}
            for r in conn.execute(
                "SELECT id, name FROM topologies ORDER BY name"
            ).fetchall()
        ]
        conn.close()
        return render_template("network/compliance_audit_index.html", topologies=topos)

    # ── Config Review ─────────────────────────────────────────────────────
    @bp.route("/config-review")
    def nc_config_review():
        """Configuration Review Assistant — upload and AI-review device configs."""
        # questions_by_role is NOT optional: the template does
        # `{{ questions_by_role | tojson }}`, and a Jinja Undefined is not JSON
        # serializable — omitting it raised TypeError and 500'd the whole page,
        # not just the question list. selectRole() also indexes it immediately.
        from tools.network.constants import CONFIG_REVIEW_QUESTIONS, CONFIG_REVIEW_ROLES
        return render_template(
            "network/config_review.html",
            roles=CONFIG_REVIEW_ROLES,
            questions_by_role=CONFIG_REVIEW_QUESTIONS,
        )

    # ── Config Review API ─────────────────────────────────────────────────
    # tools/network/config_review.py has carried the whole engine (roles,
    # questions, guided prompts, LLM prompt build, response parsing, exports)
    # and config_review.html has called these endpoints — but they were never
    # wired, so every button on the page 404'd. This is the missing layer.

    # Categories in a parsed review result whose entries are findings. Everything
    # else in the result (explanation, sample_template, topology_graph, vendor…)
    # is review-level, not a finding.
    _REVIEW_FINDING_CATEGORIES = ("security_compliance", "optimization", "remediation")

    _NC_IQE_COLLECTIONS = [
        "network.config_reviews",
        "network.config_review_findings",
        "network.nodes",
        "network.edges",
        "network.topologies",
    ]

    def _safe_json(raw, default):
        """Parse a JSON column without letting one bad row 500 the endpoint."""
        if isinstance(raw, (dict, list)):
            return raw
        try:
            return json.loads(raw) if raw else default
        except (TypeError, ValueError):
            return default

    def _flatten_review_findings(result: dict) -> list[dict]:
        """Flatten the category-keyed review result into finding rows.

        Severity defaults to 'info' for categories that don't carry one
        (optimization/remediation are advisory), and 'recommendation' is folded
        into remediation so every row has the same shape as the security ones.
        """
        out: list[dict] = []
        for category in _REVIEW_FINDING_CATEGORIES:
            for item in (result or {}).get(category) or []:
                if not isinstance(item, dict):
                    continue
                out.append({
                    "category": category,
                    "severity": item.get("severity") or "info",
                    "title": item.get("title") or "",
                    "detail": item.get("detail") or "",
                    "remediation": item.get("remediation") or item.get("recommendation") or "",
                    "sample_config_snippet": item.get("sample_config_snippet") or "",
                    "references": item.get("references") or [],
                })
        return out

    def _load_config_review(review_id: str):
        """(review, findings) for a persisted review, or (None, []) if absent."""
        conn = get_connection()
        _ph = sql_placeholder(conn)
        try:
            row = conn.execute(
                f"SELECT * FROM nc_config_reviews WHERE id={_ph}", (review_id,)
            ).fetchone()
            if not row:
                return None, []
            review = _row_to_dict(row)
            findings = [
                _row_to_dict(r) for r in conn.execute(
                    f"SELECT * FROM nc_config_review_findings WHERE review_id={_ph} "
                    "ORDER BY created_at",
                    (review_id,),
                ).fetchall()
            ]
        finally:
            conn.close()
        # The exporters read `references` and treat findings as plain dicts.
        for f in findings:
            f["references"] = _safe_json(f.get("references_json"), [])
        return review, findings

    @bp.route("/api/config-review", methods=["POST"])
    def nc_api_config_review_create():
        """Start a review: detect vendor, return the role's questions + prompts."""
        from tools.network.config_review import (
            _detect_vendor,
            _extract_hostname,
            compute_config_hash,
            generate_guided_prompts,
            get_questions,
            get_roles,
        )

        data = request.get_json(force=True, silent=True) or {}
        config_text = (data.get("config_text") or "").strip()
        role_key = (data.get("role") or "network_engineer").strip()

        if not config_text:
            return jsonify({"error": "config_text is required"}), 400
        if role_key not in get_roles():
            # 400 not 404: the request is malformed, not the resource missing.
            return jsonify({"error": f"Unknown role: {role_key}"}), 400

        vendor = _detect_vendor(config_text)
        hostname = _extract_hostname(config_text, vendor)
        review_id = str(_uuid.uuid4())
        now = _now()

        conn = get_connection()
        try:
            _ph = sql_placeholder(conn)
            conn.execute(
                "INSERT INTO nc_config_reviews (id, title, vendor, role_key, answers_json, "
                f"config_text_hash, status, result_json, created_at, updated_at) "
                f"VALUES ({_ph},{_ph},{_ph},{_ph},{_ph},{_ph},{_ph},{_ph},{_ph},{_ph})",
                (
                    review_id, hostname or f"Config review {now[:10]}", vendor, role_key,
                    json.dumps({}), compute_config_hash(config_text), "pending",
                    json.dumps({}), now, now,
                ),
            )
            conn.commit()
        finally:
            conn.close()

        _audit("CREATE", "config_review", review_id, f"{vendor} / {role_key}")
        return jsonify({
            "id": review_id,
            "vendor": vendor,
            "hostname": hostname,
            "role": role_key,
            "questions": get_questions(role_key),
            "prompts": generate_guided_prompts(role_key, vendor, hostname),
        })

    @bp.route("/api/config-review/<review_id>/analyze", methods=["POST"])
    def nc_api_config_review_analyze(review_id):
        """Run the review. The config text is NOT persisted, so the caller
        re-supplies it; only its hash is stored (see the init_db comment)."""
        from tools.network.config_review import (
            build_llm_prompt,
            parse_review_response,
        )

        data = request.get_json(force=True, silent=True) or {}
        answers = data.get("answers") or {}
        prompt_title = data.get("prompt_title") or ""
        config_text = data.get("config_text") or ""

        conn = get_connection()
        _ph = sql_placeholder(conn)
        try:
            row = conn.execute(
                f"SELECT id, vendor, role_key, title FROM nc_config_reviews WHERE id={_ph}",
                (review_id,),
            ).fetchone()
            if not row:
                conn.close()
                return jsonify({"error": "Review not found"}), 404
            review = _row_to_dict(row)
        except Exception as exc:
            conn.close()
            return jsonify({"error": str(exc)}), 500

        vendor = review.get("vendor") or "unknown"
        prompt = build_llm_prompt(
            review.get("role_key") or "network_engineer",
            config_text, vendor, answers,
            hostname=review.get("title") or "",
            selected_prompt_title=prompt_title,
        )

        raw = ""
        try:
            from tools.llm.router import LLMRouter

            resp = LLMRouter().invoke("config_review", prompt)
            raw = getattr(resp, "content", "") or ""
        except Exception as exc:
            # parse_review_response has a deterministic fallback, so an LLM
            # outage degrades to a usable review instead of a 500.
            logger.warning("config review: LLM unavailable (%s) — using fallback", exc)

        result = parse_review_response(raw, vendor)
        findings = _flatten_review_findings(result)
        now = _now()

        try:
            conn.execute(
                f"UPDATE nc_config_reviews SET status={_ph}, result_json={_ph}, "
                f"answers_json={_ph}, updated_at={_ph} WHERE id={_ph}",
                ("complete", json.dumps(result), json.dumps(answers), now, review_id),
            )
            conn.execute(
                f"DELETE FROM nc_config_review_findings WHERE review_id={_ph}", (review_id,)
            )
            for f in findings:
                conn.execute(
                    "INSERT INTO nc_config_review_findings (id, review_id, category, severity, "
                    f"title, detail, remediation, sample_config_snippet, references_json, created_at) "
                    f"VALUES ({_ph},{_ph},{_ph},{_ph},{_ph},{_ph},{_ph},{_ph},{_ph},{_ph})",
                    (
                        str(_uuid.uuid4()), review_id, f.get("category", ""),
                        f.get("severity", ""), f.get("title", ""), f.get("detail", ""),
                        f.get("remediation", ""), f.get("sample_config_snippet", ""),
                        json.dumps(f.get("references", [])), now,
                    ),
                )
            conn.commit()
        finally:
            conn.close()

        _audit("ANALYZE", "config_review", review_id, f"{len(findings)} finding(s)")
        return jsonify({"id": review_id, "status": "complete", "result": result,
                        "findings": findings})

    @bp.route("/api/config-review/<review_id>", methods=["GET"])
    def nc_api_config_review_get(review_id):
        """Persisted review + its findings."""
        conn = get_connection()
        _ph = sql_placeholder(conn)
        try:
            row = conn.execute(
                f"SELECT * FROM nc_config_reviews WHERE id={_ph}", (review_id,)
            ).fetchone()
            if not row:
                return jsonify({"error": "Review not found"}), 404
            review = _row_to_dict(row)
            review["result"] = _safe_json(review.get("result_json"), {})
            review["answers"] = _safe_json(review.get("answers_json"), {})
            findings = [
                _row_to_dict(r) for r in conn.execute(
                    f"SELECT * FROM nc_config_review_findings WHERE review_id={_ph} "
                    "ORDER BY created_at",
                    (review_id,),
                ).fetchall()
            ]
        finally:
            conn.close()
        return jsonify({"review": review, "findings": findings})

    @bp.route("/api/config-review/<review_id>/progress", methods=["GET"])
    def nc_api_config_review_progress(review_id):
        """Poll target for the page's progress bar during analyze."""
        conn = get_connection()
        _ph = sql_placeholder(conn)
        try:
            row = conn.execute(
                f"SELECT status, updated_at FROM nc_config_reviews WHERE id={_ph}",
                (review_id,),
            ).fetchone()
        finally:
            conn.close()
        if not row:
            return jsonify({"error": "Review not found"}), 404
        d = _row_to_dict(row)
        return jsonify({"id": review_id, "status": d.get("status"),
                        "updated_at": d.get("updated_at")})

    @bp.route("/api/config-review/<review_id>/export-config", methods=["POST"])
    def nc_api_config_review_export_config(review_id):
        """Starter config built from the findings' remediation snippets."""
        from tools.network.config_review import generate_export_config

        review, findings = _load_config_review(review_id)
        if review is None:
            return jsonify({"error": "Review not found"}), 404
        config = generate_export_config(review.get("vendor") or "unknown", findings)
        return jsonify({"id": review_id, "config": config})

    @bp.route("/api/config-review/<review_id>/export-topology", methods=["POST"])
    def nc_api_config_review_export_topology(review_id):
        """Topology graph synthesized from the findings."""
        from tools.network.config_review import generate_export_topology

        review, findings = _load_config_review(review_id)
        if review is None:
            return jsonify({"error": "Review not found"}), 404
        graph = generate_export_topology(findings, review.get("vendor") or "unknown")
        return jsonify({"id": review_id, "graph": graph})

    @bp.route("/api/iqe-query", methods=["POST"])
    def nc_api_iqe_query():
        """Natural-language IQE over the Network Canvas collections."""
        import importlib

        data = request.get_json(force=True, silent=True) or {}
        question = (data.get("question") or "").strip()
        execute = data.get("execute", True)
        if not question:
            return jsonify({"error": "question is required"}), 400

        try:
            importlib.import_module("tools.iqe.adapters.ndc")
        except Exception:
            pass

        try:
            from tools.iqe.nl_to_iqe import nl_to_iqe

            out = nl_to_iqe(question, collections=_NC_IQE_COLLECTIONS)
            # nl_to_iqe returns {iqe, explanation}; tolerate a bare string too.
            if isinstance(out, dict):
                iqe_str, explanation = out.get("iqe", ""), out.get("explanation", "")
            else:
                iqe_str, explanation = str(out), ""

            results: list = []
            if execute and iqe_str:
                from tools.iqe.executor import execute_query
                from tools.iqe.parser import parse as _parse

                results = execute_query(_parse(iqe_str))
            return jsonify({"iqe": iqe_str, "explanation": explanation,
                            "results": results, "row_count": len(results)})
        except Exception as exc:
            logger.warning("nc iqe-query failed: %s", exc)
            return jsonify({"error": str(exc)}), 500

    # ── Diagram Analysis ──────────────────────────────────────────────────
    @bp.route("/diagram-analysis")
    def nc_diagram_analysis():
        """Network Diagram Analysis — upload PNG/PDF/draw.io for AI analysis."""
        from tools.network.constants import DIAGRAM_ANALYSIS_INDUSTRIES
        return render_template("network/diagram_analysis.html", industries=DIAGRAM_ANALYSIS_INDUSTRIES)

    # ── Migration Phases landing (alias → hub) ────────────────────────────
    @bp.route("/migration-phases")
    def nc_migration_phases_hub():
        """Redirect bare /migration-phases to the Migration Hub."""
        from flask import redirect
        return redirect("/network/migration-hub")

    # ── PVM Predictive Vulnerability Management routes ────────────────────


    # ── PNA Predictive Network Analytics routes ───────────────────────────


    # ══════════════════════════════════════════════════════════════════════
    # Federal Network Peering — IP Address Space & Routing Policy (Step 3)
    # ══════════════════════════════════════════════════════════════════════

    @bp.route("/api/ip-space-definitions", methods=["POST"])
    @nc_login_required
    def nc_api_create_ip_space_definition():
        from tools.network.ip_address_space import create_ip_space_definition
        data = request.get_json(force=True) or {}
        required = ("initiating_party_name", "responding_party_name")
        missing = [f for f in required if not data.get(f)]
        if missing:
            return jsonify({"error": f"Missing required fields: {missing}"}), 400
        conn = get_connection()
        try:
            result = create_ip_space_definition(
                conn,
                initiating_party_name=data["initiating_party_name"],
                responding_party_name=data["responding_party_name"],
                initiating_party_org=data.get("initiating_party_org", ""),
                responding_party_org=data.get("responding_party_org", ""),
                peering_request_id=data.get("peering_request_id"),
                asn_exchange_id=data.get("asn_exchange_id"),
                initial_prefixes=data.get("initial_prefixes"),
            )
            _audit("CREATE", "ip_space_definition", result["definition_id"], conn)
            return jsonify({"ok": True, "definition": result}), 201
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500
        finally:
            conn.close()

    @bp.route("/api/ip-space-definitions", methods=["GET"])
    @nc_login_required
    def nc_api_list_ip_space_definitions():
        from tools.network.ip_address_space import list_ip_space_definitions
        conn = get_connection()
        try:
            return jsonify(list_ip_space_definitions(
                conn,
                workflow_id=request.args.get("workflow_id"),
                status=request.args.get("status"),
                peering_request_id=request.args.get("peering_request_id"),
                limit=min(int(request.args.get("limit", 100)), 500),
            ))
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500
        finally:
            conn.close()

    @bp.route("/api/ip-space-definitions/<did>", methods=["GET"])
    @nc_login_required
    def nc_api_get_ip_space_definition(did):
        from tools.network.ip_address_space import get_ip_space_definition
        conn = get_connection()
        try:
            rec = get_ip_space_definition(conn, did)
            if not rec:
                return jsonify({"error": f"Definition {did} not found"}), 404
            return jsonify(rec)
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500
        finally:
            conn.close()

    @bp.route("/api/ip-space-definitions/<did>/prefixes", methods=["POST"])
    @nc_login_required
    def nc_api_add_ip_space_prefix(did):
        from tools.network.ip_address_space import add_prefix
        data = request.get_json(force=True) or {}
        if not data.get("prefix"):
            return jsonify({"error": "prefix is required"}), 400
        conn = get_connection()
        try:
            result = add_prefix(
                conn, did,
                prefix=data["prefix"],
                party_role=data.get("party_role", "initiating"),
                prefix_type=data.get("prefix_type", "aggregate"),
                description=data.get("description", ""),
                is_customer_prefix=bool(data.get("is_customer_prefix", False)),
            )
            _audit("ADD_PREFIX", "ip_space_definition", did, conn)
            return jsonify({"ok": True, "definition": result})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500
        finally:
            conn.close()

    @bp.route("/api/ip-space-definitions/<did>/prefixes/<int:idx>", methods=["DELETE"])
    @nc_login_required
    def nc_api_remove_ip_space_prefix(did, idx):
        from tools.network.ip_address_space import remove_prefix
        conn = get_connection()
        try:
            result = remove_prefix(conn, did, idx)
            _audit("REMOVE_PREFIX", "ip_space_definition", did, conn)
            return jsonify({"ok": True, "definition": result})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500
        finally:
            conn.close()

    @bp.route("/api/ip-space-definitions/<did>/routing-policy", methods=["PUT"])
    @nc_login_required
    def nc_api_set_ip_space_routing_policy(did):
        from tools.network.ip_address_space import set_routing_policy
        data = request.get_json(force=True) or {}
        conn = get_connection()
        try:
            result = set_routing_policy(
                conn, did,
                max_prefixes_initiating=data.get("max_prefixes_initiating"),
                max_prefixes_responding=data.get("max_prefixes_responding"),
                min_prefix_length_v4=data.get("min_prefix_length_v4"),
                max_prefix_length_v4=data.get("max_prefix_length_v4"),
                min_prefix_length_v6=data.get("min_prefix_length_v6"),
                max_prefix_length_v6=data.get("max_prefix_length_v6"),
                accepted_communities=data.get("accepted_communities"),
                rejected_communities=data.get("rejected_communities"),
                local_preference=data.get("local_preference"),
                med=data.get("med"),
                no_export=data.get("no_export"),
                prefix_filter_action=data.get("prefix_filter_action"),
                notes=data.get("notes"),
            )
            _audit("SET_ROUTING_POLICY", "ip_space_definition", did, conn)
            return jsonify({"ok": True, "definition": result})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500
        finally:
            conn.close()

    @bp.route("/api/ip-space-definitions/<did>/submit", methods=["POST"])
    @nc_login_required
    def nc_api_submit_ip_space_definition(did):
        from tools.network.ip_address_space import submit_definition
        conn = get_connection()
        try:
            result = submit_definition(conn, did)
            _audit("SUBMIT", "ip_space_definition", did, conn)
            return jsonify({"ok": True, "definition": result})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500
        finally:
            conn.close()

    @bp.route("/api/ip-space-definitions/<did>/acknowledge", methods=["POST"])
    @nc_login_required
    def nc_api_acknowledge_ip_space_definition(did):
        from tools.network.ip_address_space import acknowledge_definition
        conn = get_connection()
        try:
            result = acknowledge_definition(conn, did)
            _audit("ACKNOWLEDGE", "ip_space_definition", did, conn)
            return jsonify({"ok": True, "definition": result})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500
        finally:
            conn.close()

    @bp.route("/api/ip-space-definitions/<did>/approve", methods=["POST"])
    @nc_login_required
    def nc_api_approve_ip_space_definition(did):
        from tools.network.ip_address_space import approve_definition
        data = request.get_json(force=True) or {}
        party_role = data.get("party_role")
        if not party_role:
            return jsonify({"error": "party_role is required"}), 400
        conn = get_connection()
        try:
            result = approve_definition(conn, did,
                                        party_role=party_role,
                                        notes=data.get("notes", ""))
            _audit("APPROVE", "ip_space_definition", did, conn)
            return jsonify({"ok": True, "definition": result})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500
        finally:
            conn.close()

    @bp.route("/api/ip-space-definitions/<did>/reject", methods=["POST"])
    @nc_login_required
    def nc_api_reject_ip_space_definition(did):
        from tools.network.ip_address_space import reject_definition
        data = request.get_json(force=True) or {}
        conn = get_connection()
        try:
            result = reject_definition(conn, did, reason=data.get("reason", ""))
            _audit("REJECT", "ip_space_definition", did, conn)
            return jsonify({"ok": True, "definition": result})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500
        finally:
            conn.close()

    @bp.route("/api/ip-space-definitions/<did>/document", methods=["GET"])
    @nc_login_required
    def nc_api_ip_space_definition_document(did):
        from tools.network.ip_address_space import generate_definition_document
        conn = get_connection()
        try:
            doc = generate_definition_document(conn, did)
            fmt = request.args.get("format", "text")
            if fmt == "json":
                return jsonify({"ok": True, "definition_id": did, "document": doc})
            return doc, 200, {"Content-Type": "text/plain; charset=utf-8"}
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 404
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500
        finally:
            conn.close()

    # ── Done ───────────────────────────────────────────────────────────────
